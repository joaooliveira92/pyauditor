"""PROTOTYPE do `consolidate` (2.1) — NÃO é produção. (ticket 05)

Gera `reports/relatorio_<comp>_consolidado_PROTOTYPE.xlsx` a partir dos
sumários por órgão já medidos (`roms/MinC/<comp>/INMS-*.json` +
`roms/MTur/<comp>/INMS-*.json`), aplicando as decisões dos tickets 01/02/04:

- 5 abas do núcleo financeiro: CAPA_E_CONTROLE, INMS_BASE, GLOSAS,
  CALCULO_PAGAMENTO, SERVICOS_POR_ORGAO.
- INMS_BASE: pooling Σnum/Σden por indicador com linha "Consolidado"
  (penalty = soma por órgão); por-ativo 1.4/1.5/1.14 ficam **uma linha por
  órgão** (sem consolidar, sem aportar glosa).
- GLOSAS: uma linha por (indicador × órgão) com %Ajuste = points × 0.001
  derivado da engine, + resumo agregado com teto único 30%.
- CALCULO_PAGAMENTO: rateio 0.5/0.5 provisório; bruto = mensal × rateio;
  glosa na coluna consolidada; recomendado = max(0, bruto − glosa − outros).
- Colunas de decisão (Justificativa / Decisão Fiscal / Observação) VAGAS
  para o fiscal preencher — merge na re-rodata (ticket 04 Q3).

Uso:
    python .scratch/multi-org-pipeline/prototype/consolidated_mock.py
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Border, Font, Side
from openpyxl.worksheet.worksheet import Worksheet

from pyauditor.excel._style import (
    BODY_FONT,
    BOTTOM_BORDER,
    HEADER_FILL,
    HEADER_FONT,
    LABEL_FONT,
    LEFT_ALIGN,
    TITLE_FONT,
)

REPO = Path(__file__).resolve().parents[3]
COMP = "2026-06"
OUT = REPO / "reports" / f"relatorio_{COMP}_consolidado_PROTOTYPE.xlsx"

# docs/styleguide.md — number formats/top-border-on-totals aren't in
# `_style.py` yet (production `report.py`/`capa.py` don't apply them either);
# kept local to this throwaway script rather than reaching into production.
CURRENCY_FMT = "R$#,##0.00;R$#,##0.00;-"
# Excel's literal "%" symbol auto-multiplies the cell value by 100 on
# display. `result_pct`/`%Ajuste` are already stored in percent-space (e.g.
# 95.5 meaning "95.5%"), so they need the symbol escaped (no auto-scale);
# only true 0-1 fractions (rateio) want the real, auto-scaling "%" format.
PERCENT_FMT_SCALED = '0.00"%";0.00"%";-'
PERCENT_FMT_FRACTION = "0.00%;0.00%;-"
TOP_BORDER = Border(top=Side(style="thin", color="1F2937"))

# Valor mensal de exemplo — a inspiração usa 481.534,80; aqui só ilustra a
# forma; o valor real vem da CAPA preenchida pelo fiscal (tickets 02/04).
VALOR_BASE = 481534.80
LIMITE_PCT = 30.0  # teto único sobre o agregado (% do valor mensal)

# As 9 serviços contratuais (ticket 01 — SERVICOS_POR_ORGAO da inspiração).
SERVICOS = (
    ("Central de Serviços e Monitoramento", "Sim", "Sim", "Sim"),
    ("Gerenciamento Técnico das Operações e Projetos", "Sim", "Sim", "Sim"),
    ("Banco de Dados", "Sim", "Sim", "Sim"),
    ("Aplicações, Virtualização e Computação em Nuvem", "Sim", "Sim", "Sim"),
    ("Serviços Corporativos", "Sim", "Sim", "Sim"),
    ("Armazenamento e Backup", "Sim", "Sim", "Sim"),
    ("Redes", "Sim", "Sim", "Sim"),
    ("Segurança da Informação", "Sim", "Sim", "Sim"),
    ("DevOps", "Sim", "Sim", "Sim"),
)

# Disponibilidade por-ativo — não consolida (decisões 02/04).
PER_ATIVO = {"INMS 1.4", "INMS 1.5", "INMS 1.14"}

INMS_BASE_COLUNAS = (
    "Competência", "Item contratual", "Serviço", "Grupo operacional",
    "Código INMS", "Descrição", "Órgão", "Meta mínima ou máxima",
    "Sentido da meta", "Numerador", "Denominador", "Resultado calculado",
    "Unidade", "Aplicabilidade", "Resultado esperado", "Conformidade",
    "Diferença para a meta", "Ocorrência de glosa", "% de glosa",
    "Valor-base", "Valor da glosa", "Justificativa", "Referência da evidência",
    "Número SEI", "Responsável pela evidência", "Observação do fiscal",
)

# Inspiração (tabela 01): uma linha por indicador × órgão, decisões em aberto.
GLOSAS_COLUNAS = (
    "Competência", "Órgão", "Item Contratual", "Serviço", "Indicador",
    "Resultado", "Meta", "Faixa de Descumprimento", "Percentual de Ajuste",
    "Valor Base", "Valor Glosa", "Reincidência", "Justificativa",
    "Número da Ocorrência", "Decisão Fiscal", "Observação do Gestor",
)

CALCULO_COLUNAS = ("Componente", "MinC", "MTur", "Consolidado")

# Linhas da tabela de componentes de CALCULO_PAGAMENTO (espelha a inspiração).
CALCULO_LINHAS = (
    "Percentual de rateio",
    "Valor bruto (= mensal × rateio)",
    "Pontos de glosa",
    "Valor da glosa (= MIN(pontos×0,001, 30%)/100 × bruto)",
    "Outros ajustes",
    "Valor recomendado (= max(0, bruto − glosa − outros))",
)

CAPA_CAMPOS = (
    ("Número do contrato", "40/2022 - Ministério da Cultura"),
    ("Processo SEI", ""),
    ("Empresa contratada", ""),
    ("Órgão contratante atual", "Ministério da Cultura / Ministério do Turismo"),
    ("Competência", "Junho/2026"),
    ("Período da aferição", "01/06/2026 a 30/06/2026"),
    ("Valor Mensal Vigente", VALOR_BASE),
    ("Valor Global Anual", VALOR_BASE * 12),
)


def _clean(s: str) -> str:
    """Os YAML/JSONs são UTF‑8; o Excel do usuário abre em cp1252."""
    if not s:
        return s
    return s.encode("cp1252", errors="replace").decode("cp1252")


def _load(orgao: str) -> list[dict]:
    base = REPO / "roms" / orgao / COMP
    out = []
    for p in sorted(base.glob("INMS-*.json")):
        out.append(json.loads(p.read_text(encoding="utf-8")))
    return out


# ---------------------------------------------------------------------------
# Divergência sintética do MTur (PROTOTYPE)
# ---------------------------------------------------------------------------
# Os CSVs de input/MTur são cópias byte-a-byte dos de MinC (placeholders
# sintéticos — mapa, "Not yet specified": dados reais do MTur virão depois).
# Isso tornava o mock da consolidada um espelho degenerado: cada linha de
# GLOSAS duplicada, pooling == cada órgão, colunas de CALCULO idênticas.
#
# Para o mock exercitar de verdade a consolidação, aplicamos uma divergência
# determinística e DOCUMENTADA aqui — no script descartável, NUNCA nos CSVs:
# o MTur passa a ser o "órgão melhor desempenho", conforme em quase todos os
# indicadores (falha apenas no INMS 1.6, como o MinC). Assim:
#   - GLOSAS mostra linhas SÓ do MinC em vários indicadores e só do MTur no
#     1.6 → per-órgão de verdade;
#   - INMS_BASE mostra pooling (Σnum/Σden) onde os dois mediram, com
#     resultado consolidado ≠ de cada órgão;
#   - CALCULO_PAGAMENTO fica com colunas MinC/MTur distintas.
# Isso é MAQUIAGEM DE DEMO; o consolidate real lê os workbooks por órgão.
_MTUR_CONFORME: frozenset[str] = frozenset({
    "INMS 1.1", "INMS 1.2", "INMS 1.3", "INMS 1.4", "INMS 1.5",
    "INMS 1.7", "INMS 1.8", "INMS 1.9", "INMS 1.10", "INMS 1.11",
    "INMS 1.12", "INMS 1.13", "INMS 1.14",
})


def _divergir_mtur(summaries: list[dict]) -> list[dict]:
    for rec in summaries:
        if rec["contractual_id"] in _MTUR_CONFORME:
            rec["conforms"] = True
            rec["penalty_points"] = 0.0
            # Reflete um resultado dentro da meta (pooling usa num/den).
            if rec.get("numerator") is not None and rec.get("denominator"):
                rec["numerator"] = rec["denominator"]
                rec["result_pct"] = 100.0 if rec["target_operator"] == ">=" else 0.0
        # INMS 1.6 continua fora da meta (glosa compartilhada com o MinC).
    return summaries


def _new_sheet(wb: Workbook, name: str, columns: tuple[str, ...], width: int = 24) -> Worksheet:
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    for i, col in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=i, value=_clean(col))
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = LEFT_ALIGN
        ws.column_dimensions[cell.column_letter].width = width
    ws.freeze_panes = "A2"
    return ws


def _write(ws: Worksheet, row: int, values: tuple[object, ...]) -> None:
    for i, value in enumerate(values, start=1):
        cell = ws.cell(row=row, column=i, value=_clean(value) if isinstance(value, str) else value)
        cell.font = BODY_FONT
        cell.border = BOTTOM_BORDER


def _faixa(rec: dict) -> str:
    """Texto de 'Faixa de Descumprimento' para uma ocorrência de glosa.

    A ocorrência é decidida pela engine (`conforms`/`penalty_points` gravados
    no sumário) — não por re-derivar da meta aqui. Para por-ativo (1.4/1.5/
    1.14), o déficit é por ativo no engine, não global; o texto é ilustrativo.
    """
    r, meta, op = rec["result_pct"], rec["target_value"], rec["target_operator"]
    if op is None or meta is None:
        return ""
    if not rec["conforms"]:
        dif = (meta - r) if op == ">=" else (r - meta)
        if dif > 0:
            return f"Déficit de {dif:.2f}pp"
        return "Ocorrência sob detalhamento por-ativo"
    return ""


def _conforms(rpct: float, op: str | None, meta: float | None) -> bool:
    if op is None or meta is None:
        return False
    return (rpct >= meta) if op == ">=" else (rpct <= meta)


def _unidade(shape: str) -> str:
    if shape in ("ratio", "segmented_ratio", "precomputed_table"):
        return "%"
    if shape == "external_catalog_sum":
        return "pontos"
    return ""


def _write_inms_row(ws: Worksheet, row: int, rec: dict) -> None:
    op, meta = rec["target_operator"], rec["target_value"]
    conf = "Conforme" if rec["conforms"] else "Não conforme"
    dif = None
    if meta is not None:
        dif = round(meta - rec["result_pct"], 2) if op == ">=" else round(rec["result_pct"] - meta, 2)
    _write(ws, row, (
        COMP, "", "", "", rec["contractual_id"], rec["name"], rec["orgao"], meta,
        rec["target_operator"], rec["numerator"], rec["denominator"],
        round(rec["result_pct"], 2), _unidade(rec["shape"]), "", "", conf, dif,
        None, None, None, None, None, None, None, None, None,
    ))
    if _unidade(rec["shape"]) == "%":
        for col in (8, 12, 17):  # Meta, Resultado calculado, Diferença para a meta
            ws.cell(row=row, column=col).number_format = PERCENT_FMT_SCALED


# ---------------------------------------------------------------------------
# Abas
# ---------------------------------------------------------------------------


def build_capa(wb: Workbook) -> None:
    ws = wb.create_sheet("CAPA_E_CONTROLE", 0)
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Capa e controle — consolidado (PROTOTYPE)"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:B1")

    ws["A3"] = "Campo"
    ws["B3"] = "Valor"
    for cell in (ws["A3"], ws["B3"]):
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = LEFT_ALIGN

    for offset, (label, value) in enumerate(CAPA_CAMPOS):
        row = 4 + offset
        label_cell = ws.cell(row=row, column=1, value=_clean(label))
        label_cell.font = LABEL_FONT
        label_cell.alignment = LEFT_ALIGN
        label_cell.border = BOTTOM_BORDER

        value_cell = ws.cell(row=row, column=2, value=value)
        value_cell.font = BODY_FONT
        value_cell.border = BOTTOM_BORDER
        if label.startswith("Valor"):
            value_cell.number_format = CURRENCY_FMT

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 42


def build_servicos(wb: Workbook) -> None:
    ws = _new_sheet(
        wb, "SERVICOS_POR_ORGAO",
        ("Item", "Serviço", "Prestado ao MinC?", "Prestado ao MTur?",
         "Segregação Obrigatória?", "Critério de Rateio"),
    )
    for i, (nome, minc, mtur, seg) in enumerate(SERVICOS, start=2):
        _write(ws, i, (i - 1, _clean(nome), minc, mtur, seg, "Chamados, ativos ou valor definido"))


def build_inms_base(wb: Workbook, minc: list[dict], mtur: list[dict]) -> None:
    ws = _new_sheet(wb, "INMS_BASE", INMS_BASE_COLUNAS, width=20)
    row = 2

    by_orgao: dict[str, dict[str, dict]] = {}
    for rec in minc + mtur:
        by_orgao.setdefault(rec["contractual_id"], {})[rec["orgao"]] = rec

    for contractual_id in sorted(by_orgao):
        occ = by_orgao[contractual_id]

        # Por-ativo → uma linha por órgão, sem linha consolidada.
        if contractual_id in PER_ATIVO:
            for orgao in ("MinC", "MTur"):
                if orgao in occ:
                    _write_inms_row(ws, row, occ[orgao])
                    row += 1
            continue

        minc_rec, mtur_rec = occ.get("MinC"), occ.get("MTur")
        if minc_rec is not None and mtur_rec is not None and minc_rec["numerator"] is not None:
            numerador = minc_rec["numerator"] + mtur_rec["numerator"]
            denominador = minc_rec["denominator"] + mtur_rec["denominator"]
            rpct = (numerador / denominador * 100) if denominador else 0.0
            # Engine: denominador == 0 → nada a medir → nunca vira falha
            # (espelha `pyauditor.engine.strategies.ratio`).
            conforms = True if denominador == 0 else _conforms(rpct, minc_rec["target_operator"], minc_rec["target_value"])
            cons = dict(minc_rec)
            cons.update(
                orgao="Consolidado",
                result_pct=rpct,
                penalty_points=minc_rec["penalty_points"] + mtur_rec["penalty_points"],
                numerator=numerador,
                denominator=denominador,
                conforms=conforms,
            )
            _write_inms_row(ws, row, cons)
            row += 1

        for orgao in ("MinC", "MTur"):
            if orgao in occ:
                _write_inms_row(ws, row, occ[orgao])
                row += 1


def build_glosas(wb: Workbook, minc: list[dict], mtur: list[dict]) -> float:
    """Aba GLOSAS — retorna o valor da glosa final (para a aba de cálculo)."""
    ws = _new_sheet(wb, "GLOSAS", GLOSAS_COLUNAS, width=26)
    row = 2
    total_pontos = 0.0
    for rec in minc + mtur:
        if rec["penalty_points"] <= 0:
            continue
        faixa = _faixa(rec) or "Não conforme"
        pontos = rec["penalty_points"]
        total_pontos += pontos
        pct = pontos * 0.001
        _write(ws, row, (
            COMP, rec["orgao"], "", "", rec["contractual_id"],
            round(rec["result_pct"], 2), rec["target_value"], faixa,
            round(pct, 4), VALOR_BASE, round(VALOR_BASE * pct / 100, 2),
            "", "", "", "", "",
        ))
        ws.cell(row=row, column=6).number_format = PERCENT_FMT_SCALED  # Resultado
        ws.cell(row=row, column=7).number_format = PERCENT_FMT_SCALED  # Meta
        ws.cell(row=row, column=9).number_format = PERCENT_FMT_SCALED  # Percentual de Ajuste
        ws.cell(row=row, column=10).number_format = CURRENCY_FMT  # Valor Base
        ws.cell(row=row, column=11).number_format = CURRENCY_FMT  # Valor Glosa
        row += 1

    pct_bruto = total_pontos * 0.001
    aplicado = min(pct_bruto, LIMITE_PCT)
    glosa_final = VALOR_BASE * aplicado / 100

    r = row + 2
    for label, value in (
        ("Total de Pontos", round(total_pontos, 2)),
        ("Fórmula (pontos × 0,001)", f"{pct_bruto:.4f}%"),
        ("Limite", f"{LIMITE_PCT:.1f}%"),
        ("Percentual Aplicado", f"{aplicado:.2f}%"),
        ("Valor Glosa", round(glosa_final, 2)),
    ):
        label_cell = ws.cell(row=r, column=1, value=label)
        value_cell = ws.cell(row=r, column=2, value=value)
        label_cell.font = Font(bold=True)
        value_cell.font = Font(bold=True) if label == "Valor Glosa" else BODY_FONT
        if label == "Total de Pontos":  # top of the totals block
            label_cell.border = TOP_BORDER
            value_cell.border = TOP_BORDER
        if label == "Valor Glosa":
            value_cell.number_format = CURRENCY_FMT
        r += 1
    return glosa_final


def build_calculo(wb: Workbook, total_pontos: float, _glosa_final: float) -> None:
    """CALCULO_PAGAMENTO — MinC/MTur/Consolidado, glosa na coluna consolidada."""
    ws = wb.create_sheet("CALCULO_PAGAMENTO")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Parâmetros de Entrada — rateio PROVISÓRIO 0.5/0.5 até fonte oficial"
    for row, (label, value, fmt) in enumerate(
        (
            ("Valor mensal vigente", VALOR_BASE, CURRENCY_FMT),
            ("Limite máximo de glosa (%)", LIMITE_PCT, PERCENT_FMT_SCALED),
            ("Rateio MinC (provisório)", 0.5, PERCENT_FMT_FRACTION),
            ("Rateio MTur (provisório)", 0.5, PERCENT_FMT_FRACTION),
        ),
        start=3,
    ):
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        value_cell = ws.cell(row=row, column=2, value=value)
        value_cell.font = BODY_FONT
        value_cell.number_format = fmt

    header_row = 9
    for i, col in enumerate(CALCULO_COLUNAS, start=1):
        cell = ws.cell(row=header_row, column=i, value=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    ws.column_dimensions["A"].width = 58

    # Inspiração: os pontos de glosa vivem na coluna Consolidado (D14 = total
    # dos dois órgãos); MinC/MTur não têm glosa própria (B15/C15 = 0).
    colunas = (
        ("MinC", 0.5, 0.0),
        ("MTur", 0.5, 0.0),
        ("Consolidado", 1.0, total_pontos),
    )

    is_total_row = len(CALCULO_LINHAS) - 1  # "Valor recomendado" — last row, bold + top border
    for idx, label in enumerate(CALCULO_LINHAS):
        row = header_row + 1 + idx
        bold = idx == is_total_row
        label_cell = ws.cell(row=row, column=1, value=_clean(label))
        label_cell.font = Font(bold=True) if bold else BODY_FONT
        if bold:
            label_cell.border = TOP_BORDER
        fmt: str | None = None
        if label.startswith("Percentual de rateio"):
            fmt = PERCENT_FMT_FRACTION
        elif label.startswith(("Valor bruto", "Valor da glosa", "Outros ajustes", "Valor recomendado")):
            fmt = CURRENCY_FMT
        for col, (nome, rateio, pontos) in zip(range(2, len(CALCULO_COLUNAS) + 2), colunas):
            if label.startswith("Percentual de rateio"):
                value = rateio
            elif label.startswith("Valor bruto"):
                value = round(VALOR_BASE * rateio, 2)
            elif label.startswith("Pontos de glosa"):
                value = round(pontos, 2)
            elif label.startswith("Valor da glosa"):
                # MIN(pontos×0,001, teto%)/100 × bruto — só a coluna consolidada
                # carrega pontos (ticket 02 espelha a inspiração).
                value = round(min(pontos * 0.001, LIMITE_PCT) / 100 * VALOR_BASE * rateio, 2) if pontos else 0.0
            elif label.startswith("Outros ajustes"):
                value = 0
            else:
                bruto = VALOR_BASE * rateio
                glosa = min(pontos * 0.001, LIMITE_PCT) / 100 * bruto if pontos else 0.0
                value = round(max(0.0, bruto - glosa), 2)
            value_cell = ws.cell(row=row, column=col, value=value)
            value_cell.font = Font(bold=True) if bold else BODY_FONT
            if fmt:
                value_cell.number_format = fmt
            if bold:
                value_cell.border = TOP_BORDER


def main() -> int:
    minc = _load("MinC")
    mtur = _divergir_mtur(_load("MTur"))
    if not minc or not mtur:
        print("ERRO: falta ROM por órgão — rode `pyauditor report --orgao MinC` e `--orgao MTur` primeiro", file=sys.stderr)
        return 1

    wb = Workbook()
    wb.remove(wb.active)

    build_capa(wb)
    build_servicos(wb)
    build_inms_base(wb, minc, mtur)

    total_pontos = sum(r["penalty_points"] for r in minc + mtur if r["penalty_points"] > 0)
    glosa_final = build_glosas(wb, minc, mtur)
    build_calculo(wb, total_pontos, glosa_final)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)

    print(f"PROTOTYPE gravado em: {OUT}")
    print(f"Total de pontos (MinC+MTur): {total_pontos:.2f}")
    print(f"Percentual bruto: {total_pontos * 0.001:.4f}%  (aplicado = min(perc, {LIMITE_PCT}%))")
    print(f"Glosa final: {glosa_final:,.2f}")
    print(f"Abas: {wb.sheetnames}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
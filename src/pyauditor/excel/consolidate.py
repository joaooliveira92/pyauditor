"""Builds the consolidated financial workbook for `pyauditor consolidate`
(2.1): `CAPA_E_CONTROLE` + `SERVICOS_POR_ORGAO` + `INMS_BASE` + `GLOSAS` +
`CALCULO_PAGAMENTO` — the núcleo financeiro decided in
.scratch/multi-org-pipeline tickets 01/02/04.

`consolidate` never re-runs `measure`/`report`: its precondition is that
`reports/relatorio_<comp>_MinC.xlsx` and `_MTur.xlsx` already exist (ticket
04 Q1) — mas este módulo nunca abre esses ``.xlsx`` (capa vem de
``capa.csv``, indicadores vêm dos ROM JSON ``roms/<orgao>/<comp>/*.json``),
que são o mesmo artefato já computado que ``report.py`` consumiu.
A aba ``INMS_BASE`` do report deixa colunas de glosa em branco por
indicador (spec §12: glosa é agregado mensal lá), então não serviria para
o detalhe por-(indicador x órgão) que ``GLOSAS`` precisa.
"""

from __future__ import annotations

from collections.abc import Sequence
from dataclasses import dataclass
from pathlib import Path
from typing import Final

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Font, Side

from pyauditor.codes import format_inms_code
from pyauditor.excel._style import (
    BODY_FONT,
    BOTTOM_BORDER,
    HEADER_FILL,
    HEADER_FONT,
    LABEL_FONT,
    LEFT_ALIGN,
    TITLE_FONT,
    CellValue,
)
from pyauditor.excel._style import UNIT_BY_SHAPE as _UNIT_BY_SHAPE
from pyauditor.excel._style import new_sheet as _new_sheet
from pyauditor.excel._style import write_row as _write
from pyauditor.excel.glosas import Historico, compute_glosa, saldo_anterior_pct_de
from pyauditor.excel.orgao_consolidation import with_orgao_consolidation
from pyauditor.logging import logger
from pyauditor.rom.summary import IndicatorSummary

CAPA_SHEET: Final = "CAPA_E_CONTROLE"
SERVICOS_SHEET: Final = "SERVICOS_POR_ORGAO"
INMS_BASE_SHEET: Final = "INMS_BASE"
GLOSAS_SHEET: Final = "GLOSAS"
CALCULO_SHEET: Final = "CALCULO_PAGAMENTO"

LIMITE_PCT: Final = 30.0  # teto único sobre o agregado dos dois órgãos (ticket 02)
RATEIO_PADRAO: Final = 0.5  # provisório, até fonte oficial (ticket 01/02)

# docs/styleguide.md number formats — currency and percent, zero as "-".
# `_style.py` doesn't carry these yet (production report.py/capa.py don't
# apply them either); local to this module until that becomes a shared need.
_CURRENCY_FMT: Final = "R$#,##0.00;R$#,##0.00;-"
# `result_pct`/`%Ajuste` are already stored in percent-space (95.5 meaning
# "95.5%"); the literal "%" symbol auto-multiplies by 100 on display, so it
# must be escaped for these. Only true 0-1 fractions (rateio) want the real,
# auto-scaling "%" format.
_PERCENT_FMT_SCALED: Final = '0.00"%";0.00"%";-'
_PERCENT_FMT_FRACTION: Final = "0.00%;0.00%;-"
_TOP_BORDER: Final = Border(top=Side(style="thin", color="1F2937"))

# Decisão Fiscal — o fiscal aceita a justificativa do fornecedor (anistia: a
# ocorrência sai da base de pontos) ou não aceita (glosa mantida). Ticket 02.
_DECISAO_ACEITA: Final = "aceita"

# GLOSAS: uma linha por (indicador x órgão) + resumo agregado (ticket 02).
_GLOSAS_COLUMNS: Final[tuple[str, ...]] = (
    "Competência", "Órgão", "Item Contratual", "Serviço", "Indicador",
    "Resultado", "Meta", "Faixa de Descumprimento", "Percentual de Ajuste",
    "Valor Base", "Valor Glosa", "Reincidência", "Justificativa",
    "Número da Ocorrência", "Decisão Fiscal", "Observação do Gestor",
)
# Colunas de decisão do fiscal — nunca recalculadas, só preservadas no merge.
_DECISION_COLUMNS: Final[tuple[str, ...]] = (
    "Reincidência", "Justificativa", "Número da Ocorrência",
    "Decisão Fiscal", "Observação do Gestor",
)
_DECISION_TEXT_COLUMNS: Final[tuple[int, ...]] = tuple(
    _GLOSAS_COLUMNS.index(name) + 1 for name in _DECISION_COLUMNS
)

_CALCULO_COLUMNS: Final[tuple[str, ...]] = ("Componente", "MinC", "MTur", "Consolidado")
_CALCULO_LINHAS: Final[tuple[str, ...]] = (
    "Percentual de rateio",
    "Valor bruto (= mensal x rateio)",
    "Pontos de glosa",
    "Valor da glosa (= MIN(pontos x 0,001, 30%)/100 x bruto)",
    "Outros ajustes",
    "Valor recomendado (= max(0, bruto - glosa - outros))",
)

# docs spreadsheet.md's SERVICOS_POR_ORGAO — 9 serviços contratuais fixos
# (ticket 01). Segregação/critério de rateio idênticos para os dois órgãos
# até o Termo de Referência dizer o contrário.
_SERVICOS: Final[tuple[tuple[str, str, str, str], ...]] = (
    ("Central de Serviços e Monitoramento", "Sim", "Sim", "Sim"),
    ("Gerenciamento Técnico das Operações e Projetos", "Sim", "Sim", "Sim"),
    ("Banco de Dados", "Sim", "Sim", "Sim"),
    ("Aplicações, Virtualização e Computação em Nuvem", "Sim", "Sim", "Sim"),
    ("Serviços Corporativos", "Sim", "Sim", "Sim"),
    ("Armazenamento e Backup", "Sim", "Sim", "Sim"),
    ("Redes", "Sim", "Sim", "Sim"),
    ("Segurança da Informação", "Sim", "Sim", "Sim"),
    ("DevOps", "Sim", "Sim", "Sim"),
)

_CAPA_VALOR_LABELS: Final = ("Valor mensal vigente", "Valor global anual")

RowKey = tuple[str, str]  # (contractual_id, orgao)


@dataclass(frozen=True)
class ConsolidationResult:
    workbook: Workbook
    total_pontos: float
    glosa_final: float
    warnings: tuple[str, ...]
    glosa_calculada: bool


_TRACKED_HEADER_NAMES: Final[frozenset[str]] = frozenset({"Indicador", "Órgão", *_DECISION_COLUMNS})


def _normalize_header(name: str) -> str:
    """Normaliza espaços/acentos/case para detectar renomeação leve (issue 08)."""
    import unicodedata

    # strip, collapse internal whitespace, casefold, remove acentos
    collapsed = " ".join(name.strip().split())
    folded = collapsed.casefold()
    nfkd = unicodedata.normalize("NFD", folded)
    return "".join(c for c in nfkd if not unicodedata.combining(c))


def _check_no_duplicate_headers(path: Path, header_row: Sequence[object]) -> None:
    """A hand-edited workbook with a duplicate column name (two "Indicador"
    columns, or a stray column literally named "Justificativa" inserted by a
    fiscal) would otherwise make the header→index lookup silently keep the
    *last* matching column — decision values then get read from the wrong
    column with no error. Fail loudly instead."""
    seen: set[str] = set()
    duplicates: set[str] = set()
    for name in header_row:
        if isinstance(name, str) and name in _TRACKED_HEADER_NAMES:
            (duplicates if name in seen else seen).add(name)
    if duplicates:
        raise ValueError(
            f"{path}: coluna(s) duplicada(s) em {GLOSAS_SHEET!r}: "
            f"{', '.join(sorted(duplicates))} — planilha hand-edited em formato inesperado, "
            "corrija antes de rodar consolidate de novo"
        )


def _check_renamed_headers(path: Path, header_row: Sequence[object]) -> None:
    """Se um cabeçalho esperado não existe mas há candidato próximo por
    normalização (espaço extra, acento, case), falha nomeando ambos — em
    vez de perder a decisão silenciosamente (issue 08)."""
    present = {h for h in header_row if isinstance(h, str)}
    present_normalized = {_normalize_header(h): h for h in present}
    for expected in _TRACKED_HEADER_NAMES:
        if expected not in present:
            norm_expected = _normalize_header(expected)
            candidate = present_normalized.get(norm_expected)
            if candidate is not None:
                raise ValueError(
                    f"{path}: cabeçalho esperado {expected!r} não encontrado, "
                    f"mas existe candidato próximo {candidate!r} — possível renomeação "
                    "com espaço/acentuação/case, corrija o cabeçalho"
                )
            # também tenta match por um caractere de diferença (ex. espaço extra interno)
            for actual in present:
                if _normalize_header(actual) == norm_expected:
                    raise ValueError(
                        f"{path}: cabeçalho esperado {expected!r} não encontrado, "
                        f"mas existe candidato próximo {actual!r}"
                    )


def read_existing_decisions(path: Path) -> dict[RowKey, dict[str, object]]:
    """Reads the decision columns from a previously-generated consolidado
    workbook's `GLOSAS` sheet, keyed by (indicador, órgão) — the merge
    contract's preservation source (ticket 04 Q3). Empty dict if `path`
    doesn't exist yet or has no `GLOSAS` sheet (first run).
    """
    if not path.exists():
        return {}
    workbook = load_workbook(path, data_only=True)
    try:
        if GLOSAS_SHEET not in workbook.sheetnames:
            return {}
        sheet = workbook[GLOSAS_SHEET]

        header_row = [cell.value for cell in sheet[1]]
        if not header_row:
            return {}
        _check_no_duplicate_headers(path, header_row)
        _check_renamed_headers(path, header_row)
        if "Indicador" not in header_row:
            return {}
        col_idx = {name: i + 1 for i, name in enumerate(header_row) if isinstance(name, str)}
        indicador_col = col_idx.get("Indicador")
        orgao_col = col_idx.get("Órgão")
        if indicador_col is None or orgao_col is None:
            return {}

        decisions: dict[RowKey, dict[str, object]] = {}
        for row in range(2, sheet.max_row + 1):
            indicador = sheet.cell(row=row, column=indicador_col).value
            orgao = sheet.cell(row=row, column=orgao_col).value
            if not isinstance(indicador, str) or not isinstance(orgao, str):
                continue
            values: dict[str, object] = {
                name: sheet.cell(row=row, column=idx).value
                for name, idx in col_idx.items()
                if name in _DECISION_COLUMNS
            }
            if any(v not in (None, "") for v in values.values()):
                decisions[(format_inms_code(indicador), orgao)] = values
        return decisions
    finally:
        workbook.close()


def build_capa(
    wb: Workbook,
    competencia: str,
    capa: dict[str, object],
    warnings: list[str],
    valor_base: float | None,
) -> None:
    """Embeds a consolidated CAPA_E_CONTROLE. The contract is shared, so the
    common fields come from `capa.csv` (ticket 07); the monetary value comes
    from `objetos.csv` (`valor_base`), never from a capa file.
    """
    ws = wb.create_sheet(CAPA_SHEET, 0)
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Capa e controle — consolidado"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:B1")

    ws["A3"] = "Campo"
    ws["B3"] = "Valor"
    for cell in (ws["A3"], ws["B3"]):
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = LEFT_ALIGN

    campos: tuple[tuple[str, object], ...] = (
        ("Número do contrato", capa.get("Número do contrato")),
        ("Processo SEI", capa.get("Processo SEI")),
        ("Empresa contratada", capa.get("Empresa contratada")),
        ("Órgãos contratantes", "Ministério da Cultura / Ministério do Turismo"),
        ("Competência", competencia),
        ("Valor mensal vigente", valor_base),
        ("Valor global anual", valor_base * 12 if valor_base is not None else None),
    )
    for offset, (label, value) in enumerate(campos):
        row = 4 + offset
        label_cell = ws.cell(row=row, column=1, value=label)
        label_cell.font = LABEL_FONT
        label_cell.alignment = LEFT_ALIGN
        label_cell.border = BOTTOM_BORDER

        value_cell = ws.cell(row=row, column=2, value=value)
        value_cell.font = BODY_FONT
        value_cell.border = BOTTOM_BORDER
        if label in _CAPA_VALOR_LABELS:
            value_cell.number_format = _CURRENCY_FMT

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 42


def build_servicos(wb: Workbook, itens: tuple[float, ...] | None = None) -> None:
    """SERVICOS_POR_ORGAO — os 9 serviços contratuais, agora com o valor
    mensal de cada item vindo de `objetos.csv` (ticket 07 Q4), mapeado pelo
    índice (os nomes divergem entre as fontes — o índice não)."""
    ws = _new_sheet(
        wb, SERVICOS_SHEET,
        ("Item", "Serviço", "Valor Mensal (R$)", "Prestado ao MinC?", "Prestado ao MTur?",
         "Segregação Obrigatória?", "Critério de Rateio"),
    )
    for i, (nome, minc, mtur, seg) in enumerate(_SERVICOS, start=2):
        valor = itens[i - 2] if itens is not None and i - 2 < len(itens) else None
        _write(ws, i, (i - 1, nome, valor, minc, mtur, seg, "Chamados, ativos ou valor definido"))
        if valor is not None:
            ws.cell(row=i, column=3).number_format = _CURRENCY_FMT


_INMS_BASE_COLUMNS: Final[tuple[str, ...]] = (
    "Competência", "Item contratual", "Serviço", "Grupo operacional",
    "Código INMS", "Descrição", "Órgão", "Meta mínima ou máxima",
    "Sentido da meta", "Numerador", "Denominador", "Resultado calculado",
    "Unidade", "Conformidade", "Diferença para a meta",
)

def _inms_base_row(competencia: str, summary: IndicatorSummary) -> tuple[CellValue, ...]:
    dif = None
    if summary.target_value is not None:
        dif = round(summary.target_value - summary.result_pct, 2)
    return (
        competencia, None, summary.asset, None, format_inms_code(summary.contractual_id),
        summary.name, summary.orgao, summary.target_value, summary.target_operator,
        summary.numerator, summary.denominator, round(summary.result_pct, 2),
        _UNIT_BY_SHAPE.get(summary.shape, ""),
        "Conforme" if summary.conforms else "Não conforme", dif,
    )


def build_inms_base(
    wb: Workbook, competencia: str, minc: list[IndicatorSummary], mtur: list[IndicatorSummary]
) -> None:
    """Pooling Σnum/Σden por indicador (ticket 02) — reusa a mesma regra que
    já era testada dentro de `report.py`, agora vivendo só aqui.
    """
    ws = _new_sheet(wb, INMS_BASE_SHEET, _INMS_BASE_COLUMNS, width=20)
    rows = with_orgao_consolidation(minc + mtur)
    rows.sort(key=lambda s: (s.contractual_id, s.asset or "", s.orgao))
    for row_idx, summary in enumerate(rows, start=2):
        _write(ws, row_idx, _inms_base_row(competencia, summary))
        # Percentuais: Meta / Resultado calculado / Diferença para a meta.
        if _UNIT_BY_SHAPE.get(summary.shape) == "%":
            for col in (8, 12, 15):
                ws.cell(row=row_idx, column=col).number_format = _PERCENT_FMT_SCALED


def _decision_value(decision: dict[str, object], key: str) -> CellValue:
    """`decision`'s values come from openpyxl cells (any scalar type it
    supports) — narrowed to what a fresh cell can hold, coercing anything
    unexpected (dates, etc.) to its string form rather than dropping it.
    """
    value = decision.get(key)
    if value is None or isinstance(value, str | int | float):
        return value
    return str(value)


def _faixa(summary: IndicatorSummary) -> str:
    if summary.target_operator is None or summary.target_value is None:
        return "Ocorrência sob detalhamento por-ativo" if summary.penalty_points > 0 else ""
    dif = (
        summary.target_value - summary.result_pct
        if summary.target_operator == ">="
        else summary.result_pct - summary.target_value
    )
    return f"Déficit de {dif:.2f}pp" if dif > 0 else "Não conforme"


def _is_derived(summary: IndicatorSummary) -> bool:
    return "." in summary.indicator_id


def _deduplicate(summaries: list[IndicatorSummary]) -> list[IndicatorSummary]:
    by_key: dict[tuple[str, str | None], list[IndicatorSummary]] = {}
    for s in summaries:
        by_key.setdefault((s.contractual_id, s.asset), []).append(s)
    deduped: list[IndicatorSummary] = []
    for group in by_key.values():
        if len(group) > 1 and any(_is_derived(s) for s in group):
            deduped.extend(s for s in group if _is_derived(s))
        else:
            deduped.extend(group)
    return deduped


def build_glosas(
    wb: Workbook,
    competencia: str,
    minc: list[IndicatorSummary],
    mtur: list[IndicatorSummary],
    valor_base: float | None,
    existing_decisions: dict[RowKey, dict[str, object]],
    warnings: list[str],
    *,
    historico: Historico | None = None,
    is_final_month: bool = False,
) -> tuple[float, float]:
    """GLOSAS — uma linha por (indicador x órgão) com ocorrência de glosa,
    mais o resumo agregado. Decisão do fiscal ('Aceita' = anistia) tira a
    ocorrência da base de pontos; colunas de decisão são preservadas do
    workbook anterior, nunca recalculadas (ticket 02/04 Q3).

    Reusa ``glosas.compute_glosa`` por-órgão (mesma fórmula de ``report.py``) —
    com ``saldo_anterior_pct`` e ``is_final_month`` — e soma contra teto
    por-órgão em vez de teto único sobre o agregado.
    """
    ws = _new_sheet(wb, GLOSAS_SHEET, _GLOSAS_COLUMNS, width=26)
    seen_keys: set[RowKey] = set()
    row = 2

    # Deduplicação por Categoria (issue 01): base não conta quando derivados existem
    minc_dedup = _deduplicate(minc)
    mtur_dedup = _deduplicate(mtur)

    # Primeiro escreve linhas por ocorrência (valor por ocorrência continua pct simples,
    # mas o agregado usa compute_glosa por-órgão)
    pontos_por_orgao: dict[str, float] = {"MinC": 0.0, "MTur": 0.0}
    for summary in minc_dedup + mtur_dedup:
        if summary.penalty_points <= 0:
            continue
        key: RowKey = (format_inms_code(summary.contractual_id), summary.orgao)
        seen_keys.add(key)
        decision = existing_decisions.get(key, {})
        is_amnestied = (
            str(decision.get("Decisão Fiscal") or "").strip().lower().startswith(_DECISAO_ACEITA)
        )

        pontos = summary.penalty_points
        if not is_amnestied:
            pontos_por_orgao[summary.orgao] = pontos_por_orgao.get(summary.orgao, 0.0) + pontos
        pct = pontos * 0.001
        valor_glosa = round((valor_base or 0.0) * pct / 100, 2) if valor_base is not None else None

        _write(ws, row, (
            competencia, summary.orgao, "", "", format_inms_code(summary.contractual_id),
            round(summary.result_pct, 2), summary.target_value, _faixa(summary),
            round(pct, 4), valor_base, 0.0 if is_amnestied else valor_glosa,
            _decision_value(decision, "Reincidência"), _decision_value(decision, "Justificativa"),
            _decision_value(decision, "Número da Ocorrência"),
            _decision_value(decision, "Decisão Fiscal"),
            _decision_value(decision, "Observação do Gestor"),
        ))
        ws.cell(row=row, column=6).number_format = _PERCENT_FMT_SCALED
        ws.cell(row=row, column=7).number_format = _PERCENT_FMT_SCALED
        ws.cell(row=row, column=9).number_format = _PERCENT_FMT_SCALED
        if valor_base is not None:
            ws.cell(row=row, column=10).number_format = _CURRENCY_FMT
            ws.cell(row=row, column=11).number_format = _CURRENCY_FMT
        for column in _DECISION_TEXT_COLUMNS:
            ws.cell(row=row, column=column).number_format = "@"
        row += 1

    for key in existing_decisions:
        if key not in seen_keys:
            warnings.append(
                f"decisão registrada para {key[0]}/{key[1]} mas a ocorrência não existe mais "
                "nesta rodada — preservada apenas no histórico, não reescrita"
            )

    # Agregado por-órgão via compute_glosa (com saldo_anterior e is_final_month)
    historico = historico or {}
    saldo_anterior = saldo_anterior_pct_de(historico, competencia)
    # Rateio do saldo anterior proporcional aos pontos (se ambos têm pontos, divide;
    # se só um tem, todo saldo vai para ele)
    total_pontos_bruto = sum(pontos_por_orgao.values())
    if total_pontos_bruto > 0:
        # Distribui saldo anterior proporcionalmente — mesma semântica de manter
        # o teto por-órgão mas respeitar rollover global do período
        saldo_minc = saldo_anterior * (pontos_por_orgao.get("MinC", 0.0) / total_pontos_bruto)
        saldo_mtur = saldo_anterior * (pontos_por_orgao.get("MTur", 0.0) / total_pontos_bruto)
    else:
        saldo_minc = saldo_mtur = 0.0

    glosa_minc = compute_glosa(
        pontos_por_orgao.get("MinC", 0.0), valor_base,
        is_final_month=is_final_month, saldo_anterior_pct=saldo_minc,
    )
    glosa_mtur = compute_glosa(
        pontos_por_orgao.get("MTur", 0.0), valor_base,
        is_final_month=is_final_month, saldo_anterior_pct=saldo_mtur,
    )
    # Para valor_base por-órgão (rateio), o valor da glosa consolidada é a soma
    # das glosas por-órgão calculadas sobre o mesmo valor_base proporcional?
    # Simplificação: se valor_base é único (objetos.csv total), manter soma
    # direta — reflete o Valor Glosa idêntico ao report por-órgão somado.
    total_pontos = glosa_minc.total_points + glosa_mtur.total_points
    glosa_final = (glosa_minc.valor_da_glosa or 0.0) + (glosa_mtur.valor_da_glosa or 0.0)
    # Fallback se valor_base is None: glosa_final já é 0 via compute_glosa
    pct_bruto = total_pontos * 0.001 + saldo_anterior
    aplicado = min(pct_bruto, LIMITE_PCT)

    r = row + 2
    for label, value in (
        ("Total de Pontos", round(total_pontos, 2)),
        ("Fórmula (pontos x 0,001)", f"{pct_bruto:.4f}%"),
        ("Limite", f"{LIMITE_PCT:.1f}%"),
        ("Percentual Aplicado", f"{aplicado:.2f}%"),
        ("Valor Glosa", round(glosa_final, 2) if valor_base is not None else None),
    ):
        label_cell = ws.cell(row=r, column=1, value=label)
        value_cell = ws.cell(row=r, column=2, value=value)
        bold = label in ("Total de Pontos", "Valor Glosa")
        label_cell.font = Font(bold=True) if bold else BODY_FONT
        value_cell.font = Font(bold=True) if bold else BODY_FONT
        if label == "Total de Pontos":
            label_cell.border = _TOP_BORDER
            value_cell.border = _TOP_BORDER
        if label == "Valor Glosa" and valor_base is not None:
            value_cell.number_format = _CURRENCY_FMT
        r += 1

    return total_pontos, glosa_final


def build_calculo(
    wb: Workbook,
    valor_base: float | None,
    total_pontos: float,
    rateio_minc: float = RATEIO_PADRAO,
    rateio_mtur: float = RATEIO_PADRAO,
) -> None:
    """CALCULO_PAGAMENTO — espelha a inspiração: colunas MinC/MTur/
    Consolidado, glosa só na coluna Consolidado (ticket 02)."""
    ws = wb.create_sheet(CALCULO_SHEET)
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Parâmetros de Entrada — rateio PROVISÓRIO até fonte oficial"
    base = valor_base or 0.0
    params: tuple[tuple[str, float, str], ...] = (
        ("Valor mensal vigente", base, _CURRENCY_FMT),
        ("Limite máximo de glosa (%)", LIMITE_PCT, _PERCENT_FMT_SCALED),
        ("Rateio MinC (provisório)", rateio_minc, _PERCENT_FMT_FRACTION),
        ("Rateio MTur (provisório)", rateio_mtur, _PERCENT_FMT_FRACTION),
    )
    for row, (param_label, param_value, param_fmt) in enumerate(params, start=3):
        ws.cell(row=row, column=1, value=param_label).font = Font(bold=True)
        value_cell = ws.cell(row=row, column=2, value=param_value)
        value_cell.font = BODY_FONT
        value_cell.number_format = param_fmt

    header_row = 9
    for i, col_name in enumerate(_CALCULO_COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=i, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    ws.column_dimensions["A"].width = 58

    colunas = (
        ("MinC", rateio_minc, 0.0),
        ("MTur", rateio_mtur, 0.0),
        ("Consolidado", 1.0, total_pontos),
    )
    total_row_idx = len(_CALCULO_LINHAS) - 1

    for idx, label in enumerate(_CALCULO_LINHAS):
        row = header_row + 1 + idx
        bold = idx == total_row_idx
        label_cell = ws.cell(row=row, column=1, value=label)
        label_cell.font = Font(bold=True) if bold else BODY_FONT
        if bold:
            label_cell.border = _TOP_BORDER
        fmt: str | None = None
        if label.startswith("Percentual de rateio"):
            fmt = _PERCENT_FMT_FRACTION
        elif label.startswith(
            ("Valor bruto", "Valor da glosa", "Outros ajustes", "Valor recomendado")
        ):
            fmt = _CURRENCY_FMT
        for col, (_nome, rateio, pontos) in zip(
            range(2, len(_CALCULO_COLUMNS) + 1), colunas, strict=True
        ):
            if label.startswith("Percentual de rateio"):
                value: object = rateio
            elif label.startswith("Valor bruto"):
                value = round(base * rateio, 2)
            elif label.startswith("Pontos de glosa"):
                value = round(pontos, 2)
            elif label.startswith("Valor da glosa"):
                value = (
                    round(min(pontos * 0.001, LIMITE_PCT) / 100 * base * rateio, 2)
                    if pontos
                    else 0.0
                )
            elif label.startswith("Outros ajustes"):
                value = 0
            else:
                bruto = base * rateio
                glosa = min(pontos * 0.001, LIMITE_PCT) / 100 * bruto if pontos else 0.0
                value = round(max(0.0, bruto - glosa), 2)
            value_cell = ws.cell(row=row, column=col, value=value)
            value_cell.font = Font(bold=True) if bold else BODY_FONT
            if fmt:
                value_cell.number_format = fmt
            if bold:
                value_cell.border = _TOP_BORDER


def build_consolidated_workbook(
    competencia: str,
    minc: list[IndicatorSummary],
    mtur: list[IndicatorSummary],
    capa: dict[str, object],
    existing_decisions: dict[RowKey, dict[str, object]] | None = None,
    *,
    valor_base: float | None = None,
    itens: tuple[float, ...] | None = None,
    historico: Historico | None = None,
    is_final_month: bool = False,
) -> ConsolidationResult:
    """Pure, in-memory build of the 5-sheet consolidated workbook.

    `capa` carries the contract-common fields (from ``capa.csv``, ticket 07)
    e indicadores vêm dos ROM JSON (``roms/<orgao>/<comp>/*.json``), não do
    ``.xlsx`` de ``report.py``. ``valor_base`` e ``itens`` vêm de
    ``objetos.csv``. ``glosa_calculada`` é ``valor_base is not None``.
    """
    warnings: list[str] = []
    wb = Workbook()
    default_sheet = wb.active
    assert default_sheet is not None
    wb.remove(default_sheet)

    build_capa(wb, competencia, capa, warnings, valor_base)
    build_servicos(wb, itens)
    build_inms_base(wb, competencia, minc, mtur)
    total_pontos, glosa_final = build_glosas(
        wb, competencia, minc, mtur, valor_base, existing_decisions or {}, warnings,
        historico=historico, is_final_month=is_final_month,
    )
    build_calculo(wb, valor_base, total_pontos)

    for warning in warnings:
        logger.warning(warning)

    return ConsolidationResult(
        workbook=wb, total_pontos=total_pontos, glosa_final=glosa_final,
        warnings=tuple(warnings), glosa_calculada=valor_base is not None,
    )

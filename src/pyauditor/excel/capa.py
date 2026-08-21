"""The contract's capa — now CSV (ticket 07).

`input/capa.csv` holds the contract-common fields (`COMMON_FIELD_LABELS`);
`input/capa_{orgao}.csv` the per-órgão hand-fill fields
(`ORGAO_FIELD_LABELS`). The monetary fields left the capa entirely — their
source is `input/objetos.csv` (`excel/objetos.py`) — e, desde a spec
competencia-cli-equipe §4, Competência/períodos/responsáveis também saíram:
Competência/períodos são derivados do argumento `--competência`
(`pyauditor.periodo`) e os responsáveis vêm de `input/equipe.csv`
(`excel/equipe.py`). A capa nunca mais os carrega — nem como hand-fill,
nem como fallback de capas antigas. The `CAPA_E_CONTROLE` sheet embedded in
`report`'s workbook is still Excel (a `render_capa_sheet` view over the
merged CSV fields), so `SHEET_NAME`/`render_capa_sheet` survive; the file
itself is CSV.
"""

from __future__ import annotations

import csv
from pathlib import Path
from typing import Final

from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from pyauditor.atomic_write import atomic_write
from pyauditor.excel._style import (
    BODY_FONT,
    BOTTOM_BORDER,
    HEADER_FILL,
    HEADER_FONT,
    LABEL_FONT,
    LEFT_ALIGN,
    TITLE_FONT,
)
from pyauditor.excel.equipe import RESPONSAVEL_LABELS

SHEET_NAME: Final = "CAPA_E_CONTROLE"

CAPA_DELIMITER: Final = ";"
CAPA_ENCODING: Final = "utf-8-sig"

# docs/spreadsheet.md §Aba 1 "Campos" — split in two by the CSV migration
# (ticket 07 Q1): the contract's common fields live only in `capa.csv`.
COMMON_FIELD_LABELS: Final[tuple[str, ...]] = (
    "Número do contrato",
    "Processo SEI",
    "Empresa contratada",
    "CNPJ da contratada",
    "Objeto",
    "Vigência",
)

# Per-órgão fields, one `capa_{orgao}.csv` per órgão. Órgão contratante atual
# stays here (it names the party, so it's not common). Monetários ficam fora:
# a capa nunca mais carrega "Valor mensal vigente"/"Valor global anual".
# Competência/períodos/responsáveis idem (spec competencia-cli-equipe §4) —
# vivem em DERIVED_FIELD_LABELS/EQUIPE_FIELD_LABELS abaixo.
ORGAO_FIELD_LABELS: Final[tuple[str, ...]] = (
    "Órgão contratante atual",
    "Número da Ordem de Serviço",
    "Número da nota fiscal",
    "Data de emissão da nota fiscal",
    "Versão da planilha",
    "Data da análise",
    "Situação geral da aferição",
)

# Rótulos derivados da CLI — exibidos nas planilhas com valores sempre do
# argumento `--competência` (pyauditor.periodo.mes_bounds), nunca hand-fill.
DERIVED_FIELD_LABELS: Final[tuple[str, ...]] = (
    "Competência",
    "Período inicial da aferição",
    "Período final da aferição",
)

EQUIPE_FIELD_LABELS: Final[tuple[str, ...]] = RESPONSAVEL_LABELS

# The full field list, used by the report workbook's embedded CAPA_E_CONTROLE
# (comum + derivados + por-órgão + equipe, in display order).
FIELD_LABELS: Final[tuple[str, ...]] = (
    COMMON_FIELD_LABELS + DERIVED_FIELD_LABELS + ORGAO_FIELD_LABELS + EQUIPE_FIELD_LABELS
)

# docs/spreadsheet.md §Aba 1 — "Situações possíveis"
SITUACOES: Final[tuple[str, ...]] = (
    "Em preenchimento",
    "Aguardando evidências",
    "Em análise",
    "Conforme",
    "Conforme com glosa",
    "Não conforme",
    "Aprovado para pagamento",
    "Não recomendado para pagamento",
)


def render_capa_sheet(sheet: Worksheet, values: dict[str, object] | None = None) -> None:
    """Renders the CAPA_E_CONTROLE label/value layout onto `sheet`. With no
    `values`, cells are left blank for the fiscal técnico to fill in (the
    `bootstrap` case). With `values` (as returned by `read_capa_csv_fields`),
    reproduces an existing capa's content — used by `report` to embed the
    capa as the final workbook's first sheet, so the value stays in sync
    with whatever the fiscal técnico last filled in, rather than a copy
    that can drift.
    """
    sheet.sheet_view.showGridLines = False

    sheet["A1"] = "Capa e controle do contrato"
    sheet["A1"].font = TITLE_FONT
    sheet.merge_cells("A1:B1")

    sheet["A3"] = "Campo"
    sheet["B3"] = "Valor"
    for cell in (sheet["A3"], sheet["B3"]):
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = LEFT_ALIGN

    situacao_row: int | None = None
    for offset, label in enumerate(FIELD_LABELS):
        row = 4 + offset
        label_cell = sheet.cell(row=row, column=1, value=label)
        label_cell.font = LABEL_FONT
        label_cell.alignment = LEFT_ALIGN
        label_cell.border = BOTTOM_BORDER

        value_cell = sheet.cell(row=row, column=2)
        value_cell.font = BODY_FONT
        value_cell.border = BOTTOM_BORDER
        if values is not None:
            value_cell.value = values.get(label)  # type: ignore[assignment]
        elif label == "Situação geral da aferição":
            value_cell.value = SITUACOES[0]
        if label == "Situação geral da aferição":
            situacao_row = row

    if situacao_row is not None and values is None:
        validation = DataValidation(
            type="list",
            formula1=f'"{",".join(SITUACOES)}"',
            allow_blank=False,
        )
        sheet.add_data_validation(validation)
        validation.add(sheet.cell(row=situacao_row, column=2))

    sheet.column_dimensions["A"].width = 32
    sheet.column_dimensions["B"].width = 40


def capa_csv_text(labels: tuple[str, ...]) -> str:
    """CSV template for a capa file: title line, blank line, `Campo;Valor`
    header, then one `label;value` row per field. `Situação geral da
    aferição` defaults to the first situação (matching the xlsx bootstrap).
    """
    lines = ["Capa e controle do contrato;", "", "Campo;Valor"]
    for label in labels:
        value = SITUACOES[0] if label == "Situação geral da aferição" else ""
        lines.append(f"{label};{value}")
    return "\n".join(lines) + "\n"


def bootstrap_capa_csv(path: Path, labels: tuple[str, ...]) -> bool:
    """Creates the capa CSV at `path` if it doesn't exist yet.

    Returns True if the file was created, False if it already existed
    (in which case nothing is touched — bootstrap is idempotent).
    """
    if path.exists():
        return False
    atomic_write(path, lambda p: p.write_text(capa_csv_text(labels), encoding=CAPA_ENCODING))
    return True


def read_capa_csv_fields(path: Path) -> dict[str, str]:
    """Reads every `label;value` pair from an existing capa CSV — the fiscal
    técnico fills these in by hand after `bootstrap` creates the blank rows.
    Missing labels (e.g. an older capa predating a field added later) are
    simply absent from the returned dict. Duplicate labels raise
    `ValueError` — a hand-edited file with repeated labels would otherwise
    silently keep one of them.

    Returns string values only — the capa carries no monetary data anymore
    (ticket 07); numeric-looking text (Versão "1") stays a string here.
    """
    with path.open(encoding=CAPA_ENCODING, newline="") as handle:
        rows = list(csv.reader(handle, delimiter=CAPA_DELIMITER))

    fields: dict[str, str] = {}
    duplicates: set[str] = set()
    for row in rows:
        if not row:
            continue
        label = row[0].strip()
        if not label or label in ("Campo", "Capa e controle do contrato"):
            continue
        value = row[1] if len(row) > 1 else ""
        if label in fields:
            duplicates.add(label)
        fields[label] = value
    if duplicates:
        names = ", ".join(sorted(duplicates))
        raise ValueError(
            f"{path}: rótulo(s) duplicado(s): {names} — planilha "
            "hand-edited em formato inesperado, corrija antes de continuar"
        )
    return fields
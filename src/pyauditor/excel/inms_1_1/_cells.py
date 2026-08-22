"""Primitivas de célula/fórmula da aba INMS 1.1 — extraídas de
`excel/inms_1_1_audit.py` (ticket 04 SRP).
"""

from __future__ import annotations

from datetime import date, datetime
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter as cl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet
from typing import Callable, Final

from pyauditor.excel.inms_1_1._layout import (
    BORDER,
    BODY_FONT,
    GREEN_FILL,
    HEADER_FILL,
    HEADER_FONT,
    LABEL_FONT,
    NOTE_FONT,
    RED_FILL,
    SECTION_FILL,
    SECTION_FONT,
    _AJ,
    _AM,
    _R,
    _UNLOCKED,
)

_ColumnRange = Callable[[int], str]

_CellValue = str | float | int | datetime | date | None


def _raw_range(col: int, last_row: int) -> str:
    return f"${cl(col)}$2:${cl(col)}${last_row}"


def _section_bar(sheet: Worksheet, row: int, text: str, last_col: int = 12) -> None:
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)
    cell = sheet.cell(row=row, column=1, value=text)
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center")
    sheet.row_dimensions[row].height = 20


def _label_value(
    sheet: Worksheet,
    row: int,
    label: str,
    value: _CellValue,
    *,
    fmt: str | None = None,
    fill: PatternFill | None = None,
) -> None:
    lc = sheet.cell(row=row, column=1, value=label)
    lc.font = LABEL_FONT
    lc.border = BORDER
    vc = sheet.cell(row=row, column=2, value=value)
    vc.font = BODY_FONT
    vc.border = BORDER
    if fmt:
        vc.number_format = fmt
    if fill:
        lc.fill = fill
        vc.fill = fill

def _header_row(sheet: Worksheet, row: int, headers: tuple[str, ...]) -> None:
    for idx, text in enumerate(headers, start=1):
        cell = sheet.cell(row=row, column=idx, value=text)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left", wrap_text=True, vertical="center")
    sheet.row_dimensions[row].height = 30


def _add_table(sheet: Worksheet, name: str, ref: str) -> None:
    """Tabela nativa do Excel — não `sheet.auto_filter` (único por aba, seria
    sobrescrito pela segunda tabela detalhada); cada tabela tem seu próprio
    filtro/ordenação."""
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name=None, showRowStripes=False, showFirstColumn=False, showLastColumn=False
    )
    sheet.add_table(table)


def _add_situacao_conditional_formatting(sheet: Worksheet, coordinate: str) -> None:
    """ "Meta atingida"/"Meta não atingida" -> verde/vermelho — usado na Seção
    2 (resultado consolidado) e na Seção 7 (cada metodologia de controle)."""
    sheet.conditional_formatting.add(
        coordinate,
        CellIsRule(operator="equal", formula=['"Meta atingida"'], fill=GREEN_FILL),  # type: ignore[no-untyped-call]  # openpyxl não publica stub tipado para `CellIsRule`
    )
    sheet.conditional_formatting.add(
        coordinate,
        CellIsRule(operator="equal", formula=['"Meta não atingida"'], fill=RED_FILL),  # type: ignore[no-untyped-call]  # openpyxl não publica stub tipado para `CellIsRule`
    )

def _protect_support_columns(sheet: Worksheet) -> None:
    """Oculta as colunas de apoio (R:AM) e protege a aba contra edição
    acidental das fórmulas — os únicos campos que continuam editáveis são
    os marcados com `_UNLOCKED` (justificativa/documento/evidência de
    preenchimento manual da auditoria, Seções 4 e 6). Sem senha: o
    objetivo é reduzir edição acidental das fórmulas, não impedir edição
    deliberada por quem tem o arquivo (ticket 20 / B-03). Não afeta a
    leitura das fórmulas pelo pipeline — proteção de planilha do Excel só
    bloqueia edição interativa, nunca o cálculo/leitura de valores por
    openpyxl ou por qualquer motor de recálculo.

    A coluna `AJ` (Situação dos dados) fica de fora do ocultamento: é
    preenchida em Python, não alimenta fórmula alguma, e é o indicador
    visual de linhas com data ausente/inválida exigido pelo ticket 02
    (C-02) — ocultá-la anularia a sinalização (consenso C-02 × B-03)."""
    for col in range(_R, _AM + 1):
        if col == _AJ:
            continue
        sheet.column_dimensions[cl(col)].hidden = True
    sheet.protection.sheet = True
    sheet.protection.formatCells = False
    sheet.protection.formatColumns = False
    sheet.protection.formatRows = False

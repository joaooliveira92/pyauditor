"""Seções 1–3 da aba INMS 1.1 (identificação, resumo, memória) — extraídas de
`excel/inms_1_1_audit.py` (ticket 04 SRP).
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.worksheet import Worksheet

from pyauditor.excel._safety import safe_excel_text
from pyauditor.excel.inms_1_1._cells import (
    _add_situacao_conditional_formatting,
    _CellValue,
    _header_row,
    _label_value,
    _section_bar,
)
from pyauditor.excel.inms_1_1._layout import (
    _DATE_FMT,
    _DATETIME_FMT,
    _PCT2,
    _PCT4,
    GRAY_FILL,
    LABEL_FONT,
    NOTE_FONT,
    ORANGE_FILL,
    TEAL_FILL,
    TITLE_FONT,
)
from pyauditor.periodo import PeriodoAfericao


def _write_section_1_identificacao(
    sheet: Worksheet,
    *,
    rows: list[dict[str, str]],
    contract: str,
    periodo: PeriodoAfericao | None,
    raw_csv_path: Path,
    generated_at: datetime,
) -> None:
    sheet.merge_cells('A1:L1')
    t = sheet.cell(
        row=1, column=1, value='INMS 1.1 – Incidentes atendidos dentro do prazo'
    )
    t.font = TITLE_FONT

    _section_bar(sheet, 3, 'SEÇÃO 1 · IDENTIFICAÇÃO')
    data_corte: _CellValue
    if periodo is not None:
        competencia = (
            f'{periodo.inicio:%d/%m/%Y}'
            f'a'
            f'{periodo.fim:%d/%m/%Y}'
            f'({periodo.inicio:%Y-%m})'
        )
        data_corte = periodo.fim
        data_corte_fmt = _DATE_FMT
    else:
        competencia = 'Não informado'
        data_corte = 'Não informado'
        data_corte_fmt = None
    _label_value(sheet, 4, 'Competência:', competencia, fill=GRAY_FILL)
    _label_value(
        sheet, 5, 'Contrato:', safe_excel_text(contract), fill=GRAY_FILL
    )
    _label_value(
        sheet,
        6,
        'Fonte dos dados:',
        # Só o nome do arquivo, não o caminho completo — evita expor
        # estrutura de diretório/usuário do ambiente que gerou a planilha
        # (ticket 19 / B-02).
        safe_excel_text(f'{raw_csv_path.name} ({len(rows)} registros brutos)'),
        fill=GRAY_FILL,
    )
    _label_value(
        sheet,
        7,
        'Data de geração:',
        generated_at,
        fmt=_DATETIME_FMT,
        fill=GRAY_FILL,
    )
    _label_value(
        sheet,
        8,
        'Data de corte:',
        data_corte,
        fmt=data_corte_fmt,
        fill=GRAY_FILL,
    )
    _label_value(
        sheet,
        9,
        'Responsável pela elaboração:',
        'Não informado',
        fill=GRAY_FILL,
    )


def _write_section_2_resumo(
    sheet: Worksheet,
    *,
    iap: str,
    iadp: str,
    fora: str,
    meta_value: float,
) -> int:
    """Seção de linhas fixas (11-14) — devolve `next_free_row` (16, fixo)
    para a Seção 3."""
    _section_bar(sheet, 11, 'SEÇÃO 2 · RESUMO EXECUTIVO')
    _header_row(
        sheet,
        12,
        (
            'Meta contratual',
            'IAP (abertos)',
            'IADP (dentro do prazo)',
            'Fora do prazo',
            'Resultado INMS 1.1',
            'Diferença p/ meta',
            'Situação',
        ),
    )
    sheet['A13'] = f'={meta_value}'
    sheet['B13'] = f'={iap}'
    sheet['C13'] = f'={iadp}'
    sheet['D13'] = f'={fora}'
    # B13=0 (nenhum incidente aberto no período) não é "0% de desempenho" —
    # é resultado não mensurável; guardado para não propagar #DIV/0! e para
    # não afirmar "meta não atingida" indevidamente.
    sheet['E13'] = '=IF(B13=0,"Sem ocorrências",C13/B13)'
    # Meta mínima ("`>=`") — único operador suportado por este renderer
    # (validado em `_validate_write_sheet_params`, ticket 05 / A-03).
    sheet['F13'] = '=IF(B13=0,"",E13-A13)'
    sheet['G13'] = (
        '=IF(B13=0,"Sem ocorrências",IF(E13>=A13,"Meta atingida","Meta não '
        'atingida"))'
    )

    sheet['A13'].number_format = _PCT2
    sheet['E13'].number_format = _PCT2
    sheet['F13'].number_format = _PCT2
    for c in ('A13', 'B13', 'C13', 'D13', 'E13', 'F13', 'G13'):
        sheet[c].font = Font(name='Arial', size=12, bold=True)
        sheet[c].fill = TEAL_FILL
        sheet[c].alignment = Alignment(horizontal='center')
    sheet.row_dimensions[13].height = 24

    _add_situacao_conditional_formatting(sheet, 'G13')

    sheet.merge_cells('A14:L14')
    pen_note = sheet.cell(
        row=14,
        column=1,
        value='Penalidade: '
        'Pendente '
        'de '
        'confirmação '
        'da '
        'regra '
        'de '
        'arredondamento '
        '— '
        'ver '
        'Seção '
        '9.',
    )
    pen_note.font = LABEL_FONT
    pen_note.fill = ORANGE_FILL
    return 16


def _write_section_3_memoria(sheet: Worksheet) -> int:
    """Seção de linhas fixas (16-26) — devolve `next_free_row` (28, fixo)
    para a Seção 4."""
    _section_bar(sheet, 16, 'SEÇÃO 3 · MEMÓRIA DO CÁLCULO CONSOLIDADO')
    _label_value(
        sheet, 17, 'IAP (incidentes abertos no período):', '=B13', fmt='0'
    )
    _label_value(
        sheet, 18, 'IADP (incidentes dentro do prazo):', '=C13', fmt='0'
    )
    sheet.merge_cells('A19:C19')
    sheet['A19'] = 'INMS 1.1 = IADP ÷ IAP'
    sheet['A19'].font = NOTE_FONT
    sheet.merge_cells('A20:C20')
    sheet['A20'] = '=CONCATENATE("INMS 1.1 = ",B18," ÷ ",B17)'
    sheet['A20'].font = NOTE_FONT
    _label_value(sheet, 21, 'INMS 1.1 (resultado, 4 casas):', '=E13', fmt=_PCT4)
    _label_value(sheet, 22, 'Meta contratual:', '=A13', fmt=_PCT4)
    _label_value(
        sheet,
        23,
        'Desvio em pontos percentuais (Resultado - Meta):',
        '=IF(B13=0,"",E13-A13)',
        fmt=_PCT4,
    )
    _label_value(
        sheet,
        24,
        'Quantidade mínima dentro do prazo p/ atingir a meta:',
        '=ROUNDUP(B13*A13,0)',
        fmt='0',
    )
    _label_value(
        sheet,
        25,
        'Margem em quantidade de incidentes (IADP - mínimo):',
        '=C13-B24',
        fmt='0;-0',
    )
    sheet.merge_cells('A26:L26')

    def _narrativa(desfecho: str) -> str:
        return (
            'CONCATENATE("Com ",B17," incidentes, seriam necessários pelo menos '
            '",B24,'
            f'" dentro do prazo para atingir ",TEXT(B22,"0.00%"),". O resultado'
            f'ficou {desfecho})'
        )

    # Narrativa condicional ao sinal da margem (B25 = IADP - mínimo): "abaixo"
    # só quando a margem é negativa — antes disso o texto sempre dizia
    # "abaixo do mínimo" mesmo com meta superada (ABS(B25) escondia o sinal).
    abaixo = _narrativa(
        '",ABS(B25)," incidente(s) abaixo do mínimo necessário."'
    )
    exato = _narrativa('exatamente no mínimo necessário."')
    acima = _narrativa('",B25," incidente(s) acima do mínimo necessário."')
    sheet['A26'] = (
        '=IF(B13=0,"Sem incidentes abertos no período — indicador não '
        'mensurável.",'
        f'IF(B25<0,{abaixo},IF(B25=0,{exato},{acima})))'
    )
    sheet['A26'].font = NOTE_FONT
    return 28

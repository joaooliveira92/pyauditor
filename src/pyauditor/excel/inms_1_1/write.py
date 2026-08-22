"""Fronteira pública da aba INMS 1.1: validação de params + `write_sheet`
(composição das 9 seções por cursor de linha) + `has_required_columns`.

Extraído de `excel/inms_1_1_audit.py` (ticket 04 SRP); a API pública fica
em `excel/inms_1_1_audit.py` (fachada).
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Final

from openpyxl import Workbook
from openpyxl.utils import get_column_letter as cl
from openpyxl.worksheet.worksheet import Worksheet

from pyauditor.categoria_filter import GRUPO_EXECUTOR_COLUMN
from pyauditor.config.categorias import CategoriasFile, GrupoExecutorMode, WholeIndicatorMode
from pyauditor.excel._workbook import create_sheet_atomic, force_recalc, unique_table_name
from pyauditor.excel.inms_1_1._cells import _protect_support_columns, _raw_range
from pyauditor.excel.inms_1_1._domain import (
    _REQUIRED_COLUMNS,
    _build_grupo_rows,
    _normalize_no_prazo,
    has_required_columns,
)
from pyauditor.excel.inms_1_1._layout import (
    _NO_PRAZO_COLUMN,
    _R,
    _X,
)
from pyauditor.excel.inms_1_1._raw_block import _write_raw_block
from pyauditor.excel.inms_1_1._sections_1_3 import (
    _write_section_1_identificacao,
    _write_section_2_resumo,
    _write_section_3_memoria,
)
from pyauditor.excel.inms_1_1._sections_4_5 import (
    _write_section_4_detalhamento,
    _write_section_5_subtotais,
)
from pyauditor.excel.inms_1_1._sections_6_7 import (
    _write_section_6_fora_prazo,
    _write_section_7_auditoria,
)
from pyauditor.excel.inms_1_1._sections_8_9 import (
    _write_section_8_tempo,
    _write_section_9_penalidade,
)
from pyauditor.periodo import PeriodoAfericao

__all__: Final[tuple[str, ...]] = ("has_required_columns", "write_sheet")

def _validate_write_sheet_params(
    *,
    target_operator: str,
    target_value: float,
    penalty_base_points: float,
    penalty_step_points: float,
    penalty_step_size_pct: float,
) -> None:
    """Validação na fronteira de `write_sheet()` — a função é pública e
    aceita `str`/`float` irrestritos; o Pydantic (`config/models.py`) já
    valida esses mesmos limites quando os parâmetros vêm de um YAML de
    config, mas nada impede uma chamada direta com valores fora de faixa."""
    # A2-03: a memória de cálculo (Seção 3) e a penalidade (Seção 9) só têm
    # semântica de "quantidade mínima dentro do prazo"/"diferença Meta -
    # Resultado" para meta mínima (`>=`); nenhum uso real do INMS 1.1 hoje
    # é meta-teto (`<=`), então restringimos aqui em vez de manter as duas
    # narrativas coerentes ao mesmo tempo — ver ticket 05 (A-03).
    if target_operator != ">=":
        raise ValueError(
            f"target_operator {target_operator!r} não suportado por este renderer — "
            "a aba enriquecida do INMS 1.1 assume meta mínima ('>='); para meta-teto "
            "('<='), use o renderer genérico (_write_grupo_executor_sheet)"
        )
    if not 0 <= target_value <= 100:
        raise ValueError(f"target_value {target_value!r} fora da faixa 0–100")
    if penalty_base_points < 0:
        raise ValueError(f"penalty_base_points {penalty_base_points!r} não pode ser negativo")
    if penalty_step_points < 0:
        raise ValueError(f"penalty_step_points {penalty_step_points!r} não pode ser negativo")
    if penalty_step_size_pct <= 0:
        raise ValueError(
            f"penalty_step_size_pct {penalty_step_size_pct!r} deve ser positivo — "
            "usado como divisor nas fórmulas da Seção 9"
        )


def write_sheet(
    workbook: Workbook,
    sheet_name: str,
    *,
    categorias_file: CategoriasFile,
    grupo_executor_entries: list[tuple[str, GrupoExecutorMode]],
    whole_indicator_entries: list[tuple[str, WholeIndicatorMode]],
    fieldnames: list[str],
    rows: list[dict[str, str]],
    target_operator: str,
    target_value: float,
    penalty_base_points: float,
    penalty_step_points: float,
    penalty_step_size_pct: float,
    contract: str,
    periodo: PeriodoAfericao | None,
    raw_csv_path: Path,
    generated_at: datetime,
) -> None:
    del whole_indicator_entries  # categorias.yaml não usa whole_indicator para 1.1 hoje
    del fieldnames  # validado por has_required_columns antes de chamar

    _validate_write_sheet_params(
        target_operator=target_operator,
        target_value=target_value,
        penalty_base_points=penalty_base_points,
        penalty_step_points=penalty_step_points,
        penalty_step_size_pct=penalty_step_size_pct,
    )

    rows = [{**row, _NO_PRAZO_COLUMN: _normalize_no_prazo(row)} for row in rows]

    with create_sheet_atomic(workbook, sheet_name) as sheet:
        sheet.sheet_view.showGridLines = False
        for col, width in {
            1: 42,
            2: 24,
            3: 16,
            4: 12,
            5: 14,
            6: 12,
            7: 12,
            8: 16,
            9: 14,
            10: 34,
            11: 20,
            12: 40,
        }.items():
            sheet.column_dimensions[cl(col)].width = width

        last_row = 1 + len(rows)
        real_values = {row[GRUPO_EXECUTOR_COLUMN] for row in rows}
        grupo_rows = _build_grupo_rows(categorias_file, grupo_executor_entries, real_values)
        _write_raw_block(sheet, rows, grupo_rows, last_row)

        def rng(col: int) -> str:
            return _raw_range(col, last_row)

        iap = f"ROWS({rng(_R)})"
        iadp = f'COUNTIF({rng(_X)},"S")'
        fora = f'COUNTIF({rng(_X)},"N")'
        meta_value = target_value / 100

        table_names = {
            "grupo_executor": unique_table_name(workbook, "TabelaGrupoExecutor"),
            "fora_do_prazo": unique_table_name(workbook, "TabelaForaDoPrazo"),
            "amostra_divergencias": unique_table_name(workbook, "TabelaAmostraDivergencias"),
        }

        _write_section_1_identificacao(
            sheet,
            rows=rows,
            contract=contract,
            periodo=periodo,
            raw_csv_path=raw_csv_path,
            generated_at=generated_at,
        )
        next_row = _write_section_2_resumo(
            sheet,
            iap=iap,
            iadp=iadp,
            fora=fora,
            meta_value=meta_value,
        )
        next_row = _write_section_3_memoria(sheet)
        next_row = _write_section_4_detalhamento(
            sheet,
            grupo_rows=grupo_rows,
            rng=rng,
            start_row=next_row,
            table_name=table_names["grupo_executor"],
        )
        next_row = _write_section_5_subtotais(sheet, rng=rng, start_row=next_row)
        next_row = _write_section_6_fora_prazo(
            sheet,
            rows=rows,
            rng=rng,
            start_row=next_row,
            table_name=table_names["fora_do_prazo"],
        )
        next_row = _write_section_7_auditoria(
            sheet,
            rows=rows,
            rng=rng,
            start_row=next_row,
            table_name=table_names["amostra_divergencias"],
        )
        next_row = _write_section_8_tempo(sheet, rng=rng, start_row=next_row)
        _write_section_9_penalidade(
            sheet,
            penalty_base_points=penalty_base_points,
            penalty_step_points=penalty_step_points,
            penalty_step_size_pct=penalty_step_size_pct,
            start_row=next_row,
        )

        _protect_support_columns(sheet)
        sheet.freeze_panes = "A2"

    force_recalc(workbook)

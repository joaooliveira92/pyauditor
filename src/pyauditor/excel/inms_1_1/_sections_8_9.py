"""Seções 8–9 da aba INMS 1.1 (tempo corrido, penalidade) — extraídas de
`excel/inms_1_1_audit.py` (ticket 04 SRP).
"""

from __future__ import annotations

from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.worksheet import Worksheet

from pyauditor.excel.inms_1_1._cells import (
    _ColumnRange,
    _label_value,
    _section_bar,
)
from pyauditor.excel.inms_1_1._layout import (
    _AD,
    _AF,
    _AG,
    _DUR,
    _PCT4,
    BODY_FONT,
    LABEL_FONT,
    NOTE_FONT,
    ORANGE_FILL,
    TEAL_FILL,
)


def _write_section_8_tempo(
    sheet: Worksheet, *, rng: _ColumnRange, start_row: int
) -> int:
    """Devolve `next_free_row` — a linha livre após a nota de rodapé da
    seção, usada pela Seção 9 como sua própria linha inicial."""
    s8_bar = start_row
    _section_bar(
        sheet,
        s8_bar,
        'SEÇÃO 8 · TEMPO CORRIDO MÉDIO ATÉ A RESOLUÇÃO',
        last_col=6,
    )
    _label_value(
        sheet,
        s8_bar + 1,
        'Tempo corrido médio até a resolução (todas as linhas):',
        f'=AVERAGE({rng(_AG)})',
        fmt=_DUR,
    )
    _label_value(
        sheet,
        s8_bar + 2,
        'Mediana do tempo corrido até a resolução:',
        f'=MEDIAN({rng(_AG)})',
        fmt=_DUR,
    )
    # M-01: `_AG` já devolve "" (texto, ignorado por AVERAGE/MEDIAN) para
    # linhas com data ausente/malformada ou encerramento anterior à
    # abertura — mas isso ficava implícito; expõe a contagem de linhas
    # rejeitadas para que a média/mediana acima não pareça cobrir 100% dos
    # incidentes sem dizer quantos foram excluídos.
    rejeitados_row = s8_bar + 3
    sheet.cell(
        row=rejeitados_row,
        column=1,
        value=(
            'Registros excluídos da média/mediana (data ausente/inválida ou '
            'encerramento antes da abertura):'
        ),
    ).font = LABEL_FONT
    rejeitados_cell = sheet.cell(
        row=rejeitados_row, column=2, value=f'=COUNTIF({rng(_AG)},"")'
    )
    rejeitados_cell.font = BODY_FONT
    atraso_row = s8_bar + 4
    sheet.cell(
        row=atraso_row,
        column=1,
        value='Atrasomédiodosregistrosforadoprazo(vs.limiteITSM,minutos):',
    ).font = LABEL_FONT
    # M-02: seleciona pela coluna calculada `_AD` ("No prazo (data limite
    # ITSM)"), não pela classificação do fornecedor (`_X`) — consistente
    # com o rótulo "vs. limite ITSM"; guarda contra #DIV/0! quando não há
    # nenhum registro fora do prazo por esse critério.
    atraso_cell = sheet.cell(
        row=atraso_row,
        column=2,
        value=(
            f'=IF(COUNTIF({rng(_AD)},"N")=0,"Sem'
            f'atrasos",AVERAGEIF({rng(_AD)},"N",{rng(_AF)}))'
        ),
    )
    atraso_cell.number_format = '0.0'
    atraso_cell.font = BODY_FONT
    note8 = s8_bar + 5
    sheet.merge_cells(f'A{note8}:F{note8}')
    sheet[f'A{note8}'] = (
        'O tempo corrido médio é indicador gerencial complementar — não '
        'substitui a '
        "verificação linha-a-linha do campo 'No prazo' para o cálculo do INMS "
        '1.1.'
    )
    sheet[f'A{note8}'].font = NOTE_FONT
    return note8 + 2


def _write_section_9_penalidade(
    sheet: Worksheet,
    *,
    penalty_base_points: float,
    penalty_step_points: float,
    penalty_step_size_pct: float,
    start_row: int,
) -> None:
    s9_bar = start_row
    _section_bar(
        sheet, s9_bar, 'SEÇÃO 9 · PENALIDADE (CÁLCULO PRELIMINAR)', last_col=6
    )
    _label_value(sheet, s9_bar + 1, 'Meta:', '=A13', fmt=_PCT4)
    _label_value(sheet, s9_bar + 2, 'Resultado:', '=E13', fmt=_PCT4)
    diff9_row = s9_bar + 3
    # Meta mínima ("`>=`") — único operador suportado por este renderer.
    diff_formula = '=IF(B13=0,"",A13-E13)'
    _label_value(
        sheet,
        diff9_row,
        'Diferença (Meta - Resultado):',
        diff_formula,
        fmt=_PCT4,
    )
    diffpp_row = diff9_row + 1
    _label_value(
        sheet,
        diffpp_row,
        'Diferença em pontos percentuais:',
        f'=IF(B{diff9_row}="","",B{diff9_row}*100)',
        fmt='0.0000',
    )
    base_row = diffpp_row + 1
    below_target = 'E13<A13'
    _label_value(
        sheet,
        base_row,
        'Penalidade-base:',
        f'=IF(B13=0,"Não'
        f'aplicável",IF({below_target},{penalty_base_points:g},0))',
        fmt='0',
    )
    add_row = base_row + 1
    _label_value(
        sheet,
        add_row,
        f'Adicional proporcional ({penalty_step_points:g} pontos a cada '
        f'{penalty_step_size_pct:g} p.p. — cálculo contínuo):',
        (
            f'=IF(B13=0,"Não aplicável",IF({below_target},'
            f'(B{diffpp_row}/{penalty_step_size_pct!r})*{penalty_step_points!r},0))'
        ),
        fmt='0.0000',
    )
    total_row = add_row + 1
    _label_value(
        sheet,
        total_row,
        'Total proporcional (base + adicional):',
        (
            f'=IF(OR(ISTEXT(B{base_row}),ISTEXT(B{add_row})),"Não'
            f'aplicável",B{base_row}+B{add_row})'
        ),
        fmt='0.0000',
        fill=TEAL_FILL,
    )
    scenario_row = total_row + 1
    _label_value(
        sheet,
        scenario_row,
        'Cenário — faixas completas ou iniciadas:',
        (
            f'=IF(B13=0,"Não'
            f'aplicável",IF({below_target},{penalty_base_points:g}+'
            f'CEILING(B{diffpp_row},{penalty_step_size_pct!r})/{penalty_step_size_pct!r}'
            f'*{penalty_step_points!r},0))'
        ),
        fmt='0',
        fill=ORANGE_FILL,
    )
    obs_row = scenario_row + 2
    sheet.merge_cells(f'A{obs_row}:L{obs_row}')
    sheet[f'A{obs_row}'] = (
        'Resultado sujeito à confirmação da regra de arredondamento e das '
        'disposições '
        'gerais de glosa do Termo de Referência.'
    )
    sheet[f'A{obs_row}'].font = Font(
        name='Arial', size=10, bold=True, color='9A3412'
    )
    sheet[f'A{obs_row}'].fill = ORANGE_FILL
    sheet[f'A{obs_row}'].alignment = Alignment(wrap_text=True)

    final_note_row = obs_row + 2
    sheet.merge_cells(f'A{final_note_row}:L{final_note_row}')
    sheet[f'A{final_note_row}'] = (
        'Dados de apoio às fórmulas desta aba nas colunas R:AM (estrutura de '
        'apoio, '
        'auditável) — mantidos para rastreabilidade; não excluir nem reordenar.'
    )
    sheet[f'A{final_note_row}'].font = NOTE_FONT

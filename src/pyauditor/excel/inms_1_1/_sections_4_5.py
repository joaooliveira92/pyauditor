"""Seções 4–5 da aba INMS 1.1 (detalhamento por grupo, subtotais) — extraídas de
`excel/inms_1_1_audit.py` (ticket 04 SRP).
"""

from __future__ import annotations

from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.worksheet import Worksheet

from pyauditor.excel._safety import safe_excel_text
from pyauditor.excel.inms_1_1._cells import (
    _add_table,
    _ColumnRange,
    _header_row,
    _section_bar,
)
from pyauditor.excel.inms_1_1._layout import (
    _AG,
    _AUDIT_REVIEW_LABEL,
    _DUR,
    _NIVEL_ORDER,
    _PCT2,
    _S,
    _SEM_NIVEL,
    _UNLOCKED,
    _X,
    _Z,
    BODY_FONT,
    BORDER,
    GREEN_FILL,
    LABEL_FONT,
    NOTE_FONT,
    ORANGE_FILL,
    RED_FILL,
    TEAL_FILL,
)


def _write_section_4_detalhamento(
    sheet: Worksheet,
    *,
    grupo_rows: list[tuple[str, str, str]],
    rng: _ColumnRange,
    start_row: int,
    table_name: str,
) -> int:
    """Devolve `next_free_row` — a linha livre após a tabela de grupos,
    usada pela Seção 5 como sua própria linha inicial."""
    _section_bar(
        sheet,
        start_row,
        'SEÇÃO 4 · DETALHAMENTO POR GRUPO EXECUTOR',
        last_col=12,
    )
    _header_row(
        sheet,
        start_row + 1,
        (
            'Categoria',
            'Nível',
            'Grupo executor',
            'Linhas',
            'Dentro do prazo',
            'Fora do prazo',
            '% (dentro/linhas)',
            'Tempo médio',
            'Incluído no INMS?',
            'Justificativa de exclusão',
            'Documento autorizador',
            'Observação de auditoria',
        ),
    )
    first_group_row = start_row + 2
    for offset, (grupo, nivel, categoria) in enumerate(grupo_rows):
        r = first_group_row + offset
        sheet.cell(
            row=r, column=1, value=safe_excel_text(categoria)
        ).font = BODY_FONT
        sheet.cell(
            row=r, column=2, value=safe_excel_text(nivel)
        ).font = BODY_FONT
        sheet.cell(
            row=r, column=3, value=safe_excel_text(grupo)
        ).font = BODY_FONT
        grupo_ref = f'$C${r}'
        linhas_cell = sheet.cell(
            row=r, column=4, value=f'=COUNTIF({rng(_S)},{grupo_ref})'
        )
        dentro_cell = sheet.cell(
            row=r,
            column=5,
            value=f'=COUNTIFS({rng(_S)},{grupo_ref},{rng(_X)},"S")',
        )
        fora_cell = sheet.cell(
            row=r,
            column=6,
            value=f'=COUNTIFS({rng(_S)},{grupo_ref},{rng(_X)},"N")',
        )
        pct_cell = sheet.cell(
            row=r, column=7, value=f'=IF(D{r}=0,"Sem ocorrências",E{r}/D{r})'
        )
        pct_cell.number_format = _PCT2
        tempo_cell = sheet.cell(
            row=r,
            column=8,
            value=f'=IF(D{r}=0,"—",AVERAGEIF({rng(_S)},{grupo_ref},{rng(_AG)}))',
        )
        tempo_cell.number_format = _DUR
        incluido = sheet.cell(row=r, column=9, value='Sim')
        is_audit_review = categoria == _AUDIT_REVIEW_LABEL
        if is_audit_review:
            just = (
                'Nenhuma exclusão aplicada — ausência de fundamento documental '
                'formal '
                'para exclusão do grupo.'
            )
            obs = (
                'Grupo mantido no consolidado do INMS 1.1 por força da regra '
                'contratual '
                '(denominador = total de incidentes abertos no período); '
                'reclassificação '
                'exigiria documento autorizador válido — ver Seção 9.'
            )
            for cc in (
                linhas_cell,
                dentro_cell,
                fora_cell,
                pct_cell,
                tempo_cell,
                incluido,
            ):
                cc.fill = ORANGE_FILL
        else:
            just = 'Não aplicável — nenhuma exclusão aplicada.'
            obs = '—'
        just_cell = sheet.cell(row=r, column=10, value=just)
        doc_cell = sheet.cell(row=r, column=11, value='Não informado')
        obs_cell = sheet.cell(row=r, column=12, value=obs)
        for c in (
            linhas_cell,
            dentro_cell,
            fora_cell,
            incluido,
            just_cell,
            doc_cell,
            obs_cell,
        ):
            c.font = BODY_FONT
            c.alignment = Alignment(wrap_text=True, vertical='top')
        for col in range(1, 13):
            sheet.cell(row=r, column=col).border = BORDER
        # Campos de preenchimento manual da auditoria (justificativa/documento
        # autorizador) permanecem editáveis quando a planilha for protegida
        # (ticket 20 / B-03) — os demais campos desta seção são fórmulas.
        just_cell.protection = _UNLOCKED
        doc_cell.protection = _UNLOCKED
    last_group_row = first_group_row + len(grupo_rows) - 1
    _add_table(sheet, table_name, f'A{start_row + 1}:L{last_group_row}')
    return last_group_row + 2


def _write_section_5_subtotais(
    sheet: Worksheet, *, rng: _ColumnRange, start_row: int
) -> int:
    """Devolve `next_free_row` — a linha livre após as verificações
    cruzadas da seção, usada pela Seção 6 como sua própria linha inicial."""
    sub_bar_row = start_row
    _section_bar(
        sheet,
        sub_bar_row,
        'SEÇÃO 5 · SUBTOTAIS POR NÍVEL (informação gerencial)',
        last_col=6,
    )
    _header_row(
        sheet,
        sub_bar_row + 1,
        (
            'Nível / Grupo',
            'Linhas',
            'Dentro do prazo',
            'Fora do prazo',
            '% bruto',
            'Tempo médio',
        ),
    )
    nivel_rows = list(
        range(sub_bar_row + 2, sub_bar_row + 2 + len(_NIVEL_ORDER))
    )
    for r, nivel_label in zip(nivel_rows, _NIVEL_ORDER, strict=True):
        sheet.cell(row=r, column=1, value=nivel_label).font = BODY_FONT
        linhas = sheet.cell(
            row=r, column=2, value=f'=COUNTIFS({rng(_Z)},"{nivel_label}")'
        )
        dentro = sheet.cell(
            row=r,
            column=3,
            value=f'=COUNTIFS({rng(_Z)},"{nivel_label}",{rng(_X)},"S")',
        )
        fora_c = sheet.cell(
            row=r,
            column=4,
            value=f'=COUNTIFS({rng(_Z)},"{nivel_label}",{rng(_X)},"N")',
        )
        pct = sheet.cell(
            row=r, column=5, value=f'=IF(B{r}=0,"Sem ocorrências",C{r}/B{r})'
        )
        pct.number_format = _PCT2
        tempo = sheet.cell(
            row=r,
            column=6,
            value=f'=IF(B{r}=0,"—",AVERAGEIFS({rng(_AG)},{rng(_Z)},"{nivel_label}"))',
        )
        tempo.number_format = _DUR
        for c in (linhas, dentro, fora_c, pct, tempo):
            c.font = BODY_FONT
        for col in range(1, 7):
            sheet.cell(row=r, column=col).border = BORDER

    outr = nivel_rows[-1] + 1
    sheet.cell(
        row=outr, column=1, value=f'Sem nível ({_AUDIT_REVIEW_LABEL})'
    ).font = BODY_FONT
    linhas = sheet.cell(
        row=outr, column=2, value=f'=COUNTIFS({rng(_Z)},"{_SEM_NIVEL}")'
    )
    dentro = sheet.cell(
        row=outr,
        column=3,
        value=f'=COUNTIFS({rng(_Z)},"{_SEM_NIVEL}",{rng(_X)},"S")',
    )
    fora_c = sheet.cell(
        row=outr,
        column=4,
        value=f'=COUNTIFS({rng(_Z)},"{_SEM_NIVEL}",{rng(_X)},"N")',
    )
    pct = sheet.cell(
        row=outr,
        column=5,
        value=f'=IF(B{outr}=0,"Sem ocorrências",C{outr}/B{outr})',
    )
    pct.number_format = _PCT2
    tempo = sheet.cell(
        row=outr,
        column=6,
        value=f'=IF(B{outr}=0,"—",AVERAGEIFS({rng(_AG)},{rng(_Z)},"{_SEM_NIVEL}"))',
    )
    tempo.number_format = _DUR
    for c in (linhas, dentro, fora_c, pct, tempo):
        c.font = BODY_FONT
        c.fill = ORANGE_FILL
    for col in range(1, 7):
        sheet.cell(row=outr, column=col).border = BORDER

    totr = outr + 1
    sheet.cell(row=totr, column=1, value='Total geral').font = Font(bold=True)
    tl = sheet.cell(row=totr, column=2, value=f'=SUM(B{nivel_rows[0]}:B{outr})')
    td = sheet.cell(row=totr, column=3, value=f'=SUM(C{nivel_rows[0]}:C{outr})')
    tf = sheet.cell(row=totr, column=4, value=f'=SUM(D{nivel_rows[0]}:D{outr})')
    tp = sheet.cell(
        row=totr,
        column=5,
        value=f'=IF(B{totr}=0,"Sem ocorrências",C{totr}/B{totr})',
    )
    tp.number_format = _PCT2
    for c in (tl, td, tf, tp):
        c.font = Font(bold=True)
        c.fill = TEAL_FILL
    for col in range(1, 7):
        sheet.cell(row=totr, column=col).border = BORDER

    note_row = totr + 1
    sheet.merge_cells(f'A{note_row}:F{note_row}')
    sheet[f'A{note_row}'] = (
        'O consolidado do INMS 1.1 é a soma dos numeradores/denominadores dos '
        'níveis '
        '(esta linha), não a média simples dos percentuais de N1/N2/N3.'
    )
    sheet[f'A{note_row}'].font = NOTE_FONT

    check_row = note_row + 1
    sheet[f'A{check_row}'] = 'Verificação cruzada (Seção 5 = Seção 2/3):'
    sheet[f'A{check_row}'].font = LABEL_FONT
    chk = sheet.cell(
        row=check_row,
        column=2,
        value=(
            f'=IF(AND(ISNUMBER(E{totr}),ISNUMBER(E13)),'
            f'IF(ROUND(E{totr},6)=ROUND(E13,6),"OK","DIVERGÊNCIA"),'
            f'IF(E{totr}=E13,"OK","DIVERGÊNCIA"))'
        ),
    )
    chk.font = Font(bold=True)
    sheet.conditional_formatting.add(
        chk.coordinate,
        CellIsRule(operator='equal', formula=['"OK"'], fill=GREEN_FILL),  # type: ignore[no-untyped-call]
    )
    sheet.conditional_formatting.add(
        chk.coordinate,
        CellIsRule(operator='equal', formula=['"DIVERGÊNCIA"'], fill=RED_FILL),  # type: ignore[no-untyped-call]
    )

    # IAP = dentro do prazo + fora do prazo (ticket 06 / A-04) — se algum
    # valor de "No prazo" escapasse à normalização de `_normalize_no_prazo`,
    # essa identidade quebraria silenciosamente (o registro entraria no IAP
    # sem entrar em nenhum dos dois `COUNTIF`); mantida como cinto e
    # suspensório mesmo com a validação na fronteira.
    check2_row = check_row + 1
    sheet[f'A{check2_row}'] = (
        'Verificação (IAP = dentro do prazo + fora do prazo):'
    )
    sheet[f'A{check2_row}'].font = LABEL_FONT
    chk2 = sheet.cell(
        row=check2_row,
        column=2,
        value='=IF(B13=C13+D13,"OK","DIVERGÊNCIA DE CLASSIFICAÇÃO")',
    )
    chk2.font = Font(bold=True)
    sheet.conditional_formatting.add(
        chk2.coordinate,
        CellIsRule(operator='equal', formula=['"OK"'], fill=GREEN_FILL),  # type: ignore[no-untyped-call]
    )
    sheet.conditional_formatting.add(
        chk2.coordinate,
        CellIsRule(
            operator='equal',
            formula=['"DIVERGÊNCIA DE CLASSIFICAÇÃO"'],
            fill=RED_FILL,
        ),  # type: ignore[no-untyped-call]
    )
    return check2_row + 2

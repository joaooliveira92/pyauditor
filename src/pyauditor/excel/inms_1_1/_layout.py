"""Constantes de leiaute da aba INMS 1.1 (estilo, colunas, prazo) — extraídas de
`excel/inms_1_1_audit.py` (ticket 04 SRP).
"""

from __future__ import annotations

from openpyxl.styles import Border, Font, PatternFill, Protection, Side
from typing import Final

from pyauditor.config.niveis import NIVEL_BY_CATEGORIA, NIVEL_ORDER

# Colunas do CSV bruto exigidas para o tratamento enriquecido, além das já
# exigidas pelo shape `ratio` genérico (`No prazo`, `DataHoraSolicitacao`,
# `DataHoraFim`, `Grupo_executor`).
_NUM_SOLICITACAO_COLUMN: Final[str] = "Nº Solicitacao"
_ATIVIDADE_COLUMN: Final[str] = "Atividades"
_DATA_SOLICITACAO_COLUMN: Final[str] = "DataHoraSolicitacao"
_DATA_LIMITE_COLUMN: Final[str] = "DataHoraLimite"
_DATA_FIM_COLUMN: Final[str] = "DataHoraFim"
_NO_PRAZO_COLUMN: Final[str] = "No prazo"
_TECNICO_COLUMN: Final[str] = "TecnicoExecutor"

# Prazo contratual de incidentes de criticidade alta (Anexo D / aba
# "Prazos" — input/prazos.csv: "Incidentes, Alta, 2h (horas corridas)").
# Constante local: só esta aba precisa do valor numérico para o controle
# contratual bruto; nenhum outro shape do pipeline usa "2 horas corridas".
_PRAZO_HORAS_CORRIDAS: Final[float] = 2.0

_NIVEL_BY_CATEGORIA = NIVEL_BY_CATEGORIA
_NIVEL_ORDER = NIVEL_ORDER
_AUDIT_REVIEW_LABEL: Final[str] = "Grupo sob análise de responsabilidade"
# Sentinela para grupos sem Nível (categoria "outros") — não usar "" como
# valor de célula: o Excel/openpyxl trata célula com "" como vazia, e
# VLOOKUP contra uma célula vazia devolve 0 (numérico), não "" (texto),
# quebrando os filtros COUNTIFS por Nível na Seção 5.
_SEM_NIVEL: Final[str] = "—"

_DATETIME_FMT: Final[str] = "dd/mm/yyyy hh:mm"
_DATE_FMT: Final[str] = "dd/mm/yyyy"
_PCT2: Final[str] = "0.00%"
_PCT4: Final[str] = "0.0000%"
_DUR: Final[str] = "[h]:mm"

TITLE_FONT: Final = Font(name="Arial", size=14, bold=True, color="1F2937")
SECTION_FONT: Final = Font(name="Arial", size=11, bold=True, color="FFFFFF")
SECTION_FILL: Final = PatternFill("solid", fgColor="1F2937")
HEADER_FONT: Final = Font(name="Arial", size=10, bold=True, color="FFFFFF")
HEADER_FILL: Final = PatternFill("solid", fgColor="374151")
BODY_FONT: Final = Font(name="Arial", size=10)
LABEL_FONT: Final = Font(name="Arial", size=10, bold=True)
NOTE_FONT: Final = Font(name="Arial", size=9, italic=True, color="6B7280")
GRAY_FILL: Final = PatternFill("solid", fgColor="E5E7EB")
TEAL_FILL: Final = PatternFill("solid", fgColor="CCFBF1")
GREEN_FILL: Final = PatternFill("solid", fgColor="BBF7D0")
RED_FILL: Final = PatternFill("solid", fgColor="FECACA")
ORANGE_FILL: Final = PatternFill("solid", fgColor="FED7AA")
BORDER: Final = Border(bottom=Side(style="thin", color="D1D5DB"))
# Campos de preenchimento manual (justificativa/documento/evidência) que
# devem continuar editáveis mesmo com a planilha protegida (ticket 20 / B-03).
_UNLOCKED: Final = Protection(locked=False)

# Colunas de apoio (dados brutos) — far à direita do conteúdo visível.
_R, _S, _T, _U, _V, _W, _X, _Y, _Z, _AA, _AB, _AC, _AD, _AE, _AF, _AG, _AH, _AI = range(18, 36)
_AJ = 36
_AK, _AL, _AM = 37, 38, 39

_DATA_QUALIDADE_OK: Final[str] = "OK"

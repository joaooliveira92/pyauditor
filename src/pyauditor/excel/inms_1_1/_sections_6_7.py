"""Seções 6–7 da aba INMS 1.1 (fora do prazo, auditoria) — extraídas de
`excel/inms_1_1_audit.py` (ticket 04 SRP).
"""

from __future__ import annotations

from datetime import timedelta
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.worksheet import Worksheet
from typing import Final

from pyauditor.excel._datetime import PRAZO_TOLERANCIA_MINUTOS, parse_dt
from pyauditor.excel.inms_1_1._cells import (
    _ColumnRange,
    _add_situacao_conditional_formatting,
    _add_table,
    _header_row,
    _section_bar,
)
from pyauditor.excel.inms_1_1._layout import (
    BODY_FONT,
    BORDER,
    LABEL_FONT,
    NOTE_FONT,
    ORANGE_FILL,
    RED_FILL,
    _AB,
    _AC,
    _AD,
    _AE,
    _AF,
    _AG,
    _AH,
    _AI,
    _DATA_LIMITE_COLUMN,
    _DATA_SOLICITACAO_COLUMN,
    _DATETIME_FMT,
    _NO_PRAZO_COLUMN,
    _PCT2,
    _PCT4,
    _PRAZO_HORAS_CORRIDAS,
    _R,
    _S,
    _T,
    _U,
    _V,
    _W,
    _X,
    _Y,
    _UNLOCKED,
)

def _write_section_6_fora_prazo(
    sheet: Worksheet,
    *,
    rows: list[dict[str, str]],
    rng: _ColumnRange,
    start_row: int,
    table_name: str,
) -> int:
    """Devolve `next_free_row` — a linha livre após a nota "nenhum
    incidente" ou após a tabela, usada pela Seção 7 como sua própria linha
    inicial."""
    s6_bar = start_row
    _section_bar(sheet, s6_bar, "SEÇÃO 6 · INCIDENTES FORA DO PRAZO", last_col=11)
    _header_row(
        sheet,
        s6_bar + 1,
        (
            "Nº solicitação",
            "Grupo executor",
            "Atividade",
            "Abertura",
            "Limite (ITSM)",
            "Encerramento",
            "Atraso vs. limite (min)",
            "Técnico executor",
            "Justificativa",
            "Aceite da justificativa",
            "Documento/evidência",
        ),
    )
    fora_first = s6_bar + 2
    # Tabela dimensionada ao número real de incidentes fora do prazo (não um
    # teto arbitrário) — a Seção 6 existe para evidenciar exceções, não pra
    # truncá-las quando o órgão tiver mais de um punhado.
    fora_count = sum(1 for row in rows if row[_NO_PRAZO_COLUMN] == "N")
    ah_range = rng(_AH)
    if fora_count == 0:
        sheet.merge_cells(f"A{fora_first}:K{fora_first}")
        sheet[f"A{fora_first}"] = "Nenhum incidente fora do prazo no período."
        sheet[f"A{fora_first}"].font = NOTE_FONT
        return fora_first + 2

    for n in range(1, fora_count + 1):
        r = fora_first + n - 1
        match_expr = f"MATCH({n},{ah_range},0)"
        c1 = sheet.cell(row=r, column=1, value=f'=IFERROR(INDEX({rng(_R)},{match_expr}),"")')
        c2 = sheet.cell(row=r, column=2, value=f'=IFERROR(INDEX({rng(_S)},{match_expr}),"")')
        c3 = sheet.cell(row=r, column=3, value=f'=IFERROR(INDEX({rng(_T)},{match_expr}),"")')
        c4 = sheet.cell(row=r, column=4, value=f'=IFERROR(INDEX({rng(_U)},{match_expr}),"")')
        c4.number_format = _DATETIME_FMT
        c5 = sheet.cell(row=r, column=5, value=f'=IFERROR(INDEX({rng(_V)},{match_expr}),"")')
        c5.number_format = _DATETIME_FMT
        c6 = sheet.cell(row=r, column=6, value=f'=IFERROR(INDEX({rng(_W)},{match_expr}),"")')
        c6.number_format = _DATETIME_FMT
        c7 = sheet.cell(row=r, column=7, value=f'=IFERROR(INDEX({rng(_AF)},{match_expr}),"")')
        c7.number_format = "0.0"
        c8 = sheet.cell(row=r, column=8, value=f'=IFERROR(INDEX({rng(_Y)},{match_expr}),"")')
        c9 = sheet.cell(row=r, column=9, value="Não informado")
        c10 = sheet.cell(row=r, column=10, value="Não informado")
        c11 = sheet.cell(row=r, column=11, value="Não informado")
        for c in (c1, c2, c3, c4, c5, c6, c7, c8, c9, c10, c11):
            c.font = BODY_FONT
            c.border = BORDER
        c1.fill = RED_FILL
        # Justificativa/aceite/evidência são preenchimento manual da
        # auditoria — permanecem editáveis com a planilha protegida
        # (ticket 20 / B-03).
        c9.protection = _UNLOCKED
        c10.protection = _UNLOCKED
        c11.protection = _UNLOCKED
    fora_last = fora_first + fora_count - 1
    _add_table(sheet, table_name, f"A{s6_bar + 1}:K{fora_last}")
    return fora_last + 2


def _write_section_7_auditoria(
    sheet: Worksheet,
    *,
    rows: list[dict[str, str]],
    rng: _ColumnRange,
    start_row: int,
    table_name: str,
) -> int:
    """Devolve `next_free_row` — a linha livre após a amostra (ou após a
    nota "nenhuma divergência"), usada pela Seção 8 como sua própria linha
    inicial."""
    s7_bar = start_row
    _section_bar(sheet, s7_bar, "SEÇÃO 7 · AUDITORIA DO PRAZO CONTRATUAL", last_col=7)
    sheet.cell(row=s7_bar + 1, column=1, value="Controles de resultado").font = LABEL_FONT
    _header_row(sheet, s7_bar + 2, ("Metodologia", "Resultado", "Situação"))
    ctrl_rows = [
        ("Resultado informado pelo fornecedor (campo 'No prazo')", "C13/B13"),
        (
            "Resultado reproduzido pela data limite registrada no ITSM "
            "(DataHoraFim ≤ DataHoraLimite)",
            f'COUNTIF({rng(_AD)},"S")/B13',
        ),
        (
            f"Controle contratual bruto (DataHoraFim ≤ DataHoraSolicitacao + "
            f"{_PRAZO_HORAS_CORRIDAS:g}h corridas)",
            f'COUNTIF({rng(_AE)},"S")/B13',
        ),
    ]
    first_ctrl = s7_bar + 3
    for i, (label, division) in enumerate(ctrl_rows):
        r = first_ctrl + i
        sheet.cell(row=r, column=1, value=label).font = BODY_FONT
        sheet.cell(row=r, column=1).alignment = Alignment(wrap_text=True)
        val = sheet.cell(row=r, column=2, value=f'=IF(B13=0,"Sem ocorrências",{division})')
        val.number_format = _PCT4
        sit_formula = (
            f'=IF(B13=0,"Não aplicável",IF(B{r}>=A13,"Meta atingida","Meta não atingida"))'
        )
        sit = sheet.cell(row=r, column=3, value=sit_formula)
        val.font = BODY_FONT
        sit.font = BODY_FONT
        for col in range(1, 4):
            sheet.cell(row=r, column=col).border = BORDER
        _add_situacao_conditional_formatting(sheet, sit.coordinate)

    div_header_row = first_ctrl + 3 + 1
    sheet.cell(
        row=div_header_row,
        column=1,
        value=(
            f"Limite ITSM superior ao prazo contratual bruto (abertura + "
            f"{_PRAZO_HORAS_CORRIDAS:g}h corridas) — comparação unidirecional: só "
            f"sinaliza prorrogação, não limite ITSM mais rígido que o contratual"
        ),
    ).font = LABEL_FONT
    div_count_row = div_header_row + 1
    sheet.cell(
        row=div_count_row,
        column=1,
        value="Registros com limite ITSM superior ao contratual bruto:",
    ).font = BODY_FONT
    div_count = sheet.cell(row=div_count_row, column=2, value=f'=COUNTIF({rng(_AC)},"Sim")')
    div_count.font = Font(bold=True)
    sheet.cell(row=div_count_row, column=3, value="% do total:").font = BODY_FONT
    div_pct = sheet.cell(
        row=div_count_row,
        column=4,
        value=f'=IF(B13=0,"Sem ocorrências",B{div_count_row}/B13)',
    )
    div_pct.number_format = _PCT2
    div_pct.font = Font(bold=True)
    div_pct.fill = ORANGE_FILL
    div_count.fill = ORANGE_FILL

    note_row7 = div_count_row + 1
    sheet.merge_cells(f"A{note_row7}:L{note_row7}")
    sheet[f"A{note_row7}"] = (
        "Casos em que o limite registrado no ITSM é superior ao limite contratual bruto "
        "não são presumidos válidos nem inválidos aqui: quando o percentual acima for "
        "relevante, solicite o histórico de pausa/suspensão/reclassificação do SLA ou a "
        "norma de calendário aplicada para confirmar a compatibilidade com a cláusula "
        "contratual de prazo corrido. Um limite ITSM inferior ao contratual bruto não é "
        "sinalizado — é mais rígido que o exigido, não uma violação."
    )
    sheet[f"A{note_row7}"].font = Font(name="Arial", size=10, bold=True, color="9A3412")
    sheet[f"A{note_row7}"].fill = ORANGE_FILL
    sheet[f"A{note_row7}"].alignment = Alignment(wrap_text=True, vertical="center")
    sheet.row_dimensions[note_row7].height = 48

    prazo_delta = timedelta(hours=_PRAZO_HORAS_CORRIDAS, minutes=PRAZO_TOLERANCIA_MINUTOS)
    divergentes_count = 0
    for row in rows:
        sol = parse_dt(row[_DATA_SOLICITACAO_COLUMN]).value
        lim = parse_dt(row[_DATA_LIMITE_COLUMN]).value
        if sol is not None and lim is not None and lim > sol + prazo_delta:
            divergentes_count += 1
    sample_size = min(divergentes_count, 15)

    sample_header_row = note_row7 + 2
    if sample_size == 0:
        sheet.merge_cells(f"A{sample_header_row}:F{sample_header_row}")
        sheet[f"A{sample_header_row}"] = (
            "Nenhum limite ITSM superior ao contratual bruto encontrado no período."
        )
        sheet[f"A{sample_header_row}"].font = NOTE_FONT
        return sample_header_row + 2

    sheet.cell(
        row=sample_header_row,
        column=1,
        value=(
            f"Amostra de limites ITSM superiores ao contratual bruto "
            f"({sample_size} primeiros, ordenados por ocorrência)"
        ),
    ).font = LABEL_FONT
    _header_row(
        sheet,
        sample_header_row + 1,
        (
            "Nº solicitação",
            "Abertura",
            "Limite registrado (ITSM)",
            "Limite bruto (abertura + prazo corrido)",
            "Diferença (horas)",
            "No prazo (fornecedor)",
        ),
    )
    sample_first = sample_header_row + 2
    ai_range = rng(_AI)
    for n in range(1, sample_size + 1):
        r = sample_first + n - 1
        match_expr = f"MATCH({n},{ai_range},0)"
        c1 = sheet.cell(row=r, column=1, value=f'=IFERROR(INDEX({rng(_R)},{match_expr}),"")')
        c2 = sheet.cell(row=r, column=2, value=f'=IFERROR(INDEX({rng(_U)},{match_expr}),"")')
        c2.number_format = _DATETIME_FMT
        c3 = sheet.cell(row=r, column=3, value=f'=IFERROR(INDEX({rng(_V)},{match_expr}),"")')
        c3.number_format = _DATETIME_FMT
        c4 = sheet.cell(row=r, column=4, value=f'=IFERROR(INDEX({rng(_AB)},{match_expr}),"")')
        c4.number_format = _DATETIME_FMT
        c5 = sheet.cell(
            row=r,
            column=5,
            value=(
                f'=IFERROR((INDEX({rng(_V)},{match_expr})-INDEX({rng(_AB)},{match_expr}))*24,"")'
            ),
        )
        c5.number_format = "0.00"
        c6 = sheet.cell(row=r, column=6, value=f'=IFERROR(INDEX({rng(_X)},{match_expr}),"")')
        for c in (c1, c2, c3, c4, c5, c6):
            c.font = BODY_FONT
            c.border = BORDER
    sample_last = sample_first + sample_size - 1
    _add_table(sheet, table_name, f"A{sample_header_row + 1}:F{sample_last}")
    sample_note_row = sample_last + 1
    sheet.merge_cells(f"A{sample_note_row}:F{sample_note_row}")
    sheet[f"A{sample_note_row}"] = (
        f'=CONCATENATE("Amostra limitada às {sample_size} primeiras ocorrências de ",'
        f'B{div_count_row}," registros com limite ITSM superior ao contratual bruto — '
        f'colunas de apoio desta aba (coluna AC) permitem reproduzir a lista completa.")'
    )
    sheet[f"A{sample_note_row}"].font = NOTE_FONT
    return sample_note_row + 2

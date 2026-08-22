"""Base de apoio R–AM da aba INMS 1.1 (dados brutos + fórmulas) — extraída de
`excel/inms_1_1_audit.py` (ticket 04 SRP).
"""

from __future__ import annotations

from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter as cl
from openpyxl.worksheet.worksheet import Worksheet

from pyauditor.categoria_filter import GRUPO_EXECUTOR_COLUMN
from pyauditor.excel._datetime import PRAZO_TOLERANCIA_MINUTOS, parse_dt
from pyauditor.excel._safety import safe_excel_text
from pyauditor.excel.inms_1_1._layout import (
    _AA,
    _AB,
    _AC,
    _AD,
    _AE,
    _AF,
    _AG,
    _AH,
    _AI,
    _AJ,
    _AK,
    _AL,
    _AM,
    _ATIVIDADE_COLUMN,
    _DATA_FIM_COLUMN,
    _DATA_LIMITE_COLUMN,
    _DATA_QUALIDADE_OK,
    _DATA_SOLICITACAO_COLUMN,
    _DATETIME_FMT,
    _DUR,
    _NO_PRAZO_COLUMN,
    _NUM_SOLICITACAO_COLUMN,
    _PRAZO_HORAS_CORRIDAS,
    _R,
    _S,
    _T,
    _TECNICO_COLUMN,
    _U,
    _V,
    _W,
    _X,
    _Y,
    _Z,
    BODY_FONT,
    HEADER_FILL,
    HEADER_FONT,
    NOTE_FONT,
    RED_FILL,
)


def _write_raw_block(
    sheet: Worksheet,
    rows: list[dict[str, str]],
    grupo_rows: list[tuple[str, str, str]],
    last_row: int,
) -> None:
    headers = {
        _R: 'Nº Solicitação',
        _S: 'Grupo executor',
        _T: 'Atividade',
        _U: 'DataHoraSolicitacao',
        _V: 'DataHoraLimite (ITSM)',
        _W: 'DataHoraFim',
        _X: 'No prazo (fornecedor)',
        _Y: 'TecnicoExecutor',
        _Z: 'Nível',
        _AA: 'Categoria',
        _AB: 'Limite contratual (abertura + prazo corrido)',
        _AC: 'Limite ITSM superior ao contratual bruto',
        _AD: 'No prazo (data limite ITSM)',
        _AE: 'No prazo (controle contratual bruto)',
        _AF: 'Atraso vs. limite ITSM (min)',
        _AG: 'Duração criação→resolução (dias)',
        _AH: 'Ordem — fora do prazo',
        _AI: 'Ordem — limite ITSM superior ao contratual bruto',
        _AJ: 'Situação dos dados',
    }
    note = sheet.cell(
        row=1,
        column=_R - 1,
        value=(
            f'ÁREA DE APOIO — base de dados para as fórmulas desta aba '
            f'({len(rows)} incidentes do CSV bruto). Não excluir.'
        ),
    )
    note.font = NOTE_FONT
    for col, text in headers.items():
        cell = sheet.cell(row=1, column=col, value=text)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True)
        sheet.column_dimensions[cl(col)].width = 16

    for col, text in {
        _AK: 'Grupo executor',
        _AL: 'Nível',
        _AM: 'Categoria',
    }.items():
        cell = sheet.cell(row=1, column=col, value=text)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        sheet.column_dimensions[cl(col)].width = 30
    for idx, (grupo, nivel, categoria) in enumerate(grupo_rows, start=2):
        sheet.cell(
            row=idx, column=_AK, value=safe_excel_text(grupo)
        ).font = BODY_FONT
        sheet.cell(
            row=idx, column=_AL, value=safe_excel_text(nivel)
        ).font = BODY_FONT
        sheet.cell(
            row=idx, column=_AM, value=safe_excel_text(categoria)
        ).font = BODY_FONT
    map_range = f'${cl(_AK)}$2:${cl(_AM)}${1 + len(grupo_rows)}'

    for i, row in enumerate(rows, start=2):
        num_solicitacao = safe_excel_text(row[_NUM_SOLICITACAO_COLUMN])
        sheet.cell(row=i, column=_R, value=num_solicitacao).font = BODY_FONT
        sheet.cell(
            row=i, column=_S, value=safe_excel_text(row[GRUPO_EXECUTOR_COLUMN])
        ).font = BODY_FONT
        sheet.cell(
            row=i, column=_T, value=safe_excel_text(row[_ATIVIDADE_COLUMN])
        ).font = BODY_FONT

        solicitacao = parse_dt(row[_DATA_SOLICITACAO_COLUMN])
        limite = parse_dt(row[_DATA_LIMITE_COLUMN])
        fim = parse_dt(row[_DATA_FIM_COLUMN])

        u_cell = sheet.cell(row=i, column=_U, value=solicitacao.value)
        u_cell.number_format = _DATETIME_FMT
        v_cell = sheet.cell(row=i, column=_V, value=limite.value)
        v_cell.number_format = _DATETIME_FMT
        w_cell = sheet.cell(row=i, column=_W, value=fim.value)
        w_cell.number_format = _DATETIME_FMT
        for c in (u_cell, v_cell, w_cell):
            c.font = BODY_FONT

        if solicitacao.is_malformed:
            qualidade = 'Data de abertura inválida'
        elif solicitacao.is_blank:
            qualidade = 'Data de abertura ausente'
        elif limite.is_malformed:
            qualidade = 'Data limite inválida'
        elif fim.is_malformed:
            qualidade = 'Data de encerramento inválida'
        elif (
            solicitacao.value is not None
            and fim.value is not None
            and fim.value < solicitacao.value
        ):
            qualidade = 'Encerramento anterior à abertura'
        else:
            qualidade = _DATA_QUALIDADE_OK
        aj_cell = sheet.cell(row=i, column=_AJ, value=qualidade)
        aj_cell.font = BODY_FONT
        if qualidade != _DATA_QUALIDADE_OK:
            aj_cell.fill = RED_FILL

        sheet.cell(
            row=i, column=_X, value=safe_excel_text(row[_NO_PRAZO_COLUMN])
        ).font = BODY_FONT
        sheet.cell(
            row=i, column=_Y, value=safe_excel_text(row[_TECNICO_COLUMN])
        ).font = BODY_FONT

        sc, uc, vc, wc, abc = (
            f'{cl(_S)}{i}',
            f'{cl(_U)}{i}',
            f'{cl(_V)}{i}',
            f'{cl(_W)}{i}',
            f'{cl(_AB)}{i}',
        )
        z_cell = sheet.cell(
            row=i,
            column=_Z,
            value=f'=IFERROR(VLOOKUP({sc},{map_range},2,FALSE),"")',
        )
        aa_cell = sheet.cell(
            row=i,
            column=_AA,
            value=f'=IFERROR(VLOOKUP({sc},{map_range},3,FALSE),"")',
        )
        ab_cell = sheet.cell(
            row=i,
            column=_AB,
            value=f'=IF({uc}="","",{uc}+{_PRAZO_HORAS_CORRIDAS}/24)',
        )
        ab_cell.number_format = _DATETIME_FMT
        ac_cell = sheet.cell(
            row=i,
            column=_AC,
            value=(
                f'=IF(OR({uc}="",{vc}=""),"Requer análise",'
                f'IF({vc}>{abc}+{PRAZO_TOLERANCIA_MINUTOS}/1440,"Sim","Não"))'
            ),
        )
        ad_cell = sheet.cell(
            row=i,
            column=_AD,
            value=f'=IF(OR({wc}="",{vc}=""),"",IF({wc}<={vc},"S","N"))',
        )
        ae_cell = sheet.cell(
            row=i,
            column=_AE,
            value=f'=IF(OR({wc}="",{abc}=""),"",IF({wc}<={abc},"S","N"))',
        )
        af_cell = sheet.cell(
            row=i,
            column=_AF,
            value=f'=IF(AND({wc}<>"",{vc}<>"",{wc}>{vc}),({wc}-{vc})*1440,0)',
        )
        af_cell.number_format = '0.0'
        ag_cell = sheet.cell(
            row=i,
            column=_AG,
            value=f'=IF(OR({uc}="",{wc}="",{wc}<{uc}),"",{wc}-{uc})',
        )
        ag_cell.number_format = _DUR
        for c in (z_cell, aa_cell, ac_cell, ad_cell, ae_cell):
            c.font = BODY_FONT

        xc, acc = f'{cl(_X)}{i}', f'{cl(_AC)}{i}'
        ah_cell = sheet.cell(
            row=i, column=_AH, value=f'=IF({xc}="N",COUNTIF($X$2:{xc},"N"),"")'
        )
        ai_cell = sheet.cell(
            row=i,
            column=_AI,
            value=f'=IF({acc}="Sim",COUNTIF($AC$2:{acc},"Sim"),"")',
        )
        for c in (ah_cell, ai_cell):
            c.font = BODY_FONT

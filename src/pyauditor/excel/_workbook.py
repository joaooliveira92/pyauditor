"""Utilitários de workbook genéricos (não específicos de nenhum INMS),
reusáveis por qualquer renderer de `excel/` que grave fórmulas, tabelas
nativas ou abas em múltiplas etapas: `force_recalc()` garante que um leitor
externo (openpyxl `data_only=True`, conversor headless, serviço de extração)
recalcule as fórmulas em vez de mostrar cache velho; `unique_table_name()`
evita colisão de nome de `Table` quando o mesmo renderer é chamado mais de uma
vez no mesmo workbook (nomes de tabela são únicos por pasta de trabalho, não
por aba); `create_sheet_atomic()` evita deixar uma aba parcialmente escrita no
workbook se a escrita falhar no meio do caminho.
"""

from __future__ import annotations

from collections.abc import Iterator
from contextlib import contextmanager

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet


def force_recalc(workbook: Workbook) -> None:
    """Força recálculo completo na próxima abertura — sem isso, um leitor
    que não recalcula (ex.: `openpyxl(data_only=True)` sem o arquivo ter
    sido reaberto/recalculado por Excel/LibreOffice antes) pode devolver
    células vazias ou com cache desatualizado para fórmulas gravadas aqui."""
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = 'auto'


def unique_table_name(workbook: Workbook, base_name: str) -> str:
    """Devolve `base_name` se ainda não usado por nenhuma tabela do
    workbook; caso contrário, sufixa `_2`, `_3`, ... até encontrar um nome
    livre. Nomes de `Table` do openpyxl precisam ser únicos na pasta de
    trabalho inteira, não só na aba — reusar este renderer mais de uma vez
    no mesmo workbook, sem isso, faz a criação da segunda tabela falhar."""
    existing = {
        table.name
        for sheet in workbook.worksheets
        for table in sheet.tables.values()
    }
    if base_name not in existing:
        return base_name
    suffix = 2
    while f'{base_name}_{suffix}' in existing:
        suffix += 1
    return f'{base_name}_{suffix}'


@contextmanager
def create_sheet_atomic(
    workbook: Workbook, sheet_name: str
) -> Iterator[Worksheet]:
    """Cria a aba `sheet_name` e a remove do workbook se qualquer exceção
    escapar do bloco `with` — evita deixar uma aba parcialmente preenchida
    (criada, mas com escrita interrompida por um erro) no arquivo final."""
    sheet = workbook.create_sheet(sheet_name)
    try:
        yield sheet
    except Exception:
        workbook.remove(sheet)
        raise

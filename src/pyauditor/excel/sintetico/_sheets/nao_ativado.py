"""Aba "não ativado" do sintetico (spec §14.4) — extraída de
`excel/sintetico/workbook.py` (ticket 04 SRP).
"""

from __future__ import annotations

from openpyxl import Workbook

from pyauditor.excel._style import LABEL_FONT
from pyauditor.excel.sintetico._sheets._shared import COLUMNS, NAO_ATIVADO_TEXT


def _write_nao_ativado_sheet(workbook: Workbook, sheet_name: str) -> None:
    sheet = workbook.create_sheet(sheet_name)
    sheet.sheet_view.showGridLines = False
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(COLUMNS))
    cell = sheet.cell(row=2, column=1, value=NAO_ATIVADO_TEXT)
    cell.font = LABEL_FONT

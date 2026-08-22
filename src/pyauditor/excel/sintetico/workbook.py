"""`sintetico.xlsx` (spec §14.4, ticket 05) — um workbook por órgão/
competência, uma aba por INMS com entrada em `categorias.yaml`, mais as abas
verbatim de `input/{capa,equipe,prazos}.csv`.

Este módulo é o **dispatcher**: decide o renderer por INMS (config herdada,
colunas do CSV, shape), acumula warnings e executa o `atomic_write` final.
Os renderers por-shape vivem em `excel/sintetico/_sheets/`, os verbatim em
`_sheets/verbatim_sheets.py`, e a aritmética em `_stats.py` (ticket 04 SRP).

Contagens são brutas/pré-quality-gate — conferência rápida, não substitui o
ROM da categoria. A única exceção é `Tempo médio criação→resolução`,
restrito às linhas aprovadas pelo quality gate, conforme a spec.

INMS 1.1 tem renderer dedicado em `excel/inms_1_1_audit.py` (fachada do
pacote `inms_1_1/`): resumo executivo, memória de cálculo, incidentes fora
do prazo e auditoria do prazo contratual via fórmulas do Excel sobre os
dados brutos embutidos na própria aba. Só entra em jogo quando o CSV bruto
tem as colunas de detalhe de produção (`Nº Solicitacao`, `Atividades`,
`DataHoraLimite`, `TecnicoExecutor` — ver
`inms_1_1_audit.has_required_columns`).
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Final

from openpyxl import Workbook

from pyauditor.atomic_write import atomic_write
from pyauditor.categoria_filter import GRUPO_EXECUTOR_COLUMN, base_config_stem
from pyauditor.config.categorias import (
    CategoriasFile,
    GrupoExecutorMode,
    WholeIndicatorMode,
)
from pyauditor.config.manifest import DatasetManifest
from pyauditor.config.models import (
    PrecomputedTableCalculation,
    RatioCalculation,
)
from pyauditor.engine.pipeline import load_config, measurement_source
from pyauditor.excel import inms_1_1_audit
from pyauditor.excel.equipe import EQUIPE_DELIMITER, EQUIPE_ENCODING
from pyauditor.excel.prazos import (
    PRAZOS_DELIMITER,
    PRAZOS_ENCODING,
    PRAZOS_SHEET_NAME,
)
from pyauditor.excel.sintetico._sheets._shared import (
    CAPA_SHEET_NAME,
    EQUIPE_SHEET_NAME,
)
from pyauditor.excel.sintetico._sheets.grupo_executor import (
    _write_grupo_executor_sheet,
    _write_whole_indicator_sheet,
)
from pyauditor.excel.sintetico._sheets.multi_ativo import (
    _write_multi_ativo_sheet,
)
from pyauditor.excel.sintetico._sheets.nao_ativado import (
    _write_nao_ativado_sheet,
)
from pyauditor.excel.sintetico._sheets.precomputed import (
    _write_precomputed_table_sheet,
)
from pyauditor.excel.sintetico._sheets.ratio_aggregate import (
    _write_ratio_aggregate_sheet,
)
from pyauditor.excel.sintetico._sheets.verbatim_sheets import (
    _write_capa_sheet,
    _write_csv_verbatim_sheet,
)
from pyauditor.periodo import PeriodoAfericao

__all__: Final[tuple[str, ...]] = ('write_sintetico_workbook',)

_INMS_1_1: Final[str] = '1.1'
_INMS_1_14: Final[str] = '1.14'


def write_sintetico_workbook(
    categorias_file: CategoriasFile,
    config_dir: Path,
    competencia_data_dir: Path,
    output_path: Path,
    *,
    manifest: DatasetManifest | None = None,
    periodo: PeriodoAfericao | None = None,
    strict: bool = False,
    prazos_path: Path | None = None,
    capa_path: Path | None = None,
    equipe_path: Path | None = None,
    objetos_path: Path | None = None,
    generated_at: datetime | None = None,
) -> list[str]:
    """Constrói e grava `sintetico.xlsx` de um órgão/competência. Devolve
    warnings (nunca lança por causa do problema de um único INMS — um
    config ruim ou um dataset genuinamente ausente pula/degrada a aba
    daquele INMS em vez de derrubar o workbook inteiro).

    `generated_at` (default `datetime.now()`, resolvido uma única vez aqui)
    é repassado à aba enriquecida do INMS 1.1 — injetável para permitir
    teste determinístico sem depender do relógio real (ticket 18 / B-01)."""
    warnings: list[str] = []
    generated_at = generated_at if generated_at is not None else datetime.now()

    per_inms: dict[
        str, list[tuple[str, GrupoExecutorMode | WholeIndicatorMode]]
    ] = {}
    for categoria_key, categoria in categorias_file.categorias.items():
        for inms_key, entry in categoria.inms.items():
            per_inms.setdefault(inms_key, []).append((categoria_key, entry))

    workbook = Workbook()
    default_sheet = workbook.active
    if default_sheet is None:
        raise RuntimeError('workbook novo sem aba ativa (openpyxl)')
    workbook.remove(default_sheet)

    if capa_path is not None:
        _write_capa_sheet(workbook, capa_path, objetos_path, warnings)
    elif objetos_path is not None:
        warnings.append(
            f'sintetico.xlsx: '
            f'capa_path '
            f'não '
            f'informado '
            f'— '
            f'dados '
            f'de '
            f'{objetos_path} '
            f"não anexados (dependem da aba '{CAPA_SHEET_NAME}')"
        )
    if equipe_path is not None:
        _write_csv_verbatim_sheet(
            workbook,
            EQUIPE_SHEET_NAME,
            equipe_path,
            delimiter=EQUIPE_DELIMITER,
            encoding=EQUIPE_ENCODING,
            warnings=warnings,
        )
    if prazos_path is not None:
        _write_csv_verbatim_sheet(
            workbook,
            PRAZOS_SHEET_NAME,
            prazos_path,
            delimiter=PRAZOS_DELIMITER,
            encoding=PRAZOS_ENCODING,
            warnings=warnings,
        )
    sheets_before_inms = set(workbook.sheetnames)

    for inms_key in sorted(per_inms, key=lambda k: int(k.split('.')[1])):
        entries = per_inms[inms_key]
        sheet_name = f'INMS {inms_key}'

        try:
            base_stem = base_config_stem(inms_key)
            base_config = load_config(config_dir / f'{base_stem}.yaml')
        except (OSError, ValueError) as exc:
            warning = (
                f'sintetico.xlsx:INMS{inms_key}:falhaaocarregarconfigbase:{exc}'
            )
            warnings.append(warning)
            continue

        # Backbone (ticket 03): resolve→valida→lê→filtra→gates, um só lugar
        # para os quatro reimplementadores do pipeline de medição. Sintetico
        # nunca emite WARN de janela vazia (a mesma passada do bruto coube
        # ao split) — `emit_period_filter_logs=False`. `period_column`
        # ausente/não presente no header é erro acionável (spec §2 ponto 1,
        # issue 01 item 5) — degrada esta aba como os demais erros do loop.
        try:
            bundle = measurement_source(
                base_config,
                competencia_data_dir,
                manifest,
                config_path=config_dir / f'{base_stem}.yaml',
                periodo=periodo,
                strict=strict,
                emit_period_filter_logs=False,
            )
        except FileNotFoundError:
            _write_nao_ativado_sheet(workbook, sheet_name)
            continue
        except (OSError, ValueError) as exc:
            warnings.append(f'sintetico.xlsx: INMS {inms_key}: {exc}')
            continue

        raw_csv_path = bundle.csv_path
        fieldnames = bundle.fieldnames
        rows = bundle.rows
        accepted_ids = bundle.accepted_ids

        whole_indicator_entries = [
            (ck, e) for ck, e in entries if isinstance(e, WholeIndicatorMode)
        ]

        if inms_key == _INMS_1_14:
            calculation = base_config.calculation
            if not isinstance(calculation, PrecomputedTableCalculation) or (
                calculation.name_column is None
            ):
                warning = (
                    f'sintetico.xlsx: INMS {inms_key}: config base não é '
                    "'precomputed_table' com 'name_column' — aba não gerada"
                )
                warnings.append(warning)
                continue
            _write_multi_ativo_sheet(
                workbook,
                sheet_name,
                categorias_file,
                whole_indicator_entries,
                calculation.name_column,
                fieldnames,
                rows,
                accepted_ids,
            )
            continue

        grupo_executor_entries = [
            (ck, e) for ck, e in entries if isinstance(e, GrupoExecutorMode)
        ]

        if grupo_executor_entries:
            if GRUPO_EXECUTOR_COLUMN not in fieldnames:
                warning = (
                    f'sintetico.xlsx: INMS {inms_key}: {raw_csv_path} não tem '
                    f'coluna '
                    f"'{GRUPO_EXECUTOR_COLUMN}' — aba não gerada"
                )
                warnings.append(warning)
                continue
            if (
                inms_key == _INMS_1_1
                and base_config.target is not None
                and base_config.penalty is not None
                and inms_1_1_audit.has_required_columns(fieldnames)
            ):
                # Aba enriquecida (resumo executivo, memória de cálculo,
                # fora-do-prazo, auditoria de prazo) — só quando o CSV bruto
                # tem as colunas de detalhe (produção real, ambos os
                # órgãos); CSVs minimalistas (ex. fixtures de teste) caem no
                # renderer genérico abaixo, sem degradar silenciosamente.
                try:
                    inms_1_1_audit.write_sheet(
                        workbook,
                        sheet_name,
                        categorias_file=categorias_file,
                        grupo_executor_entries=grupo_executor_entries,
                        whole_indicator_entries=whole_indicator_entries,
                        fieldnames=fieldnames,
                        rows=rows,
                        target_operator=base_config.target.operator,
                        target_value=base_config.target.value,
                        penalty_base_points=base_config.penalty.base_points,
                        penalty_step_points=base_config.penalty.step_points,
                        penalty_step_size_pct=base_config.penalty.step_size_pct,
                        contract=base_config.scope.contract,
                        periodo=periodo,
                        raw_csv_path=raw_csv_path,
                        generated_at=generated_at,
                    )
                except ValueError as exc:
                    # Falha de validação/qualidade de dado específica da aba
                    # enriquecida não deve derrubar o restante do workbook —
                    # degrada para o renderer genérico, como as demais
                    # falhas por-INMS deste loop.
                    warning = (
                        f'sintetico.xlsx: INMS {inms_key}: falha ao gerar aba '
                        f'enriquecida '
                        f'({exc}) — usando renderer genérico'
                    )
                    warnings.append(warning)
                    _write_grupo_executor_sheet(
                        workbook,
                        sheet_name,
                        categorias_file,
                        grupo_executor_entries,
                        whole_indicator_entries,
                        fieldnames,
                        rows,
                        accepted_ids,
                    )
            else:
                _write_grupo_executor_sheet(
                    workbook,
                    sheet_name,
                    categorias_file,
                    grupo_executor_entries,
                    whole_indicator_entries,
                    fieldnames,
                    rows,
                    accepted_ids,
                )
        elif isinstance(base_config.calculation, PrecomputedTableCalculation):
            if base_config.target is None:
                raise ValueError(
                    'precomputed exige `target` no sintetico'
                )
            _write_precomputed_table_sheet(
                workbook,
                sheet_name,
                categorias_file,
                whole_indicator_entries,
                base_config.calculation,
                base_config.target.operator,
                base_config.target.value,
                rows,
            )
        elif (
            isinstance(base_config.calculation, RatioCalculation)
            and base_config.calculation.aggregation == 'sum'
            and base_config.calculation.sum_numerator_subtract_column
            is not None
            and base_config.calculation.denominator_filter is not None
        ):
            if base_config.target is None:
                raise ValueError(
                    'ratio_aggregate exige `target` no sintetico'
                )
            _write_ratio_aggregate_sheet(
                workbook,
                sheet_name,
                categorias_file,
                whole_indicator_entries,
                base_config.calculation,
                base_config.target.operator,
                base_config.target.value,
                rows,
            )
        else:
            _write_whole_indicator_sheet(
                workbook,
                sheet_name,
                categorias_file,
                whole_indicator_entries,
                fieldnames,
                rows,
                accepted_ids,
            )

    if set(workbook.sheetnames) == sheets_before_inms:
        # Nada pôde ser medido (todo INMS de categorias.yaml falhou ao
        # carregar/resolver) — um xlsx sem nenhuma aba de INMS não é um
        # arquivo válido pra este propósito; não sobrescreve um
        # sintetico.xlsx de rerun anterior com um arquivo vazio/quebrado.
        return warnings

    atomic_write(output_path, workbook.save)
    return warnings

"""Builds the consolidated financial workbook for `pyauditor consolidate`
(2.1): `CAPA_E_CONTROLE` + `SERVICOS_POR_ORGAO` + `INMS_BASE` + `GLOSAS` +
`CALCULO_PAGAMENTO` — the núcleo financeiro decided in
.scratch/multi-org-pipeline tickets 01/02/04.

`consolidate` never re-runs `measure`/`report`: its precondition is that
`reports/relatorio_<comp>_MinC.xlsx` and `_MTur.xlsx` already exist (ticket
04 Q1) — mas este módulo nunca abre esses ``.xlsx`` (capa vem de
``capa.csv``, indicadores vêm dos ROM JSON ``roms/<orgao>/<comp>/*.json``),
que são o mesmo artefato já computado que ``report.py`` consumiu.
A aba ``INMS_BASE`` do report deixa colunas de glosa em branco por
indicador (spec §12: glosa é agregado mensal lá), então não serviria para
o detalhe por-(indicador x órgão) que ``GLOSAS`` precisa.
"""

from __future__ import annotations

from dataclasses import dataclass
from typing import Final

from openpyxl import Workbook
from openpyxl.styles import Border, Font, Side

from pyauditor.codes import contractual_sort_key, format_inms_code
from pyauditor.excel._style import (
    BODY_FONT,
    BOTTOM_BORDER,
    HEADER_FILL,
    HEADER_FONT,
    LABEL_FONT,
    LEFT_ALIGN,
    TITLE_FONT,
    CellValue,
)
from pyauditor.excel._style import UNIT_BY_SHAPE as _UNIT_BY_SHAPE
from pyauditor.excel._style import new_sheet as _new_sheet
from pyauditor.excel._style import write_row as _write
from pyauditor.excel.consolidate._decisions_io import (
    RowKey,
    read_existing_decisions,
)
from pyauditor.excel.consolidate._glosa_calcs import (
    accumulate_pontos_por_orgao,
    compute_aggregation,
    is_amnestied,
)
from pyauditor.excel.equipe import RESPONSAVEL_LABELS
from pyauditor.excel.glosas import (
    CAP_PCT,
    POINTS_TO_PERCENT,
    Historico,
    compute_glosa,
)
from pyauditor.excel.inms_base import inms_base_fields
from pyauditor.excel.orgao_consolidation import with_orgao_consolidation
from pyauditor.logging import logger
from pyauditor.periodo import PeriodoAfericao, format_date_br
from pyauditor.rom.dedup import deduplicate_summaries
from pyauditor.rom.summary import IndicatorSummary

__all__ = (
    'ConsolidationResult',
    'build_consolidated_workbook',
    'build_glosas',
    'build_inms_base',
    'read_existing_decisions',
)

CAPA_SHEET: Final = 'CAPA_E_CONTROLE'
SERVICOS_SHEET: Final = 'SERVICOS_POR_ORGAO'
INMS_BASE_SHEET: Final = 'INMS_BASE'
GLOSAS_SHEET: Final = 'GLOSAS'
CALCULO_SHEET: Final = 'CALCULO_PAGAMENTO'

RATEIO_PADRAO: Final = 0.5  # provisório, até fonte oficial (ticket 01/02)

# docs/styleguide.md number formats — currency and percent, zero as "-".
# `_style.py` doesn't carry these yet (production report.py/capa.py don't
# apply them either); local to this module until that becomes a shared need.
_CURRENCY_FMT: Final = 'R$#,##0.00;R$#,##0.00;-'
# `result_pct`/`%Ajuste` are already stored in percent-space (95.5 meaning
# "95.5%"); the literal "%" symbol auto-multiplies by 100 on display, so it
# must be escaped for these. Only true 0-1 fractions (rateio) want the real,
# auto-scaling "%" format.
_PERCENT_FMT_SCALED: Final = '0.00"%";0.00"%";-'
_PERCENT_FMT_FRACTION: Final = '0.00%;0.00%;-'
_TOP_BORDER: Final = Border(top=Side(style='thin', color='1F2937'))

# Decisão Fiscal — o fiscal aceita a justificativa do fornecedor (anistia: a
# ocorrência sai da base de pontos), ver `_glosa_calcs.is_amnestied`. Ticket 02.

# GLOSAS: uma linha por (indicador x órgão) + resumo agregado (ticket 02).
_GLOSAS_COLUMNS: Final[tuple[str, ...]] = (
    'Competência',
    'Órgão',
    'Item Contratual',
    'Serviço',
    'Indicador',
    'Resultado',
    'Meta',
    'Faixa de Descumprimento',
    'Percentual de Ajuste',
    'Valor Base',
    'Valor Glosa',
    'Reincidência',
    'Justificativa',
    'Número da Ocorrência',
    'Decisão Fiscal',
    'Observação do Gestor',
)
# Colunas de decisão do fiscal — nunca recalculadas, só preservadas no merge.
_DECISION_COLUMNS: Final[tuple[str, ...]] = (
    'Reincidência',
    'Justificativa',
    'Número da Ocorrência',
    'Decisão Fiscal',
    'Observação do Gestor',
)
_DECISION_TEXT_COLUMNS: Final[tuple[int, ...]] = tuple(
    _GLOSAS_COLUMNS.index(name) + 1 for name in _DECISION_COLUMNS
)

_CALCULO_COLUMNS: Final[tuple[str, ...]] = (
    'Componente',
    'MinC',
    'MTur',
    'Consolidado',
)
_FATOR_PCT_TEXTO: Final[str] = f'{POINTS_TO_PERCENT:g}'.replace('.', ',')
_CALCULO_LINHAS: Final[tuple[str, ...]] = (
    'Percentual de rateio',
    'Valor bruto (= mensal x rateio)',
    'Pontos de glosa',
    f'Valordaglosa(=MIN(pontosx{_FATOR_PCT_TEXTO},{CAP_PCT:g}%)/100xbruto)',
    'Outros ajustes',
    'Valor recomendado (= max(0, bruto - glosa - outros))',
)

# docs spreadsheet.md's SERVICOS_POR_ORGAO — 9 serviços contratuais fixos
# (ticket 01). Segregação/critério de rateio idênticos para os dois órgãos
# até o Termo de Referência dizer o contrário.
_SERVICOS: Final[tuple[tuple[str, str, str, str], ...]] = (
    ('Central de Serviços e Monitoramento', 'Sim', 'Sim', 'Sim'),
    ('Gerenciamento Técnico das Operações e Projetos', 'Sim', 'Sim', 'Sim'),
    ('Banco de Dados', 'Sim', 'Sim', 'Sim'),
    ('Aplicações, Virtualização e Computação em Nuvem', 'Sim', 'Sim', 'Sim'),
    ('Serviços Corporativos', 'Sim', 'Sim', 'Sim'),
    ('Armazenamento e Backup', 'Sim', 'Sim', 'Sim'),
    ('Redes', 'Sim', 'Sim', 'Sim'),
    ('Segurança da Informação', 'Sim', 'Sim', 'Sim'),
    ('DevOps', 'Sim', 'Sim', 'Sim'),
)

_CAPA_VALOR_LABELS: Final = ('Valor mensal vigente', 'Valor global anual')


@dataclass(frozen=True)
class ConsolidationResult:
    workbook: Workbook
    total_pontos: float
    glosa_final: float
    warnings: tuple[str, ...]
    glosa_calculada: bool


def build_capa(
    wb: Workbook,
    competencia: str,
    capa: dict[str, object],
    warnings: list[str],
    valor_base: float | None,
    *,
    periodo: PeriodoAfericao | None = None,
    responsaveis: dict[str, str] | None = None,
) -> None:
    """Embeds a consolidated CAPA_E_CONTROLE. The contract is shared, so the
    common fields come from `capa.csv` (ticket 07); the monetary value comes
    from `objetos.csv` (`valor_base`), never from a capa file. Competência/
    períodos vêm da CLI (`competencia`/`periodo`) e os 4 responsáveis, de
    `equipe.csv` (`responsaveis`) — spec competencia-cli-equipe §4/§6.
    """
    ws = wb.create_sheet(CAPA_SHEET, 0)
    ws.sheet_view.showGridLines = False

    ws['A1'] = 'Capa e controle — consolidado'
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:B1')

    ws['A3'] = 'Campo'
    ws['B3'] = 'Valor'
    for cell in (ws['A3'], ws['B3']):
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = LEFT_ALIGN

    linhas_periodo: tuple[tuple[str, object], ...] = (
        (
            ('Período inicial da aferição', format_date_br(periodo.inicio)),
            ('Período final da aferição', format_date_br(periodo.fim)),
        )
        if periodo is not None
        else ()
    )
    campos_equipe: dict[str, str] = responsaveis or {}
    linhas_responsaveis: tuple[tuple[str, object], ...] = tuple(
        (label, campos_equipe.get(label, '')) for label in RESPONSAVEL_LABELS
    )

    campos: tuple[tuple[str, object], ...] = (
        ('Número do contrato', capa.get('Número do contrato')),
        ('Processo SEI', capa.get('Processo SEI')),
        ('Empresa contratada', capa.get('Empresa contratada')),
        (
            'Órgãos contratantes',
            'Ministério da Cultura / Ministério do Turismo',
        ),
        ('Competência', competencia),
        *linhas_periodo,
        ('Valor mensal vigente', valor_base),
        (
            'Valor global anual',
            valor_base * 12 if valor_base is not None else None,
        ),
        *linhas_responsaveis,
    )
    for offset, (label, value) in enumerate(campos):
        row = 4 + offset
        label_cell = ws.cell(row=row, column=1, value=label)
        label_cell.font = LABEL_FONT
        label_cell.alignment = LEFT_ALIGN
        label_cell.border = BOTTOM_BORDER

        value_cell = ws.cell(row=row, column=2, value=value)
        value_cell.font = BODY_FONT
        value_cell.border = BOTTOM_BORDER
        if label in _CAPA_VALOR_LABELS:
            value_cell.number_format = _CURRENCY_FMT

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 42


def build_servicos(
    wb: Workbook, itens: tuple[float, ...] | None = None
) -> None:
    """SERVICOS_POR_ORGAO — os 9 serviços contratuais, agora com o valor
    mensal de cada item vindo de `objetos.csv` (ticket 07 Q4), mapeado pelo
    índice (os nomes divergem entre as fontes — o índice não)."""
    ws = _new_sheet(
        wb,
        SERVICOS_SHEET,
        (
            'Item',
            'Serviço',
            'Valor Mensal (R$)',
            'Prestado ao MinC?',
            'Prestado ao MTur?',
            'Segregação Obrigatória?',
            'Critério de Rateio',
        ),
    )
    for i, (nome, minc, mtur, seg) in enumerate(_SERVICOS, start=2):
        valor = (
            itens[i - 2] if itens is not None and i - 2 < len(itens) else None
        )
        _write(
            ws,
            i,
            (
                i - 1,
                nome,
                valor,
                minc,
                mtur,
                seg,
                'Chamados, ativos ou valor definido',
            ),
        )
        if valor is not None:
            ws.cell(row=i, column=3).number_format = _CURRENCY_FMT


_INMS_BASE_COLUMNS: Final[tuple[str, ...]] = (
    'Competência',
    'Item contratual',
    'Serviço',
    'Grupo operacional',
    'Código INMS',
    'Descrição',
    'Órgão',
    'Meta mínima ou máxima',
    'Sentido da meta',
    'Numerador',
    'Denominador',
    'Resultado calculado',
    'Unidade',
    'Conformidade',
    'Diferença para a meta',
)


def _inms_base_row(
    competencia: str, summary: IndicatorSummary
) -> tuple[CellValue, ...]:
    row = inms_base_fields(summary, competencia, grupo_operacional=None)
    return (
        row.competencia,
        None,
        row.servico,
        row.grupo_operacional,
        row.codigo_inms,
        row.descricao,
        row.orgao,
        row.meta,
        row.sentido,
        row.numerador,
        row.denominador,
        row.resultado,
        row.unidade,
        row.conformidade,
        row.diferenca,
    )


def build_inms_base(
    wb: Workbook,
    competencia: str,
    minc: list[IndicatorSummary],
    mtur: list[IndicatorSummary],
) -> None:
    """Pooling Σnum/Σden por indicador (ticket 02) — reusa a mesma regra que
    já era testada dentro de `report.py`, agora vivendo só aqui.
    """
    ws = _new_sheet(wb, INMS_BASE_SHEET, _INMS_BASE_COLUMNS, width=20)
    rows = with_orgao_consolidation(minc + mtur)
    rows.sort(
        key=lambda s: (
            contractual_sort_key(s.contractual_id),
            s.asset or '',
            s.orgao,
        )
    )
    for row_idx, summary in enumerate(rows, start=2):
        _write(ws, row_idx, _inms_base_row(competencia, summary))
        # Percentuais: Meta / Resultado calculado / Diferença para a meta.
        if _UNIT_BY_SHAPE.get(summary.shape) == '%':
            for col in (8, 12, 15):
                ws.cell(
                    row=row_idx, column=col
                ).number_format = _PERCENT_FMT_SCALED


def _decision_value(decision: dict[str, object], key: str) -> CellValue:
    """`decision`'s values come from openpyxl cells (any scalar type it
    supports) — narrowed to what a fresh cell can hold, coercing anything
    unexpected (dates, etc.) to its string form rather than dropping it.
    """
    value = decision.get(key)
    if value is None or isinstance(value, str | int | float):
        return value
    return str(value)


def _faixa(summary: IndicatorSummary) -> str:
    if summary.target_operator is None or summary.target_value is None:
        return (
            'Ocorrência sob detalhamento por-ativo'
            if summary.penalty_points > 0
            else ''
        )
    dif = (
        summary.target_value - summary.result_pct
        if summary.target_operator == '>='
        else summary.result_pct - summary.target_value
    )
    return f'Déficit de {dif:.2f}pp' if dif > 0 else 'Não conforme'


def build_glosas(
    wb: Workbook,
    competencia: str,
    minc: list[IndicatorSummary],
    mtur: list[IndicatorSummary],
    valor_base: float | None,
    existing_decisions: dict[RowKey, dict[str, object]],
    warnings: list[str],
    *,
    historico: Historico | None = None,
    is_final_month: bool = False,
) -> tuple[float, float]:
    """GLOSAS — uma linha por (indicador x órgão) com ocorrência de glosa,
    mais o resumo agregado. Decisão do fiscal ('Aceita' = anistia) tira a
    ocorrência da base de pontos; colunas de decisão são preservadas do
    workbook anterior, nunca recalculadas (ticket 02/04 Q3).

    Reusa ``glosas.compute_glosa`` por-órgão (mesma fórmula de ``report.py``) —
    com ``saldo_anterior_pct`` e ``is_final_month`` — e soma contra teto
    por-órgão em vez de teto único sobre o agregado.
    """
    ws = _new_sheet(wb, GLOSAS_SHEET, _GLOSAS_COLUMNS, width=26)
    row = 2

    # Aritmética (dedup, acúmulo de pontos por órgão, agregação financeira)
    # vive em `_glosa_calcs` — pura, sem planilha (ticket 04 SRP).
    pontos_por_orgao, seen_keys = accumulate_pontos_por_orgao(
        minc,
        mtur,
        existing_decisions,
    )

    # Primeiro escreve linhas por ocorrência (valor por ocorrência continua pct
    # simples,
    # mas o agregado usa compute_glosa por-órgão)
    for summary in deduplicate_summaries(minc) + deduplicate_summaries(mtur):
        if summary.penalty_points <= 0:
            continue
        key: RowKey = (format_inms_code(summary.contractual_id), summary.orgao)
        decision = existing_decisions.get(key, {})
        amnestied = is_amnestied(decision)

        pontos = summary.penalty_points
        pct = pontos * POINTS_TO_PERCENT
        valor_glosa = (
            round((valor_base or 0.0) * pct / 100, 2)
            if valor_base is not None
            else None
        )

        _write(
            ws,
            row,
            (
                competencia,
                summary.orgao,
                '',
                '',
                format_inms_code(summary.contractual_id),
                round(summary.result_pct, 2),
                summary.target_value,
                _faixa(summary),
                round(pct, 4),
                valor_base,
                0.0 if amnestied else valor_glosa,
                _decision_value(decision, 'Reincidência'),
                _decision_value(decision, 'Justificativa'),
                _decision_value(decision, 'Número da Ocorrência'),
                _decision_value(decision, 'Decisão Fiscal'),
                _decision_value(decision, 'Observação do Gestor'),
            ),
        )
        ws.cell(row=row, column=6).number_format = _PERCENT_FMT_SCALED
        ws.cell(row=row, column=7).number_format = _PERCENT_FMT_SCALED
        ws.cell(row=row, column=9).number_format = _PERCENT_FMT_SCALED
        if valor_base is not None:
            ws.cell(row=row, column=10).number_format = _CURRENCY_FMT
            ws.cell(row=row, column=11).number_format = _CURRENCY_FMT
        for column in _DECISION_TEXT_COLUMNS:
            ws.cell(row=row, column=column).number_format = '@'
        row += 1

    for key in existing_decisions:
        if key not in seen_keys:
            warnings.append(
                f'decisão registrada para {key[0]}/{key[1]} mas a ocorrência '
                f'não existe mais '
                'nesta rodada — preservada apenas no histórico, não reescrita'
            )

    aggregation = compute_aggregation(
        pontos_por_orgao=pontos_por_orgao,
        valor_base=valor_base,
        competencia=competencia,
        historico=historico,
        is_final_month=is_final_month,
    )
    total_pontos = aggregation.total_pontos
    glosa_final = aggregation.glosa_final
    pct_bruto = aggregation.pct_bruto
    aplicado = aggregation.aplicado

    r = row + 2
    for label, value in (
        ('Total de Pontos', round(total_pontos, 2)),
        (f'Fórmula (pontos x {_FATOR_PCT_TEXTO})', f'{pct_bruto:.4f}%'),
        ('Limite', f'{CAP_PCT:.1f}%'),
        ('Percentual Aplicado', f'{aplicado:.2f}%'),
        (
            'Valor Glosa',
            round(glosa_final, 2) if valor_base is not None else None,
        ),
    ):
        label_cell = ws.cell(row=r, column=1, value=label)
        value_cell = ws.cell(row=r, column=2, value=value)
        bold = label in ('Total de Pontos', 'Valor Glosa')
        label_cell.font = Font(bold=True) if bold else BODY_FONT
        value_cell.font = Font(bold=True) if bold else BODY_FONT
        if label == 'Total de Pontos':
            label_cell.border = _TOP_BORDER
            value_cell.border = _TOP_BORDER
        if label == 'Valor Glosa' and valor_base is not None:
            value_cell.number_format = _CURRENCY_FMT
        r += 1

    return total_pontos, glosa_final


def _glosa_bruto(pontos: float, bruto: float) -> float:
    """Valor da glosa sobre um bruto — a mesma aritmética de
    ``glosas.compute_glosa`` (fonte única, ticket 09), sem rollover
    (``is_final_month=True``) porque as células de ``CALCULO_PAGAMENTO``
    computam o ajuste do mês corrente, não o saldo rolado."""
    glosa = compute_glosa(pontos, bruto, is_final_month=True)
    return glosa.valor_da_glosa or 0.0


def build_calculo(
    wb: Workbook,
    valor_base: float | None,
    total_pontos: float,
    rateio_minc: float = RATEIO_PADRAO,
    rateio_mtur: float = RATEIO_PADRAO,
) -> None:
    """CALCULO_PAGAMENTO — espelha a inspiração: colunas MinC/MTur/
    Consolidado, glosa só na coluna Consolidado (ticket 02)."""
    ws = wb.create_sheet(CALCULO_SHEET)
    ws.sheet_view.showGridLines = False

    ws['A1'] = 'Parâmetros de Entrada — rateio PROVISÓRIO até fonte oficial'
    base = valor_base or 0.0
    params: tuple[tuple[str, float, str], ...] = (
        ('Valor mensal vigente', base, _CURRENCY_FMT),
        ('Limite máximo de glosa (%)', CAP_PCT, _PERCENT_FMT_SCALED),
        ('Rateio MinC (provisório)', rateio_minc, _PERCENT_FMT_FRACTION),
        ('Rateio MTur (provisório)', rateio_mtur, _PERCENT_FMT_FRACTION),
    )
    for row, (param_label, param_value, param_fmt) in enumerate(
        params, start=3
    ):
        ws.cell(row=row, column=1, value=param_label).font = Font(bold=True)
        value_cell = ws.cell(row=row, column=2, value=param_value)
        value_cell.font = BODY_FONT
        value_cell.number_format = param_fmt

    header_row = 9
    for i, col_name in enumerate(_CALCULO_COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=i, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    ws.column_dimensions['A'].width = 58

    colunas = (
        ('MinC', rateio_minc, 0.0),
        ('MTur', rateio_mtur, 0.0),
        ('Consolidado', 1.0, total_pontos),
    )
    total_row_idx = len(_CALCULO_LINHAS) - 1

    for idx, label in enumerate(_CALCULO_LINHAS):
        row = header_row + 1 + idx
        bold = idx == total_row_idx
        label_cell = ws.cell(row=row, column=1, value=label)
        label_cell.font = Font(bold=True) if bold else BODY_FONT
        if bold:
            label_cell.border = _TOP_BORDER
        fmt: str | None = None
        if label.startswith('Percentual de rateio'):
            fmt = _PERCENT_FMT_FRACTION
        elif label.startswith(
            (
                'Valor bruto',
                'Valor da glosa',
                'Outros ajustes',
                'Valor recomendado',
            )
        ):
            fmt = _CURRENCY_FMT
        for col, (_nome, rateio, pontos) in zip(
            range(2, len(_CALCULO_COLUMNS) + 1), colunas, strict=True
        ):
            if label.startswith('Percentual de rateio'):
                value: object = rateio
            elif label.startswith('Valor bruto'):
                value = round(base * rateio, 2)
            elif label.startswith('Pontos de glosa'):
                value = round(pontos, 2)
            elif label.startswith('Valor da glosa'):
                value = round(_glosa_bruto(pontos, base * rateio), 2)
            elif label.startswith('Outros ajustes'):
                value = 0
            else:
                bruto = base * rateio
                glosa = _glosa_bruto(pontos, bruto)
                value = round(max(0.0, bruto - glosa), 2)
            value_cell = ws.cell(row=row, column=col, value=value)
            value_cell.font = Font(bold=True) if bold else BODY_FONT
            if fmt:
                value_cell.number_format = fmt
            if bold:
                value_cell.border = _TOP_BORDER


def build_consolidated_workbook(
    competencia: str,
    minc: list[IndicatorSummary],
    mtur: list[IndicatorSummary],
    capa: dict[str, object],
    existing_decisions: dict[RowKey, dict[str, object]] | None = None,
    *,
    valor_base: float | None = None,
    itens: tuple[float, ...] | None = None,
    historico: Historico | None = None,
    is_final_month: bool = False,
    periodo: PeriodoAfericao | None = None,
    responsaveis: dict[str, str] | None = None,
) -> ConsolidationResult:
    """Pure, in-memory build of the 5-sheet consolidated workbook.

    `capa` carries the contract-common fields (from ``capa.csv``, ticket 07)
    e indicadores vêm dos ROM JSON (``roms/<orgao>/<comp>/*.json``), não do
    ``.xlsx`` de ``report.py``. ``valor_base`` e ``itens`` vêm de
    ``objetos.csv``. ``glosa_calculada`` é ``valor_base is not None``.
    ``periodo``/``responsaveis`` alimentam a capa (CLI + equipe.csv, §4/§6).
    """
    warnings: list[str] = []
    wb = Workbook()
    default_sheet = wb.active
    if default_sheet is None:
        raise RuntimeError('workbook novo sem aba ativa (openpyxl)')
    wb.remove(default_sheet)

    build_capa(
        wb,
        competencia,
        capa,
        warnings,
        valor_base,
        periodo=periodo,
        responsaveis=responsaveis,
    )
    build_servicos(wb, itens)
    build_inms_base(wb, competencia, minc, mtur)
    total_pontos, glosa_final = build_glosas(
        wb,
        competencia,
        minc,
        mtur,
        valor_base,
        existing_decisions or {},
        warnings,
        historico=historico,
        is_final_month=is_final_month,
    )
    build_calculo(wb, valor_base, total_pontos)

    for warning in warnings:
        logger.warning(warning)

    return ConsolidationResult(
        workbook=wb,
        total_pontos=total_pontos,
        glosa_final=glosa_final,
        warnings=tuple(warnings),
        glosa_calculada=valor_base is not None,
    )

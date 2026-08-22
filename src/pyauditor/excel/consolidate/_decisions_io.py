"""I/O das decisões fiscais gravadas no consolidado anterior (issue 08,
ticket 04 SRP): lê as colunas de decisão da aba `GLOSAS` de um workbook já
gerado, validando cabeçalhos hand-edited com erros acionáveis.
"""

from __future__ import annotations

from collections.abc import Sequence
from pathlib import Path
from typing import Final

from openpyxl import load_workbook

from pyauditor.codes import format_inms_code

__all__: Final[tuple[str, ...]] = ("RowKey", "read_existing_decisions")

GLOSAS_SHEET: Final[str] = "GLOSAS"

RowKey = tuple[str, str]  # (contractual_id, orgao)

# Colunas de decisão do fiscal — nunca recalculadas, só preservadas no merge.
_DECISION_COLUMNS: Final[tuple[str, ...]] = (
    "Reincidência",
    "Justificativa",
    "Número da Ocorrência",
    "Decisão Fiscal",
    "Observação do Gestor",
)

_TRACKED_HEADER_NAMES: Final[frozenset[str]] = frozenset({"Indicador", "Órgão", *_DECISION_COLUMNS})


def _normalize_header(name: str) -> str:
    """Normaliza espaços/acentos/case para detectar renomeação leve (issue 08)."""
    import unicodedata

    # strip, collapse internal whitespace, casefold, remove acentos
    collapsed = " ".join(name.strip().split())
    folded = collapsed.casefold()
    nfkd = unicodedata.normalize("NFD", folded)
    return "".join(c for c in nfkd if not unicodedata.combining(c))


def _check_no_duplicate_headers(path: Path, header_row: Sequence[object]) -> None:
    """A hand-edited workbook with a duplicate column name (two "Indicador"
    columns, or a stray column literally named "Justificativa" inserted by a
    fiscal) would otherwise make the header→index lookup silently keep the
    *last* matching column — decision values then get read from the wrong
    column with no error. Fail loudly instead."""
    seen: set[str] = set()
    duplicates: set[str] = set()
    for name in header_row:
        if isinstance(name, str) and name in _TRACKED_HEADER_NAMES:
            (duplicates if name in seen else seen).add(name)
    if duplicates:
        raise ValueError(
            f"{path}: coluna(s) duplicada(s) em {GLOSAS_SHEET!r}: "
            f"{', '.join(sorted(duplicates))} — planilha hand-edited em formato inesperado, "
            "corrija antes de rodar consolidate de novo"
        )


def _check_renamed_headers(path: Path, header_row: Sequence[object]) -> None:
    """Se um cabeçalho esperado não existe mas há candidato próximo por
    normalização (espaço extra, acento, case), falha nomeando ambos — em
    vez de perder a decisão silenciosamente (issue 08)."""
    present = {h for h in header_row if isinstance(h, str)}
    present_normalized = {_normalize_header(h): h for h in present}
    for expected in _TRACKED_HEADER_NAMES:
        if expected not in present:
            norm_expected = _normalize_header(expected)
            candidate = present_normalized.get(norm_expected)
            if candidate is not None:
                raise ValueError(
                    f"{path}: cabeçalho esperado {expected!r} não encontrado, "
                    f"mas existe candidato próximo {candidate!r} — possível renomeação "
                    "com espaço/acentuação/case, corrija o cabeçalho"
                )
            # também tenta match por um caractere de diferença (ex. espaço extra interno)
            for actual in present:
                if _normalize_header(actual) == norm_expected:
                    raise ValueError(
                        f"{path}: cabeçalho esperado {expected!r} não encontrado, "
                        f"mas existe candidato próximo {actual!r}"
                    )


def read_existing_decisions(path: Path) -> dict[RowKey, dict[str, object]]:
    """Reads the decision columns from a previously-generated consolidado
    workbook's `GLOSAS` sheet, keyed by (indicador, órgão) — the merge
    contract's preservation source (ticket 04 Q3). Empty dict if `path`
    doesn't exist yet or has no `GLOSAS` sheet (first run).
    """
    if not path.exists():
        return {}
    workbook = load_workbook(path, data_only=True)
    try:
        if GLOSAS_SHEET not in workbook.sheetnames:
            return {}
        sheet = workbook[GLOSAS_SHEET]

        header_row = [cell.value for cell in sheet[1]]
        if not header_row:
            return {}
        _check_no_duplicate_headers(path, header_row)
        _check_renamed_headers(path, header_row)
        if "Indicador" not in header_row:
            return {}
        col_idx = {name: i + 1 for i, name in enumerate(header_row) if isinstance(name, str)}
        indicador_col = col_idx.get("Indicador")
        orgao_col = col_idx.get("Órgão")
        if indicador_col is None or orgao_col is None:
            return {}

        decisions: dict[RowKey, dict[str, object]] = {}
        for row in range(2, sheet.max_row + 1):
            indicador = sheet.cell(row=row, column=indicador_col).value
            orgao = sheet.cell(row=row, column=orgao_col).value
            if not isinstance(indicador, str) or not isinstance(orgao, str):
                continue
            values: dict[str, object] = {
                name: sheet.cell(row=row, column=idx).value
                for name, idx in col_idx.items()
                if name in _DECISION_COLUMNS
            }
            if any(v not in (None, "") for v in values.values()):
                decisions[(format_inms_code(indicador), orgao)] = values
        return decisions
    finally:
        workbook.close()
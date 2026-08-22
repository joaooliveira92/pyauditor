"""Aba "INMS 1.1" enriquecida (resumo executivo, memória de cálculo,
detalhamento por grupo/nível, incidentes fora do prazo e auditoria do prazo
contratual) — chamada por `sintetico.py` no lugar do
`_write_grupo_executor_sheet` genérico quando o INMS é "1.1" e o CSV bruto
tem as colunas de detalhe necessárias (`Nº Solicitacao`, `Atividades`,
`DataHoraLimite`, `TecnicoExecutor`, além de `No prazo`/`DataHoraSolicitacao`/
`DataHoraFim`/`Grupo_executor`, já exigidas pelo shape `ratio`). Se alguma
faltar (ex.: fixtures de teste com CSV minimalista), o chamador cai de volta
no renderer genérico — nenhum outro INMS é afetado.

Todos os valores derivados (totais, percentuais, resultado, situação,
penalidade, tempos, atrasos) são fórmulas do Excel sobre uma base de dados
de apoio nas colunas R:AM da própria aba (uma linha por incidente do CSV
bruto), não valores pré-calculados em Python — para que a aba continue
rastreável/reproduzível quando reaberta sem o pipeline.

`write_sheet` monta a aba seção por seção (Seção 1 · Identificação ...
Seção 9 · Penalidade); cada seção 1-3 ocupa linhas fixas (o conteúdo não
depende do número de incidentes), e cada seção 4-9 recebe a linha inicial
da seção anterior e devolve a próxima linha livre, encadeando o cursor de
linha — a ordem de chamada em `write_sheet` é a própria ordem visual da
aba.
"""

from __future__ import annotations

from collections.abc import Callable
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Final

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter as cl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

from pyauditor.categoria_filter import GRUPO_EXECUTOR_COLUMN, compute_categoria_values
from pyauditor.config.categorias import CategoriasFile, GrupoExecutorMode, WholeIndicatorMode
from pyauditor.excel._datetime import PRAZO_TOLERANCIA_MINUTOS, parse_dt
from pyauditor.excel._safety import safe_excel_text
from pyauditor.excel._workbook import create_sheet_atomic, force_recalc, unique_table_name
from pyauditor.periodo import PeriodoAfericao

__all__: Final[tuple[str, ...]] = ("has_required_columns", "write_sheet")

# Colunas do CSV bruto exigidas para o tratamento enriquecido, além das já
# exigidas pelo shape `ratio` genérico (`No prazo`, `DataHoraSolicitacao`,
# `DataHoraFim`, `Grupo_executor`).
_NUM_SOLICITACAO_COLUMN: Final[str] = "Nº Solicitacao"
_ATIVIDADE_COLUMN: Final[str] = "Atividades"
_DATA_SOLICITACAO_COLUMN: Final[str] = "DataHoraSolicitacao"
_DATA_LIMITE_COLUMN: Final[str] = "DataHoraLimite"
_DATA_FIM_COLUMN: Final[str] = "DataHoraFim"
_NO_PRAZO_COLUMN: Final[str] = "No prazo"
_TECNICO_COLUMN: Final[str] = "TecnicoExecutor"

_REQUIRED_COLUMNS: Final[tuple[str, ...]] = (
    _NUM_SOLICITACAO_COLUMN,
    _ATIVIDADE_COLUMN,
    _DATA_SOLICITACAO_COLUMN,
    _DATA_LIMITE_COLUMN,
    _DATA_FIM_COLUMN,
    _NO_PRAZO_COLUMN,
    _TECNICO_COLUMN,
    GRUPO_EXECUTOR_COLUMN,
)

# Prazo contratual de incidentes de criticidade alta (Anexo D / aba
# "Prazos" — input/prazos.csv: "Incidentes, Alta, 2h (horas corridas)").
# Constante local: só esta aba precisa do valor numérico para o controle
# contratual bruto; nenhum outro shape do pipeline usa "2 horas corridas".
_PRAZO_HORAS_CORRIDAS: Final[float] = 2.0

_NIVEL_BY_CATEGORIA: Final[dict[str, str]] = {
    "ATENDIMENTO_N1": "N1",
    "ATENDIMENTO_N2": "N2",
    "OPERACAO_N3": "N3",
    "MONITORAMENTO_NOC_SOC": "N3",
}
_NIVEL_ORDER: Final[tuple[str, ...]] = ("N1", "N2", "N3")
_AUDIT_REVIEW_LABEL: Final[str] = "Grupo sob análise de responsabilidade"
# Sentinela para grupos sem Nível (categoria "outros") — não usar "" como
# valor de célula: o Excel/openpyxl trata célula com "" como vazia, e
# VLOOKUP contra uma célula vazia devolve 0 (numérico), não "" (texto),
# quebrando os filtros COUNTIFS por Nível na Seção 5.
_SEM_NIVEL: Final[str] = "—"

_DATETIME_FMT: Final[str] = "dd/mm/yyyy hh:mm"
_DATE_FMT: Final[str] = "dd/mm/yyyy"
_PCT2: Final[str] = "0.00%"
_PCT4: Final[str] = "0.0000%"
_DUR: Final[str] = "[h]:mm"

TITLE_FONT: Final = Font(name="Arial", size=14, bold=True, color="1F2937")
SECTION_FONT: Final = Font(name="Arial", size=11, bold=True, color="FFFFFF")
SECTION_FILL: Final = PatternFill("solid", fgColor="1F2937")
HEADER_FONT: Final = Font(name="Arial", size=10, bold=True, color="FFFFFF")
HEADER_FILL: Final = PatternFill("solid", fgColor="374151")
BODY_FONT: Final = Font(name="Arial", size=10)
LABEL_FONT: Final = Font(name="Arial", size=10, bold=True)
NOTE_FONT: Final = Font(name="Arial", size=9, italic=True, color="6B7280")
GRAY_FILL: Final = PatternFill("solid", fgColor="E5E7EB")
TEAL_FILL: Final = PatternFill("solid", fgColor="CCFBF1")
GREEN_FILL: Final = PatternFill("solid", fgColor="BBF7D0")
RED_FILL: Final = PatternFill("solid", fgColor="FECACA")
ORANGE_FILL: Final = PatternFill("solid", fgColor="FED7AA")
BORDER: Final = Border(bottom=Side(style="thin", color="D1D5DB"))
# Campos de preenchimento manual (justificativa/documento/evidência) que
# devem continuar editáveis mesmo com a planilha protegida (ticket 20 / B-03).
_UNLOCKED: Final = Protection(locked=False)

# Colunas de apoio (dados brutos) — far à direita do conteúdo visível.
_R, _S, _T, _U, _V, _W, _X, _Y, _Z, _AA, _AB, _AC, _AD, _AE, _AF, _AG, _AH, _AI = range(18, 36)
_AJ = 36
_AK, _AL, _AM = 37, 38, 39

_DATA_QUALIDADE_OK: Final[str] = "OK"

# Assinatura do range de uma coluna de apoio (`R2:R{last_row}` etc.), fixada
# uma vez por chamada de `write_sheet` a partir de `last_row` — repassada
# como parâmetro `rng` às seções que montam fórmulas sobre a base de apoio.
_ColumnRange = Callable[[int], str]


def has_required_columns(fieldnames: list[str]) -> bool:
    return all(column in fieldnames for column in _REQUIRED_COLUMNS)


def _raw_range(col: int, last_row: int) -> str:
    return f"${cl(col)}$2:${cl(col)}${last_row}"


def _section_bar(sheet: Worksheet, row: int, text: str, last_col: int = 12) -> None:
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)
    cell = sheet.cell(row=row, column=1, value=text)
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center")
    sheet.row_dimensions[row].height = 20


_CellValue = str | float | int | datetime | date | None


def _label_value(
    sheet: Worksheet,
    row: int,
    label: str,
    value: _CellValue,
    *,
    fmt: str | None = None,
    fill: PatternFill | None = None,
) -> None:
    lc = sheet.cell(row=row, column=1, value=label)
    lc.font = LABEL_FONT
    lc.border = BORDER
    vc = sheet.cell(row=row, column=2, value=value)
    vc.font = BODY_FONT
    vc.border = BORDER
    if fmt:
        vc.number_format = fmt
    if fill:
        lc.fill = fill
        vc.fill = fill


def _header_row(sheet: Worksheet, row: int, headers: tuple[str, ...]) -> None:
    for idx, text in enumerate(headers, start=1):
        cell = sheet.cell(row=row, column=idx, value=text)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left", wrap_text=True, vertical="center")
    sheet.row_dimensions[row].height = 30


def _add_table(sheet: Worksheet, name: str, ref: str) -> None:
    """Tabela nativa do Excel — não `sheet.auto_filter` (único por aba, seria
    sobrescrito pela segunda tabela detalhada); cada tabela tem seu próprio
    filtro/ordenação."""
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name=None, showRowStripes=False, showFirstColumn=False, showLastColumn=False
    )
    sheet.add_table(table)


def _add_situacao_conditional_formatting(sheet: Worksheet, coordinate: str) -> None:
    """"Meta atingida"/"Meta não atingida" -> verde/vermelho — usado na Seção
    2 (resultado consolidado) e na Seção 7 (cada metodologia de controle)."""
    sheet.conditional_formatting.add(
        coordinate,
        CellIsRule(operator="equal", formula=['"Meta atingida"'], fill=GREEN_FILL),  # type: ignore[no-untyped-call]
    )
    sheet.conditional_formatting.add(
        coordinate,
        CellIsRule(operator="equal", formula=['"Meta não atingida"'], fill=RED_FILL),  # type: ignore[no-untyped-call]
    )


def _protect_support_columns(sheet: Worksheet) -> None:
    """Oculta as colunas de apoio (R:AM) e protege a aba contra edição
    acidental das fórmulas — os únicos campos que continuam editáveis são
    os marcados com `_UNLOCKED` (justificativa/documento/evidência de
    preenchimento manual da auditoria, Seções 4 e 6). Sem senha: o
    objetivo é reduzir edição acidental das fórmulas, não impedir edição
    deliberada por quem tem o arquivo (ticket 20 / B-03). Não afeta a
    leitura das fórmulas pelo pipeline — proteção de planilha do Excel só
    bloqueia edição interativa, nunca o cálculo/leitura de valores por
    openpyxl ou por qualquer motor de recálculo."""
    for col in range(_R, _AM + 1):
        sheet.column_dimensions[cl(col)].hidden = True
    sheet.protection.sheet = True
    sheet.protection.formatCells = False
    sheet.protection.formatColumns = False
    sheet.protection.formatRows = False


def _build_grupo_rows(
    categorias_file: CategoriasFile,
    grupo_executor_entries: list[tuple[str, GrupoExecutorMode]],
    real_values: set[str],
) -> list[tuple[str, str, str]]:
    """(grupo, nível, categoria_label) — mesma resolução de
    `compute_categoria_values` usada por `_write_grupo_executor_sheet`, para
    que a aba enriquecida cubra exatamente os mesmos grupos/categorias por
    órgão, sem hardcodar nomes de grupo."""
    per_categoria_values, outros_values = compute_categoria_values(
        grupo_executor_entries, real_values
    )
    result: list[tuple[str, str, str]] = []
    for categoria_key, effective_values in per_categoria_values.items():
        categoria = categorias_file.categorias[categoria_key]
        nivel = _NIVEL_BY_CATEGORIA.get(categoria_key, _SEM_NIVEL)
        for grupo in sorted(effective_values):
            result.append((grupo, nivel, categoria.label))
    for grupo in sorted(outros_values):
        result.append((grupo, _SEM_NIVEL, _AUDIT_REVIEW_LABEL))
    # Nota (ticket 14 / M-05): não há checagem extra de unicidade aqui —
    # `compute_categoria_values`, chamada acima, já valida que nenhum grupo
    # pertence a mais de uma categoria (in_values sobrepostos ou
    # catch_all_contains sobrepondo in_values de outra categoria) e lança
    # `ValueError` antes de `per_categoria_values`/`outros_values` existirem.
    # Duplicar essa validação aqui seria código morto.
    return result


def _normalize_no_prazo(row: dict[str, str]) -> str:
    """Normaliza o campo "No prazo" do CSV bruto para exatamente "S" ou "N"
    — sem isso, variações como "s", "Sim", "N/A" ou vazio entram no
    denominador (IAP, via `COUNTA`/`ROWS`) mas não em nenhum dos dois
    `COUNTIF` (dentro/fora do prazo), quebrando silenciosamente a
    identidade IAP = dentro + fora. Local a este módulo: não há hoje outro
    renderer ITSM reusando esse vocabulário S/N — ver nota de reuso em
    `spec.md`."""
    normalized = row[_NO_PRAZO_COLUMN].strip().upper()
    if normalized not in {"S", "N"}:
        num_solicitacao = row.get(_NUM_SOLICITACAO_COLUMN, "?")
        raise ValueError(
            f"Nº Solicitação {num_solicitacao!r}: valor de 'No prazo' inválido "
            f"{row[_NO_PRAZO_COLUMN]!r} — esperado apenas 'S' ou 'N'"
        )
    return normalized


def _write_raw_block(
    sheet: Worksheet,
    rows: list[dict[str, str]],
    grupo_rows: list[tuple[str, str, str]],
    last_row: int,
) -> None:
    headers = {
        _R: "Nº Solicitação",
        _S: "Grupo executor",
        _T: "Atividade",
        _U: "DataHoraSolicitacao",
        _V: "DataHoraLimite (ITSM)",
        _W: "DataHoraFim",
        _X: "No prazo (fornecedor)",
        _Y: "TecnicoExecutor",
        _Z: "Nível",
        _AA: "Categoria",
        _AB: "Limite contratual (abertura + prazo corrido)",
        _AC: "Limite ITSM superior ao contratual bruto",
        _AD: "No prazo (data limite ITSM)",
        _AE: "No prazo (controle contratual bruto)",
        _AF: "Atraso vs. limite ITSM (min)",
        _AG: "Duração criação→resolução (dias)",
        _AH: "Ordem — fora do prazo",
        _AI: "Ordem — limite ITSM superior ao contratual bruto",
        _AJ: "Situação dos dados",
    }
    note = sheet.cell(
        row=1,
        column=_R - 1,
        value=(
            f"ÁREA DE APOIO — base de dados para as fórmulas desta aba "
            f"({len(rows)} incidentes do CSV bruto). Não excluir."
        ),
    )
    note.font = NOTE_FONT
    for col, text in headers.items():
        cell = sheet.cell(row=1, column=col, value=text)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True)
        sheet.column_dimensions[cl(col)].width = 16

    for col, text in {_AK: "Grupo executor", _AL: "Nível", _AM: "Categoria"}.items():
        cell = sheet.cell(row=1, column=col, value=text)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        sheet.column_dimensions[cl(col)].width = 30
    for idx, (grupo, nivel, categoria) in enumerate(grupo_rows, start=2):
        sheet.cell(row=idx, column=_AK, value=safe_excel_text(grupo)).font = BODY_FONT
        sheet.cell(row=idx, column=_AL, value=safe_excel_text(nivel)).font = BODY_FONT
        sheet.cell(row=idx, column=_AM, value=safe_excel_text(categoria)).font = BODY_FONT
    map_range = f"${cl(_AK)}$2:${cl(_AM)}${1 + len(grupo_rows)}"

    for i, row in enumerate(rows, start=2):
        num_solicitacao = safe_excel_text(row[_NUM_SOLICITACAO_COLUMN])
        sheet.cell(row=i, column=_R, value=num_solicitacao).font = BODY_FONT
        sheet.cell(
            row=i, column=_S, value=safe_excel_text(row[GRUPO_EXECUTOR_COLUMN])
        ).font = BODY_FONT
        sheet.cell(
            row=i, column=_T, value=safe_excel_text(row[_ATIVIDADE_COLUMN])
        ).font = BODY_FONT

        solicitacao = parse_dt(row[_DATA_SOLICITACAO_COLUMN])
        limite = parse_dt(row[_DATA_LIMITE_COLUMN])
        fim = parse_dt(row[_DATA_FIM_COLUMN])

        u_cell = sheet.cell(row=i, column=_U, value=solicitacao.value)
        u_cell.number_format = _DATETIME_FMT
        v_cell = sheet.cell(row=i, column=_V, value=limite.value)
        v_cell.number_format = _DATETIME_FMT
        w_cell = sheet.cell(row=i, column=_W, value=fim.value)
        w_cell.number_format = _DATETIME_FMT
        for c in (u_cell, v_cell, w_cell):
            c.font = BODY_FONT

        if solicitacao.is_malformed:
            qualidade = "Data de abertura inválida"
        elif limite.is_malformed:
            qualidade = "Data limite inválida"
        elif fim.is_malformed:
            qualidade = "Data de encerramento inválida"
        elif (
            solicitacao.value is not None
            and fim.value is not None
            and fim.value < solicitacao.value
        ):
            qualidade = "Encerramento anterior à abertura"
        else:
            qualidade = _DATA_QUALIDADE_OK
        aj_cell = sheet.cell(row=i, column=_AJ, value=qualidade)
        aj_cell.font = BODY_FONT
        if qualidade != _DATA_QUALIDADE_OK:
            aj_cell.fill = RED_FILL

        sheet.cell(row=i, column=_X, value=safe_excel_text(row[_NO_PRAZO_COLUMN])).font = BODY_FONT
        sheet.cell(
            row=i, column=_Y, value=safe_excel_text(row[_TECNICO_COLUMN])
        ).font = BODY_FONT

        sc, uc, vc, wc, abc = (
            f"{cl(_S)}{i}", f"{cl(_U)}{i}", f"{cl(_V)}{i}", f"{cl(_W)}{i}", f"{cl(_AB)}{i}",
        )
        z_cell = sheet.cell(
            row=i, column=_Z, value=f'=IFERROR(VLOOKUP({sc},{map_range},2,FALSE),"")'
        )
        aa_cell = sheet.cell(
            row=i, column=_AA, value=f'=IFERROR(VLOOKUP({sc},{map_range},3,FALSE),"")'
        )
        ab_cell = sheet.cell(
            row=i, column=_AB, value=f'=IF({uc}="","",{uc}+{_PRAZO_HORAS_CORRIDAS}/24)'
        )
        ab_cell.number_format = _DATETIME_FMT
        ac_cell = sheet.cell(
            row=i,
            column=_AC,
            value=(
                f'=IF(OR({uc}="",{vc}=""),"Requer análise",'
                f'IF({vc}>{abc}+{PRAZO_TOLERANCIA_MINUTOS}/1440,"Sim","Não"))'
            ),
        )
        ad_cell = sheet.cell(
            row=i, column=_AD, value=f'=IF(OR({wc}="",{vc}=""),"",IF({wc}<={vc},"S","N"))'
        )
        ae_cell = sheet.cell(
            row=i, column=_AE, value=f'=IF(OR({wc}="",{abc}=""),"",IF({wc}<={abc},"S","N"))'
        )
        af_cell = sheet.cell(
            row=i,
            column=_AF,
            value=f'=IF(AND({wc}<>"",{vc}<>"",{wc}>{vc}),({wc}-{vc})*1440,0)',
        )
        af_cell.number_format = "0.0"
        ag_cell = sheet.cell(
            row=i,
            column=_AG,
            value=f'=IF(OR({uc}="",{wc}="",{wc}<{uc}),"",{wc}-{uc})',
        )
        ag_cell.number_format = _DUR
        for c in (z_cell, aa_cell, ac_cell, ad_cell, ae_cell):
            c.font = BODY_FONT

        xc, acc = f"{cl(_X)}{i}", f"{cl(_AC)}{i}"
        ah_cell = sheet.cell(
            row=i, column=_AH, value=f'=IF({xc}="N",COUNTIF($X$2:{xc},"N"),"")'
        )
        ai_cell = sheet.cell(
            row=i, column=_AI, value=f'=IF({acc}="Sim",COUNTIF($AC$2:{acc},"Sim"),"")'
        )
        for c in (ah_cell, ai_cell):
            c.font = BODY_FONT


def _write_section_1_identificacao(
    sheet: Worksheet,
    *,
    rows: list[dict[str, str]],
    contract: str,
    periodo: PeriodoAfericao | None,
    raw_csv_path: Path,
    generated_at: datetime,
) -> None:
    sheet.merge_cells("A1:L1")
    t = sheet.cell(row=1, column=1, value="INMS 1.1 – Incidentes atendidos dentro do prazo")
    t.font = TITLE_FONT

    _section_bar(sheet, 3, "SEÇÃO 1 · IDENTIFICAÇÃO")
    data_corte: _CellValue
    if periodo is not None:
        competencia = (
            f"{periodo.inicio:%d/%m/%Y} a {periodo.fim:%d/%m/%Y} "
            f"({periodo.inicio:%Y-%m})"
        )
        data_corte = periodo.fim
        data_corte_fmt = _DATE_FMT
    else:
        competencia = "Não informado"
        data_corte = "Não informado"
        data_corte_fmt = None
    _label_value(sheet, 4, "Competência:", competencia, fill=GRAY_FILL)
    _label_value(sheet, 5, "Contrato:", safe_excel_text(contract), fill=GRAY_FILL)
    _label_value(
        sheet, 6, "Fonte dos dados:",
        # Só o nome do arquivo, não o caminho completo — evita expor
        # estrutura de diretório/usuário do ambiente que gerou a planilha
        # (ticket 19 / B-02).
        safe_excel_text(f"{raw_csv_path.name} ({len(rows)} registros brutos)"), fill=GRAY_FILL,
    )
    _label_value(
        sheet, 7, "Data de geração:", generated_at, fmt=_DATETIME_FMT, fill=GRAY_FILL
    )
    _label_value(sheet, 8, "Data de corte:", data_corte, fmt=data_corte_fmt, fill=GRAY_FILL)
    _label_value(sheet, 9, "Responsável pela elaboração:", "Não informado", fill=GRAY_FILL)


def _write_section_2_resumo(
    sheet: Worksheet,
    *,
    iap: str,
    iadp: str,
    fora: str,
    meta_value: float,
) -> int:
    """Seção de linhas fixas (11-14) — devolve `next_free_row` (16, fixo)
    para a Seção 3."""
    _section_bar(sheet, 11, "SEÇÃO 2 · RESUMO EXECUTIVO")
    _header_row(
        sheet, 12,
        ("Meta contratual", "IAP (abertos)", "IADP (dentro do prazo)", "Fora do prazo",
         "Resultado INMS 1.1", "Diferença p/ meta", "Situação"),
    )
    sheet["A13"] = f"={meta_value}"
    sheet["B13"] = f"={iap}"
    sheet["C13"] = f"={iadp}"
    sheet["D13"] = f"={fora}"
    # B13=0 (nenhum incidente aberto no período) não é "0% de desempenho" —
    # é resultado não mensurável; guardado para não propagar #DIV/0! e para
    # não afirmar "meta não atingida" indevidamente.
    sheet["E13"] = '=IF(B13=0,"Sem ocorrências",C13/B13)'
    # Meta mínima ("`>=`") — único operador suportado por este renderer
    # (validado em `_validate_write_sheet_params`, ticket 05 / A-03).
    sheet["F13"] = '=IF(B13=0,"",E13-A13)'
    sheet["G13"] = (
        '=IF(B13=0,"Sem ocorrências",'
        'IF(E13>=A13,"Meta atingida","Meta não atingida"))'
    )

    sheet["A13"].number_format = _PCT2
    sheet["E13"].number_format = _PCT2
    sheet["F13"].number_format = _PCT2
    for c in ("A13", "B13", "C13", "D13", "E13", "F13", "G13"):
        sheet[c].font = Font(name="Arial", size=12, bold=True)
        sheet[c].fill = TEAL_FILL
        sheet[c].alignment = Alignment(horizontal="center")
    sheet.row_dimensions[13].height = 24

    _add_situacao_conditional_formatting(sheet, "G13")

    sheet.merge_cells("A14:L14")
    pen_note = sheet.cell(
        row=14, column=1,
        value="Penalidade: Pendente de confirmação da regra de arredondamento — ver Seção 9.",
    )
    pen_note.font = LABEL_FONT
    pen_note.fill = ORANGE_FILL
    return 16


def _write_section_3_memoria(sheet: Worksheet) -> int:
    """Seção de linhas fixas (16-26) — devolve `next_free_row` (28, fixo)
    para a Seção 4."""
    _section_bar(sheet, 16, "SEÇÃO 3 · MEMÓRIA DO CÁLCULO CONSOLIDADO")
    _label_value(sheet, 17, "IAP (incidentes abertos no período):", "=B13", fmt="0")
    _label_value(sheet, 18, "IADP (incidentes dentro do prazo):", "=C13", fmt="0")
    sheet.merge_cells("A19:C19")
    sheet["A19"] = "INMS 1.1 = IADP ÷ IAP"
    sheet["A19"].font = NOTE_FONT
    sheet.merge_cells("A20:C20")
    sheet["A20"] = '=CONCATENATE("INMS 1.1 = ",B18," ÷ ",B17)'
    sheet["A20"].font = NOTE_FONT
    _label_value(sheet, 21, "INMS 1.1 (resultado, 4 casas):", "=E13", fmt=_PCT4)
    _label_value(sheet, 22, "Meta contratual:", "=A13", fmt=_PCT4)
    _label_value(
        sheet, 23, "Desvio em pontos percentuais (Resultado - Meta):",
        '=IF(B13=0,"",E13-A13)', fmt=_PCT4,
    )
    _label_value(
        sheet, 24, "Quantidade mínima dentro do prazo p/ atingir a meta:",
        "=ROUNDUP(B13*A13,0)", fmt="0",
    )
    _label_value(
        sheet, 25, "Margem em quantidade de incidentes (IADP - mínimo):", "=C13-B24", fmt="0;-0"
    )
    sheet.merge_cells("A26:L26")

    def _narrativa(desfecho: str) -> str:
        return (
            'CONCATENATE("Com ",B17," incidentes, seriam necessários pelo menos ",B24,'
            f'" dentro do prazo para atingir ",TEXT(B22,"0.00%"),". O resultado ficou {desfecho})'
        )

    # Narrativa condicional ao sinal da margem (B25 = IADP - mínimo): "abaixo"
    # só quando a margem é negativa — antes disso o texto sempre dizia
    # "abaixo do mínimo" mesmo com meta superada (ABS(B25) escondia o sinal).
    abaixo = _narrativa('",ABS(B25)," incidente(s) abaixo do mínimo necessário."')
    exato = _narrativa('exatamente no mínimo necessário."')
    acima = _narrativa('",B25," incidente(s) acima do mínimo necessário."')
    sheet["A26"] = (
        '=IF(B13=0,"Sem incidentes abertos no período — indicador não mensurável.",'
        f"IF(B25<0,{abaixo},IF(B25=0,{exato},{acima})))"
    )
    sheet["A26"].font = NOTE_FONT
    return 28


def _write_section_4_detalhamento(
    sheet: Worksheet,
    *,
    grupo_rows: list[tuple[str, str, str]],
    rng: _ColumnRange,
    start_row: int,
    table_name: str,
) -> int:
    """Devolve `next_free_row` — a linha livre após a tabela de grupos,
    usada pela Seção 5 como sua própria linha inicial."""
    _section_bar(sheet, start_row, "SEÇÃO 4 · DETALHAMENTO POR GRUPO EXECUTOR", last_col=12)
    _header_row(
        sheet, start_row + 1,
        ("Categoria", "Nível", "Grupo executor", "Linhas", "Dentro do prazo", "Fora do prazo",
         "% (dentro/linhas)", "Tempo médio", "Incluído no INMS?", "Justificativa de exclusão",
         "Documento autorizador", "Observação de auditoria"),
    )
    first_group_row = start_row + 2
    for offset, (grupo, nivel, categoria) in enumerate(grupo_rows):
        r = first_group_row + offset
        sheet.cell(row=r, column=1, value=safe_excel_text(categoria)).font = BODY_FONT
        sheet.cell(row=r, column=2, value=safe_excel_text(nivel)).font = BODY_FONT
        sheet.cell(row=r, column=3, value=safe_excel_text(grupo)).font = BODY_FONT
        grupo_ref = f"$C${r}"
        linhas_cell = sheet.cell(row=r, column=4, value=f"=COUNTIF({rng(_S)},{grupo_ref})")
        dentro_cell = sheet.cell(
            row=r, column=5, value=f'=COUNTIFS({rng(_S)},{grupo_ref},{rng(_X)},"S")'
        )
        fora_cell = sheet.cell(
            row=r, column=6, value=f'=COUNTIFS({rng(_S)},{grupo_ref},{rng(_X)},"N")'
        )
        pct_cell = sheet.cell(
            row=r, column=7, value=f'=IF(D{r}=0,"Sem ocorrências",E{r}/D{r})'
        )
        pct_cell.number_format = _PCT2
        tempo_cell = sheet.cell(
            row=r, column=8,
            value=f'=IF(D{r}=0,"—",AVERAGEIF({rng(_S)},{grupo_ref},{rng(_AG)}))',
        )
        tempo_cell.number_format = _DUR
        incluido = sheet.cell(row=r, column=9, value="Sim")
        is_audit_review = categoria == _AUDIT_REVIEW_LABEL
        if is_audit_review:
            just = (
                "Nenhuma exclusão aplicada — ausência de fundamento documental formal "
                "para exclusão do grupo."
            )
            obs = (
                "Grupo mantido no consolidado do INMS 1.1 por força da regra contratual "
                "(denominador = total de incidentes abertos no período); reclassificação "
                "exigiria documento autorizador válido — ver Seção 9."
            )
            for cc in (linhas_cell, dentro_cell, fora_cell, pct_cell, tempo_cell, incluido):
                cc.fill = ORANGE_FILL
        else:
            just = "Não aplicável — nenhuma exclusão aplicada."
            obs = "—"
        just_cell = sheet.cell(row=r, column=10, value=just)
        doc_cell = sheet.cell(row=r, column=11, value="Não informado")
        obs_cell = sheet.cell(row=r, column=12, value=obs)
        for c in (linhas_cell, dentro_cell, fora_cell, incluido, just_cell, doc_cell, obs_cell):
            c.font = BODY_FONT
            c.alignment = Alignment(wrap_text=True, vertical="top")
        for col in range(1, 13):
            sheet.cell(row=r, column=col).border = BORDER
        # Campos de preenchimento manual da auditoria (justificativa/documento
        # autorizador) permanecem editáveis quando a planilha for protegida
        # (ticket 20 / B-03) — os demais campos desta seção são fórmulas.
        just_cell.protection = _UNLOCKED
        doc_cell.protection = _UNLOCKED
    last_group_row = first_group_row + len(grupo_rows) - 1
    _add_table(sheet, table_name, f"A{start_row + 1}:L{last_group_row}")
    return last_group_row + 2


def _write_section_5_subtotais(
    sheet: Worksheet, *, rng: _ColumnRange, start_row: int
) -> int:
    """Devolve `next_free_row` — a linha livre após as verificações
    cruzadas da seção, usada pela Seção 6 como sua própria linha inicial."""
    sub_bar_row = start_row
    _section_bar(
        sheet, sub_bar_row, "SEÇÃO 5 · SUBTOTAIS POR NÍVEL (informação gerencial)", last_col=6
    )
    _header_row(sheet, sub_bar_row + 1, ("Nível / Grupo", "Linhas", "Dentro do prazo",
                                          "Fora do prazo", "% bruto", "Tempo médio"))
    nivel_rows = list(range(sub_bar_row + 2, sub_bar_row + 2 + len(_NIVEL_ORDER)))
    for r, nivel_label in zip(nivel_rows, _NIVEL_ORDER, strict=True):
        sheet.cell(row=r, column=1, value=nivel_label).font = BODY_FONT
        linhas = sheet.cell(row=r, column=2, value=f'=COUNTIFS({rng(_Z)},"{nivel_label}")')
        dentro = sheet.cell(
            row=r, column=3, value=f'=COUNTIFS({rng(_Z)},"{nivel_label}",{rng(_X)},"S")'
        )
        fora_c = sheet.cell(
            row=r, column=4, value=f'=COUNTIFS({rng(_Z)},"{nivel_label}",{rng(_X)},"N")'
        )
        pct = sheet.cell(row=r, column=5, value=f'=IF(B{r}=0,"Sem ocorrências",C{r}/B{r})')
        pct.number_format = _PCT2
        tempo = sheet.cell(
            row=r, column=6,
            value=f'=IF(B{r}=0,"—",AVERAGEIFS({rng(_AG)},{rng(_Z)},"{nivel_label}"))',
        )
        tempo.number_format = _DUR
        for c in (linhas, dentro, fora_c, pct, tempo):
            c.font = BODY_FONT
        for col in range(1, 7):
            sheet.cell(row=r, column=col).border = BORDER

    outr = nivel_rows[-1] + 1
    sheet.cell(row=outr, column=1, value=f"Sem nível ({_AUDIT_REVIEW_LABEL})").font = BODY_FONT
    linhas = sheet.cell(row=outr, column=2, value=f'=COUNTIFS({rng(_Z)},"{_SEM_NIVEL}")')
    dentro = sheet.cell(
        row=outr, column=3, value=f'=COUNTIFS({rng(_Z)},"{_SEM_NIVEL}",{rng(_X)},"S")'
    )
    fora_c = sheet.cell(
        row=outr, column=4, value=f'=COUNTIFS({rng(_Z)},"{_SEM_NIVEL}",{rng(_X)},"N")'
    )
    pct = sheet.cell(row=outr, column=5, value=f'=IF(B{outr}=0,"Sem ocorrências",C{outr}/B{outr})')
    pct.number_format = _PCT2
    tempo = sheet.cell(
        row=outr, column=6,
        value=f'=IF(B{outr}=0,"—",AVERAGEIFS({rng(_AG)},{rng(_Z)},"{_SEM_NIVEL}"))',
    )
    tempo.number_format = _DUR
    for c in (linhas, dentro, fora_c, pct, tempo):
        c.font = BODY_FONT
        c.fill = ORANGE_FILL
    for col in range(1, 7):
        sheet.cell(row=outr, column=col).border = BORDER

    totr = outr + 1
    sheet.cell(row=totr, column=1, value="Total geral").font = Font(bold=True)
    tl = sheet.cell(row=totr, column=2, value=f"=SUM(B{nivel_rows[0]}:B{outr})")
    td = sheet.cell(row=totr, column=3, value=f"=SUM(C{nivel_rows[0]}:C{outr})")
    tf = sheet.cell(row=totr, column=4, value=f"=SUM(D{nivel_rows[0]}:D{outr})")
    tp = sheet.cell(
        row=totr, column=5, value=f'=IF(B{totr}=0,"Sem ocorrências",C{totr}/B{totr})'
    )
    tp.number_format = _PCT2
    for c in (tl, td, tf, tp):
        c.font = Font(bold=True)
        c.fill = TEAL_FILL
    for col in range(1, 7):
        sheet.cell(row=totr, column=col).border = BORDER

    note_row = totr + 1
    sheet.merge_cells(f"A{note_row}:F{note_row}")
    sheet[f"A{note_row}"] = (
        "O consolidado do INMS 1.1 é a soma dos numeradores/denominadores dos níveis "
        "(esta linha), não a média simples dos percentuais de N1/N2/N3."
    )
    sheet[f"A{note_row}"].font = NOTE_FONT

    check_row = note_row + 1
    sheet[f"A{check_row}"] = "Verificação cruzada (Seção 5 = Seção 2/3):"
    sheet[f"A{check_row}"].font = LABEL_FONT
    chk = sheet.cell(
        row=check_row,
        column=2,
        value=(
            f'=IF(AND(ISNUMBER(E{totr}),ISNUMBER(E13)),'
            f'IF(ROUND(E{totr},6)=ROUND(E13,6),"OK","DIVERGÊNCIA"),'
            f'IF(E{totr}=E13,"OK","DIVERGÊNCIA"))'
        ),
    )
    chk.font = Font(bold=True)
    sheet.conditional_formatting.add(
        chk.coordinate, CellIsRule(operator="equal", formula=['"OK"'], fill=GREEN_FILL)  # type: ignore[no-untyped-call]
    )
    sheet.conditional_formatting.add(
        chk.coordinate, CellIsRule(operator="equal", formula=['"DIVERGÊNCIA"'], fill=RED_FILL)  # type: ignore[no-untyped-call]
    )

    # IAP = dentro do prazo + fora do prazo (ticket 06 / A-04) — se algum
    # valor de "No prazo" escapasse à normalização de `_normalize_no_prazo`,
    # essa identidade quebraria silenciosamente (o registro entraria no IAP
    # sem entrar em nenhum dos dois `COUNTIF`); mantida como cinto e
    # suspensório mesmo com a validação na fronteira.
    check2_row = check_row + 1
    sheet[f"A{check2_row}"] = "Verificação (IAP = dentro do prazo + fora do prazo):"
    sheet[f"A{check2_row}"].font = LABEL_FONT
    chk2 = sheet.cell(
        row=check2_row,
        column=2,
        value='=IF(B13=C13+D13,"OK","DIVERGÊNCIA DE CLASSIFICAÇÃO")',
    )
    chk2.font = Font(bold=True)
    sheet.conditional_formatting.add(
        chk2.coordinate, CellIsRule(operator="equal", formula=['"OK"'], fill=GREEN_FILL)  # type: ignore[no-untyped-call]
    )
    sheet.conditional_formatting.add(
        chk2.coordinate,
        CellIsRule(operator="equal", formula=['"DIVERGÊNCIA DE CLASSIFICAÇÃO"'], fill=RED_FILL),  # type: ignore[no-untyped-call]
    )
    return check2_row + 2


def _write_section_6_fora_prazo(
    sheet: Worksheet,
    *,
    rows: list[dict[str, str]],
    rng: _ColumnRange,
    start_row: int,
    table_name: str,
) -> int:
    """Devolve `next_free_row` — a linha livre após a nota "nenhum
    incidente" ou após a tabela, usada pela Seção 7 como sua própria linha
    inicial."""
    s6_bar = start_row
    _section_bar(sheet, s6_bar, "SEÇÃO 6 · INCIDENTES FORA DO PRAZO", last_col=11)
    _header_row(
        sheet, s6_bar + 1,
        ("Nº solicitação", "Grupo executor", "Atividade", "Abertura", "Limite (ITSM)",
         "Encerramento", "Atraso vs. limite (min)", "Técnico executor", "Justificativa",
         "Aceite da justificativa", "Documento/evidência"),
    )
    fora_first = s6_bar + 2
    # Tabela dimensionada ao número real de incidentes fora do prazo (não um
    # teto arbitrário) — a Seção 6 existe para evidenciar exceções, não pra
    # truncá-las quando o órgão tiver mais de um punhado.
    fora_count = sum(1 for row in rows if row[_NO_PRAZO_COLUMN] == "N")
    ah_range = rng(_AH)
    if fora_count == 0:
        sheet.merge_cells(f"A{fora_first}:K{fora_first}")
        sheet[f"A{fora_first}"] = "Nenhum incidente fora do prazo no período."
        sheet[f"A{fora_first}"].font = NOTE_FONT
        return fora_first + 2

    for n in range(1, fora_count + 1):
        r = fora_first + n - 1
        match_expr = f"MATCH({n},{ah_range},0)"
        c1 = sheet.cell(row=r, column=1, value=f'=IFERROR(INDEX({rng(_R)},{match_expr}),"")')
        c2 = sheet.cell(row=r, column=2, value=f'=IFERROR(INDEX({rng(_S)},{match_expr}),"")')
        c3 = sheet.cell(row=r, column=3, value=f'=IFERROR(INDEX({rng(_T)},{match_expr}),"")')
        c4 = sheet.cell(row=r, column=4, value=f'=IFERROR(INDEX({rng(_U)},{match_expr}),"")')
        c4.number_format = _DATETIME_FMT
        c5 = sheet.cell(row=r, column=5, value=f'=IFERROR(INDEX({rng(_V)},{match_expr}),"")')
        c5.number_format = _DATETIME_FMT
        c6 = sheet.cell(row=r, column=6, value=f'=IFERROR(INDEX({rng(_W)},{match_expr}),"")')
        c6.number_format = _DATETIME_FMT
        c7 = sheet.cell(row=r, column=7, value=f'=IFERROR(INDEX({rng(_AF)},{match_expr}),"")')
        c7.number_format = "0.0"
        c8 = sheet.cell(row=r, column=8, value=f'=IFERROR(INDEX({rng(_Y)},{match_expr}),"")')
        c9 = sheet.cell(row=r, column=9, value="Não informado")
        c10 = sheet.cell(row=r, column=10, value="Não informado")
        c11 = sheet.cell(row=r, column=11, value="Não informado")
        for c in (c1, c2, c3, c4, c5, c6, c7, c8, c9, c10, c11):
            c.font = BODY_FONT
            c.border = BORDER
        c1.fill = RED_FILL
        # Justificativa/aceite/evidência são preenchimento manual da
        # auditoria — permanecem editáveis com a planilha protegida
        # (ticket 20 / B-03).
        c9.protection = _UNLOCKED
        c10.protection = _UNLOCKED
        c11.protection = _UNLOCKED
    fora_last = fora_first + fora_count - 1
    _add_table(sheet, table_name, f"A{s6_bar + 1}:K{fora_last}")
    return fora_last + 2


def _write_section_7_auditoria(
    sheet: Worksheet,
    *,
    rows: list[dict[str, str]],
    rng: _ColumnRange,
    start_row: int,
    table_name: str,
) -> int:
    """Devolve `next_free_row` — a linha livre após a amostra (ou após a
    nota "nenhuma divergência"), usada pela Seção 8 como sua própria linha
    inicial."""
    s7_bar = start_row
    _section_bar(sheet, s7_bar, "SEÇÃO 7 · AUDITORIA DO PRAZO CONTRATUAL", last_col=7)
    sheet.cell(row=s7_bar + 1, column=1, value="Controles de resultado").font = LABEL_FONT
    _header_row(sheet, s7_bar + 2, ("Metodologia", "Resultado", "Situação"))
    ctrl_rows = [
        ("Resultado informado pelo fornecedor (campo 'No prazo')", "C13/B13"),
        (
            "Resultado reproduzido pela data limite registrada no ITSM "
            "(DataHoraFim ≤ DataHoraLimite)",
            f'COUNTIF({rng(_AD)},"S")/B13',
        ),
        (
            f"Controle contratual bruto (DataHoraFim ≤ DataHoraSolicitacao + "
            f"{_PRAZO_HORAS_CORRIDAS:g}h corridas)",
            f'COUNTIF({rng(_AE)},"S")/B13',
        ),
    ]
    first_ctrl = s7_bar + 3
    for i, (label, division) in enumerate(ctrl_rows):
        r = first_ctrl + i
        sheet.cell(row=r, column=1, value=label).font = BODY_FONT
        sheet.cell(row=r, column=1).alignment = Alignment(wrap_text=True)
        val = sheet.cell(row=r, column=2, value=f'=IF(B13=0,"Sem ocorrências",{division})')
        val.number_format = _PCT4
        sit_formula = (
            f'=IF(B13=0,"Não aplicável",'
            f'IF(B{r}>=A13,"Meta atingida","Meta não atingida"))'
        )
        sit = sheet.cell(row=r, column=3, value=sit_formula)
        val.font = BODY_FONT
        sit.font = BODY_FONT
        for col in range(1, 4):
            sheet.cell(row=r, column=col).border = BORDER
        _add_situacao_conditional_formatting(sheet, sit.coordinate)

    div_header_row = first_ctrl + 3 + 1
    sheet.cell(
        row=div_header_row, column=1,
        value=(
            f"Limite ITSM superior ao prazo contratual bruto (abertura + "
            f"{_PRAZO_HORAS_CORRIDAS:g}h corridas) — comparação unidirecional: só "
            f"sinaliza prorrogação, não limite ITSM mais rígido que o contratual"
        ),
    ).font = LABEL_FONT
    div_count_row = div_header_row + 1
    sheet.cell(
        row=div_count_row, column=1,
        value="Registros com limite ITSM superior ao contratual bruto:",
    ).font = BODY_FONT
    div_count = sheet.cell(row=div_count_row, column=2, value=f'=COUNTIF({rng(_AC)},"Sim")')
    div_count.font = Font(bold=True)
    sheet.cell(row=div_count_row, column=3, value="% do total:").font = BODY_FONT
    div_pct = sheet.cell(
        row=div_count_row,
        column=4,
        value=f'=IF(B13=0,"Sem ocorrências",B{div_count_row}/B13)',
    )
    div_pct.number_format = _PCT2
    div_pct.font = Font(bold=True)
    div_pct.fill = ORANGE_FILL
    div_count.fill = ORANGE_FILL

    note_row7 = div_count_row + 1
    sheet.merge_cells(f"A{note_row7}:L{note_row7}")
    sheet[f"A{note_row7}"] = (
        "Casos em que o limite registrado no ITSM é superior ao limite contratual bruto "
        "não são presumidos válidos nem inválidos aqui: quando o percentual acima for "
        "relevante, solicite o histórico de pausa/suspensão/reclassificação do SLA ou a "
        "norma de calendário aplicada para confirmar a compatibilidade com a cláusula "
        "contratual de prazo corrido. Um limite ITSM inferior ao contratual bruto não é "
        "sinalizado — é mais rígido que o exigido, não uma violação."
    )
    sheet[f"A{note_row7}"].font = Font(name="Arial", size=10, bold=True, color="9A3412")
    sheet[f"A{note_row7}"].fill = ORANGE_FILL
    sheet[f"A{note_row7}"].alignment = Alignment(wrap_text=True, vertical="center")
    sheet.row_dimensions[note_row7].height = 48

    prazo_delta = timedelta(
        hours=_PRAZO_HORAS_CORRIDAS, minutes=PRAZO_TOLERANCIA_MINUTOS
    )
    divergentes_count = 0
    for row in rows:
        sol = parse_dt(row[_DATA_SOLICITACAO_COLUMN]).value
        lim = parse_dt(row[_DATA_LIMITE_COLUMN]).value
        if sol is not None and lim is not None and lim > sol + prazo_delta:
            divergentes_count += 1
    sample_size = min(divergentes_count, 15)

    sample_header_row = note_row7 + 2
    if sample_size == 0:
        sheet.merge_cells(f"A{sample_header_row}:F{sample_header_row}")
        sheet[f"A{sample_header_row}"] = (
            "Nenhum limite ITSM superior ao contratual bruto encontrado no período."
        )
        sheet[f"A{sample_header_row}"].font = NOTE_FONT
        return sample_header_row + 2

    sheet.cell(
        row=sample_header_row, column=1,
        value=(
            f"Amostra de limites ITSM superiores ao contratual bruto "
            f"({sample_size} primeiros, ordenados por ocorrência)"
        ),
    ).font = LABEL_FONT
    _header_row(
        sheet, sample_header_row + 1,
        ("Nº solicitação", "Abertura", "Limite registrado (ITSM)",
         "Limite bruto (abertura + prazo corrido)", "Diferença (horas)",
         "No prazo (fornecedor)"),
    )
    sample_first = sample_header_row + 2
    ai_range = rng(_AI)
    for n in range(1, sample_size + 1):
        r = sample_first + n - 1
        match_expr = f"MATCH({n},{ai_range},0)"
        c1 = sheet.cell(row=r, column=1, value=f'=IFERROR(INDEX({rng(_R)},{match_expr}),"")')
        c2 = sheet.cell(row=r, column=2, value=f'=IFERROR(INDEX({rng(_U)},{match_expr}),"")')
        c2.number_format = _DATETIME_FMT
        c3 = sheet.cell(row=r, column=3, value=f'=IFERROR(INDEX({rng(_V)},{match_expr}),"")')
        c3.number_format = _DATETIME_FMT
        c4 = sheet.cell(row=r, column=4, value=f'=IFERROR(INDEX({rng(_AB)},{match_expr}),"")')
        c4.number_format = _DATETIME_FMT
        c5 = sheet.cell(
            row=r, column=5,
            value=(
                f'=IFERROR((INDEX({rng(_V)},{match_expr})-'
                f'INDEX({rng(_AB)},{match_expr}))*24,"")'
            ),
        )
        c5.number_format = "0.00"
        c6 = sheet.cell(row=r, column=6, value=f'=IFERROR(INDEX({rng(_X)},{match_expr}),"")')
        for c in (c1, c2, c3, c4, c5, c6):
            c.font = BODY_FONT
            c.border = BORDER
    sample_last = sample_first + sample_size - 1
    _add_table(sheet, table_name, f"A{sample_header_row + 1}:F{sample_last}")
    sample_note_row = sample_last + 1
    sheet.merge_cells(f"A{sample_note_row}:F{sample_note_row}")
    sheet[f"A{sample_note_row}"] = (
        f'=CONCATENATE("Amostra limitada às {sample_size} primeiras ocorrências de ",'
        f'B{div_count_row}," registros com limite ITSM superior ao contratual bruto — '
        f'colunas de apoio desta aba (coluna AC) permitem reproduzir a lista completa.")'
    )
    sheet[f"A{sample_note_row}"].font = NOTE_FONT
    return sample_note_row + 2


def _write_section_8_tempo(sheet: Worksheet, *, rng: _ColumnRange, start_row: int) -> int:
    """Devolve `next_free_row` — a linha livre após a nota de rodapé da
    seção, usada pela Seção 9 como sua própria linha inicial."""
    s8_bar = start_row
    _section_bar(sheet, s8_bar, "SEÇÃO 8 · TEMPO CORRIDO MÉDIO ATÉ A RESOLUÇÃO", last_col=6)
    _label_value(
        sheet, s8_bar + 1, "Tempo corrido médio até a resolução (todas as linhas):",
        f"=AVERAGE({rng(_AG)})", fmt=_DUR,
    )
    _label_value(
        sheet, s8_bar + 2, "Mediana do tempo corrido até a resolução:",
        f"=MEDIAN({rng(_AG)})", fmt=_DUR,
    )
    # M-01: `_AG` já devolve "" (texto, ignorado por AVERAGE/MEDIAN) para
    # linhas com data ausente/malformada ou encerramento anterior à
    # abertura — mas isso ficava implícito; expõe a contagem de linhas
    # rejeitadas para que a média/mediana acima não pareça cobrir 100% dos
    # incidentes sem dizer quantos foram excluídos.
    rejeitados_row = s8_bar + 3
    sheet.cell(
        row=rejeitados_row, column=1,
        value=(
            "Registros excluídos da média/mediana (data ausente/inválida ou "
            "encerramento antes da abertura):"
        ),
    ).font = LABEL_FONT
    rejeitados_cell = sheet.cell(
        row=rejeitados_row, column=2, value=f'=COUNTIF({rng(_AG)},"")'
    )
    rejeitados_cell.font = BODY_FONT
    atraso_row = s8_bar + 4
    sheet.cell(
        row=atraso_row, column=1,
        value="Atraso médio dos registros fora do prazo (vs. limite ITSM, minutos):",
    ).font = LABEL_FONT
    # M-02: seleciona pela coluna calculada `_AD` ("No prazo (data limite
    # ITSM)"), não pela classificação do fornecedor (`_X`) — consistente
    # com o rótulo "vs. limite ITSM"; guarda contra #DIV/0! quando não há
    # nenhum registro fora do prazo por esse critério.
    atraso_cell = sheet.cell(
        row=atraso_row,
        column=2,
        value=(
            f'=IF(COUNTIF({rng(_AD)},"N")=0,"Sem atrasos",'
            f'AVERAGEIF({rng(_AD)},"N",{rng(_AF)}))'
        ),
    )
    atraso_cell.number_format = "0.0"
    atraso_cell.font = BODY_FONT
    note8 = s8_bar + 5
    sheet.merge_cells(f"A{note8}:F{note8}")
    sheet[f"A{note8}"] = (
        "O tempo corrido médio é indicador gerencial complementar — não substitui a "
        "verificação linha-a-linha do campo 'No prazo' para o cálculo do INMS 1.1."
    )
    sheet[f"A{note8}"].font = NOTE_FONT
    return note8 + 2


def _write_section_9_penalidade(
    sheet: Worksheet,
    *,
    penalty_base_points: float,
    penalty_step_points: float,
    penalty_step_size_pct: float,
    start_row: int,
) -> None:
    s9_bar = start_row
    _section_bar(sheet, s9_bar, "SEÇÃO 9 · PENALIDADE (CÁLCULO PRELIMINAR)", last_col=6)
    _label_value(sheet, s9_bar + 1, "Meta:", "=A13", fmt=_PCT4)
    _label_value(sheet, s9_bar + 2, "Resultado:", "=E13", fmt=_PCT4)
    diff9_row = s9_bar + 3
    # Meta mínima ("`>=`") — único operador suportado por este renderer.
    diff_formula = '=IF(B13=0,"",A13-E13)'
    _label_value(sheet, diff9_row, "Diferença (Meta - Resultado):", diff_formula, fmt=_PCT4)
    diffpp_row = diff9_row + 1
    _label_value(
        sheet, diffpp_row, "Diferença em pontos percentuais:",
        f'=IF(B{diff9_row}="","",B{diff9_row}*100)',
        fmt="0.0000",
    )
    base_row = diffpp_row + 1
    below_target = "E13<A13"
    _label_value(
        sheet, base_row, "Penalidade-base:",
        f'=IF(B13=0,"Não aplicável",IF({below_target},{penalty_base_points:g},0))', fmt="0",
    )
    add_row = base_row + 1
    _label_value(
        sheet, add_row,
        f"Adicional proporcional ({penalty_step_points:g} pontos a cada "
        f"{penalty_step_size_pct:g} p.p. — cálculo contínuo):",
        (
            f'=IF(B13=0,"Não aplicável",IF({below_target},'
            f"(B{diffpp_row}/{penalty_step_size_pct!r})*{penalty_step_points!r},0))"
        ),
        fmt="0.0000",
    )
    total_row = add_row + 1
    _label_value(
        sheet, total_row, "Total proporcional (base + adicional):",
        (
            f"=IF(OR(ISTEXT(B{base_row}),ISTEXT(B{add_row})),"
            f'"Não aplicável",B{base_row}+B{add_row})'
        ),
        fmt="0.0000", fill=TEAL_FILL,
    )
    scenario_row = total_row + 1
    _label_value(
        sheet, scenario_row, "Cenário — faixas completas ou iniciadas:",
        (
            f'=IF(B13=0,"Não aplicável",IF({below_target},{penalty_base_points:g}+'
            f"CEILING(B{diffpp_row},{penalty_step_size_pct!r})/{penalty_step_size_pct!r}"
            f"*{penalty_step_points!r},0))"
        ),
        fmt="0", fill=ORANGE_FILL,
    )
    obs_row = scenario_row + 2
    sheet.merge_cells(f"A{obs_row}:L{obs_row}")
    sheet[f"A{obs_row}"] = (
        "Resultado sujeito à confirmação da regra de arredondamento e das disposições "
        "gerais de glosa do Termo de Referência."
    )
    sheet[f"A{obs_row}"].font = Font(name="Arial", size=10, bold=True, color="9A3412")
    sheet[f"A{obs_row}"].fill = ORANGE_FILL
    sheet[f"A{obs_row}"].alignment = Alignment(wrap_text=True)

    final_note_row = obs_row + 2
    sheet.merge_cells(f"A{final_note_row}:L{final_note_row}")
    sheet[f"A{final_note_row}"] = (
        "Dados de apoio às fórmulas desta aba nas colunas R:AM (estrutura de apoio, "
        "auditável) — mantidos para rastreabilidade; não excluir nem reordenar."
    )
    sheet[f"A{final_note_row}"].font = NOTE_FONT


def _validate_write_sheet_params(
    *,
    target_operator: str,
    target_value: float,
    penalty_base_points: float,
    penalty_step_points: float,
    penalty_step_size_pct: float,
) -> None:
    """Validação na fronteira de `write_sheet()` — a função é pública e
    aceita `str`/`float` irrestritos; o Pydantic (`config/models.py`) já
    valida esses mesmos limites quando os parâmetros vêm de um YAML de
    config, mas nada impede uma chamada direta com valores fora de faixa."""
    # A2-03: a memória de cálculo (Seção 3) e a penalidade (Seção 9) só têm
    # semântica de "quantidade mínima dentro do prazo"/"diferença Meta -
    # Resultado" para meta mínima (`>=`); nenhum uso real do INMS 1.1 hoje
    # é meta-teto (`<=`), então restringimos aqui em vez de manter as duas
    # narrativas coerentes ao mesmo tempo — ver ticket 05 (A-03).
    if target_operator != ">=":
        raise ValueError(
            f"target_operator {target_operator!r} não suportado por este renderer — "
            "a aba enriquecida do INMS 1.1 assume meta mínima ('>='); para meta-teto "
            "('<='), use o renderer genérico (_write_grupo_executor_sheet)"
        )
    if not 0 <= target_value <= 100:
        raise ValueError(f"target_value {target_value!r} fora da faixa 0–100")
    if penalty_base_points < 0:
        raise ValueError(f"penalty_base_points {penalty_base_points!r} não pode ser negativo")
    if penalty_step_points < 0:
        raise ValueError(f"penalty_step_points {penalty_step_points!r} não pode ser negativo")
    if penalty_step_size_pct <= 0:
        raise ValueError(
            f"penalty_step_size_pct {penalty_step_size_pct!r} deve ser positivo — "
            "usado como divisor nas fórmulas da Seção 9"
        )


def write_sheet(
    workbook: Workbook,
    sheet_name: str,
    *,
    categorias_file: CategoriasFile,
    grupo_executor_entries: list[tuple[str, GrupoExecutorMode]],
    whole_indicator_entries: list[tuple[str, WholeIndicatorMode]],
    fieldnames: list[str],
    rows: list[dict[str, str]],
    target_operator: str,
    target_value: float,
    penalty_base_points: float,
    penalty_step_points: float,
    penalty_step_size_pct: float,
    contract: str,
    periodo: PeriodoAfericao | None,
    raw_csv_path: Path,
    generated_at: datetime,
) -> None:
    del whole_indicator_entries  # categorias.yaml não usa whole_indicator para 1.1 hoje
    del fieldnames  # validado por has_required_columns antes de chamar

    _validate_write_sheet_params(
        target_operator=target_operator,
        target_value=target_value,
        penalty_base_points=penalty_base_points,
        penalty_step_points=penalty_step_points,
        penalty_step_size_pct=penalty_step_size_pct,
    )

    rows = [{**row, _NO_PRAZO_COLUMN: _normalize_no_prazo(row)} for row in rows]

    with create_sheet_atomic(workbook, sheet_name) as sheet:
        sheet.sheet_view.showGridLines = False
        for col, width in {
            1: 42, 2: 24, 3: 16, 4: 12, 5: 14, 6: 12, 7: 12, 8: 16, 9: 14, 10: 34, 11: 20, 12: 40,
        }.items():
            sheet.column_dimensions[cl(col)].width = width

        last_row = 1 + len(rows)
        real_values = {row[GRUPO_EXECUTOR_COLUMN] for row in rows}
        grupo_rows = _build_grupo_rows(categorias_file, grupo_executor_entries, real_values)
        _write_raw_block(sheet, rows, grupo_rows, last_row)

        def rng(col: int) -> str:
            return _raw_range(col, last_row)

        iap = f"ROWS({rng(_R)})"
        iadp = f'COUNTIF({rng(_X)},"S")'
        fora = f'COUNTIF({rng(_X)},"N")'
        meta_value = target_value / 100

        table_names = {
            "grupo_executor": unique_table_name(workbook, "TabelaGrupoExecutor"),
            "fora_do_prazo": unique_table_name(workbook, "TabelaForaDoPrazo"),
            "amostra_divergencias": unique_table_name(workbook, "TabelaAmostraDivergencias"),
        }

        _write_section_1_identificacao(
            sheet,
            rows=rows,
            contract=contract,
            periodo=periodo,
            raw_csv_path=raw_csv_path,
            generated_at=generated_at,
        )
        next_row = _write_section_2_resumo(
            sheet, iap=iap, iadp=iadp, fora=fora, meta_value=meta_value,
        )
        next_row = _write_section_3_memoria(sheet)
        next_row = _write_section_4_detalhamento(
            sheet, grupo_rows=grupo_rows, rng=rng, start_row=next_row,
            table_name=table_names["grupo_executor"],
        )
        next_row = _write_section_5_subtotais(sheet, rng=rng, start_row=next_row)
        next_row = _write_section_6_fora_prazo(
            sheet, rows=rows, rng=rng, start_row=next_row,
            table_name=table_names["fora_do_prazo"],
        )
        next_row = _write_section_7_auditoria(
            sheet, rows=rows, rng=rng, start_row=next_row,
            table_name=table_names["amostra_divergencias"],
        )
        next_row = _write_section_8_tempo(sheet, rng=rng, start_row=next_row)
        _write_section_9_penalidade(
            sheet,
            penalty_base_points=penalty_base_points,
            penalty_step_points=penalty_step_points,
            penalty_step_size_pct=penalty_step_size_pct,
            start_row=next_row,
        )

        _protect_support_columns(sheet)
        sheet.freeze_panes = "A2"

    force_recalc(workbook)

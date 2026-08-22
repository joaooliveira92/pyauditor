"""`sintetico.xlsx` (spec §14.4, ticket 05) — um workbook por órgão/
competência, uma aba por INMS com entrada em `categorias.yaml` (exclui
1.8/1.10, que não têm entrada nenhuma), mais três abas opcionais que
reproduzem `input/{capa,equipe,prazos}.csv` verbatim (cabeçalho + linhas,
sem processamento) — "Capa", "Equipe" e "Prazos", nessa ordem, cada uma só
quando o respectivo `*_path` é passado. `objetos.csv` não vira aba própria:
a pedido do usuário, é anexado ao final da aba "Capa" (depois de "Término
da vigência", a última linha de `capa.csv`), separado por uma linha em
branco — só quando `capa_path` também é passado. Como são escritas antes
do laço de INMS, acabam sendo as primeiras abas do workbook.

Contagens são brutas/pré-quality-gate — conferência rápida, não substitui o
ROM da categoria. A única exceção é `Tempo médio criação→resolução`,
restrito às linhas aprovadas pelo quality gate, conforme a spec.

As colunas assumidas (`No prazo`, `DataHoraSolicitacao`, `DataHoraFim`)
valem para os datasets linha-a-linha (INMS 1.1-1.3, 1.7ish) contra os quais
`split` foi construído. Dois shapes de `calculation` têm renderer dedicado
em vez de degradar pra "—":

- `precomputed_table` (INMS 1.4/1.5/1.9/1.13 — 1.14 é o caso multi-ativo x
  categoria abaixo, e 1.8/1.10 não têm entrada em `categorias.yaml`): uma
  linha por `name_column`, com `result_column`/`penalty_column` lidos
  direto do CSV e "Meta atingida?" derivada de `target`, não da coluna
  crua `atingiu_meta` (que não faz parte do schema).
- `ratio` com `aggregation: sum` + `sum_numerator_subtract_column` (hoje só
  INMS 1.6): uma linha por valor distinto da coluna de `denominator_filter`
  (ex. "Acordo de Nível de Serviço"), somando os CSVs elegíveis por grupo —
  mesma aritmética de `RatioStrategy._aggregate`.

Os demais `whole_indicator` (ex. INMS 1.11/1.12, `ratio` sobre CDR de
telefonia com colunas de data/prazo próprias) continuam degradando pra
"—" — fog, não coberto ainda.

INMS 1.14 (spec §14.5, ticket 06) é o único caso de composição multi-ativo x
categoria: sua aba troca a coluna `Grupo executor` por `Ativo`, lida do
`name_column` da config `precomputed_table` (mesma dataset da medição real,
sem recalcular nada), e agrupa/subtotaliza por Categoria em vez de Nível
(as duas categorias que o 1.14 integra são ambas N3 — subtotalizar por
Nível não distinguiria nada).

INMS 1.1 tem um terceiro renderer dedicado, em `excel/inms_1_1_audit.py`:
resumo executivo, memória de cálculo, incidentes fora do prazo e auditoria
de prazo contratual, tudo via fórmulas do Excel sobre os dados brutos
embutidos na própria aba (não valores pré-calculados em Python). Só entra
em jogo quando o CSV bruto tem as colunas de detalhe de produção (`Nº
Solicitacao`, `Atividades`, `DataHoraLimite`, `TecnicoExecutor` — ver
`inms_1_1_audit.has_required_columns`); CSVs minimalistas (fixtures de
teste) caem no `_write_grupo_executor_sheet` genérico abaixo, como antes.
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from math import isnan
from pathlib import Path
from typing import Final

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from pyauditor.atomic_write import atomic_write
from pyauditor.categoria_filter import (
    GRUPO_EXECUTOR_COLUMN,
    base_config_stem,
    compute_categoria_values,
)
from pyauditor.config.categorias import (
    CategoriasFile,
    GrupoExecutorMode,
    WholeIndicatorMode,
)
from pyauditor.config.manifest import DatasetManifest
from pyauditor.config.models import PrecomputedTableCalculation, RatioCalculation
from pyauditor.engine.pipeline import load_config, measurement_source
from pyauditor.engine.strategies import filter_rows, meets_target, parse_decimal, safe_pct
from pyauditor.excel import inms_1_1_audit
from pyauditor.excel._csv_verbatim import read_csv_verbatim
from pyauditor.excel._style import LABEL_FONT, new_sheet, write_row
from pyauditor.excel.capa import CAPA_DELIMITER, CAPA_ENCODING
from pyauditor.excel.equipe import EQUIPE_DELIMITER, EQUIPE_ENCODING
from pyauditor.excel.objetos import OBJETOS_DELIMITER, OBJETOS_ENCODING
from pyauditor.excel.prazos import PRAZOS_DELIMITER, PRAZOS_ENCODING, PRAZOS_SHEET_NAME
from pyauditor.periodo import PeriodoAfericao

__all__: Final[tuple[str, ...]] = ("write_sintetico_workbook",)

_CAPA_SHEET_NAME: Final[str] = "Capa"
_EQUIPE_SHEET_NAME: Final[str] = "Equipe"

_INMS_1_1: Final[str] = "1.1"
_INMS_1_14: Final[str] = "1.14"
_OUTROS_LABEL: Final[str] = "outros (não contabilizado na meta)"
_WHOLE_INDICATOR_LABEL: Final[str] = "(indicador inteiro)"
_NAO_ATIVADO_TEXT: Final[str] = "Esse serviço não foi requisitado no período selecionado."

_COLUMNS: Final[tuple[str, ...]] = (
    "Categoria",
    "Nível",
    "Grupo executor",
    "Linhas",
    "Dentro do prazo",
    "Fora do prazo",
    "% bruto",
    "Tempo médio criação→resolução",
)
_SUBTOTAL_COLUMNS: Final[tuple[str, ...]] = (
    "Nível",
    "Linhas",
    "Dentro do prazo",
    "Fora do prazo",
    "% bruto",
    "Tempo médio criação→resolução",
)
_NIVEL_ORDER: Final[tuple[str, ...]] = ("N1", "N2", "N3")
_NIVEL_BY_CATEGORIA: Final[dict[str, str]] = {
    "ATENDIMENTO_N1": "N1",
    "ATENDIMENTO_N2": "N2",
    "OPERACAO_N3": "N3",
    "MONITORAMENTO_NOC_SOC": "N3",
}

_ATIVO_COLUMNS: Final[tuple[str, ...]] = (
    "Categoria",
    "Nível",
    "Ativo",
    "Linhas",
    "Dentro do prazo",
    "Fora do prazo",
    "% bruto",
    "Tempo médio criação→resolução",
)
_ATIVO_SUBTOTAL_COLUMNS: Final[tuple[str, ...]] = (
    "Categoria",
    "Linhas",
    "Dentro do prazo",
    "Fora do prazo",
    "% bruto",
    "Tempo médio criação→resolução",
)
# Ordem fixa da spec §14.5: bloco NOC/SOC antes do bloco Operação N3,
# independente da ordem de declaração em categorias.yaml.
_INMS_1_14_CATEGORIA_ORDER: Final[tuple[str, ...]] = ("MONITORAMENTO_NOC_SOC", "OPERACAO_N3")

_NO_PRAZO_COLUMN: Final[str] = "No prazo"
_DATA_SOLICITACAO_COLUMN: Final[str] = "DataHoraSolicitacao"
_DATA_FIM_COLUMN: Final[str] = "DataHoraFim"
_DATAHORA_FORMAT: Final[str] = "%d/%m/%Y %H:%M"


def _parse_datahora(raw: str) -> datetime | None:
    raw = raw.strip()
    if not raw:
        return None
    try:
        return datetime.strptime(raw, _DATAHORA_FORMAT)
    except ValueError:
        return None


@dataclass(frozen=True, slots=True)
class _Stats:
    linhas: int
    dentro: int | None
    fora: int | None
    duracao_total_segundos: float
    duracao_contagem: int


def _compute_stats(
    rows: list[dict[str, str]], fieldnames: list[str], accepted_ids: set[int]
) -> _Stats:
    dentro: int | None = None
    fora: int | None = None
    if _NO_PRAZO_COLUMN in fieldnames:
        dentro = sum(1 for row in rows if row.get(_NO_PRAZO_COLUMN) == "S")
        fora = sum(1 for row in rows if row.get(_NO_PRAZO_COLUMN) == "N")

    duracao_total = 0.0
    duracao_contagem = 0
    if _DATA_SOLICITACAO_COLUMN in fieldnames and _DATA_FIM_COLUMN in fieldnames:
        for row in rows:
            if id(row) not in accepted_ids:
                continue
            inicio = _parse_datahora(row.get(_DATA_SOLICITACAO_COLUMN, ""))
            fim = _parse_datahora(row.get(_DATA_FIM_COLUMN, ""))
            if inicio is None or fim is None:
                continue
            duracao_total += (fim - inicio).total_seconds()
            duracao_contagem += 1

    return _Stats(
        linhas=len(rows),
        dentro=dentro,
        fora=fora,
        duracao_total_segundos=duracao_total,
        duracao_contagem=duracao_contagem,
    )


def _format_duracao(segundos: float) -> str:
    total_minutos = round(segundos / 60)
    dias, resto_minutos = divmod(total_minutos, 24 * 60)
    horas = resto_minutos // 60
    return f"{dias}d {horas:02d}h"


def _fmt_pt_br(value: float, *, decimals: int = 1) -> str:
    """Separador decimal pt-BR (`,`) — mesma convenção de
    `orchestration/summary.py.fmt_pt_br`, duplicada aqui em miniatura pra
    `excel/` não depender de `orchestration/` (que por sua vez depende de
    `cli/split.py`, o chamador deste módulo — importar de volta ciclaria)."""
    return f"{value:.{decimals}f}".replace(".", ",")


def _format_pct_bruto(dentro: int | None, fora: int | None) -> str:
    if dentro is None or fora is None or (dentro + fora) == 0:
        return "—"
    pct = dentro / (dentro + fora) * 100
    return f"{_fmt_pt_br(pct)}%"


def _format_row(stats: _Stats) -> tuple[int, str | int, str | int, str, str]:
    dentro_display: str | int = stats.dentro if stats.dentro is not None else "—"
    fora_display: str | int = stats.fora if stats.fora is not None else "—"
    pct_display = _format_pct_bruto(stats.dentro, stats.fora)
    tempo_display = (
        _format_duracao(stats.duracao_total_segundos / stats.duracao_contagem)
        if stats.duracao_contagem > 0
        else "—"
    )
    return stats.linhas, dentro_display, fora_display, pct_display, tempo_display


@dataclass(frozen=True, slots=True)
class _NivelAccumulator:
    linhas: int = 0
    dentro: int = 0
    fora: int = 0
    tem_prazo: bool = False
    duracao_total_segundos: float = 0.0
    duracao_contagem: int = 0

    def add(self, stats: _Stats) -> _NivelAccumulator:
        return _NivelAccumulator(
            linhas=self.linhas + stats.linhas,
            dentro=self.dentro + (stats.dentro or 0),
            fora=self.fora + (stats.fora or 0),
            tem_prazo=self.tem_prazo or stats.dentro is not None,
            duracao_total_segundos=self.duracao_total_segundos + stats.duracao_total_segundos,
            duracao_contagem=self.duracao_contagem + stats.duracao_contagem,
        )


def _write_nao_ativado_sheet(workbook: Workbook, sheet_name: str) -> None:
    sheet = workbook.create_sheet(sheet_name)
    sheet.sheet_view.showGridLines = False
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(_COLUMNS))
    cell = sheet.cell(row=2, column=1, value=_NAO_ATIVADO_TEXT)
    cell.font = LABEL_FONT


def _write_subtotals(
    sheet: Worksheet, start_row: int, accumulators: dict[str, _NivelAccumulator]
) -> None:
    header_cell = sheet.cell(row=start_row, column=1, value="Subtotais por Nível")
    header_cell.font = LABEL_FONT
    for col_idx, column in enumerate(_SUBTOTAL_COLUMNS, start=1):
        cell = sheet.cell(row=start_row + 1, column=col_idx, value=column)
        cell.font = LABEL_FONT

    row_idx = start_row + 2
    for nivel in _NIVEL_ORDER:
        acc = accumulators.get(nivel)
        if acc is None:
            continue
        dentro = acc.dentro if acc.tem_prazo else None
        fora = acc.fora if acc.tem_prazo else None
        pct_display = _format_pct_bruto(dentro, fora)
        tempo_display = (
            _format_duracao(acc.duracao_total_segundos / acc.duracao_contagem)
            if acc.duracao_contagem > 0
            else "—"
        )
        write_row(
            sheet,
            row_idx,
            (
                nivel,
                acc.linhas,
                dentro if dentro is not None else "—",
                fora if fora is not None else "—",
                pct_display,
                tempo_display,
            ),
            expected_columns=len(_SUBTOTAL_COLUMNS),
        )
        row_idx += 1


def _write_ativo_subtotals(
    sheet: Worksheet,
    start_row: int,
    order: list[str],
    labels: dict[str, str],
    accumulators: dict[str, _NivelAccumulator],
) -> None:
    header_cell = sheet.cell(row=start_row, column=1, value="Subtotais por Categoria")
    header_cell.font = LABEL_FONT
    for col_idx, column in enumerate(_ATIVO_SUBTOTAL_COLUMNS, start=1):
        cell = sheet.cell(row=start_row + 1, column=col_idx, value=column)
        cell.font = LABEL_FONT

    row_idx = start_row + 2
    for categoria_key in order:
        acc = accumulators.get(categoria_key)
        if acc is None:
            continue
        dentro = acc.dentro if acc.tem_prazo else None
        fora = acc.fora if acc.tem_prazo else None
        pct_display = _format_pct_bruto(dentro, fora)
        tempo_display = (
            _format_duracao(acc.duracao_total_segundos / acc.duracao_contagem)
            if acc.duracao_contagem > 0
            else "—"
        )
        write_row(
            sheet,
            row_idx,
            (
                labels[categoria_key],
                acc.linhas,
                dentro if dentro is not None else "—",
                fora if fora is not None else "—",
                pct_display,
                tempo_display,
            ),
            expected_columns=len(_ATIVO_SUBTOTAL_COLUMNS),
        )
        row_idx += 1


def _write_multi_ativo_sheet(
    workbook: Workbook,
    sheet_name: str,
    categorias_file: CategoriasFile,
    entries: list[tuple[str, WholeIndicatorMode]],
    ativo_column: str,
    fieldnames: list[str],
    rows: list[dict[str, str]],
    accepted_ids: set[int],
) -> None:
    """INMS 1.14 (spec §14.5): produto cartesiano ativo x categoria — as 6
    medições por ativo (já presentes no CSV, uma linha por ativo) são
    duplicadas/rotuladas sob cada categoria à qual o INMS pertence, sem
    recalcular nada. Bloco NOC/SOC antes do bloco Operação N3."""
    sheet = new_sheet(workbook, sheet_name, _ATIVO_COLUMNS)
    row_idx = 2
    accumulators: dict[str, _NivelAccumulator] = {}
    labels: dict[str, str] = {}

    # Ordem de primeira aparição no CSV (Anexo D já lista os 6 ativos numa
    # ordem fixa — File Server, Telefonia, Mensageria, etc. — preservada
    # aqui em vez de reordenar alfabeticamente).
    ativos = list(dict.fromkeys(row[ativo_column] for row in rows if row.get(ativo_column)))

    ordered_entries = sorted(
        entries,
        key=lambda item: (
            _INMS_1_14_CATEGORIA_ORDER.index(item[0])
            if item[0] in _INMS_1_14_CATEGORIA_ORDER
            else len(_INMS_1_14_CATEGORIA_ORDER),
            item[0],
        ),
    )

    for categoria_key, _entry in ordered_entries:
        categoria = categorias_file.categorias[categoria_key]
        nivel = _NIVEL_BY_CATEGORIA.get(categoria_key)
        labels[categoria_key] = categoria.label
        for ativo in ativos:
            ativo_rows = [row for row in rows if row.get(ativo_column) == ativo]
            stats = _compute_stats(ativo_rows, fieldnames, accepted_ids)
            linhas, dentro, fora, pct, tempo = _format_row(stats)
            write_row(
                sheet,
                row_idx,
                (
                    categoria.label,
                    nivel or "",
                    ativo,
                    linhas,
                    dentro,
                    fora,
                    pct,
                    tempo,
                ),
            )
            row_idx += 1
            current = accumulators.get(categoria_key, _NivelAccumulator())
            accumulators[categoria_key] = current.add(stats)

    if accumulators:
        order = [categoria_key for categoria_key, _entry in ordered_entries]
        _write_ativo_subtotals(sheet, row_idx + 1, order, labels, accumulators)


def _write_grupo_executor_sheet(
    workbook: Workbook,
    sheet_name: str,
    categorias_file: CategoriasFile,
    grupo_executor_entries: list[tuple[str, GrupoExecutorMode]],
    whole_indicator_entries: list[tuple[str, WholeIndicatorMode]],
    fieldnames: list[str],
    rows: list[dict[str, str]],
    accepted_ids: set[int],
) -> None:
    sheet = new_sheet(workbook, sheet_name, _COLUMNS)
    row_idx = 2
    accumulators: dict[str, _NivelAccumulator] = {}

    real_values = {row[GRUPO_EXECUTOR_COLUMN] for row in rows}
    per_categoria_values, outros_values = compute_categoria_values(
        grupo_executor_entries, real_values
    )

    def _emit(
        categoria_label: str, nivel: str | None, grupo: str, group_rows: list[dict[str, str]]
    ) -> None:
        nonlocal row_idx
        stats = _compute_stats(group_rows, fieldnames, accepted_ids)
        linhas, dentro, fora, pct, tempo = _format_row(stats)
        write_row(
            sheet, row_idx, (categoria_label, nivel or "", grupo, linhas, dentro, fora, pct, tempo)
        )
        row_idx += 1
        if nivel is not None:
            current = accumulators.get(nivel, _NivelAccumulator())
            accumulators[nivel] = current.add(stats)

    for categoria_key, effective_values in per_categoria_values.items():
        categoria = categorias_file.categorias[categoria_key]
        nivel = _NIVEL_BY_CATEGORIA.get(categoria_key)
        for grupo in sorted(effective_values):
            group_rows = [row for row in rows if row[GRUPO_EXECUTOR_COLUMN] == grupo]
            _emit(categoria.label, nivel, grupo, group_rows)

    for categoria_key, _entry in whole_indicator_entries:
        categoria = categorias_file.categorias[categoria_key]
        nivel = _NIVEL_BY_CATEGORIA.get(categoria_key)
        _emit(categoria.label, nivel, _WHOLE_INDICATOR_LABEL, rows)

    for grupo in sorted(outros_values):
        group_rows = [row for row in rows if row[GRUPO_EXECUTOR_COLUMN] == grupo]
        _emit(_OUTROS_LABEL, None, grupo, group_rows)

    if accumulators:
        _write_subtotals(sheet, row_idx + 1, accumulators)


def _write_whole_indicator_sheet(
    workbook: Workbook,
    sheet_name: str,
    categorias_file: CategoriasFile,
    entries: list[tuple[str, WholeIndicatorMode]],
    fieldnames: list[str],
    rows: list[dict[str, str]],
    accepted_ids: set[int],
) -> None:
    sheet = new_sheet(workbook, sheet_name, _COLUMNS)
    stats = _compute_stats(rows, fieldnames, accepted_ids)
    linhas, dentro, fora, pct, tempo = _format_row(stats)
    row_idx = 2
    for categoria_key, _entry in entries:
        categoria = categorias_file.categorias[categoria_key]
        nivel = _NIVEL_BY_CATEGORIA.get(categoria_key)
        write_row(
            sheet,
            row_idx,
            (
                categoria.label,
                nivel or "",
                _WHOLE_INDICATOR_LABEL,
                linhas,
                dentro,
                fora,
                pct,
                tempo,
            ),
        )
        row_idx += 1


_PRECOMPUTED_COLUMNS: Final[tuple[str, ...]] = (
    "Categoria",
    "Nível",
    "Item",
    "Resultado",
    "Meta atingida?",
    "Penalidade",
)


def _meta_atingida_display(value: float, target_operator: str, target_value: float) -> str:
    return "Sim" if meets_target(value, target_operator, target_value) else "Não"


def _write_precomputed_table_sheet(
    workbook: Workbook,
    sheet_name: str,
    categorias_file: CategoriasFile,
    entries: list[tuple[str, WholeIndicatorMode]],
    calculation: PrecomputedTableCalculation,
    target_operator: str,
    target_value: float,
    rows: list[dict[str, str]],
) -> None:
    """INMS 1.4/1.5/1.9/1.13: uma linha por `name_column` (o mesmo dado que a
    medição oficial lê — `PrecomputedTableStrategy`), em vez do colapso
    "(indicador inteiro)"/traços. "Meta atingida?" reusa `meets_target`
    contra `target`, não a coluna crua `atingiu_meta` do CSV."""
    sheet = new_sheet(workbook, sheet_name, _PRECOMPUTED_COLUMNS)
    row_idx = 2
    for categoria_key, _entry in entries:
        categoria = categorias_file.categorias[categoria_key]
        nivel = _NIVEL_BY_CATEGORIA.get(categoria_key)
        for row in rows:
            raw_value = row.get(calculation.result_column, "")
            if not raw_value.strip():
                continue  # linhas vazias/placeholder (';' sobrando), sem medição
            value = parse_decimal(raw_value)
            if isnan(value):
                continue

            name = row.get(calculation.name_column, "") if calculation.name_column else ""
            resultado_display = (
                f"{_fmt_pt_br(value)}%"
                if calculation.result_is_percent
                else _fmt_pt_br(value, decimals=0)
            )
            meta_display = (
                _meta_atingida_display(value, target_operator, target_value)
                if calculation.result_is_percent
                else "—"
            )
            penalidade_raw = (
                row.get(calculation.penalty_column, "") if calculation.penalty_column else ""
            )
            penalidade_value = (
                parse_decimal(penalidade_raw) if penalidade_raw.strip() else float("nan")
            )
            penalidade_display = (
                "—" if isnan(penalidade_value) else _fmt_pt_br(penalidade_value, decimals=0)
            )

            write_row(
                sheet,
                row_idx,
                (
                    categoria.label,
                    nivel or "",
                    name,
                    resultado_display,
                    meta_display,
                    penalidade_display,
                ),
            )
            row_idx += 1


def _write_ratio_aggregate_sheet(
    workbook: Workbook,
    sheet_name: str,
    categorias_file: CategoriasFile,
    entries: list[tuple[str, WholeIndicatorMode]],
    calculation: RatioCalculation,
    target_operator: str,
    target_value: float,
    rows: list[dict[str, str]],
) -> None:
    """INMS 1.6 (e qualquer futuro `ratio`/`sum` com
    `sum_numerator_subtract_column`): uma linha por valor distinto da coluna
    de `denominator_filter` (ex. "Acordo de Nível de Serviço"), agregando os
    CSVs elegíveis por grupo com a mesma aritmética de
    `RatioStrategy._aggregate`, em vez do colapso "(indicador inteiro)"."""
    assert calculation.denominator_filter is not None
    assert calculation.sum_numerator_column is not None
    assert calculation.sum_numerator_subtract_column is not None
    group_column = calculation.denominator_filter.column
    numerator_column = calculation.sum_numerator_column
    subtract_column = calculation.sum_numerator_subtract_column

    columns = (
        "Categoria",
        "Nível",
        group_column,
        numerator_column,
        subtract_column,
        "% resultado",
        "Meta atingida?",
    )
    sheet = new_sheet(workbook, sheet_name, columns)
    row_idx = 2

    eligible_rows = filter_rows(rows, calculation.denominator_filter)
    grupos = list(
        dict.fromkeys(row[group_column] for row in eligible_rows if row.get(group_column))
    )

    for categoria_key, _entry in entries:
        categoria = categorias_file.categorias[categoria_key]
        nivel = _NIVEL_BY_CATEGORIA.get(categoria_key)
        for grupo in grupos:
            grupo_rows = [row for row in eligible_rows if row[group_column] == grupo]
            total = sum(parse_decimal(row.get(numerator_column, "") or "0") for row in grupo_rows)
            subtraido = sum(
                parse_decimal(row.get(subtract_column, "") or "0") for row in grupo_rows
            )
            pct = safe_pct(total - subtraido, total)
            meta_display = _meta_atingida_display(pct, target_operator, target_value)
            write_row(
                sheet,
                row_idx,
                (
                    categoria.label,
                    nivel or "",
                    grupo,
                    int(total),
                    int(subtraido),
                    f"{_fmt_pt_br(pct)}%",
                    meta_display,
                ),
            )
            row_idx += 1


def _write_csv_verbatim_sheet(
    workbook: Workbook,
    sheet_name: str,
    path: Path,
    *,
    delimiter: str,
    encoding: str,
    warnings: list[str],
) -> tuple[Worksheet, int] | None:
    """Aba de reprodução verbatim de um CSV bruto de `input/`
    (`capa.csv`/`equipe.csv`/`prazos.csv`) — cabeçalho e linhas exatamente
    como estão no arquivo, sem processamento. Devolve `(sheet, última linha
    escrita)` em caso de sucesso (usado por `_write_capa_sheet` pra saber
    onde continuar), ou `None` se a aba não foi gerada."""
    try:
        header, rows = read_csv_verbatim(path, delimiter=delimiter, encoding=encoding)
    except FileNotFoundError:
        warnings.append(f"sintetico.xlsx: {path} não encontrado — aba '{sheet_name}' não gerada")
        return None
    except (OSError, ValueError) as exc:
        warnings.append(
            f"sintetico.xlsx: falha ao ler {path}: {exc} — aba '{sheet_name}' não gerada"
        )
        return None

    sheet = new_sheet(workbook, sheet_name, tuple(header))
    row_idx = 1
    for row_idx, row in enumerate(rows, start=2):
        write_row(sheet, row_idx, tuple(row))
    return sheet, row_idx


def _write_capa_sheet(
    workbook: Workbook,
    capa_path: Path,
    objetos_path: Path | None,
    warnings: list[str],
) -> None:
    """Aba "Capa": reprodução verbatim de `capa.csv`. Quando `objetos_path`
    é passado, `objetos.csv` é anexado logo abaixo (a pedido do usuário: não
    vira aba própria, fica junto da capa, depois da última linha — "Término
    da vigência" no CSV atual), separado por uma linha em branco e o próprio
    cabeçalho de `objetos.csv`."""
    written = _write_csv_verbatim_sheet(
        workbook,
        _CAPA_SHEET_NAME,
        capa_path,
        delimiter=CAPA_DELIMITER,
        encoding=CAPA_ENCODING,
        warnings=warnings,
    )
    if written is None or objetos_path is None:
        return
    sheet, last_row = written

    try:
        objetos_header, objetos_rows = read_csv_verbatim(
            objetos_path, delimiter=OBJETOS_DELIMITER, encoding=OBJETOS_ENCODING
        )
    except FileNotFoundError:
        warnings.append(
            f"sintetico.xlsx: {objetos_path} não encontrado — dados de objetos "
            f"não anexados à aba '{_CAPA_SHEET_NAME}'"
        )
        return
    except (OSError, ValueError) as exc:
        warnings.append(
            f"sintetico.xlsx: falha ao ler {objetos_path}: {exc} — dados de objetos "
            f"não anexados à aba '{_CAPA_SHEET_NAME}'"
        )
        return

    header_row = last_row + 2  # 1 linha em branco de separação
    objetos_columns = len(objetos_header)
    write_row(sheet, header_row, tuple(objetos_header), expected_columns=objetos_columns)
    for row_idx, row in enumerate(objetos_rows, start=header_row + 1):
        write_row(sheet, row_idx, tuple(row), expected_columns=objetos_columns)


def write_sintetico_workbook(
    categorias_file: CategoriasFile,
    config_dir: Path,
    competencia_data_dir: Path,
    output_path: Path,
    *,
    manifest: DatasetManifest | None = None,
    periodo: PeriodoAfericao | None = None,
    strict: bool = False,
    prazos_path: Path | None = None,
    capa_path: Path | None = None,
    equipe_path: Path | None = None,
    objetos_path: Path | None = None,
    generated_at: datetime | None = None,
) -> list[str]:
    """Constrói e grava `sintetico.xlsx` de um órgão/competência. Devolve
    warnings (nunca lança por causa do problema de um único INMS — um
    config ruim ou um dataset genuinamente ausente pula/degrada a aba
    daquele INMS em vez de derrubar o workbook inteiro).

    `generated_at` (default `datetime.now()`, resolvido uma única vez aqui)
    é repassado à aba enriquecida do INMS 1.1 — injetável para permitir
    teste determinístico sem depender do relógio real (ticket 18 / B-01)."""
    warnings: list[str] = []
    generated_at = generated_at if generated_at is not None else datetime.now()

    per_inms: dict[str, list[tuple[str, GrupoExecutorMode | WholeIndicatorMode]]] = {}
    for categoria_key, categoria in categorias_file.categorias.items():
        for inms_key, entry in categoria.inms.items():
            per_inms.setdefault(inms_key, []).append((categoria_key, entry))

    workbook = Workbook()
    default_sheet = workbook.active
    assert default_sheet is not None
    workbook.remove(default_sheet)

    if capa_path is not None:
        _write_capa_sheet(workbook, capa_path, objetos_path, warnings)
    elif objetos_path is not None:
        warnings.append(
            f"sintetico.xlsx: capa_path não informado — dados de {objetos_path} "
            f"não anexados (dependem da aba '{_CAPA_SHEET_NAME}')"
        )
    if equipe_path is not None:
        _write_csv_verbatim_sheet(
            workbook,
            _EQUIPE_SHEET_NAME,
            equipe_path,
            delimiter=EQUIPE_DELIMITER,
            encoding=EQUIPE_ENCODING,
            warnings=warnings,
        )
    if prazos_path is not None:
        _write_csv_verbatim_sheet(
            workbook,
            PRAZOS_SHEET_NAME,
            prazos_path,
            delimiter=PRAZOS_DELIMITER,
            encoding=PRAZOS_ENCODING,
            warnings=warnings,
        )
    sheets_before_inms = set(workbook.sheetnames)

    for inms_key in sorted(per_inms, key=lambda k: int(k.split(".")[1])):
        entries = per_inms[inms_key]
        sheet_name = f"INMS {inms_key}"

        try:
            base_stem = base_config_stem(inms_key)
            base_config = load_config(config_dir / f"{base_stem}.yaml")
        except (OSError, ValueError) as exc:
            warning = f"sintetico.xlsx: INMS {inms_key}: falha ao carregar config base: {exc}"
            warnings.append(warning)
            continue

        # Backbone (ticket 03): resolve→valida→lê→filtra→gates, um só lugar
        # para os quatro reimplementadores do pipeline de medição. Sintetico
        # nunca emite WARN de janela vazia (a mesma passada do bruto coube
        # ao split) — `emit_empty_window_warning=False`. `period_column`
        # ausente/não presente no header é erro acionável (spec §2 ponto 1,
        # issue 01 item 5) — degrada esta aba como os demais erros do loop.
        try:
            bundle = measurement_source(
                base_config,
                competencia_data_dir,
                manifest,
                config_path=config_dir / f"{base_stem}.yaml",
                periodo=periodo,
                strict=strict,
                emit_empty_window_warning=False,
            )
        except FileNotFoundError:
            _write_nao_ativado_sheet(workbook, sheet_name)
            continue
        except (OSError, ValueError) as exc:
            warnings.append(f"sintetico.xlsx: INMS {inms_key}: {exc}")
            continue

        raw_csv_path = bundle.csv_path
        fieldnames = bundle.fieldnames
        rows = bundle.rows
        accepted_ids = bundle.accepted_ids

        whole_indicator_entries = [
            (ck, e) for ck, e in entries if isinstance(e, WholeIndicatorMode)
        ]

        if inms_key == _INMS_1_14:
            calculation = base_config.calculation
            if not isinstance(calculation, PrecomputedTableCalculation) or (
                calculation.name_column is None
            ):
                warning = (
                    f"sintetico.xlsx: INMS {inms_key}: config base não é "
                    "'precomputed_table' com 'name_column' — aba não gerada"
                )
                warnings.append(warning)
                continue
            _write_multi_ativo_sheet(
                workbook,
                sheet_name,
                categorias_file,
                whole_indicator_entries,
                calculation.name_column,
                fieldnames,
                rows,
                accepted_ids,
            )
            continue

        grupo_executor_entries = [(ck, e) for ck, e in entries if isinstance(e, GrupoExecutorMode)]

        if grupo_executor_entries:
            if GRUPO_EXECUTOR_COLUMN not in fieldnames:
                warning = (
                    f"sintetico.xlsx: INMS {inms_key}: {raw_csv_path} não tem coluna "
                    f"'{GRUPO_EXECUTOR_COLUMN}' — aba não gerada"
                )
                warnings.append(warning)
                continue
            if (
                inms_key == _INMS_1_1
                and base_config.target is not None
                and base_config.penalty is not None
                and inms_1_1_audit.has_required_columns(fieldnames)
            ):
                # Aba enriquecida (resumo executivo, memória de cálculo,
                # fora-do-prazo, auditoria de prazo) — só quando o CSV bruto
                # tem as colunas de detalhe (produção real, ambos os
                # órgãos); CSVs minimalistas (ex. fixtures de teste) caem no
                # renderer genérico abaixo, sem degradar silenciosamente.
                try:
                    inms_1_1_audit.write_sheet(
                        workbook,
                        sheet_name,
                        categorias_file=categorias_file,
                        grupo_executor_entries=grupo_executor_entries,
                        whole_indicator_entries=whole_indicator_entries,
                        fieldnames=fieldnames,
                        rows=rows,
                        target_operator=base_config.target.operator,
                        target_value=base_config.target.value,
                        penalty_base_points=base_config.penalty.base_points,
                        penalty_step_points=base_config.penalty.step_points,
                        penalty_step_size_pct=base_config.penalty.step_size_pct,
                        contract=base_config.scope.contract,
                        periodo=periodo,
                        raw_csv_path=raw_csv_path,
                        generated_at=generated_at,
                    )
                except ValueError as exc:
                    # Falha de validação/qualidade de dado específica da aba
                    # enriquecida (ex.: "No prazo" com valor fora de S/N,
                    # target_operator "<=" não suportado, grupo mapeado em
                    # mais de uma categoria) não deve derrubar o restante do
                    # workbook — degrada para o renderer genérico, como as
                    # demais falhas por-INMS deste loop.
                    warning = (
                        f"sintetico.xlsx: INMS {inms_key}: falha ao gerar aba enriquecida "
                        f"do INMS 1.1 ({exc}) — usando renderer genérico"
                    )
                    warnings.append(warning)
                    _write_grupo_executor_sheet(
                        workbook,
                        sheet_name,
                        categorias_file,
                        grupo_executor_entries,
                        whole_indicator_entries,
                        fieldnames,
                        rows,
                        accepted_ids,
                    )
            else:
                _write_grupo_executor_sheet(
                    workbook,
                    sheet_name,
                    categorias_file,
                    grupo_executor_entries,
                    whole_indicator_entries,
                    fieldnames,
                    rows,
                    accepted_ids,
                )
        elif isinstance(base_config.calculation, PrecomputedTableCalculation):
            assert base_config.target is not None
            _write_precomputed_table_sheet(
                workbook,
                sheet_name,
                categorias_file,
                whole_indicator_entries,
                base_config.calculation,
                base_config.target.operator,
                base_config.target.value,
                rows,
            )
        elif (
            isinstance(base_config.calculation, RatioCalculation)
            and base_config.calculation.aggregation == "sum"
            and base_config.calculation.sum_numerator_subtract_column is not None
            and base_config.calculation.denominator_filter is not None
        ):
            assert base_config.target is not None
            _write_ratio_aggregate_sheet(
                workbook,
                sheet_name,
                categorias_file,
                whole_indicator_entries,
                base_config.calculation,
                base_config.target.operator,
                base_config.target.value,
                rows,
            )
        else:
            _write_whole_indicator_sheet(
                workbook,
                sheet_name,
                categorias_file,
                whole_indicator_entries,
                fieldnames,
                rows,
                accepted_ids,
            )

    if set(workbook.sheetnames) == sheets_before_inms:
        # Nada pôde ser medido (todo INMS de categorias.yaml falhou ao
        # carregar/resolver) — um xlsx sem nenhuma aba de INMS não é um
        # arquivo válido pra este propósito (as abas Capa/Equipe/Prazos
        # sozinhas não contam); não sobrescreve um sintetico.xlsx de rerun
        # anterior com um arquivo vazio/quebrado.
        return warnings

    atomic_write(output_path, workbook.save)
    return warnings

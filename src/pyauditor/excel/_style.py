"""Provide shared styles and worksheet-writing utilities.

The styles defined here follow ``docs/styleguide.md``:

- Arial 10-point font for body cells;
- bold Arial headers with a dark background;
- thin bottom borders for body rows;
- no decorative fills;
- hidden worksheet gridlines.

This module is the canonical source of worksheet styles, shape-to-unit
mappings, and common sheet-writing behavior used by ``capa.py``,
``report.py``, and ``consolidate.py``.
"""

from __future__ import annotations

from types import MappingProxyType
from typing import Final, Mapping, TypeAlias

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

__all__: Final[tuple[str, ...]] = (
    "BODY_FONT",
    "BOTTOM_BORDER",
    "CellValue",
    "HEADER_FILL",
    "HEADER_FONT",
    "LABEL_FONT",
    "LEFT_ALIGN",
    "TITLE_FONT",
    "UNIT_BY_SHAPE",
    "new_sheet",
    "write_row",
)

TITLE_FONT: Final = Font(name="Arial", size=14, bold=True)
LABEL_FONT: Final = Font(name="Arial", size=10, bold=True)
BODY_FONT: Final = Font(name="Arial", size=10)
HEADER_FONT: Final = Font(
    name="Arial",
    size=10,
    bold=True,
    color="FFFFFFFF",
)
HEADER_FILL: Final = PatternFill(
    start_color="FF1F2937",
    end_color="FF1F2937",
    fill_type="solid",
)
BOTTOM_BORDER: Final = Border(
    bottom=Side(style="thin", color="FFD1D5DB"),
)
LEFT_ALIGN: Final = Alignment(horizontal="left")

CellValue: TypeAlias = str | float | int | None

UNIT_BY_SHAPE: Final[Mapping[str, str]] = MappingProxyType(
    {
        "ratio": "%",
        "segmented_ratio": "%",
        "count_difference": "unidades",
        "external_catalog_sum": "pontos",
    }
)


def new_sheet(
    workbook: Workbook,
    name: str,
    columns: tuple[str, ...],
    width: int = 24,
) -> Worksheet:
    """Create and initialize a consistently styled worksheet.

    The function modifies ``workbook`` by creating a worksheet, writing its
    headers in the first row, setting a uniform column width, hiding gridlines,
    and freezing the header row.

    Args:
        workbook: Workbook that will own the new worksheet.
        name: Exact worksheet name. It must be non-empty and unique within the
            workbook.
        columns: Non-empty sequence of column headings.
        width: Positive display width applied to every declared column.

    Returns:
        The newly created and initialized worksheet.

    Raises:
        ValueError: If ``name`` is empty, a worksheet with the same name
            already exists, ``columns`` is empty, a column heading is empty,
            or ``width`` is not positive.
    """
    if not name.strip():
        raise ValueError("Worksheet name must not be empty.")

    if name in workbook.sheetnames:
        raise ValueError(f"Worksheet already exists: {name!r}.")

    if not columns:
        raise ValueError("Worksheet must declare at least one column.")

    if any(not column.strip() for column in columns):
        raise ValueError("Worksheet column headings must not be empty.")

    if isinstance(width, bool) or width <= 0:
        raise ValueError("Worksheet column width must be a positive integer.")

    sheet: Worksheet = workbook.create_sheet(title=name)
    sheet.sheet_view.showGridLines = False

    for column_index, column_name in enumerate(columns, start=1):
        cell = sheet.cell(row=1, column=column_index, value=column_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = LEFT_ALIGN
        sheet.column_dimensions[cell.column_letter].width = width

    sheet.freeze_panes = "A2"
    return sheet


def write_row(
    sheet: Worksheet,
    row_idx: int,
    values: tuple[CellValue, ...],
    *,
    expected_columns: int | None = None,
) -> None:
    """Write and style one complete worksheet body row.

    By default, the number of values must match the number of columns already
    declared on the worksheet (its header row). Existing cells at ``row_idx``
    are overwritten.

    Args:
        sheet: Worksheet that receives the row.
        row_idx: One-based destination row. Body rows must start after the
            header row.
        values: Cell values in the same order as the worksheet columns.
        expected_columns: Column count to validate ``values`` against, instead
            of the worksheet's declared header width. Needed when a sheet
            carries more than one differently-shaped row block (e.g. a second
            verbatim CSV block appended below the sheet's own header).

    Raises:
        ValueError: If ``row_idx`` refers to the header row or an earlier row,
            or if the number of values differs from the expected column
            count.
    """
    if isinstance(row_idx, bool) or row_idx < 2:
        raise ValueError("Body row index must be an integer greater than 1.")

    if expected_columns is None:
        expected_columns = sheet.max_column
    if len(values) != expected_columns:
        raise ValueError(
            "Row value count does not match the worksheet schema: "
            f"expected {expected_columns}, received {len(values)}."
        )

    for column_index, value in enumerate(values, start=1):
        cell = sheet.cell(row=row_idx, column=column_index, value=value)
        cell.font = BODY_FONT
        cell.border = BOTTOM_BORDER
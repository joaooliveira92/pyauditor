import json
from pathlib import Path
from typing import Any
from unittest.mock import patch

from openpyxl import load_workbook

from pyauditor.cli.consolidate import run_consolidate
from pyauditor.cli.report import run_report
from pyauditor.excel.capa import (
    COMMON_FIELD_LABELS,
    ORGAO_FIELD_LABELS,
    bootstrap_capa_csv,
)
from pyauditor.excel.consolidate import (
    CAPA_SHEET as CAPA_E_CONTROLE_SHEET,
)
from pyauditor.excel.consolidate import (
    GLOSAS_SHEET,
    build_consolidated_workbook,
)
from pyauditor.excel.equipe import RESPONSAVEL_LABELS

OBJETOS_CSV = """Item,Categoria,Valor
1,Central de Serviços,"R$ 148.205,54"
2,GT dos Projetos e Operações,"R$ 77.654,90"
3,Banco de Dados,"R$ 43.888,89"
4,"Aplicações, virtualização","R$ 59.694,54"
5,Serviços Corporativos,"R$ 21.035,21"
6,Armazenamento e Backup,"R$ 16.145,94"
7,Redes,"R$ 31.382,28"
8,"Segurança da Informação","R$ 34.143,44"
9,DevOps,"R$ 28.912,84"
"""


def _write_summary(
    roms_dir: Path,
    competencia: str,
    orgao: str,
    indicator_id: str,
    contractual_id: str,
    *,
    penalty_points: float = 0.0,
    conforms: bool = True,
) -> None:
    target_dir = roms_dir / orgao / competencia
    target_dir.mkdir(parents=True, exist_ok=True)
    summary = {
        'indicator_id': indicator_id,
        'contractual_id': contractual_id,
        'name': f'Indicador {contractual_id}',
        'asset': None,
        'orgao': orgao,
        'shape': 'ratio',
        'target_operator': '>=',
        'target_value': 98.0,
        'result_pct': 97.71,
        'conforms': conforms,
        'penalty_points': penalty_points,
        'numerator': 171,
        'denominator': 175,
        'hard_failure': False,
    }
    (target_dir / f'{indicator_id}.json').write_text(
        json.dumps(summary), encoding='utf-8'
    )
    (target_dir / f'{indicator_id}.md').write_text('# ROM', encoding='utf-8')


def _scaffold_capas(tmp_path: Path) -> None:
    bootstrap_capa_csv(tmp_path / 'capa.csv', COMMON_FIELD_LABELS)
    for orgao in ('MinC', 'MTur'):
        bootstrap_capa_csv(tmp_path / f'capa_{orgao}.csv', ORGAO_FIELD_LABELS)
    (tmp_path / 'objetos.csv').write_text(OBJETOS_CSV, encoding='utf-8-sig')


def _build_orgao_report(tmp_path: Path, orgao: str) -> None:
    roms_dir = tmp_path / 'roms'
    _write_summary(
        roms_dir,
        '2026-06',
        orgao,
        f'INMS-1.6-{orgao}',
        'INMS 1.6',
        penalty_points=100.0,
        conforms=False,
    )

    output_path = tmp_path / 'reports' / f'relatorio_2026-06_{orgao}.xlsx'
    exit_code = run_report(
        '2026-06',
        tmp_path / 'capa.csv',
        roms_dir / orgao,
        output_path,
        config_dir=tmp_path / 'configs' / orgao,
        expected_orgao=orgao,
        data_dir=tmp_path,
    )
    assert exit_code.status == 'done'


def test_run_consolidate_forwards_is_final_month_to_workbook(
    tmp_path: Path,
) -> None:
    """Ticket 10: `consolidate --final-month` chega até
    `build_consolidated_workbook` — o rollover de glosa desliga no mês final
    como no `report`."""
    _scaffold_capas(tmp_path)
    _build_orgao_report(tmp_path, 'MinC')
    _build_orgao_report(tmp_path, 'MTur')
    output_path = tmp_path / 'reports' / 'relatorio_2026-06_consolidado.xlsx'
    captured: dict[str, object] = {}

    from unittest.mock import patch

    import pyauditor.cli.consolidate as consolidate_mod

    real_build = build_consolidated_workbook

    def _spy(*args: Any, **kwargs: Any) -> Any:
        captured.update(kwargs)
        return real_build(*args, **kwargs)

    with patch.object(
        consolidate_mod, 'build_consolidated_workbook', side_effect=_spy
    ):
        exit_code = run_consolidate(
            '2026-06',
            tmp_path / 'reports',
            tmp_path / 'roms',
            output_path,
            data_dir=tmp_path,
            is_final_month=True,
        )

    assert exit_code.status == 'done'
    assert captured.get('is_final_month') is True


def test_run_consolidate_rejects_malformed_competencia(tmp_path: Path) -> None:
    result = run_consolidate(
        '../../etc',
        tmp_path / 'reports',
        tmp_path / 'roms',
        tmp_path / 'reports' / 'out.xlsx',
        data_dir=tmp_path,
    )

    assert result.status == 'error'
    assert result.error_message is not None
    assert 'competência inválida' in result.error_message


def test_run_consolidate_converts_unexpected_exception_to_error_result(
    tmp_path: Path,
) -> None:
    _scaffold_capas(tmp_path)
    _build_orgao_report(tmp_path, 'MinC')
    _build_orgao_report(tmp_path, 'MTur')

    with patch(
        'pyauditor.cli.consolidate.build_consolidated_workbook',
        side_effect=ValueError('boom'),
    ):
        result = run_consolidate(
            '2026-06',
            tmp_path / 'reports',
            tmp_path / 'roms',
            tmp_path / 'reports' / 'relatorio_2026-06_consolidado.xlsx',
            data_dir=tmp_path,
        )

    assert result.status == 'error'
    assert result.error_message is not None
    assert 'boom' in result.error_message


def test_run_consolidate_os_error_message_has_actionable_hint(
    tmp_path: Path,
) -> None:
    # Ticket 11: mensagem de falha de escrita ganha dica acionável.
    _scaffold_capas(tmp_path)
    _build_orgao_report(tmp_path, 'MinC')
    _build_orgao_report(tmp_path, 'MTur')

    with patch(
        'pyauditor.cli.consolidate.atomic_write',
        side_effect=PermissionError('Permission denied'),
    ):
        result = run_consolidate(
            '2026-06',
            tmp_path / 'reports',
            tmp_path / 'roms',
            tmp_path / 'reports' / 'relatorio_2026-06_consolidado.xlsx',
            data_dir=tmp_path,
        )

    assert result.status == 'error'
    assert result.error_message is not None
    assert 'aberto em outro programa' in result.error_message


def test_run_consolidate_fails_when_a_report_is_missing(tmp_path: Path) -> None:
    _scaffold_capas(tmp_path)
    _build_orgao_report(tmp_path, 'MinC')
    # MTur report never built.

    exit_code = run_consolidate(
        '2026-06',
        tmp_path / 'reports',
        tmp_path / 'roms',
        tmp_path / 'reports' / 'relatorio_2026-06_consolidado.xlsx',
        data_dir=tmp_path,
    )

    assert exit_code.status == 'error'
    assert not (
        tmp_path / 'reports' / 'relatorio_2026-06_consolidado.xlsx'
    ).exists()


def test_run_consolidate_builds_workbook_from_both_orgaos(
    tmp_path: Path,
) -> None:
    _scaffold_capas(tmp_path)
    _build_orgao_report(tmp_path, 'MinC')
    _build_orgao_report(tmp_path, 'MTur')
    output_path = tmp_path / 'reports' / 'relatorio_2026-06_consolidado.xlsx'

    exit_code = run_consolidate(
        '2026-06',
        tmp_path / 'reports',
        tmp_path / 'roms',
        output_path,
        data_dir=tmp_path,
    )

    assert exit_code.status == 'done'
    assert output_path.exists()
    workbook = load_workbook(output_path)
    sheet = workbook[GLOSAS_SHEET]
    orgaos = {str(sheet.cell(row=r, column=2).value) for r in (2, 3)}
    assert orgaos == {'MinC', 'MTur'}


def test_run_consolidate_capa_tem_competencia_periodo_e_responsaveis(
    tmp_path: Path,
) -> None:
    """Spec competencia-cli-equipe §4/§6 — a capa do consolidado recebe
    Competência/períodos derivados da CLI e responsáveis do `equipe.csv`."""
    _scaffold_capas(tmp_path)
    primeira_funcao = RESPONSAVEL_LABELS[0]
    (tmp_path / 'equipe.csv').write_text(
        f'FUNÇÃO,NOME,SIAPE\n{primeira_funcao},Maria Souza,123456\n',
        encoding='utf-8-sig',
    )
    _build_orgao_report(tmp_path, 'MinC')
    _build_orgao_report(tmp_path, 'MTur')
    output_path = tmp_path / 'reports' / 'relatorio_2026-06_consolidado.xlsx'

    exit_code = run_consolidate(
        '2026-06',
        tmp_path / 'reports',
        tmp_path / 'roms',
        output_path,
        data_dir=tmp_path,
    )

    assert exit_code.status == 'done'
    workbook = load_workbook(output_path)
    sheet = workbook[CAPA_E_CONTROLE_SHEET]
    capa_rows = {
        sheet.cell(row=r, column=1).value: sheet.cell(row=r, column=2).value
        for r in range(4, sheet.max_row + 1)
        if sheet.cell(row=r, column=1).value
    }
    assert capa_rows['Competência'] == '2026-06'
    assert capa_rows['Período inicial da aferição'] == '01/06/2026'
    assert capa_rows['Período final da aferição'] == '30/06/2026'
    assert capa_rows[primeira_funcao] == 'Maria Souza (123456)'


def test_run_consolidate_never_regenerates_the_orgao_reports(
    tmp_path: Path,
) -> None:
    _scaffold_capas(tmp_path)
    _build_orgao_report(tmp_path, 'MinC')
    _build_orgao_report(tmp_path, 'MTur')
    report_path = tmp_path / 'reports' / 'relatorio_2026-06_MinC.xlsx'
    original_mtime = report_path.stat().st_mtime_ns

    run_consolidate(
        '2026-06',
        tmp_path / 'reports',
        tmp_path / 'roms',
        tmp_path / 'reports' / 'relatorio_2026-06_consolidado.xlsx',
        data_dir=tmp_path,
    )

    assert report_path.stat().st_mtime_ns == original_mtime


def test_run_consolidate_rerun_preserves_fiscal_decision(
    tmp_path: Path,
) -> None:
    _scaffold_capas(tmp_path)
    _build_orgao_report(tmp_path, 'MinC')
    _build_orgao_report(tmp_path, 'MTur')
    output_path = tmp_path / 'reports' / 'relatorio_2026-06_consolidado.xlsx'

    run_consolidate(
        '2026-06',
        tmp_path / 'reports',
        tmp_path / 'roms',
        output_path,
        data_dir=tmp_path,
    )

    workbook = load_workbook(output_path)
    sheet = workbook[GLOSAS_SHEET]
    header = [cell.value for cell in sheet[1]]
    decisao_col = header.index('Decisão Fiscal') + 1
    sheet.cell(row=2, column=decisao_col, value='Aceita')
    workbook.save(output_path)

    run_consolidate(
        '2026-06',
        tmp_path / 'reports',
        tmp_path / 'roms',
        output_path,
        data_dir=tmp_path,
    )

    reread = load_workbook(output_path)
    reread_sheet = reread[GLOSAS_SHEET]
    decisoes = {
        str(reread_sheet.cell(row=r, column=decisao_col).value) for r in (2, 3)
    }
    assert 'Aceita' in decisoes


def test_run_consolidate_without_objetos_is_glosa_nao_calculada(
    tmp_path: Path,
) -> None:
    _scaffold_capas(tmp_path)
    _build_orgao_report(tmp_path, 'MinC')
    _build_orgao_report(tmp_path, 'MTur')
    (tmp_path / 'objetos.csv').unlink()
    output_path = tmp_path / 'reports' / 'relatorio_2026-06_consolidado.xlsx'

    exit_code = run_consolidate(
        '2026-06',
        tmp_path / 'reports',
        tmp_path / 'roms',
        output_path,
        data_dir=tmp_path,
    )

    assert exit_code.status == 'done'
    assert exit_code.glosa_calculada is False


def test_run_consolidate_malformed_objetos_is_hard_failure(
    tmp_path: Path,
) -> None:
    _scaffold_capas(tmp_path)
    _build_orgao_report(tmp_path, 'MinC')
    _build_orgao_report(tmp_path, 'MTur')
    (tmp_path / 'objetos.csv').write_text('a;b\n1;2\n', encoding='utf-8-sig')

    result = run_consolidate(
        '2026-06',
        tmp_path / 'reports',
        tmp_path / 'roms',
        tmp_path / 'reports' / 'relatorio_2026-06_consolidado.xlsx',
        data_dir=tmp_path,
    )

    assert result.status == 'error'
    assert result.error_message is not None
    assert 'objetos.csv' in result.error_message


def test_run_consolidate_fails_when_both_reports_are_missing(
    tmp_path: Path,
) -> None:
    _scaffold_capas(tmp_path)
    # Nenhum dos dois relatórios de órgão foi construído.

    exit_code = run_consolidate(
        '2026-06',
        tmp_path / 'reports',
        tmp_path / 'roms',
        tmp_path / 'reports' / 'relatorio_2026-06_consolidado.xlsx',
        data_dir=tmp_path,
    )

    assert exit_code.status == 'error'
    assert exit_code.error_message is not None
    assert 'MinC' in exit_code.error_message
    assert 'MTur' in exit_code.error_message
    assert not (
        tmp_path / 'reports' / 'relatorio_2026-06_consolidado.xlsx'
    ).exists()


def test_run_consolidate_succeeds_with_one_orgao_as_rascunho(
    tmp_path: Path,
) -> None:
    # Ticket 08: `consolidate` só checa existência de arquivo, não
    # `publicable` — um relatório rascunho (capa incompleta) ainda satisfaz
    # `check_consolidate_ready` e entra na consolidação normalmente.
    _scaffold_capas(tmp_path)
    _build_orgao_report(tmp_path, 'MinC')
    _build_orgao_report(tmp_path, 'MTur')
    # MinC vira rascunho: capa_MinC.csv sem fiscais preenchidos — já é o caso
    # (bootstrap_capa_csv deixa tudo em branco), então basta reler o result.
    minc_result = run_report(
        '2026-06',
        tmp_path / 'capa.csv',
        tmp_path / 'roms' / 'MinC',
        tmp_path / 'reports' / 'relatorio_2026-06_MinC.xlsx',
        config_dir=tmp_path / 'configs' / 'MinC',
        expected_orgao='MinC',
        data_dir=tmp_path,
    )
    assert minc_result.publicable is False
    output_path = tmp_path / 'reports' / 'relatorio_2026-06_consolidado.xlsx'

    exit_code = run_consolidate(
        '2026-06',
        tmp_path / 'reports',
        tmp_path / 'roms',
        output_path,
        data_dir=tmp_path,
    )

    assert exit_code.status == 'done'
    assert output_path.exists()


def test_run_consolidate_without_objetos_cells_are_none_not_zero(
    tmp_path: Path,
) -> None:
    # Ticket 01 — a garantia central: "não calculada" nunca vira 0.00, nem no
    # consolidado. Célula vazia (None), não numérica.
    _scaffold_capas(tmp_path)
    _build_orgao_report(tmp_path, 'MinC')
    _build_orgao_report(tmp_path, 'MTur')
    (tmp_path / 'objetos.csv').unlink()
    output_path = tmp_path / 'reports' / 'relatorio_2026-06_consolidado.xlsx'

    exit_code = run_consolidate(
        '2026-06',
        tmp_path / 'reports',
        tmp_path / 'roms',
        output_path,
        data_dir=tmp_path,
    )

    assert exit_code.status == 'done'
    assert exit_code.glosa_calculada is False
    workbook = load_workbook(output_path)
    sheet = workbook[GLOSAS_SHEET]
    for row in (2, 3):
        valor_base = sheet.cell(row=row, column=10).value  # "Valor Base"
        valor_glosa = sheet.cell(row=row, column=11).value  # "Valor Glosa"
        assert valor_base is None
        assert valor_glosa is None
        assert valor_base != 0.0
        assert valor_glosa != 0.0


def test_run_consolidate_preserves_decision_despite_new_penalty_value(
    tmp_path: Path,
) -> None:
    # "Conflito entre decisão anterior e nova apuração": a decisão do fiscal
    # (coluna preservada) não é sobrescrita mesmo quando a apuração seguinte
    # recalcula um Percentual de Ajuste/Valor Glosa diferente para a mesma
    # ocorrência (indicador x órgão).
    _scaffold_capas(tmp_path)
    _build_orgao_report(tmp_path, 'MinC')
    _build_orgao_report(tmp_path, 'MTur')
    output_path = tmp_path / 'reports' / 'relatorio_2026-06_consolidado.xlsx'
    run_consolidate(
        '2026-06',
        tmp_path / 'reports',
        tmp_path / 'roms',
        output_path,
        data_dir=tmp_path,
    )

    workbook = load_workbook(output_path)
    sheet = workbook[GLOSAS_SHEET]
    header = [cell.value for cell in sheet[1]]
    decisao_col = header.index('Decisão Fiscal') + 1
    ajuste_col = header.index('Percentual de Ajuste') + 1
    sheet.cell(row=2, column=decisao_col, value='Aceita')
    workbook.save(output_path)

    # Nova apuração do MinC com penalidade diferente — o Percentual de
    # Ajuste recalculado diverge do que existia quando a decisão foi tomada.
    _write_summary(
        tmp_path / 'roms',
        '2026-06',
        'MinC',
        'INMS-1.6-MinC',
        'INMS 1.6',
        penalty_points=250.0,
        conforms=False,
    )
    run_report(
        '2026-06',
        tmp_path / 'capa.csv',
        tmp_path / 'roms' / 'MinC',
        tmp_path / 'reports' / 'relatorio_2026-06_MinC.xlsx',
        config_dir=tmp_path / 'configs' / 'MinC',
        expected_orgao='MinC',
        data_dir=tmp_path,
    )

    run_consolidate(
        '2026-06',
        tmp_path / 'reports',
        tmp_path / 'roms',
        output_path,
        data_dir=tmp_path,
    )

    reread = load_workbook(output_path)
    reread_sheet = reread[GLOSAS_SHEET]
    ajustes = {
        str(reread_sheet.cell(row=r, column=ajuste_col).value) for r in (2, 3)
    }
    decisoes = {
        str(reread_sheet.cell(row=r, column=decisao_col).value) for r in (2, 3)
    }
    assert 'Aceita' in decisoes  # decisão preservada apesar da nova apuração
    # MinC recalculado diverge de MTur (inalterado) — não travou no valor antigo
    assert len(ajustes) == 2


def test_run_consolidate_corrupt_ever_output_is_error(tmp_path: Path) -> None:
    """Um consolidado pré-existente corrompido (a preservar decisões) é falha
    técnica — o consolidate não regenera decisões às cegas sobre um arquivo
    ilegível."""
    _scaffold_capas(tmp_path)
    _build_orgao_report(tmp_path, 'MinC')
    _build_orgao_report(tmp_path, 'MTur')
    output_path = tmp_path / 'reports' / 'relatorio_2026-06_consolidado.xlsx'
    run_consolidate(
        '2026-06',
        tmp_path / 'reports',
        tmp_path / 'roms',
        output_path,
        data_dir=tmp_path,
    )
    output_path.write_bytes(b'not a real xlsx file')

    result = run_consolidate(
        '2026-06',
        tmp_path / 'reports',
        tmp_path / 'roms',
        output_path,
        data_dir=tmp_path,
    )

    assert result.status == 'error'
    assert result.error_message is not None

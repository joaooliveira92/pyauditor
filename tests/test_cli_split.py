from pathlib import Path

import yaml
from openpyxl import load_workbook

from pyauditor.cli.split import check_split_ready, run_split
from pyauditor.engine.pipeline import discover_config_files, load_rows, measure

_BASE_CONFIG_YAML = """\
indicator:
  id: INMS-01
  contractual_id: "INMS 1.1"
  name: Incidentes atendidos dentro do prazo

scope:
  contract: "40/2022 - Ministério da Cultura"
  orgao: MinC

source:
  csv: inms-01.csv
  delimiter: ";"
  encoding: utf-8
  period_column: "DataHoraFim"

quality_gates:
  checks:
    - type: not_null
      column: "DataHoraFim"

calculation:
  shape: ratio
  aggregation: count_distinct
  numerator_filter:
    column: "No prazo"
    equals: "S"

target:
  operator: ">="
  value: 98.0

penalty:
  base_points: 165
  step_points: 20
  step_size_pct: 0.1

acceptance_test:
  expected:
    shape: ratio
    numerator: 1
    denominator: 1
    result_pct: 100.0
    conforms: true
    penalty_points: 0.0
"""

_CATEGORIAS_YAML = """\
categorias:
  ATENDIMENTO_N1:
    label: "Atendimento Remoto aos Usuários"
    inms:
      "1.1": {mode: grupo_executor, in_values: ["N1"]}
  OPERACAO_N3:
    label: "Operação e Sustentação da Infraestrutura de TI"
    inms:
      "1.1": {mode: grupo_executor, catch_all_contains: "(CIT)"}
      "1.9": {mode: whole_indicator}
"""

_RAW_CSV = (
    "Nº Solicitacao;DataHoraFim;No prazo;Grupo_executor\n"
    "1;2026-06-01;S;N1\n"
    "2;2026-06-02;S;N1\n"
    "3;2026-06-03;S;(CIT) - Infra\n"
    "4;2026-06-04;N;(CIT) - Infra\n"
    "5;2026-06-05;S;Grupo Desconhecido\n"
)


def _write_fixture(tmp_path: Path, categorias_yaml: str = _CATEGORIAS_YAML) -> tuple[Path, Path]:
    config_dir = tmp_path / "configs"
    data_dir = tmp_path / "input"
    config_dir.mkdir(parents=True)
    competencia_dir = data_dir / "2026" / "06"
    competencia_dir.mkdir(parents=True)
    (config_dir / "inms-01.yaml").write_text(_BASE_CONFIG_YAML, encoding="utf-8")
    (config_dir / "categorias.yaml").write_text(categorias_yaml, encoding="utf-8")
    (competencia_dir / "inms-01.csv").write_text(_RAW_CSV, encoding="utf-8")
    return config_dir, data_dir


def test_run_split_writes_filtered_csv_and_derived_config_for_in_values(tmp_path: Path) -> None:
    config_dir, data_dir = _write_fixture(tmp_path)

    result = run_split("2026-06", config_dir, data_dir, expected_orgao="MinC")

    assert result.status == "done"
    n1_outcome = next(o for o in result.categorias if o.categoria == "ATENDIMENTO_N1")
    assert n1_outcome.row_count == 2
    rows = load_rows(n1_outcome.csv_path, ";", "utf-8")
    assert {r["Grupo_executor"] for r in rows} == {"N1"}

    assert n1_outcome.config_path is not None
    derived = yaml.safe_load(n1_outcome.config_path.read_text(encoding="utf-8"))
    assert derived["indicator"]["id"] == "INMS-01.ATENDIMENTO_N1"
    assert derived["source"]["csv"] == "_split/1.1/ATENDIMENTO_N1.csv"
    assert "dataset" not in derived["source"]
    assert "acceptance_test" not in derived
    # quality_gates/calculation/target/penalty copiados do base sem mudança.
    assert derived["calculation"] == {
        "shape": "ratio",
        "aggregation": "count_distinct",
        "numerator_filter": {"column": "No prazo", "equals": "S"},
    }
    assert derived["penalty"] == {"base_points": 165, "step_points": 20, "step_size_pct": 0.1}


def test_run_split_catch_all_excludes_values_claimed_by_in_values(tmp_path: Path) -> None:
    config_dir, data_dir = _write_fixture(tmp_path)

    result = run_split("2026-06", config_dir, data_dir, expected_orgao="MinC")

    n3_outcome = next(o for o in result.categorias if o.categoria == "OPERACAO_N3")
    assert n3_outcome.row_count == 2
    rows = load_rows(n3_outcome.csv_path, ";", "utf-8")
    assert {r["Grupo_executor"] for r in rows} == {"(CIT) - Infra"}


def test_run_split_outros_always_written_and_warns_when_nonempty(tmp_path: Path) -> None:
    config_dir, data_dir = _write_fixture(tmp_path)

    result = run_split("2026-06", config_dir, data_dir, expected_orgao="MinC")

    outros_outcome = next(o for o in result.categorias if o.categoria == "outros")
    assert outros_outcome.config_path is None
    assert outros_outcome.row_count == 1
    rows = load_rows(outros_outcome.csv_path, ";", "utf-8")
    assert {r["Grupo_executor"] for r in rows} == {"Grupo Desconhecido"}
    assert any(
        "INMS 1.1 (MinC/2026-06), categoria outros: 1 linha(s) não classificada(s) em "
        "nenhuma categoria — revisar categorias.yaml" in w
        for w in result.warnings
    )


def test_run_split_outros_empty_no_warning(tmp_path: Path) -> None:
    categorias_yaml = """\
categorias:
  ATENDIMENTO_N1:
    label: "Atendimento Remoto aos Usuários"
    inms:
      "1.1":
        mode: grupo_executor
        in_values: ["N1", "(CIT) - Infra", "Grupo Desconhecido"]
"""
    config_dir, data_dir = _write_fixture(tmp_path, categorias_yaml)

    result = run_split("2026-06", config_dir, data_dir, expected_orgao="MinC")

    outros_outcome = next(o for o in result.categorias if o.categoria == "outros")
    assert outros_outcome.row_count == 0
    assert outros_outcome.csv_path.exists()
    assert result.warnings == ()


def test_run_split_whole_indicator_produces_no_artifact(tmp_path: Path) -> None:
    config_dir, data_dir = _write_fixture(tmp_path)

    result = run_split("2026-06", config_dir, data_dir, expected_orgao="MinC")

    assert all(o.inms != "1.9" for o in result.categorias)
    assert not (data_dir / "2026" / "06" / "_split" / "1.9").exists()


def test_run_split_is_idempotent_on_rerun(tmp_path: Path) -> None:
    config_dir, data_dir = _write_fixture(tmp_path)

    first = run_split("2026-06", config_dir, data_dir, expected_orgao="MinC")
    second = run_split("2026-06", config_dir, data_dir, expected_orgao="MinC")

    assert first.status == "done"
    assert second.status == "done"
    assert len(first.categorias) == len(second.categorias)
    # Bruto original nunca é tocado.
    raw = (data_dir / "2026" / "06" / "inms-01.csv").read_text(encoding="utf-8")
    assert raw == _RAW_CSV


def test_run_split_invalid_competencia_is_error(tmp_path: Path) -> None:
    config_dir, data_dir = _write_fixture(tmp_path)

    result = run_split("2026/06", config_dir, data_dir, expected_orgao="MinC")

    assert result.status == "error"
    assert result.error_message is not None


def test_measure_discovers_derived_config_after_split(tmp_path: Path) -> None:
    """Depois de `split`, `measure` (sem mudança de código) acha e mede as
    configs derivadas via o glob não-recursivo já existente."""
    config_dir, data_dir = _write_fixture(tmp_path)
    split_result = run_split("2026-06", config_dir, data_dir, expected_orgao="MinC")
    assert split_result.status == "done"

    configs = discover_config_files(config_dir, expected_orgao="MinC")
    ids = {config.indicator.id for _, _, config in configs}
    assert "INMS-01.ATENDIMENTO_N1" in ids
    assert "INMS-01.OPERACAO_N3" in ids

    derived = next(
        config for _, _, config in configs if config.indicator.id == "INMS-01.ATENDIMENTO_N1"
    )
    competencia_data_dir = data_dir / "2026" / "06"
    result = measure(derived, data_dir=competencia_data_dir)
    assert result.calculation.result_pct == 100.0


def test_run_split_writes_sintetico_xlsx_when_report_dir_given(tmp_path: Path) -> None:
    config_dir, data_dir = _write_fixture(tmp_path)
    report_dir = tmp_path / "reports"

    result = run_split(
        "2026-06", config_dir, data_dir, expected_orgao="MinC", report_dir=report_dir
    )

    assert result.status == "done"
    sintetico_path = report_dir / "2026-06" / "sintetico.xlsx"
    assert sintetico_path.exists()
    wb = load_workbook(sintetico_path)
    assert set(wb.sheetnames) == {"INMS 1.1"}


def test_run_split_skips_sintetico_xlsx_when_report_dir_omitted(tmp_path: Path) -> None:
    config_dir, data_dir = _write_fixture(tmp_path)

    result = run_split("2026-06", config_dir, data_dir, expected_orgao="MinC")

    assert result.status == "done"
    assert not (tmp_path / "reports").exists()


def test_check_split_ready_always_satisfied() -> None:
    assert check_split_ready().satisfied is True


def test_run_split_missing_base_config_is_error(tmp_path: Path) -> None:
    config_dir, data_dir = _write_fixture(tmp_path)
    (config_dir / "inms-01.yaml").unlink()

    result = run_split("2026-06", config_dir, data_dir, expected_orgao="MinC")

    assert result.status == "error"


def test_run_split_missing_raw_csv_is_error(tmp_path: Path) -> None:
    config_dir, data_dir = _write_fixture(tmp_path)
    (data_dir / "2026" / "06" / "inms-01.csv").unlink()

    result = run_split("2026-06", config_dir, data_dir, expected_orgao="MinC")

    assert result.status == "error"


def test_run_split_missing_grupo_executor_column_is_error(tmp_path: Path) -> None:
    config_dir, data_dir = _write_fixture(tmp_path)
    (data_dir / "2026" / "06" / "inms-01.csv").write_text(
        "Nº Solicitacao;DataHoraFim;No prazo\n1;2026-06-01;S\n", encoding="utf-8"
    )

    result = run_split("2026-06", config_dir, data_dir, expected_orgao="MinC")

    assert result.status == "error"


def test_run_split_missing_categorias_yaml_is_error(tmp_path: Path) -> None:
    config_dir, data_dir = _write_fixture(tmp_path)
    (config_dir / "categorias.yaml").unlink()

    result = run_split("2026-06", config_dir, data_dir, expected_orgao="MinC")

    assert result.status == "error"


# Spec competencia-cli-equipe §2/§3 — split é o primeiro a filtrar pela janela
# da competência; row_count/warnings refletem o pós-filtro.


def test_run_split_filtra_janela_antes_da_segregacao(tmp_path: Path) -> None:
    from datetime import date

    from pyauditor.periodo import PeriodoAfericao

    config_dir, data_dir = _write_fixture(tmp_path)
    (data_dir / "2026" / "06" / "inms-01.csv").write_text(
        "Nº Solicitacao;DataHoraFim;No prazo;Grupo_executor\n"
        "1;01/06/2026 10:00;S;N1\n"
        "2;02/06/2026 10:00;S;N1\n"
        "3;20/05/2026 10:00;S;(CIT) - Infra\n"
        "4;04/06/2026 10:00;N;(CIT) - Infra\n"
        "5;05/06/2026 10:00;S;Grupo Desconhecido\n",
        encoding="utf-8",
    )
    periodo = PeriodoAfericao(date(2026, 6, 1), date(2026, 6, 30))

    result = run_split(
        "2026-06", config_dir, data_dir, expected_orgao="MinC", periodo=periodo
    )

    assert result.status == "done"
    n1 = next(o for o in result.categorias if o.categoria == "ATENDIMENTO_N1")
    assert n1.row_count == 2  # só linhas dentro da janela chegam à categoria
    rows = load_rows(n1.csv_path, ";", "utf-8")
    assert {r["DataHoraFim"] for r in rows} == {"01/06/2026 10:00", "02/06/2026 10:00"}


def test_run_split_janela_vazia_warns_e_segue(tmp_path: Path) -> None:
    from datetime import date

    from pyauditor.periodo import PeriodoAfericao

    config_dir, data_dir = _write_fixture(tmp_path)
    (data_dir / "2026" / "06" / "inms-01.csv").write_text(
        "Nº Solicitacao;DataHoraFim;No prazo;Grupo_executor\n"
        "1;20/05/2026 10:00;S;N1\n",
        encoding="utf-8",
    )

    result = run_split(
        "2026-06", config_dir, data_dir, expected_orgao="MinC",
        periodo=PeriodoAfericao(date(2026, 6, 1), date(2026, 6, 30)),
    )

    assert result.status == "done"  # janela vazia não é falha técnica
    assert any("nenhuma linha no período" in w for w in result.warnings)


def test_run_split_sem_period_column_e_erro_com_periodo(tmp_path: Path) -> None:
    """§2 ponto 3 — dataset bruto sem `source.period_column` declarado é falha
    técnica quando há janela para aplicar."""
    from datetime import date

    from pyauditor.periodo import PeriodoAfericao

    config_dir, data_dir = _write_fixture(tmp_path)
    yaml_sem_periodo = _BASE_CONFIG_YAML.replace('  period_column: "DataHoraFim"\n', "")
    (config_dir / "inms-01.yaml").write_text(yaml_sem_periodo, encoding="utf-8")

    result = run_split(
        "2026-06", config_dir, data_dir, expected_orgao="MinC",
        periodo=PeriodoAfericao(date(2026, 6, 1), date(2026, 6, 30)),
    )

    assert result.status == "error"
    assert result.categorias == ()
    assert result.error_message is not None

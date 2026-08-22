import csv
from pathlib import Path

from openpyxl import load_workbook

from pyauditor.config.categorias import load_categorias
from pyauditor.excel.sintetico import write_sintetico_workbook

_INMS_01_CONFIG = """\
indicator:
  id: INMS-01
  contractual_id: "INMS 1.1"
  name: Incidentes atendidos dentro do prazo

scope:
  contract: "40/2022 - Ministério da Cultura"
  orgao: MinC

source:
  csv: inms-01.csv
  delimiter: ";"
  encoding: utf-8
  period_column: "DataHoraFim"

quality_gates:
  checks:
    - type: not_null
      column: "DataHoraFim"

calculation:
  shape: ratio
  aggregation: count_distinct
  numerator_filter:
    column: "No prazo"
    equals: "S"

target:
  operator: ">="
  value: 98.0

penalty:
  base_points: 165
  step_points: 20
  step_size_pct: 0.1
"""

_INMS_09_CONFIG = """\
indicator:
  id: INMS-09
  contractual_id: "INMS 1.9"
  name: Mudanças atendidas dentro do prazo

scope:
  contract: "40/2022 - Ministério da Cultura"
  orgao: MinC

source:
  csv: inms-09.csv
  delimiter: ";"
  encoding: utf-8

quality_gates:
  checks: []

calculation:
  shape: ratio
  aggregation: count_distinct
  numerator_filter:
    column: "Situação"
    equals: "No prazo"

target:
  operator: ">="
  value: 95.0

penalty:
  base_points: 100
  step_points: 10
  step_size_pct: 1.0
"""

_INMS_04_CONFIG = """\
indicator:
  id: INMS-04
  contractual_id: "INMS 1.4"
  name: Disponibilidade de sistema crítico

scope:
  contract: "40/2022 - Ministério da Cultura"
  orgao: MinC

source:
  csv: inms-04.csv
  delimiter: ";"
  encoding: utf-8

quality_gates:
  checks: []

calculation:
  shape: ratio
  aggregation: count_distinct
  numerator_filter:
    column: "Situação"
    equals: "Disponível"

target:
  operator: ">="
  value: 99.0

penalty:
  base_points: 100
  step_points: 10
  step_size_pct: 1.0
"""

_INMS_14_CONFIG = """\
indicator:
  id: INMS-14
  contractual_id: "INMS 1.14"
  name: Índice de disponibilidade de serviços de infraestrutura

scope:
  contract: "40/2022 - Ministério da Cultura"
  orgao: MinC

source:
  csv: inms-14.csv
  delimiter: ";"
  encoding: utf-8

quality_gates:
  checks: []

calculation:
  shape: precomputed_table
  result_column: inms_1_14_percentual
  name_column: servico_infraestrutura
  numerator_column: numerador_inms_horas
  denominator_column: base_apuracao_horas
  penalty_column: penalidade_pontos

target:
  operator: ">="
  value: 99.5
"""

_CATEGORIAS_YAML = """\
categorias:
  ATENDIMENTO_N1:
    label: "Atendimento Remoto aos Usuários"
    inms:
      "1.1": {mode: grupo_executor, in_values: ["N1"]}
  ATENDIMENTO_N2:
    label: "Atendimento Presencial aos Usuários"
    inms:
      "1.1": {mode: grupo_executor, in_values: ["N2"]}
  OPERACAO_N3:
    label: "Operação e Sustentação da Infraestrutura de TI"
    inms:
      "1.1": {mode: grupo_executor, catch_all_contains: "(CIT)"}
      "1.9": {mode: whole_indicator}
  MONITORAMENTO_NOC_SOC:
    label: "Monitoramento de Ambiente (NOC/SOC)"
    inms:
      "1.4": {mode: whole_indicator}
"""

_CATEGORIAS_YAML_COM_1_14 = """\
categorias:
  ATENDIMENTO_N1:
    label: "Atendimento Remoto aos Usuários"
    inms:
      "1.1": {mode: grupo_executor, in_values: ["N1"]}
  ATENDIMENTO_N2:
    label: "Atendimento Presencial aos Usuários"
    inms:
      "1.1": {mode: grupo_executor, in_values: ["N2"]}
  OPERACAO_N3:
    label: "Operação e Sustentação da Infraestrutura de TI"
    inms:
      "1.1": {mode: grupo_executor, catch_all_contains: "(CIT)"}
      "1.9": {mode: whole_indicator}
      "1.14": {mode: whole_indicator}
  MONITORAMENTO_NOC_SOC:
    label: "Monitoramento de Ambiente (NOC/SOC)"
    inms:
      "1.4": {mode: whole_indicator}
      "1.14": {mode: whole_indicator}
"""

# Nº Solicitacao;DataHoraSolicitacao;DataHoraFim;No prazo;Grupo_executor
# 1 (N1, S, 2h) / 2 (N1, N, 24h) / 3 (N2, S, 1h) / 4 ((CIT) - Infra, S, 12h) /
# 5 (Grupo Desconhecido, N, sem DataHoraFim -> rejeitada pelo quality gate).
_INMS_01_RAW_CSV = (
    "Nº Solicitacao;DataHoraSolicitacao;DataHoraFim;No prazo;Grupo_executor\n"
    "1;01/06/2026 08:00;01/06/2026 10:00;S;N1\n"
    "2;01/06/2026 09:00;02/06/2026 09:00;N;N1\n"
    "3;01/06/2026 08:00;01/06/2026 09:00;S;N2\n"
    "4;01/06/2026 08:00;01/06/2026 20:00;S;(CIT) - Infra\n"
    "5;01/06/2026 08:00;;N;Grupo Desconhecido\n"
)

_INMS_09_RAW_CSV = "Chamado;Situação\n1;No prazo\n2;No prazo\n3;Fora do prazo\n"

# Anexo D, Tabela 28: 6 ativos nomeados, um por linha — mesmo formato do
# /input real (`disponibilidade_infra`), sem colunas de fila de atendimento
# (`No prazo`/`DataHoraSolicitacao`/`DataHoraFim`), que degradam pra "—".
_INMS_14_RAW_CSV = (
    "servico_infraestrutura;numerador_inms_horas;base_apuracao_horas;"
    "inms_1_14_percentual;penalidade_pontos\n"
    "File Server;715;720;99,3;250\n"
    "Telefonia;720;720;100,0;0\n"
    "Mensageria;720;720;100,0;0\n"
    "Servidores de impressão;720;720;100,0;0\n"
    "WI-FI;718;720;99,7;0\n"
    "Rede;720;720;100,0;0\n"
)


def _write_fixture(
    tmp_path: Path, *, include_inms_04_csv: bool, include_inms_14: bool = False
) -> tuple[Path, Path]:
    config_dir = tmp_path / "configs"
    data_dir = tmp_path / "input" / "2026" / "06"
    config_dir.mkdir(parents=True)
    data_dir.mkdir(parents=True)

    (config_dir / "inms-01.yaml").write_text(_INMS_01_CONFIG, encoding="utf-8")
    (config_dir / "inms-09.yaml").write_text(_INMS_09_CONFIG, encoding="utf-8")
    (config_dir / "inms-04.yaml").write_text(_INMS_04_CONFIG, encoding="utf-8")
    categorias_yaml = _CATEGORIAS_YAML_COM_1_14 if include_inms_14 else _CATEGORIAS_YAML
    (config_dir / "categorias.yaml").write_text(categorias_yaml, encoding="utf-8")

    (data_dir / "inms-01.csv").write_text(_INMS_01_RAW_CSV, encoding="utf-8")
    (data_dir / "inms-09.csv").write_text(_INMS_09_RAW_CSV, encoding="utf-8")
    if include_inms_04_csv:
        (data_dir / "inms-04.csv").write_text("Situação\nDisponível\n", encoding="utf-8")
    if include_inms_14:
        (config_dir / "inms-14.yaml").write_text(_INMS_14_CONFIG, encoding="utf-8")
        (data_dir / "inms-14.csv").write_text(_INMS_14_RAW_CSV, encoding="utf-8")

    return config_dir, data_dir


def test_sheet_names_cover_every_inms_with_categoria_entry(tmp_path: Path) -> None:
    config_dir, data_dir = _write_fixture(tmp_path, include_inms_04_csv=True)
    categorias_file = load_categorias(config_dir / "categorias.yaml")
    output_path = tmp_path / "sintetico.xlsx"

    warnings = write_sintetico_workbook(categorias_file, config_dir, data_dir, output_path)

    assert warnings == []
    wb = load_workbook(output_path)
    assert set(wb.sheetnames) == {"INMS 1.1", "INMS 1.9", "INMS 1.4"}


def test_grupo_executor_sheet_has_one_row_per_categoria_x_grupo_executor(
    tmp_path: Path,
) -> None:
    config_dir, data_dir = _write_fixture(tmp_path, include_inms_04_csv=True)
    categorias_file = load_categorias(config_dir / "categorias.yaml")
    output_path = tmp_path / "sintetico.xlsx"

    write_sintetico_workbook(categorias_file, config_dir, data_dir, output_path)

    wb = load_workbook(output_path)
    sheet = wb["INMS 1.1"]
    assert tuple(cell.value for cell in sheet[1]) == (
        "Categoria",
        "Nível",
        "Grupo executor",
        "Linhas",
        "Dentro do prazo",
        "Fora do prazo",
        "% bruto",
        "Tempo médio criação→resolução",
    )

    rows_by_grupo = {row[2].value: row for row in sheet.iter_rows(min_row=2, max_row=5)}

    n1_row = rows_by_grupo["N1"]
    assert [c.value for c in n1_row] == [
        "Atendimento Remoto aos Usuários",
        "N1",
        "N1",
        2,
        1,
        1,
        "50,0%",
        "0d 13h",
    ]

    n2_row = rows_by_grupo["N2"]
    assert [c.value for c in n2_row] == [
        "Atendimento Presencial aos Usuários",
        "N2",
        "N2",
        1,
        1,
        0,
        "100,0%",
        "0d 01h",
    ]

    n3_row = rows_by_grupo["(CIT) - Infra"]
    assert [c.value for c in n3_row] == [
        "Operação e Sustentação da Infraestrutura de TI",
        "N3",
        "(CIT) - Infra",
        1,
        1,
        0,
        "100,0%",
        "0d 12h",
    ]

    outros_row = rows_by_grupo["Grupo Desconhecido"]
    assert [c.value for c in outros_row] == [
        "outros (não contabilizado na meta)",
        None,
        "Grupo Desconhecido",
        1,
        0,
        1,
        "0,0%",
        "—",
    ]


def test_grupo_executor_sheet_has_subtotals_by_nivel(tmp_path: Path) -> None:
    config_dir, data_dir = _write_fixture(tmp_path, include_inms_04_csv=True)
    categorias_file = load_categorias(config_dir / "categorias.yaml")
    output_path = tmp_path / "sintetico.xlsx"

    write_sintetico_workbook(categorias_file, config_dir, data_dir, output_path)

    wb = load_workbook(output_path)
    sheet = wb["INMS 1.1"]
    header_row = next(
        row[0].row for row in sheet.iter_rows(min_row=1) if row[0].value == "Subtotais por Nível"
    )
    subtotal_rows = {
        row[0].value: row[:6]
        for row in sheet.iter_rows(min_row=header_row + 2, max_row=header_row + 4)
    }

    assert [c.value for c in subtotal_rows["N1"]] == ["N1", 2, 1, 1, "50,0%", "0d 13h"]
    assert [c.value for c in subtotal_rows["N2"]] == ["N2", 1, 1, 0, "100,0%", "0d 01h"]
    assert [c.value for c in subtotal_rows["N3"]] == ["N3", 1, 1, 0, "100,0%", "0d 12h"]


def test_whole_indicator_sheet_is_single_row_no_subtotals(tmp_path: Path) -> None:
    config_dir, data_dir = _write_fixture(tmp_path, include_inms_04_csv=True)
    categorias_file = load_categorias(config_dir / "categorias.yaml")
    output_path = tmp_path / "sintetico.xlsx"

    write_sintetico_workbook(categorias_file, config_dir, data_dir, output_path)

    wb = load_workbook(output_path)
    sheet = wb["INMS 1.9"]
    data_rows = list(sheet.iter_rows(min_row=2))
    assert len(data_rows) == 1
    assert [c.value for c in data_rows[0]] == [
        "Operação e Sustentação da Infraestrutura de TI",
        "N3",
        "(indicador inteiro)",
        3,
        "—",
        "—",
        "—",
        "—",
    ]
    assert not any(cell.value == "Subtotais por Nível" for row in sheet.iter_rows() for cell in row)


def test_inms_1_14_sheet_uses_ativo_column_with_12_rows_in_categoria_blocks(
    tmp_path: Path,
) -> None:
    config_dir, data_dir = _write_fixture(tmp_path, include_inms_04_csv=True, include_inms_14=True)
    categorias_file = load_categorias(config_dir / "categorias.yaml")
    output_path = tmp_path / "sintetico.xlsx"

    warnings = write_sintetico_workbook(categorias_file, config_dir, data_dir, output_path)

    assert warnings == []
    wb = load_workbook(output_path)
    assert "INMS 1.14" in wb.sheetnames
    sheet = wb["INMS 1.14"]

    assert tuple(cell.value for cell in sheet[1]) == (
        "Categoria",
        "Nível",
        "Ativo",
        "Linhas",
        "Dentro do prazo",
        "Fora do prazo",
        "% bruto",
        "Tempo médio criação→resolução",
    )

    data_rows = list(sheet.iter_rows(min_row=2, max_row=13))
    assert len(data_rows) == 12

    # Bloco NOC/SOC (linhas 2-7) antes do bloco Operação N3 (linhas 8-13).
    noc_soc_rows = [row[2].value for row in data_rows[:6]]
    operacao_n3_rows = [row[2].value for row in data_rows[6:]]
    assert (
        noc_soc_rows
        == operacao_n3_rows
        == [
            "File Server",
            "Telefonia",
            "Mensageria",
            "Servidores de impressão",
            "WI-FI",
            "Rede",
        ]
    )

    first_row = data_rows[0]
    assert [cell.value for cell in first_row] == [
        "Monitoramento de Ambiente (NOC/SOC)",
        "N3",
        "File Server",
        1,
        "—",
        "—",
        "—",
        "—",
    ]
    eighth_row = data_rows[6]
    assert [cell.value for cell in eighth_row] == [
        "Operação e Sustentação da Infraestrutura de TI",
        "N3",
        "File Server",
        1,
        "—",
        "—",
        "—",
        "—",
    ]


def test_inms_1_14_sheet_has_subtotals_by_categoria_not_nivel(tmp_path: Path) -> None:
    config_dir, data_dir = _write_fixture(tmp_path, include_inms_04_csv=True, include_inms_14=True)
    categorias_file = load_categorias(config_dir / "categorias.yaml")
    output_path = tmp_path / "sintetico.xlsx"

    write_sintetico_workbook(categorias_file, config_dir, data_dir, output_path)

    wb = load_workbook(output_path)
    sheet = wb["INMS 1.14"]
    header_row = next(
        row[0].row
        for row in sheet.iter_rows(min_row=1)
        if row[0].value == "Subtotais por Categoria"
    )
    assert tuple(cell.value for cell in sheet[header_row + 1][:6]) == (
        "Categoria",
        "Linhas",
        "Dentro do prazo",
        "Fora do prazo",
        "% bruto",
        "Tempo médio criação→resolução",
    )
    subtotal_rows = [
        [c.value for c in row[:6]]
        for row in sheet.iter_rows(min_row=header_row + 2, max_row=header_row + 3)
    ]
    assert subtotal_rows == [
        ["Monitoramento de Ambiente (NOC/SOC)", 6, "—", "—", "—", "—"],
        ["Operação e Sustentação da Infraestrutura de TI", 6, "—", "—", "—", "—"],
    ]


def test_nao_ativado_when_raw_csv_missing(tmp_path: Path) -> None:
    config_dir, data_dir = _write_fixture(tmp_path, include_inms_04_csv=False)
    categorias_file = load_categorias(config_dir / "categorias.yaml")
    output_path = tmp_path / "sintetico.xlsx"

    warnings = write_sintetico_workbook(categorias_file, config_dir, data_dir, output_path)

    assert warnings == []
    wb = load_workbook(output_path)
    sheet = wb["INMS 1.4"]
    assert sheet.cell(row=2, column=1).value == (
        "Esse serviço não foi requisitado no período selecionado."
    )


def test_periodo_filtra_antes_dos_gates_e_falha_sem_period_column(tmp_path: Path) -> None:
    """Spec competencia-cli-equipe §2/§ issue 01 item 5 — sintetico aplica a
    mesma janela do split: linha de maio some das contagens; config sem
    `period_column` é erro acionável, aba daquele INMS é pulada (mesmo
    tratamento dos demais erros do loop) em vez do workbook inteiro cair."""
    from datetime import date

    from pyauditor.periodo import PeriodoAfericao

    config_dir, data_dir = _write_fixture(tmp_path, include_inms_04_csv=True)
    (data_dir / "inms-01.csv").write_text(
        "Nº Solicitacao;DataHoraSolicitacao;DataHoraFim;No prazo;Grupo_executor\n"
        "1;01/06/2026 08:00;01/06/2026 10:00;S;N1\n"
        "2;20/05/2026 09:00;20/05/2026 11:00;N;N1\n"
        "3;02/06/2026 08:00;02/06/2026 09:00;S;N2\n",
        encoding="utf-8",
    )
    categorias_file = load_categorias(config_dir / "categorias.yaml")
    output_path = tmp_path / "sintetico.xlsx"

    warnings = write_sintetico_workbook(
        categorias_file,
        config_dir,
        data_dir,
        output_path,
        periodo=PeriodoAfericao(date(2026, 6, 1), date(2026, 6, 30)),
    )

    # INMS 1.9/1.4 não declaram period_column → erro acionável, aba pulada.
    assert any("INMS 1.9" in w and "period_column" in w and "não declarado" in w for w in warnings)
    wb = load_workbook(output_path)
    assert "INMS 1.9" not in wb.sheetnames
    assert "INMS 1.4" not in wb.sheetnames
    sheet = wb["INMS 1.1"]
    rows_by_grupo = {row[2].value: row for row in sheet.iter_rows(min_row=2)}
    n1 = [c.value for c in rows_by_grupo["N1"]]
    assert n1[3:6] == [1, 1, 0]  # linha de maio descartada antes da segregação
    n2 = [c.value for c in rows_by_grupo["N2"]]
    assert n2[3:6] == [1, 1, 0]


_INMS_04_PRECOMPUTED_CONFIG = """\
indicator:
  id: INMS-04
  contractual_id: "INMS 1.4"
  name: Disponibilidade de sistema crítico

scope:
  contract: "40/2022 - Ministério da Cultura"
  orgao: MinC

source:
  csv: inms-04.csv
  delimiter: ";"
  encoding: utf-8

quality_gates:
  checks: []

calculation:
  shape: precomputed_table
  result_column: inms_1_4_percentual
  name_column: sistema_servico_nome
  penalty_column: penalidade_pontos

target:
  operator: ">="
  value: 99.5
"""

_INMS_04_PRECOMPUTED_RAW_CSV = (
    "sistema_servico_nome;inms_1_4_percentual;penalidade_pontos\n"
    "Barramento de Integracao;99,451;1000\n"
    "Servico de Correio;99,7744;0\n"
    ";;\n"  # linha vazia/placeholder — sem medição, deve ser pulada
)

_CATEGORIAS_YAML_1_4_ONLY = """\
categorias:
  MONITORAMENTO_NOC_SOC:
    label: "Monitoramento de Ambiente (NOC/SOC)"
    inms:
      "1.4": {mode: whole_indicator}
"""


def test_precomputed_table_sheet_has_one_row_per_item_with_meta_from_target(
    tmp_path: Path,
) -> None:
    config_dir = tmp_path / "configs"
    data_dir = tmp_path / "input" / "2026" / "06"
    config_dir.mkdir(parents=True)
    data_dir.mkdir(parents=True)
    (config_dir / "inms-04.yaml").write_text(_INMS_04_PRECOMPUTED_CONFIG, encoding="utf-8")
    (config_dir / "categorias.yaml").write_text(_CATEGORIAS_YAML_1_4_ONLY, encoding="utf-8")
    (data_dir / "inms-04.csv").write_text(_INMS_04_PRECOMPUTED_RAW_CSV, encoding="utf-8")
    categorias_file = load_categorias(config_dir / "categorias.yaml")
    output_path = tmp_path / "sintetico.xlsx"

    warnings = write_sintetico_workbook(categorias_file, config_dir, data_dir, output_path)

    assert warnings == []
    wb = load_workbook(output_path)
    sheet = wb["INMS 1.4"]
    assert tuple(cell.value for cell in sheet[1]) == (
        "Categoria",
        "Nível",
        "Item",
        "Resultado",
        "Meta atingida?",
        "Penalidade",
    )
    data_rows = [[c.value for c in row] for row in sheet.iter_rows(min_row=2)]
    assert data_rows == [
        [
            "Monitoramento de Ambiente (NOC/SOC)",
            "N3",
            "Barramento de Integracao",
            "99,5%",
            "Não",
            "1000",
        ],
        [
            "Monitoramento de Ambiente (NOC/SOC)",
            "N3",
            "Servico de Correio",
            "99,8%",
            "Sim",
            "0",
        ],
    ]


_INMS_06_CONFIG = """\
indicator:
  id: INMS-06
  contractual_id: "INMS 1.6"
  name: Eficácia no tratamento de chamados

scope:
  contract: "40/2022 - Ministério da Cultura"
  orgao: MinC

source:
  csv: inms-06.csv
  delimiter: ","
  encoding: utf-8
  unfilterable: true

quality_gates:
  checks: []

calculation:
  shape: ratio
  aggregation: sum
  denominator_filter:
    column: "Acordo de Nível de Serviço"
    not_equals: TOTAIS
  sum_numerator_column: "Total de Chamados"
  sum_numerator_subtract_column: "Total de Chamados Reabertos"

target:
  operator: ">="
  value: 97.0

penalty:
  step_points: 200
  step_size_pct: 0.5
"""

_INMS_06_RAW_CSV = (
    "Acordo de Nível de Serviço,Total de Chamados,Total de Chamados Reabertos\n"
    "SLA A,10,1\n"
    "SLA A,10,0\n"
    "SLA B,5,0\n"
    "TOTAIS,25,1\n"
)

_CATEGORIAS_YAML_1_6_ONLY = """\
categorias:
  OPERACAO_N3:
    label: "Operação e Sustentação da Infraestrutura de TI"
    inms:
      "1.6": {mode: whole_indicator}
"""


def test_ratio_sum_subtract_sheet_aggregates_one_row_per_group_excluding_totais(
    tmp_path: Path,
) -> None:
    config_dir = tmp_path / "configs"
    data_dir = tmp_path / "input" / "2026" / "06"
    config_dir.mkdir(parents=True)
    data_dir.mkdir(parents=True)
    (config_dir / "inms-06.yaml").write_text(_INMS_06_CONFIG, encoding="utf-8")
    (config_dir / "categorias.yaml").write_text(_CATEGORIAS_YAML_1_6_ONLY, encoding="utf-8")
    (data_dir / "inms-06.csv").write_text(_INMS_06_RAW_CSV, encoding="utf-8")
    categorias_file = load_categorias(config_dir / "categorias.yaml")
    output_path = tmp_path / "sintetico.xlsx"

    warnings = write_sintetico_workbook(categorias_file, config_dir, data_dir, output_path)

    assert warnings == []
    wb = load_workbook(output_path)
    sheet = wb["INMS 1.6"]
    assert tuple(cell.value for cell in sheet[1]) == (
        "Categoria",
        "Nível",
        "Acordo de Nível de Serviço",
        "Total de Chamados",
        "Total de Chamados Reabertos",
        "% resultado",
        "Meta atingida?",
    )
    data_rows = [[c.value for c in row] for row in sheet.iter_rows(min_row=2)]
    # TOTAIS excluída (denominator_filter); SLA A soma as 2 linhas (20/1).
    assert data_rows == [
        [
            "Operação e Sustentação da Infraestrutura de TI",
            "N3",
            "SLA A",
            20,
            1,
            "95,0%",
            "Não",
        ],
        [
            "Operação e Sustentação da Infraestrutura de TI",
            "N3",
            "SLA B",
            5,
            0,
            "100,0%",
            "Sim",
        ],
    ]


_PRAZOS_RAW_CSV = (
    "Demanda,Criticidade,Prazo máximo para atendimento\n"
    "Incidentes,Alta,2h (horas corridas)\n"
    "Requisições,Alta,4h (horas corridas)\n"
    "Requisições,Média,8h (horas corridas)\n"
    "Requisições,Baixa,16h (horas corridas)\n"
    "Projetos,-,Prazo acordado entre Contratante e Contratada\n"
)


def test_prazos_sheet_is_first_and_reproduces_csv_verbatim(tmp_path: Path) -> None:
    config_dir, data_dir = _write_fixture(tmp_path, include_inms_04_csv=True)
    categorias_file = load_categorias(config_dir / "categorias.yaml")
    output_path = tmp_path / "sintetico.xlsx"
    prazos_path = tmp_path / "prazos.csv"
    prazos_path.write_text(_PRAZOS_RAW_CSV, encoding="utf-8")

    warnings = write_sintetico_workbook(
        categorias_file, config_dir, data_dir, output_path, prazos_path=prazos_path
    )

    assert warnings == []
    wb = load_workbook(output_path)
    assert wb.sheetnames[0] == "Prazos"
    sheet = wb["Prazos"]
    rows = [[c.value for c in row] for row in sheet.iter_rows()]
    assert rows == [line.split(",") for line in _PRAZOS_RAW_CSV.strip("\n").split("\n")]


def test_prazos_sheet_missing_file_warns_and_skips(tmp_path: Path) -> None:
    config_dir, data_dir = _write_fixture(tmp_path, include_inms_04_csv=True)
    categorias_file = load_categorias(config_dir / "categorias.yaml")
    output_path = tmp_path / "sintetico.xlsx"
    prazos_path = tmp_path / "prazos.csv"  # não existe

    warnings = write_sintetico_workbook(
        categorias_file, config_dir, data_dir, output_path, prazos_path=prazos_path
    )

    assert len(warnings) == 1
    assert "Prazos" in warnings[0]
    wb = load_workbook(output_path)
    assert "Prazos" not in wb.sheetnames


def test_prazos_sheet_omitted_when_prazos_path_not_given(tmp_path: Path) -> None:
    config_dir, data_dir = _write_fixture(tmp_path, include_inms_04_csv=True)
    categorias_file = load_categorias(config_dir / "categorias.yaml")
    output_path = tmp_path / "sintetico.xlsx"

    warnings = write_sintetico_workbook(categorias_file, config_dir, data_dir, output_path)

    assert warnings == []
    wb = load_workbook(output_path)
    assert "Prazos" not in wb.sheetnames


_CAPA_RAW_CSV = "Campo;Valor\nNúmero do contrato;40/2022\nProcesso SEI;72031.010172/2020-97\n"
_EQUIPE_RAW_CSV = (
    "FUNÇÃO,NOME,SIAPE\n"
    "Gestor do Contrato,Thiago Augusto Arcanjo Lima,1500967\n"
    "Fiscal Técnico,João Antônio Carvalho Monteiro de Oliveira,1499628\n"
)
_OBJETOS_RAW_CSV = (
    "Item,Categoria,Valor\n"
    '1,Central de Serviços," R$  148.205,54 "\n'
    '2,GT dos Projetos e Operações," R$  77.654,90 "\n'
)


def test_capa_equipe_prazos_sheets_come_first_in_order(tmp_path: Path) -> None:
    config_dir, data_dir = _write_fixture(tmp_path, include_inms_04_csv=True)
    categorias_file = load_categorias(config_dir / "categorias.yaml")
    output_path = tmp_path / "sintetico.xlsx"
    capa_path = tmp_path / "capa.csv"
    capa_path.write_text(_CAPA_RAW_CSV, encoding="utf-8-sig")
    equipe_path = tmp_path / "equipe.csv"
    equipe_path.write_text(_EQUIPE_RAW_CSV, encoding="utf-8-sig")
    prazos_path = tmp_path / "prazos.csv"
    prazos_path.write_text(_PRAZOS_RAW_CSV, encoding="utf-8")

    warnings = write_sintetico_workbook(
        categorias_file,
        config_dir,
        data_dir,
        output_path,
        capa_path=capa_path,
        equipe_path=equipe_path,
        prazos_path=prazos_path,
    )

    assert warnings == []
    wb = load_workbook(output_path)
    assert wb.sheetnames[:3] == ["Capa", "Equipe", "Prazos"]
    assert "Objetos" not in wb.sheetnames

    capa_rows = [[c.value for c in row] for row in wb["Capa"].iter_rows()]
    assert capa_rows == [line.split(";") for line in _CAPA_RAW_CSV.strip("\n").split("\n")]

    equipe_rows = [[c.value for c in row] for row in wb["Equipe"].iter_rows()]
    assert equipe_rows == [line.split(",") for line in _EQUIPE_RAW_CSV.strip("\n").split("\n")]


def test_objetos_is_appended_below_capa_not_a_separate_sheet(tmp_path: Path) -> None:
    config_dir, data_dir = _write_fixture(tmp_path, include_inms_04_csv=True)
    categorias_file = load_categorias(config_dir / "categorias.yaml")
    output_path = tmp_path / "sintetico.xlsx"
    capa_path = tmp_path / "capa.csv"
    capa_path.write_text(_CAPA_RAW_CSV, encoding="utf-8-sig")
    objetos_path = tmp_path / "objetos.csv"
    objetos_path.write_text(_OBJETOS_RAW_CSV, encoding="utf-8-sig")

    warnings = write_sintetico_workbook(
        categorias_file,
        config_dir,
        data_dir,
        output_path,
        capa_path=capa_path,
        objetos_path=objetos_path,
    )

    assert warnings == []
    wb = load_workbook(output_path)
    assert "Objetos" not in wb.sheetnames
    sheet = wb["Capa"]
    rows = [[c.value for c in row] for row in sheet.iter_rows()]
    # A grade tem 3 colunas (largura de objetos.csv) — linhas da capa (2
    # colunas) ficam com a 3ª célula vazia (None) por causa disso.
    capa_lines = [[*line.split(";"), None] for line in _CAPA_RAW_CSV.strip("\n").split("\n")]
    objetos_lines = list(csv.reader(_OBJETOS_RAW_CSV.strip("\n").split("\n")))
    # Última linha da capa ("Processo SEI..." nesse fixture), 1 linha em
    # branco, depois cabeçalho + linhas de objetos.csv verbatim.
    assert rows == [*capa_lines, [None, None, None], *objetos_lines]


def test_objetos_not_appended_when_capa_path_missing(tmp_path: Path) -> None:
    config_dir, data_dir = _write_fixture(tmp_path, include_inms_04_csv=True)
    categorias_file = load_categorias(config_dir / "categorias.yaml")
    output_path = tmp_path / "sintetico.xlsx"
    objetos_path = tmp_path / "objetos.csv"
    objetos_path.write_text(_OBJETOS_RAW_CSV, encoding="utf-8-sig")

    warnings = write_sintetico_workbook(
        categorias_file, config_dir, data_dir, output_path, objetos_path=objetos_path
    )

    assert len(warnings) == 1
    assert "objetos" in warnings[0].lower()
    wb = load_workbook(output_path)
    assert "Capa" not in wb.sheetnames
    assert "Objetos" not in wb.sheetnames


def test_capa_sheet_missing_file_warns_and_skips(tmp_path: Path) -> None:
    config_dir, data_dir = _write_fixture(tmp_path, include_inms_04_csv=True)
    categorias_file = load_categorias(config_dir / "categorias.yaml")
    output_path = tmp_path / "sintetico.xlsx"
    capa_path = tmp_path / "capa.csv"  # não existe

    warnings = write_sintetico_workbook(
        categorias_file, config_dir, data_dir, output_path, capa_path=capa_path
    )

    assert len(warnings) == 1
    assert "Capa" in warnings[0]
    wb = load_workbook(output_path)
    assert "Capa" not in wb.sheetnames


def test_equipe_sheet_omitted_when_equipe_path_not_given(tmp_path: Path) -> None:
    config_dir, data_dir = _write_fixture(tmp_path, include_inms_04_csv=True)
    categorias_file = load_categorias(config_dir / "categorias.yaml")
    output_path = tmp_path / "sintetico.xlsx"

    warnings = write_sintetico_workbook(categorias_file, config_dir, data_dir, output_path)

    assert warnings == []
    wb = load_workbook(output_path)
    assert "Equipe" not in wb.sheetnames

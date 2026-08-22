"""Caracterização da aba enriquecida do INMS 1.1 (`inms_1_1_audit.write_sheet`),
acionada por `sintetico.py` quando o CSV bruto tem as colunas de detalhe
(`Nº Solicitacao`, `Atividades`, `DataHoraLimite`, `TecnicoExecutor`). Sem
isso, nenhum teste do repositório passava pelo caminho enriquecido — só o
renderer genérico (`test_excel_sintetico.py`). Estes testes fixam o
comportamento observável (estrutura de seções, fórmulas-chave) para permitir
refatoração segura de `inms_1_1_audit.py`.
"""

from datetime import date, datetime
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from pyauditor.config.categorias import GrupoExecutorMode, load_categorias
from pyauditor.excel import inms_1_1_audit
from pyauditor.excel.inms_1_1._domain import _build_grupo_rows
from pyauditor.excel.inms_1_1._layout import _SEM_NIVEL
from pyauditor.excel.sintetico import write_sintetico_workbook
from pyauditor.periodo import PeriodoAfericao

_INMS_01_CONFIG = """\
indicator:
  id: INMS-01
  contractual_id: "INMS 1.1"
  name: Incidentes atendidos dentro do prazo

scope:
  contract: "40/2022 - Ministério da Cultura"
  orgao: MinC

source:
  csv: inms-01.csv
  delimiter: ";"
  encoding: utf-8
  period_column: "DataHoraFim"

quality_gates:
  checks:
    - type: not_null
      column: "DataHoraFim"

calculation:
  shape: ratio
  aggregation: count_distinct
  numerator_filter:
    column: "No prazo"
    equals: "S"

target:
  operator: ">="
  value: 98.0

penalty:
  base_points: 165
  step_points: 20
  step_size_pct: 0.1
"""

_CATEGORIAS_YAML = """\
categorias:
  ATENDIMENTO_N1:
    label: "Atendimento Remoto aos Usuários"
    inms:
      "1.1": {mode: grupo_executor, in_values: ["N1"]}
  ATENDIMENTO_N2:
    label: "Atendimento Presencial aos Usuários"
    inms:
      "1.1": {mode: grupo_executor, in_values: ["N2"]}
  OPERACAO_N3:
    label: "Operação e Sustentação da Infraestrutura de TI"
    inms:
      "1.1": {mode: grupo_executor, catch_all_contains: "(CIT)"}
"""

# Todas as colunas exigidas por `inms_1_1_audit.has_required_columns`, ao
# contrário da fixture minimalista de `test_excel_sintetico.py`.
_INMS_01_RAW_CSV_ENRIQUECIDO = (
    "Nº Solicitacao;Atividades;DataHoraSolicitacao;DataHoraLimite;DataHoraFim;"
    "No prazo;Grupo_executor;TecnicoExecutor\n"
    "1;Reset de senha;01/06/2026 08:00;01/06/2026 10:00;01/06/2026 09:30;"
    "S;N1;Fulano\n"
    "2;Instalação de software;01/06/2026 09:00;01/06/2026 11:00;02/06/2026 09:00;"
    "N;N1;Fulano\n"
    "3;Acesso a sistema;01/06/2026 08:00;01/06/2026 10:00;01/06/2026 09:00;"
    "S;N2;Beltrano\n"
    "4;Restart de serviço;01/06/2026 08:00;01/06/2026 10:00;01/06/2026 20:00;"
    "S;(CIT) - Infra;Ciclano\n"
)


def _write_fixture(tmp_path: Path) -> tuple[Path, Path]:
    config_dir = tmp_path / "configs"
    data_dir = tmp_path / "input" / "2026" / "06"
    config_dir.mkdir(parents=True)
    data_dir.mkdir(parents=True)

    (config_dir / "inms-01.yaml").write_text(_INMS_01_CONFIG, encoding="utf-8")
    (config_dir / "categorias.yaml").write_text(_CATEGORIAS_YAML, encoding="utf-8")
    (data_dir / "inms-01.csv").write_text(_INMS_01_RAW_CSV_ENRIQUECIDO, encoding="utf-8")

    return config_dir, data_dir


def test_enriched_sheet_is_used_when_raw_csv_has_detail_columns(tmp_path: Path) -> None:
    config_dir, data_dir = _write_fixture(tmp_path)
    categorias_file = load_categorias(config_dir / "categorias.yaml")
    output_path = tmp_path / "sintetico.xlsx"
    periodo = PeriodoAfericao(date(2026, 6, 1), date(2026, 6, 30))

    warnings = write_sintetico_workbook(
        categorias_file, config_dir, data_dir, output_path, periodo=periodo
    )

    assert warnings == []
    wb = load_workbook(output_path)
    sheet = wb["INMS 1.1"]

    assert sheet["A1"].value == "INMS 1.1 – Incidentes atendidos dentro do prazo"

    section_bars = {
        cell.value
        for row in sheet.iter_rows(min_col=1, max_col=1)
        for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("SEÇÃO")
    }
    assert section_bars == {
        "SEÇÃO 1 · IDENTIFICAÇÃO",
        "SEÇÃO 2 · RESUMO EXECUTIVO",
        "SEÇÃO 3 · MEMÓRIA DO CÁLCULO CONSOLIDADO",
        "SEÇÃO 4 · DETALHAMENTO POR GRUPO EXECUTOR",
        "SEÇÃO 5 · SUBTOTAIS POR NÍVEL (informação gerencial)",
        "SEÇÃO 6 · INCIDENTES FORA DO PRAZO",
        "SEÇÃO 7 · AUDITORIA DO PRAZO CONTRATUAL",
        "SEÇÃO 8 · TEMPO CORRIDO MÉDIO ATÉ A RESOLUÇÃO",
        "SEÇÃO 9 · PENALIDADE (CÁLCULO PRELIMINAR)",
    }

    # Seção 2 — resumo executivo: fórmulas dependem só da base de apoio.
    assert sheet["A13"].value == "=0.98"
    assert sheet["B13"].value == "=ROWS($R$2:$R$5)"
    assert sheet["C13"].value == '=COUNTIF($X$2:$X$5,"S")'
    assert sheet["D13"].value == '=COUNTIF($X$2:$X$5,"N")'
    assert sheet["E13"].value == '=IF(B13=0,"Sem ocorrências",C13/B13)'

    # Seção 4 — uma linha por grupo executor real do CSV.
    grupo_col = [cell.value for cell in sheet["C"][29:32]]
    assert grupo_col == ["N1", "N2", "(CIT) - Infra"]

    # Base de apoio (colunas R:AM) tem uma linha por incidente do CSV bruto.
    assert sheet["R2"].value == "1"
    assert sheet["R5"].value == "4"


def test_enriched_sheet_lists_out_of_deadline_incidents_in_section_6(
    tmp_path: Path,
) -> None:
    config_dir, data_dir = _write_fixture(tmp_path)
    categorias_file = load_categorias(config_dir / "categorias.yaml")
    output_path = tmp_path / "sintetico.xlsx"
    periodo = PeriodoAfericao(date(2026, 6, 1), date(2026, 6, 30))

    write_sintetico_workbook(
        categorias_file, config_dir, data_dir, output_path, periodo=periodo
    )

    wb = load_workbook(output_path)
    sheet = wb["INMS 1.1"]

    s6_bar_row = next(
        cell.row
        for row in sheet.iter_rows(min_col=1, max_col=1)
        for cell in row
        if cell.value == "SEÇÃO 6 · INCIDENTES FORA DO PRAZO"
    )
    header_row = s6_bar_row + 1
    assert sheet.cell(row=header_row, column=1).value == "Nº solicitação"
    first_data_row = header_row + 1
    assert sheet.cell(row=first_data_row, column=1).value == (
        '=IFERROR(INDEX($R$2:$R$5,MATCH(1,$AH$2:$AH$5,0)),"")'
    )


def test_enriched_sheet_sanitizes_formula_injection_from_raw_csv(tmp_path: Path) -> None:
    config_dir, data_dir = _write_fixture(tmp_path)
    malicious_csv = (
        "Nº Solicitacao;Atividades;DataHoraSolicitacao;DataHoraLimite;DataHoraFim;"
        "No prazo;Grupo_executor;TecnicoExecutor\n"
        '=HYPERLINK("https://exemplo.invalid","clique");'
        "+Atividade maliciosa;01/06/2026 08:00;01/06/2026 10:00;01/06/2026 09:30;"
        "S;N1;-Fulano\n"
    )
    (data_dir / "inms-01.csv").write_text(malicious_csv, encoding="utf-8")
    categorias_file = load_categorias(config_dir / "categorias.yaml")
    output_path = tmp_path / "sintetico.xlsx"
    periodo = PeriodoAfericao(date(2026, 6, 1), date(2026, 6, 30))

    write_sintetico_workbook(
        categorias_file, config_dir, data_dir, output_path, periodo=periodo
    )

    wb = load_workbook(output_path)
    sheet = wb["INMS 1.1"]

    # Coluna de apoio: valores maliciosos gravados como texto literal
    # (prefixados com apóstrofo), não interpretados como fórmula.
    assert sheet["R2"].value == '\'=HYPERLINK("https://exemplo.invalid","clique")'
    assert sheet["T2"].value == "'+Atividade maliciosa"
    assert sheet["Y2"].value == "'-Fulano"


def test_enriched_sheet_flags_invalid_and_inconsistent_dates(tmp_path: Path) -> None:
    config_dir, data_dir = _write_fixture(tmp_path)
    csv_with_bad_dates = (
        "Nº Solicitacao;Atividades;DataHoraSolicitacao;DataHoraLimite;DataHoraFim;"
        "No prazo;Grupo_executor;TecnicoExecutor\n"
        # Linha 2: data de abertura malformada.
        "1;Reset de senha;31/02/2026 08:00;01/06/2026 10:00;01/06/2026 09:30;S;N1;Fulano\n"
        # Linha 3: data de encerramento vazia (incidente ainda em aberto) — OK.
        "2;Instalação;01/06/2026 08:00;01/06/2026 10:00;;S;N1;Fulano\n"
        # Linha 4: encerramento anterior à abertura.
        "3;Restart;01/06/2026 12:00;01/06/2026 14:00;01/06/2026 10:00;N;N1;Fulano\n"
        # Linha 5: tudo válido.
        "4;Acesso;01/06/2026 08:00;01/06/2026 10:00;01/06/2026 09:00;S;N1;Fulano\n"
        # Linha 6: data de abertura ausente (mas com encerramento) — sinalizada.
        "5;Acesso;;01/06/2026 10:00;01/06/2026 09:00;S;N1;Fulano\n"
    )
    (data_dir / "inms-01.csv").write_text(csv_with_bad_dates, encoding="utf-8")
    categorias_file = load_categorias(config_dir / "categorias.yaml")
    output_path = tmp_path / "sintetico.xlsx"
    periodo = PeriodoAfericao(date(2026, 6, 1), date(2026, 6, 30))

    write_sintetico_workbook(
        categorias_file, config_dir, data_dir, output_path, periodo=periodo
    )

    wb = load_workbook(output_path)
    sheet = wb["INMS 1.1"]

    # Coluna AJ (situação dos dados).
    assert sheet["AJ2"].value == "Data de abertura inválida"
    assert sheet["AJ3"].value == "OK"
    assert sheet["AJ4"].value == "Encerramento anterior à abertura"
    assert sheet["AJ5"].value == "OK"
    assert sheet["AJ6"].value == "Data de abertura ausente"

    # U2 vazia (data de abertura malformada -> não gravada) protege AB2/AG2
    # contra tratar célula vazia como zero.
    assert sheet["U2"].value is None
    assert sheet["AB2"].value == '=IF(U2="","",U2+2.0/24)'
    assert sheet["AG2"].value == '=IF(OR(U2="",W2="",W2<U2),"",W2-U2)'

    # Linha 4: encerramento < abertura -> AG fica vazio, não negativo.
    assert sheet["AG4"].value == '=IF(OR(U4="",W4="",W4<U4),"",W4-U4)'


def test_prazo_tolerancia_minutos_used_consistently(tmp_path: Path) -> None:
    from pyauditor.excel._datetime import PRAZO_TOLERANCIA_MINUTOS

    config_dir, data_dir = _write_fixture(tmp_path)
    categorias_file = load_categorias(config_dir / "categorias.yaml")
    output_path = tmp_path / "sintetico.xlsx"
    periodo = PeriodoAfericao(date(2026, 6, 1), date(2026, 6, 30))

    write_sintetico_workbook(
        categorias_file, config_dir, data_dir, output_path, periodo=periodo
    )

    wb = load_workbook(output_path)
    sheet = wb["INMS 1.1"]

    # Coluna AC (divergência de prazo) usa a mesma constante de tolerância
    # que o cálculo Python da amostra da Seção 7 — sem valor duplicado.
    assert f"{PRAZO_TOLERANCIA_MINUTOS}/1440" in sheet["AC2"].value


def test_zero_denominator_formulas_are_guarded_against_div0(tmp_path: Path) -> None:
    """A-01: nenhuma fórmula que divide por B13 (IAP) ou pelo total de linhas
    da Seção 5 pode gerar #DIV/0! quando o denominador for zero — todas devem
    ter guarda explícita ("Sem ocorrências"/"Não aplicável")."""
    config_dir, data_dir = _write_fixture(tmp_path)
    categorias_file = load_categorias(config_dir / "categorias.yaml")
    output_path = tmp_path / "sintetico.xlsx"
    periodo = PeriodoAfericao(date(2026, 6, 1), date(2026, 6, 30))

    write_sintetico_workbook(
        categorias_file, config_dir, data_dir, output_path, periodo=periodo
    )

    wb = load_workbook(output_path)
    sheet = wb["INMS 1.1"]

    def formula(coord: str) -> str:
        value = sheet[coord].value
        assert isinstance(value, str)
        return value

    # Seção 2.
    assert 'IF(B13=0,"Sem ocorrências"' in formula("E13")
    assert 'IF(B13=0,""' in formula("F13")
    assert 'IF(B13=0,"Sem ocorrências"' in formula("G13")

    # Seção 3.
    assert 'IF(B13=0,""' in formula("B23")

    # Seção 5 — linha "Total geral" (label fixo) e verificação cruzada.
    total_row = next(
        cell.row
        for row in sheet.iter_rows(min_col=1, max_col=1)
        for cell in row
        if cell.value == "Total geral"
    )
    assert 'IF(B' in formula(f"E{total_row}") and "Sem ocorrências" in formula(f"E{total_row}")
    check_label_row = next(
        cell.row
        for row in sheet.iter_rows(min_col=1, max_col=1)
        for cell in row
        if cell.value == "Verificação cruzada (Seção 5 = Seção 2/3):"
    )
    assert "ISNUMBER" in formula(f"B{check_label_row}")

    # Seção 7 — três metodologias de controle e % de divergência.
    controles_row = next(
        cell.row
        for row in sheet.iter_rows(min_col=1, max_col=1)
        for cell in row
        if cell.value == "Controles de resultado"
    )
    for offset in range(2, 5):
        r = controles_row + offset
        assert 'IF(B13=0,"Sem ocorrências"' in formula(f"B{r}")
        assert 'IF(B13=0,"Não aplicável"' in formula(f"C{r}")
    divergentes_label_row = next(
        cell.row
        for row in sheet.iter_rows(min_col=1, max_col=1)
        for cell in row
        if cell.value == "Registros com limite ITSM superior ao contratual bruto:"
    )
    assert 'IF(B13=0,"Sem ocorrências"' in formula(f"D{divergentes_label_row}")

    # Seção 9 — penalidade não pode ser calculada como se a meta tivesse
    # sido descumprida quando não há incidentes.
    penalidade_base_row = next(
        cell.row
        for row in sheet.iter_rows(min_col=1, max_col=1)
        for cell in row
        if cell.value == "Penalidade-base:"
    )
    assert 'IF(B13=0,"Não aplicável"' in formula(f"B{penalidade_base_row}")


def test_section_3_narrativa_reflects_actual_margin_sign(tmp_path: Path) -> None:
    """A-02: a narrativa da Seção 3 (célula A26) não pode dizer "abaixo do
    mínimo" quando a meta foi superada."""
    config_dir, data_dir = _write_fixture(tmp_path)
    categorias_file = load_categorias(config_dir / "categorias.yaml")
    output_path = tmp_path / "sintetico.xlsx"
    periodo = PeriodoAfericao(date(2026, 6, 1), date(2026, 6, 30))

    write_sintetico_workbook(
        categorias_file, config_dir, data_dir, output_path, periodo=periodo
    )

    wb = load_workbook(output_path)
    sheet = wb["INMS 1.1"]
    formula = sheet["A26"].value
    assert isinstance(formula, str)

    assert formula.startswith(
        '=IF(B13=0,"Sem incidentes abertos no período — indicador não mensurável.",'
    )
    assert "IF(B25<0,CONCATENATE(" in formula
    assert "IF(B25=0,CONCATENATE(" in formula
    # As três narrativas (abaixo/exato/acima) coexistem na fórmula, cada uma
    # sob a guarda IF correta — não há mais um único texto fixo com ABS().
    assert 'ABS(B25)," incidente(s) abaixo do mínimo necessário."' in formula
    assert 'exatamente no mínimo necessário."' in formula
    assert '",B25," incidente(s) acima do mínimo necessário."' in formula
    assert formula.count("CONCATENATE(") == 3


def _direct_write_sheet_kwargs(categorias_file: object) -> dict[str, object]:
    """Kwargs mínimos e válidos para chamar `inms_1_1_audit.write_sheet`
    diretamente (sem passar por `write_sintetico_workbook`/config YAML) —
    usado por testes que precisam de controle fino sobre `rows` ou que
    exercitam o mesmo workbook duas vezes."""
    from pyauditor.config.categorias import CategoriasFile

    assert isinstance(categorias_file, CategoriasFile)
    grupo_executor_entries: list[tuple[str, GrupoExecutorMode]] = [
        (key, entry)
        for key, categoria in categorias_file.categorias.items()
        for inms_key, entry in categoria.inms.items()
        if inms_key == "1.1" and isinstance(entry, GrupoExecutorMode)
    ]
    rows = [
        {
            "Nº Solicitacao": "1",
            "Atividades": "Reset de senha",
            "DataHoraSolicitacao": "01/06/2026 08:00",
            "DataHoraLimite": "01/06/2026 10:00",
            "DataHoraFim": "01/06/2026 09:30",
            "No prazo": "S",
            "Grupo_executor": "N1",
            "TecnicoExecutor": "Fulano",
        },
    ]
    return dict(
        categorias_file=categorias_file,
        grupo_executor_entries=grupo_executor_entries,
        whole_indicator_entries=[],
        fieldnames=list(rows[0].keys()),
        rows=rows,
        target_operator=">=",
        target_value=98.0,
        penalty_base_points=165.0,
        penalty_step_points=20.0,
        penalty_step_size_pct=0.1,
        contract="40/2022 - Ministério da Cultura",
        periodo=None,
        raw_csv_path=Path("/tmp/fake-dir/inms-01.csv"),
        generated_at=datetime(2026, 6, 1, 12, 0),
    )


def test_target_operator_le_is_rejected(tmp_path: Path) -> None:
    """A-03: este renderer só tem semântica coerente (Seção 3/9) para meta
    mínima ('>=') — outro operador deve falhar explicitamente, não gerar
    fórmulas/labels contraditórios."""
    config_dir, _data_dir = _write_fixture(tmp_path)
    categorias_file = load_categorias(config_dir / "categorias.yaml")
    kwargs = _direct_write_sheet_kwargs(categorias_file)
    kwargs["target_operator"] = "<="

    with pytest.raises(ValueError, match="target_operator"):
        inms_1_1_audit.write_sheet(Workbook(), "INMS 1.1", **kwargs)  # type: ignore[arg-type]


@pytest.mark.parametrize(
    ("penalty_kwarg", "bad_value", "match"),
    [
        ("target_value", 150.0, "target_value"),
        ("target_value", -1.0, "target_value"),
        ("penalty_base_points", -1.0, "penalty_base_points"),
        ("penalty_step_points", -1.0, "penalty_step_points"),
        ("penalty_step_size_pct", 0.0, "penalty_step_size_pct"),
        ("penalty_step_size_pct", -0.1, "penalty_step_size_pct"),
    ],
)
def test_out_of_range_params_are_rejected(
    tmp_path: Path, penalty_kwarg: str, bad_value: float, match: str
) -> None:
    """A-07: parâmetros fora de faixa devem falhar explicitamente na
    fronteira, não gerar #DIV/0! ou penalidade sem sentido de domínio."""
    config_dir, _data_dir = _write_fixture(tmp_path)
    categorias_file = load_categorias(config_dir / "categorias.yaml")
    kwargs = _direct_write_sheet_kwargs(categorias_file)
    kwargs[penalty_kwarg] = bad_value

    with pytest.raises(ValueError, match=match):
        inms_1_1_audit.write_sheet(Workbook(), "INMS 1.1", **kwargs)  # type: ignore[arg-type]


@pytest.mark.parametrize("raw_value", ["s", "n", " S ", " N ", "Sim", "Não", "N/A", ""])
def test_no_prazo_invalid_or_variant_values(tmp_path: Path, raw_value: str) -> None:
    """A-04: valores diferentes de 'S'/'N' (após strip+upper) devem falhar
    explicitamente — sem isso, entram no IAP (denominador) sem entrar em
    nenhum dos dois COUNTIF de classificação."""
    config_dir, _data_dir = _write_fixture(tmp_path)
    categorias_file = load_categorias(config_dir / "categorias.yaml")
    kwargs = _direct_write_sheet_kwargs(categorias_file)
    rows = kwargs["rows"]
    assert isinstance(rows, list)
    rows[0]["No prazo"] = raw_value

    if raw_value.strip().upper() in {"S", "N"}:
        inms_1_1_audit.write_sheet(Workbook(), "INMS 1.1", **kwargs)  # type: ignore[arg-type]
    else:
        with pytest.raises(ValueError, match="No prazo"):
            inms_1_1_audit.write_sheet(Workbook(), "INMS 1.1", **kwargs)  # type: ignore[arg-type]


def test_no_prazo_lowercase_and_spaces_are_normalized(tmp_path: Path) -> None:
    config_dir, _data_dir = _write_fixture(tmp_path)
    categorias_file = load_categorias(config_dir / "categorias.yaml")
    kwargs = _direct_write_sheet_kwargs(categorias_file)
    rows = kwargs["rows"]
    assert isinstance(rows, list)
    rows[0]["No prazo"] = " s "

    workbook = Workbook()
    inms_1_1_audit.write_sheet(workbook, "INMS 1.1", **kwargs)  # type: ignore[arg-type]

    assert workbook["INMS 1.1"]["X2"].value == "S"


def test_iap_denominator_uses_rows_not_counta(tmp_path: Path) -> None:
    """A-05: registro com Nº Solicitação vazio ou duplicado não pode
    divergir da contagem de linhas brutas (nota da Seção 1 vs. IAP)."""
    config_dir, data_dir = _write_fixture(tmp_path)
    csv_with_blank_and_duplicate_id = (
        "Nº Solicitacao;Atividades;DataHoraSolicitacao;DataHoraLimite;DataHoraFim;"
        "No prazo;Grupo_executor;TecnicoExecutor\n"
        ";Reset de senha;01/06/2026 08:00;01/06/2026 10:00;01/06/2026 09:30;S;N1;Fulano\n"
        "1;Instalação;01/06/2026 08:00;01/06/2026 10:00;01/06/2026 09:00;S;N1;Fulano\n"
        "1;Acesso;01/06/2026 08:00;01/06/2026 10:00;01/06/2026 09:00;N;N1;Fulano\n"
    )
    (data_dir / "inms-01.csv").write_text(csv_with_blank_and_duplicate_id, encoding="utf-8")
    categorias_file = load_categorias(config_dir / "categorias.yaml")
    output_path = tmp_path / "sintetico.xlsx"
    periodo = PeriodoAfericao(date(2026, 6, 1), date(2026, 6, 30))

    write_sintetico_workbook(
        categorias_file, config_dir, data_dir, output_path, periodo=periodo
    )

    wb = load_workbook(output_path)
    sheet = wb["INMS 1.1"]
    assert sheet["B13"].value == "=ROWS($R$2:$R$4)"
    assert sheet["B6"].value == "inms-01.csv (3 registros brutos)"


def test_force_recalc_flags_are_set(tmp_path: Path) -> None:
    """A-06: sem `fullCalcOnLoad`, um leitor com `data_only=True` (ou
    conversor headless) pode devolver células vazias em vez do resultado
    calculado das fórmulas desta aba."""
    config_dir, data_dir = _write_fixture(tmp_path)
    categorias_file = load_categorias(config_dir / "categorias.yaml")
    output_path = tmp_path / "sintetico.xlsx"
    periodo = PeriodoAfericao(date(2026, 6, 1), date(2026, 6, 30))

    write_sintetico_workbook(
        categorias_file, config_dir, data_dir, output_path, periodo=periodo
    )

    wb = load_workbook(output_path)
    assert wb.calculation.fullCalcOnLoad is True
    assert wb.calculation.forceFullCalc is True


def test_section_8_exposes_rejected_record_count(tmp_path: Path) -> None:
    """M-01: registros com data ausente/inválida ou encerramento anterior
    à abertura são excluídos da média/mediana — a contagem de excluídos
    precisa ficar visível, não só implícita na fórmula `_AG`."""
    config_dir, data_dir = _write_fixture(tmp_path)
    csv_with_bad_dates = (
        "Nº Solicitacao;Atividades;DataHoraSolicitacao;DataHoraLimite;DataHoraFim;"
        "No prazo;Grupo_executor;TecnicoExecutor\n"
        "1;Reset de senha;31/02/2026 08:00;01/06/2026 10:00;01/06/2026 09:30;S;N1;Fulano\n"
        "2;Instalação;01/06/2026 08:00;01/06/2026 10:00;;S;N1;Fulano\n"
        "3;Acesso;01/06/2026 08:00;01/06/2026 10:00;01/06/2026 09:00;S;N1;Fulano\n"
    )
    (data_dir / "inms-01.csv").write_text(csv_with_bad_dates, encoding="utf-8")
    categorias_file = load_categorias(config_dir / "categorias.yaml")
    output_path = tmp_path / "sintetico.xlsx"
    periodo = PeriodoAfericao(date(2026, 6, 1), date(2026, 6, 30))

    write_sintetico_workbook(
        categorias_file, config_dir, data_dir, output_path, periodo=periodo
    )

    wb = load_workbook(output_path)
    sheet = wb["INMS 1.1"]
    rejeitados_cell = next(
        cell
        for row in sheet.iter_rows(min_col=1, max_col=1)
        for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("Registros excluídos da média")
    )
    assert isinstance(rejeitados_cell.row, int)
    rejeitados_row = rejeitados_cell.row
    assert sheet.cell(row=rejeitados_row, column=2).value == '=COUNTIF($AG$2:$AG$4,"")'


def test_atraso_medio_uses_itsm_calculated_column(tmp_path: Path) -> None:
    """M-02: o rótulo diz "vs. limite ITSM" — o critério de seleção do
    AVERAGEIF precisa usar a coluna calculada (_AD), não a classificação do
    fornecedor (_X), para não contradizer o próprio rótulo."""
    config_dir, data_dir = _write_fixture(tmp_path)
    categorias_file = load_categorias(config_dir / "categorias.yaml")
    output_path = tmp_path / "sintetico.xlsx"
    periodo = PeriodoAfericao(date(2026, 6, 1), date(2026, 6, 30))

    write_sintetico_workbook(
        categorias_file, config_dir, data_dir, output_path, periodo=periodo
    )

    wb = load_workbook(output_path)
    sheet = wb["INMS 1.1"]
    atraso_cell = next(
        cell
        for row in sheet.iter_rows(min_col=1, max_col=1)
        for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("Atraso médio dos registros")
    )
    assert isinstance(atraso_cell.row, int)
    formula = sheet.cell(row=atraso_cell.row, column=2).value
    assert isinstance(formula, str)
    assert "$AD$2:$AD$5" in formula
    assert "$X$2:$X$5" not in formula
    assert 'IF(COUNTIF($AD$2:$AD$5,"N")=0,"Sem atrasos"' in formula


def test_limite_itsm_column_renamed_unidirectional(tmp_path: Path) -> None:
    """M-03: a coluna/rótulo deixam explícito que só a prorrogação indevida
    (limite ITSM > contratual bruto) é sinalizada — um limite ITSM mais
    rígido não é uma divergência no sentido contratual."""
    config_dir, data_dir = _write_fixture(tmp_path)
    categorias_file = load_categorias(config_dir / "categorias.yaml")
    output_path = tmp_path / "sintetico.xlsx"
    periodo = PeriodoAfericao(date(2026, 6, 1), date(2026, 6, 30))

    write_sintetico_workbook(
        categorias_file, config_dir, data_dir, output_path, periodo=periodo
    )

    wb = load_workbook(output_path)
    sheet = wb["INMS 1.1"]
    assert sheet["AC1"].value == "Limite ITSM superior ao contratual bruto"


def test_limite_itsm_menor_que_contratual_bruto_nao_e_sinalizado(tmp_path: Path) -> None:
    """M-03: limite ITSM mais rígido que o contratual bruto (2h corridas) não
    é uma divergência no sentido unidirecional adotado — a amostra da Seção 7
    (calculada em Python, não em fórmula) deve continuar vazia."""
    config_dir, data_dir = _write_fixture(tmp_path)
    csv_com_limite_itsm_mais_rigido = (
        "Nº Solicitacao;Atividades;DataHoraSolicitacao;DataHoraLimite;DataHoraFim;"
        "No prazo;Grupo_executor;TecnicoExecutor\n"
        # Limite ITSM (1h) é mais rígido que o contratual bruto (abertura + 2h).
        "1;Reset de senha;01/06/2026 08:00;01/06/2026 09:00;01/06/2026 08:30;"
        "S;N1;Fulano\n"
    )
    (data_dir / "inms-01.csv").write_text(csv_com_limite_itsm_mais_rigido, encoding="utf-8")
    categorias_file = load_categorias(config_dir / "categorias.yaml")
    output_path = tmp_path / "sintetico.xlsx"
    periodo = PeriodoAfericao(date(2026, 6, 1), date(2026, 6, 30))

    write_sintetico_workbook(
        categorias_file, config_dir, data_dir, output_path, periodo=periodo
    )

    wb = load_workbook(output_path)
    sheet = wb["INMS 1.1"]
    sample_row = next(
        cell.row
        for row in sheet.iter_rows(min_col=1, max_col=1)
        for cell in row
        if isinstance(cell.value, str)
        and cell.value.startswith("Nenhum limite ITSM superior ao contratual bruto")
    )
    assert isinstance(sample_row, int)
    assert sample_row > 0


def test_duplicate_grupo_with_different_categories_raises() -> None:
    """M-05: se o mesmo grupo executor aparecesse mapeado em duas
    categorias diferentes, o VLOOKUP da aba mascararia isso silenciosamente
    (sempre a primeira ocorrência) — deve falhar explicitamente antes."""
    with pytest.raises(ValueError, match="categorias devem ser disjuntas"):
        _raise_duplicate_grupo_conflict()


def _raise_duplicate_grupo_conflict() -> None:
    from pyauditor.config.categorias import CategoriaConfig, CategoriasFile

    categorias_file = CategoriasFile(
        categorias={
            "ATENDIMENTO_N1": CategoriaConfig(
                label="Atendimento Remoto aos Usuários",
                inms={"1.1": GrupoExecutorMode(mode="grupo_executor", in_values=["N1"])},
            ),
            "ATENDIMENTO_N2": CategoriaConfig(
                label="Atendimento Presencial aos Usuários",
                inms={"1.1": GrupoExecutorMode(mode="grupo_executor", in_values=["N1"])},
            ),
        }
    )
    grupo_executor_entries = [
        (key, entry)
        for key, categoria in categorias_file.categorias.items()
        for inms_key, entry in categoria.inms.items()
        if inms_key == "1.1" and isinstance(entry, GrupoExecutorMode)
    ]
    _build_grupo_rows(
        categorias_file, grupo_executor_entries, real_values={"N1"}
    )


def test_write_sheet_twice_in_same_workbook_uses_unique_table_names(tmp_path: Path) -> None:
    """M-06: nomes de `Table` são únicos por workbook, não por aba — chamar
    este renderer duas vezes no mesmo workbook (dois contratos/abas) não
    pode falhar por colisão de nome de tabela."""
    config_dir, _data_dir = _write_fixture(tmp_path)
    categorias_file = load_categorias(config_dir / "categorias.yaml")
    kwargs = _direct_write_sheet_kwargs(categorias_file)
    workbook = Workbook()

    inms_1_1_audit.write_sheet(workbook, "INMS 1.1 - A", **kwargs)  # type: ignore[arg-type]
    inms_1_1_audit.write_sheet(workbook, "INMS 1.1 - B", **kwargs)  # type: ignore[arg-type]

    table_names = {
        table.name for sheet in workbook.worksheets for table in sheet.tables.values()
    }
    assert "TabelaGrupoExecutor" in table_names
    assert "TabelaGrupoExecutor_2" in table_names


def test_write_sheet_atomic_rollback_on_duplicate_grupo(tmp_path: Path) -> None:
    """M-08: uma falha no meio da escrita (aqui, a validação de grupo
    duplicado do ticket 14) não pode deixar a aba parcialmente criada no
    workbook."""
    config_dir, _data_dir = _write_fixture(tmp_path)
    categorias_file_ok = load_categorias(config_dir / "categorias.yaml")
    kwargs = _direct_write_sheet_kwargs(categorias_file_ok)

    from pyauditor.config.categorias import CategoriaConfig, CategoriasFile

    categorias_file_conflito = CategoriasFile(
        categorias={
            "ATENDIMENTO_N1": CategoriaConfig(
                label="Atendimento Remoto aos Usuários",
                inms={"1.1": GrupoExecutorMode(mode="grupo_executor", in_values=["N1"])},
            ),
            "ATENDIMENTO_N2": CategoriaConfig(
                label="Atendimento Presencial aos Usuários",
                inms={"1.1": GrupoExecutorMode(mode="grupo_executor", in_values=["N1"])},
            ),
        }
    )
    kwargs["categorias_file"] = categorias_file_conflito
    kwargs["grupo_executor_entries"] = [
        (key, entry)
        for key, categoria in categorias_file_conflito.categorias.items()
        for inms_key, entry in categoria.inms.items()
        if inms_key == "1.1" and isinstance(entry, GrupoExecutorMode)
    ]

    workbook = Workbook()
    sheets_before = set(workbook.sheetnames)

    with pytest.raises(ValueError, match="categorias devem ser disjuntas"):
        inms_1_1_audit.write_sheet(workbook, "INMS 1.1", **kwargs)  # type: ignore[arg-type]

    assert set(workbook.sheetnames) == sheets_before


def test_generated_at_is_injectable(tmp_path: Path) -> None:
    """B-01: a data de geração precisa ser injetável para permitir teste
    determinístico, em vez de depender de `datetime.now()`."""
    config_dir, data_dir = _write_fixture(tmp_path)
    categorias_file = load_categorias(config_dir / "categorias.yaml")
    output_path = tmp_path / "sintetico.xlsx"
    periodo = PeriodoAfericao(date(2026, 6, 1), date(2026, 6, 30))
    fixed_now = datetime(2020, 1, 1, 10, 30)

    write_sintetico_workbook(
        categorias_file, config_dir, data_dir, output_path, periodo=periodo,
        generated_at=fixed_now,
    )

    wb = load_workbook(output_path)
    assert wb["INMS 1.1"]["B7"].value == fixed_now


def test_source_note_does_not_expose_full_path(tmp_path: Path) -> None:
    """B-02: o caminho completo do CSV bruto pode revelar usuário/estrutura
    de diretório interna — só o nome do arquivo deve aparecer na aba."""
    config_dir, data_dir = _write_fixture(tmp_path)
    categorias_file = load_categorias(config_dir / "categorias.yaml")
    output_path = tmp_path / "sintetico.xlsx"
    periodo = PeriodoAfericao(date(2026, 6, 1), date(2026, 6, 30))

    write_sintetico_workbook(
        categorias_file, config_dir, data_dir, output_path, periodo=periodo
    )

    wb = load_workbook(output_path)
    note = wb["INMS 1.1"]["B6"].value
    assert isinstance(note, str)
    assert str(data_dir) not in note
    assert "inms-01.csv" in note


def test_support_columns_are_hidden_and_sheet_is_protected(tmp_path: Path) -> None:
    """B-03: colunas de apoio (R:AM) ocultas e planilha protegida — só os
    campos de justificativa/evidência de preenchimento manual continuam
    editáveis."""
    config_dir, data_dir = _write_fixture(tmp_path)
    categorias_file = load_categorias(config_dir / "categorias.yaml")
    output_path = tmp_path / "sintetico.xlsx"
    periodo = PeriodoAfericao(date(2026, 6, 1), date(2026, 6, 30))

    write_sintetico_workbook(
        categorias_file, config_dir, data_dir, output_path, periodo=periodo
    )

    wb = load_workbook(output_path)
    sheet = wb["INMS 1.1"]
    assert sheet.column_dimensions["R"].hidden is True
    assert sheet.column_dimensions["AM"].hidden is True
    # C-02 × B-03: a coluna de qualidade dos dados (AJ) fica visível de
    # propósito — não é fonte de fórmula, é o indicador visual de linhas
    # com data ausente/inválida que o ticket C-02 exige na planilha.
    assert sheet.column_dimensions["AJ"].hidden is not True
    assert sheet.protection.sheet is True
    # Coluna 10 (Justificativa de exclusão) da Seção 4 continua editável.
    justificativa_row = next(
        cell.row
        for row in sheet.iter_rows(min_col=1, max_col=1)
        for cell in row
        if cell.value == "SEÇÃO 4 · DETALHAMENTO POR GRUPO EXECUTOR"
    ) + 2
    assert sheet.cell(row=justificativa_row, column=10).protection.locked is False
    # Uma célula de fórmula comum continua bloqueada.
    assert sheet["A13"].protection.locked is True


def test_section_5_nivel_rows_use_body_font_consistently(tmp_path: Path) -> None:
    """B-04: `pct` e `tempo` das linhas N1/N2/N3 da Seção 5 recebiam a
    fonte padrão do openpyxl (não `BODY_FONT`), ao contrário dos demais
    campos da mesma seção."""
    config_dir, data_dir = _write_fixture(tmp_path)
    categorias_file = load_categorias(config_dir / "categorias.yaml")
    output_path = tmp_path / "sintetico.xlsx"
    periodo = PeriodoAfericao(date(2026, 6, 1), date(2026, 6, 30))

    write_sintetico_workbook(
        categorias_file, config_dir, data_dir, output_path, periodo=periodo
    )

    wb = load_workbook(output_path)
    sheet = wb["INMS 1.1"]
    nivel_row = next(
        cell.row
        for row in sheet.iter_rows(min_col=1, max_col=1)
        for cell in row
        if cell.value == "N1"
    )
    assert sheet.cell(row=nivel_row, column=5).font.name == "Arial"
    assert sheet.cell(row=nivel_row, column=6).font.name == "Arial"


def test_categoria_sem_nivel_configurado_usa_sem_nivel_sentinela(tmp_path: Path) -> None:
    """M-07: uma categoria conhecida (com `in_values` mapeado) mas ausente de
    `_NIVEL_BY_CATEGORIA` não pode cair no valor vazio `""` — o comentário do
    próprio código explica que célula vazia quebra os filtros `COUNTIFS` por
    Nível; o sentinela `_SEM_NIVEL` ("—") precisa ser usado também aqui, não
    só para o grupo "outros"."""
    config_dir, data_dir = _write_fixture(tmp_path)
    categorias_com_categoria_desconhecida = """\
categorias:
  ATENDIMENTO_N1:
    label: "Atendimento Remoto aos Usuários"
    inms:
      "1.1": {mode: grupo_executor, in_values: ["N1"]}
  SUPORTE_ESPECIAL:
    label: "Suporte Especializado"
    inms:
      "1.1": {mode: grupo_executor, in_values: ["N2"]}
"""
    (config_dir / "categorias.yaml").write_text(
        categorias_com_categoria_desconhecida, encoding="utf-8"
    )
    categorias_file = load_categorias(config_dir / "categorias.yaml")
    output_path = tmp_path / "sintetico.xlsx"
    periodo = PeriodoAfericao(date(2026, 6, 1), date(2026, 6, 30))

    write_sintetico_workbook(
        categorias_file, config_dir, data_dir, output_path, periodo=periodo
    )

    wb = load_workbook(output_path)
    sheet = wb["INMS 1.1"]
    linha_n2 = next(
        cell.row
        for row in sheet.iter_rows(min_col=1, max_col=3)
        for cell in row
        if cell.column == 3 and cell.value == "N2"
    )
    assert sheet.cell(row=linha_n2, column=2).value == _SEM_NIVEL

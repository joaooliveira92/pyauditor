"""`excel/_workbook.py` — utilitários genéricos de workbook (recálculo
forçado, nome de tabela único, criação atômica de aba), reusáveis por
qualquer renderer de `excel/`.
"""

import pytest
from openpyxl import Workbook
from openpyxl.worksheet.table import Table

from pyauditor.excel._workbook import (
    create_sheet_atomic,
    force_recalc,
    unique_table_name,
)


def test_force_recalc_sets_full_recalculation_flags() -> None:
    workbook = Workbook()

    force_recalc(workbook)

    assert workbook.calculation.fullCalcOnLoad is True
    assert workbook.calculation.forceFullCalc is True
    assert workbook.calculation.calcMode == "auto"


def test_unique_table_name_returns_base_name_when_free() -> None:
    workbook = Workbook()

    assert unique_table_name(workbook, "TabelaGrupoExecutor") == "TabelaGrupoExecutor"


def test_unique_table_name_suffixes_on_collision() -> None:
    workbook = Workbook()
    sheet1 = workbook.create_sheet("Aba1")
    sheet1["A1"] = "x"
    sheet1.add_table(Table(displayName="TabelaGrupoExecutor", ref="A1:A1"))

    assert unique_table_name(workbook, "TabelaGrupoExecutor") == "TabelaGrupoExecutor_2"


def test_unique_table_name_skips_multiple_collisions() -> None:
    workbook = Workbook()
    for n, name in enumerate(("TabelaGrupoExecutor", "TabelaGrupoExecutor_2"), start=1):
        sheet = workbook.create_sheet(f"Aba{n}")
        sheet["A1"] = "x"
        sheet.add_table(Table(displayName=name, ref="A1:A1"))

    assert unique_table_name(workbook, "TabelaGrupoExecutor") == "TabelaGrupoExecutor_3"


def test_create_sheet_atomic_yields_usable_sheet() -> None:
    workbook = Workbook()

    with create_sheet_atomic(workbook, "Nova Aba") as sheet:
        sheet["A1"] = "ok"

    assert workbook["Nova Aba"]["A1"].value == "ok"


def test_create_sheet_atomic_removes_sheet_on_failure() -> None:
    workbook = Workbook()
    sheets_before = set(workbook.sheetnames)

    with (
        pytest.raises(ValueError, match="boom"),
        create_sheet_atomic(workbook, "Nova Aba") as sheet,
    ):
        sheet["A1"] = "parcial"
        raise ValueError("boom")

    assert set(workbook.sheetnames) == sheets_before
    assert "Nova Aba" not in workbook.sheetnames

from pathlib import Path

from openpyxl import Workbook, load_workbook

from pyauditor.excel.capa import FIELD_LABELS, SHEET_NAME, SITUACOES, bootstrap_capa, read_valor_mensal_vigente


def test_bootstrap_creates_capa_with_expected_fields(tmp_path: Path) -> None:
    capa_path = tmp_path / "capa.xlsx"

    created = bootstrap_capa(capa_path)

    assert created is True
    assert capa_path.exists()

    workbook = load_workbook(capa_path)
    assert SHEET_NAME in workbook.sheetnames
    sheet = workbook[SHEET_NAME]

    labels_in_sheet = [sheet.cell(row=4 + i, column=1).value for i in range(len(FIELD_LABELS))]
    assert labels_in_sheet == list(FIELD_LABELS)


def test_bootstrap_defaults_situacao_to_first_option(tmp_path: Path) -> None:
    capa_path = tmp_path / "capa.xlsx"
    bootstrap_capa(capa_path)

    workbook = load_workbook(capa_path)
    sheet = workbook[SHEET_NAME]

    situacao_row = 4 + FIELD_LABELS.index("Situação geral da aferição")
    assert sheet.cell(row=situacao_row, column=2).value == SITUACOES[0]


def test_bootstrap_is_idempotent_never_overwrites(tmp_path: Path) -> None:
    capa_path = tmp_path / "capa.xlsx"
    bootstrap_capa(capa_path)

    # Simulate a fiscal técnico having filled in real data.
    workbook = load_workbook(capa_path)
    sheet = workbook[SHEET_NAME]
    sheet.cell(row=4, column=2, value="40/2022 - Ministério Cultura")
    workbook.save(capa_path)
    mtime_before = capa_path.stat().st_mtime_ns

    created_again = bootstrap_capa(capa_path)

    assert created_again is False
    assert capa_path.stat().st_mtime_ns == mtime_before
    workbook_after = load_workbook(capa_path)
    assert workbook_after[SHEET_NAME].cell(row=4, column=2).value == "40/2022 - Ministério Cultura"


def test_bootstrap_creates_parent_directories(tmp_path: Path) -> None:
    capa_path = tmp_path / "nested" / "dir" / "capa.xlsx"

    created = bootstrap_capa(capa_path)

    assert created is True
    assert capa_path.exists()


def test_read_valor_mensal_vigente_none_when_blank(tmp_path: Path) -> None:
    capa_path = tmp_path / "capa.xlsx"
    bootstrap_capa(capa_path)

    assert read_valor_mensal_vigente(capa_path) is None


def test_read_valor_mensal_vigente_reads_fiscal_filled_value(tmp_path: Path) -> None:
    capa_path = tmp_path / "capa.xlsx"
    bootstrap_capa(capa_path)

    workbook = load_workbook(capa_path)
    sheet = workbook[SHEET_NAME]
    row = 4 + FIELD_LABELS.index("Valor mensal vigente")
    sheet.cell(row=row, column=2, value=125_000.50)
    workbook.save(capa_path)

    assert read_valor_mensal_vigente(capa_path) == 125_000.50


def test_read_valor_mensal_vigente_none_when_sheet_missing(tmp_path: Path) -> None:
    capa_path = tmp_path / "capa.xlsx"
    workbook = Workbook()
    workbook.save(capa_path)

    assert read_valor_mensal_vigente(capa_path) is None

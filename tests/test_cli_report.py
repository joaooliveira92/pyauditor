import json
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

from pyauditor.cli.report import missing_publication_fields, run_report
from pyauditor.excel.capa import COMMON_FIELD_LABELS, ORGAO_FIELD_LABELS, bootstrap_capa_csv
from pyauditor.excel.capa import SHEET_NAME as CAPA_SHEET_NAME
from pyauditor.excel.equipe import RESPONSAVEL_LABELS
from pyauditor.excel.report import GLOSAS_SHEET, INMS_BASE_SHEET

OBJETOS_CSV = """Item,Categoria,Valor
1,Central de ServiÃ§os,"R$ 148.205,54"
2,GT dos Projetos e OperaÃ§Ãµes,"R$ 77.654,90"
3,Banco de Dados,"R$ 43.888,89"
4,"AplicaÃ§Ãµes, virtualizaÃ§Ã£o","R$ 59.694,54"
5,ServiÃ§os Corporativos,"R$ 21.035,21"
6,Armazenamento e Backup,"R$ 16.145,94"
7,Redes,"R$ 31.382,28"
8,"SeguranÃ§a da InformaÃ§Ã£o","R$ 34.143,44"
9,DevOps,"R$ 28.912,84"
"""


def _write_summary(
    roms_dir: Path, competencia: str, indicator_id: str, contractual_id: str
) -> None:
    target_dir = roms_dir / competencia
    target_dir.mkdir(parents=True, exist_ok=True)
    summary = {
        "indicator_id": indicator_id,
        "contractual_id": contractual_id,
        "name": f"Indicador {contractual_id}",
        "asset": None,
        "orgao": "MinC",
        "shape": "ratio",
        "target_operator": ">=",
        "target_value": 98.0,
        "result_pct": 97.71,
        "conforms": False,
        "penalty_points": 222.14,
        "numerator": 171,
        "denominator": 175,
        "hard_failure": False,
    }
    (target_dir / f"{indicator_id}.json").write_text(json.dumps(summary), encoding="utf-8")
    (target_dir / f"{indicator_id}.md").write_text("# ROM", encoding="utf-8")


def _write_summary_with_points(
    roms_dir: Path, competencia: str, indicator_id: str, contractual_id: str, penalty_points: float
) -> None:
    target_dir = roms_dir / competencia
    target_dir.mkdir(parents=True, exist_ok=True)
    summary = {
        "indicator_id": indicator_id,
        "contractual_id": contractual_id,
        "name": f"Indicador {contractual_id}",
        "asset": None,
        "orgao": "MinC",
        "shape": "ratio",
        "target_operator": ">=",
        "target_value": 98.0,
        "result_pct": 97.71,
        "conforms": False,
        "penalty_points": penalty_points,
        "numerator": 171,
        "denominator": 175,
        "hard_failure": False,
    }
    (target_dir / f"{indicator_id}.json").write_text(json.dumps(summary), encoding="utf-8")
    (target_dir / f"{indicator_id}.md").write_text("# ROM", encoding="utf-8")


def _scaffold_capas(tmp_path: Path) -> Path:
    """Cria capa.csv (comum) + capa_MinC.csv + objetos.csv e devolve o comum."""
    comum = tmp_path / "capa.csv"
    bootstrap_capa_csv(comum, COMMON_FIELD_LABELS)
    bootstrap_capa_csv(tmp_path / "capa_MinC.csv", ORGAO_FIELD_LABELS)
    (tmp_path / "objetos.csv").write_text(OBJETOS_CSV, encoding="utf-8-sig")
    return comum


def test_sidecar_legado_sem_campos_de_periodo_carrega_com_defaults(tmp_path: Path) -> None:
    """Spec §8 — sidecar novo lê sidecar antigo: chaves ausentes caem nos
    defaults None e o relatório segue de pé."""
    from pyauditor.rom.loading import load_summaries

    roms_dir = tmp_path / "roms"
    _write_summary(roms_dir, "2026-06", "INMS-1.1", "INMS 1.1")

    summaries = load_summaries(roms_dir / "2026-06")

    assert len(summaries) == 1
    assert summaries[0].dropped_out_of_period is None
    assert summaries[0].undated_dropped is None


def test_run_report_persists_glosa_historico(tmp_path: Path) -> None:
    comum = _scaffold_capas(tmp_path)
    roms_dir = tmp_path / "roms"
    # 40000 pontos -> 40% bruto, capped 30%, 10 p.p. de saldo rolado
    _write_summary_with_points(roms_dir, "2026-06", "INMS-1.1", "INMS 1.1", 40_000.0)
    output_path = tmp_path / "reports" / "relatorio_2026-06.xlsx"

    exit_code = run_report(
        "2026-06",
        comum,
        roms_dir,
        output_path,
        config_dir=tmp_path / "configs",
        expected_orgao="MinC",
        data_dir=tmp_path,
    )

    assert exit_code.status == "done"
    historico = json.loads((roms_dir / "glosa_historico.json").read_text(encoding="utf-8"))
    assert historico["2026-06"]["saldo_rolado_pct"] == 10.0
    assert historico["2026-06"]["teto_atingido"] is True


def test_run_report_next_competencia_consumes_rollover(tmp_path: Path) -> None:
    comum = _scaffold_capas(tmp_path)
    roms_dir = tmp_path / "roms"
    _write_summary_with_points(roms_dir, "2026-06", "INMS-1.1", "INMS 1.1", 40_000.0)
    run_report(
        "2026-06",
        comum,
        roms_dir,
        tmp_path / "reports" / "relatorio_2026-06.xlsx",
        config_dir=tmp_path / "configs",
        expected_orgao="MinC",
        data_dir=tmp_path,
    )

    # Julho sem nenhuma penalidade nova â€” sÃ³ o saldo rolado de junho (10 p.p.)
    _write_summary_with_points(roms_dir, "2026-07", "INMS-1.1", "INMS 1.1", 0.0)
    output_path = tmp_path / "reports" / "relatorio_2026-07.xlsx"
    exit_code = run_report(
        "2026-07",
        comum,
        roms_dir,
        output_path,
        config_dir=tmp_path / "configs",
        expected_orgao="MinC",
        data_dir=tmp_path,
    )

    assert exit_code.status == "done"
    workbook = load_workbook(output_path)
    glosas_sheet = workbook[GLOSAS_SHEET]
    assert glosas_sheet.cell(row=2, column=3).value == 10.0  # saldo recebido do mÃªs anterior
    assert glosas_sheet.cell(row=2, column=4).value == 10.0  # percentual de ajuste


def test_run_report_final_month_does_not_roll_over(tmp_path: Path) -> None:
    comum = _scaffold_capas(tmp_path)
    roms_dir = tmp_path / "roms"
    _write_summary_with_points(roms_dir, "2026-06", "INMS-1.1", "INMS 1.1", 40_000.0)
    output_path = tmp_path / "reports" / "relatorio_2026-06.xlsx"

    exit_code = run_report(
        "2026-06",
        comum,
        roms_dir,
        output_path,
        config_dir=tmp_path / "configs",
        is_final_month=True,
        expected_orgao="MinC",
        data_dir=tmp_path,
    )

    assert exit_code.status == "done"
    historico = json.loads((roms_dir / "glosa_historico.json").read_text(encoding="utf-8"))
    assert historico["2026-06"]["saldo_rolado_pct"] == 0.0


def test_run_report_rejects_malformed_competencia(tmp_path: Path) -> None:
    result = run_report(
        "../../etc",
        tmp_path / "capa.csv",
        tmp_path / "roms",
        tmp_path / "reports" / "out.xlsx",
        config_dir=tmp_path / "configs",
        data_dir=tmp_path,
    )

    assert result.status == "error"
    assert result.error_message is not None
    assert "competência inválida" in result.error_message


def test_run_report_converts_unexpected_exception_to_error_result(tmp_path: Path) -> None:
    comum = _scaffold_capas(tmp_path)
    roms_dir = tmp_path / "roms"
    _write_summary(roms_dir, "2026-06", "INMS-1.1", "INMS 1.1")

    with patch("pyauditor.cli.report.build_report", side_effect=ValueError("boom")):
        result = run_report(
            "2026-06",
            comum,
            roms_dir,
            tmp_path / "reports" / "out.xlsx",
            config_dir=tmp_path / "configs",
            expected_orgao="MinC",
            data_dir=tmp_path,
        )

    assert result.status == "error"
    assert result.error_message is not None
    assert "boom" in result.error_message


def test_run_report_os_error_message_has_actionable_hint(tmp_path: Path) -> None:
    # Ticket 11: mensagem de falha de escrita ganha dica acionável.
    comum = _scaffold_capas(tmp_path)
    roms_dir = tmp_path / "roms"
    _write_summary(roms_dir, "2026-06", "INMS-1.1", "INMS 1.1")

    with patch(
        "pyauditor.cli.report.build_report", side_effect=PermissionError("Permission denied")
    ):
        result = run_report(
            "2026-06",
            comum,
            roms_dir,
            tmp_path / "reports" / "out.xlsx",
            config_dir=tmp_path / "configs",
            expected_orgao="MinC",
            data_dir=tmp_path,
        )

    assert result.status == "error"
    assert result.error_message is not None
    assert "aberto em outro programa" in result.error_message


def test_run_report_fails_without_capa(tmp_path: Path) -> None:
    comum = tmp_path / "capa.csv"
    roms_dir = tmp_path / "roms"
    _write_summary(roms_dir, "2026-06", "INMS-1.1", "INMS 1.1")

    exit_code = run_report(
        "2026-06",
        comum,
        roms_dir,
        tmp_path / "reports" / "out.xlsx",
        config_dir=tmp_path / "configs",
        expected_orgao="MinC",
        data_dir=tmp_path,
    )

    assert exit_code.status == "error"
    assert not (tmp_path / "reports" / "out.xlsx").exists()


def test_run_report_fails_without_roms(tmp_path: Path) -> None:
    comum = _scaffold_capas(tmp_path)
    roms_dir = tmp_path / "roms"  # never populated

    exit_code = run_report(
        "2026-06",
        comum,
        roms_dir,
        tmp_path / "reports" / "out.xlsx",
        config_dir=tmp_path / "configs",
        expected_orgao="MinC",
        data_dir=tmp_path,
    )

    assert exit_code.status == "error"


def test_run_report_names_the_offending_file_for_a_malformed_summary(tmp_path: Path) -> None:
    comum = _scaffold_capas(tmp_path)
    roms_dir = tmp_path / "roms"
    _write_summary(roms_dir, "2026-06", "INMS-1.1", "INMS 1.1")
    (roms_dir / "2026-06" / "INMS-1.1.json").write_text("not json", encoding="utf-8")

    result = run_report(
        "2026-06",
        comum,
        roms_dir,
        tmp_path / "reports" / "out.xlsx",
        config_dir=tmp_path / "configs",
        expected_orgao="MinC",
        data_dir=tmp_path,
    )

    assert result.status == "error"
    assert result.error_message is not None
    assert "INMS-1.1.json" in result.error_message


def test_run_report_builds_workbook_from_summaries(tmp_path: Path) -> None:
    comum = _scaffold_capas(tmp_path)
    roms_dir = tmp_path / "roms"
    _write_summary(roms_dir, "2026-06", "INMS-1.1", "INMS 1.1")
    output_path = tmp_path / "reports" / "relatorio.xlsx"

    exit_code = run_report(
        "2026-06",
        comum,
        roms_dir,
        output_path,
        config_dir=tmp_path / "configs",
        expected_orgao="MinC",
        data_dir=tmp_path,
    )

    assert exit_code.status == "done"
    assert output_path.exists()

    workbook = load_workbook(output_path)
    assert INMS_BASE_SHEET in workbook.sheetnames
    sheet = workbook[INMS_BASE_SHEET]
    assert sheet.cell(row=2, column=5).value == "INMS 1.01"


def test_run_report_capa_do_orgao_exibe_derivados_da_cli_e_equipe(tmp_path: Path) -> None:
    """Spec competencia-cli-equipe §4/§6 — a CAPA_E_CONTROLE do relatório do
    órgão exibe Competência/períodos derivados do argumento da CLI e os
    responsáveis vindos do equipe.csv."""
    comum = _scaffold_capas(tmp_path)
    primeira_funcao = RESPONSAVEL_LABELS[0]
    (tmp_path / "equipe.csv").write_text(
        f"FUNÇÃO,NOME,SIAPE\n{primeira_funcao},Maria Souza,123456\n", encoding="utf-8-sig"
    )
    roms_dir = tmp_path / "roms"
    _write_summary(roms_dir, "2026-06", "INMS-1.1", "INMS 1.1")
    output_path = tmp_path / "reports" / "relatorio.xlsx"

    exit_code = run_report(
        "2026-06",
        comum,
        roms_dir,
        output_path,
        config_dir=tmp_path / "configs",
        expected_orgao="MinC",
        data_dir=tmp_path,
    )

    assert exit_code.status == "done"
    workbook = load_workbook(output_path)
    sheet = workbook[CAPA_SHEET_NAME]
    capa_rows = {
        sheet.cell(row=r, column=1).value: sheet.cell(row=r, column=2).value
        for r in range(4, sheet.max_row + 1)
        if sheet.cell(row=r, column=1).value
    }
    assert capa_rows["Competência"] == "2026-06"
    assert capa_rows["Período inicial da aferição"] == "01/06/2026"
    assert capa_rows["Período final da aferição"] == "30/06/2026"
    assert capa_rows[primeira_funcao] == "Maria Souza (123456)"


def test_run_report_is_regenerated_not_cumulative(tmp_path: Path) -> None:
    comum = _scaffold_capas(tmp_path)
    roms_dir = tmp_path / "roms"
    _write_summary(roms_dir, "2026-06", "INMS-1.1", "INMS 1.1")
    output_path = tmp_path / "reports" / "relatorio.xlsx"

    run_report(
        "2026-06",
        comum,
        roms_dir,
        output_path,
        config_dir=tmp_path / "configs",
        expected_orgao="MinC",
        data_dir=tmp_path,
    )
    first_size = output_path.stat().st_size

    # Second indicator appears for the same competÃªncia â€” rerun must reflect
    # exactly the current ROMs, not append to the previous run's workbook.
    _write_summary(roms_dir, "2026-06", "INMS-1.2", "INMS 1.2")
    run_report(
        "2026-06",
        comum,
        roms_dir,
        output_path,
        config_dir=tmp_path / "configs",
        expected_orgao="MinC",
        data_dir=tmp_path,
    )

    workbook = load_workbook(output_path)
    sheet = workbook[INMS_BASE_SHEET]
    codes = [str(sheet.cell(row=r, column=5).value) for r in range(2, sheet.max_row + 1)]
    assert sorted(codes) == ["INMS 1.01", "INMS 1.02"]
    assert output_path.stat().st_size != first_size or len(codes) == 2


def test_run_report_reads_valor_base_from_objetos_for_glosas(tmp_path: Path) -> None:
    comum = _scaffold_capas(tmp_path)
    roms_dir = tmp_path / "roms"
    _write_summary(roms_dir, "2026-06", "INMS-1.1", "INMS 1.1")  # penalty_points=222.14
    output_path = tmp_path / "reports" / "relatorio.xlsx"

    exit_code = run_report(
        "2026-06",
        comum,
        roms_dir,
        output_path,
        config_dir=tmp_path / "configs",
        expected_orgao="MinC",
        data_dir=tmp_path,
    )

    assert exit_code.status == "done"
    workbook = load_workbook(output_path)
    glosas_sheet = workbook[GLOSAS_SHEET]
    # Valor-base vem do TOTAL MENSAL de objetos.csv (461.063,58)
    assert glosas_sheet.cell(row=2, column=5).value == 461063.58
    assert glosas_sheet.cell(row=2, column=6).value == round(461063.58 * 0.22214 / 100, 2)


def test_run_report_without_objetos_marks_glosa_nao_calculada(tmp_path: Path) -> None:
    comum = _scaffold_capas(tmp_path)
    (tmp_path / "objetos.csv").unlink()
    roms_dir = tmp_path / "roms"
    _write_summary(roms_dir, "2026-06", "INMS-1.1", "INMS 1.1")
    output_path = tmp_path / "reports" / "relatorio.xlsx"

    exit_code = run_report(
        "2026-06",
        comum,
        roms_dir,
        output_path,
        config_dir=tmp_path / "configs",
        expected_orgao="MinC",
        data_dir=tmp_path,
    )

    assert exit_code.status == "done"
    assert exit_code.glosa_calculada is False
    workbook = load_workbook(output_path)
    glosas_sheet = workbook[GLOSAS_SHEET]
    assert glosas_sheet.cell(row=2, column=5).value is None  # Valor-base
    assert glosas_sheet.cell(row=2, column=6).value is None  # Valor da glosa


def test_run_report_malformed_objetos_is_hard_failure(tmp_path: Path) -> None:
    comum = _scaffold_capas(tmp_path)
    (tmp_path / "objetos.csv").write_text("a;b\n1;2\n", encoding="utf-8-sig")
    roms_dir = tmp_path / "roms"
    _write_summary(roms_dir, "2026-06", "INMS-1.1", "INMS 1.1")

    result = run_report(
        "2026-06",
        comum,
        roms_dir,
        tmp_path / "reports" / "out.xlsx",
        config_dir=tmp_path / "configs",
        expected_orgao="MinC",
        data_dir=tmp_path,
    )

    assert result.status == "error"
    assert result.error_message is not None
    assert "objetos.csv" in result.error_message


def test_missing_publication_fields_empty_capa_returns_all_fields() -> None:
    assert len(missing_publication_fields({})) == 4


def test_missing_publication_fields_complete_capa_returns_empty() -> None:
    campos: dict[str, object] = {
        "Fiscal técnico": "Fulano",
        "Fiscal requisitante": "Beltrano",
        "Fiscal administrativo": "Ciclano",
        "Gestor do contrato": "Sicrano",
    }

    assert missing_publication_fields(campos) == ()


def test_missing_publication_fields_reports_only_missing_ones() -> None:
    campos: dict[str, object] = {
        "Fiscal técnico": "Fulano",
        "Gestor do contrato": "",
    }  # vazio conta como ausente

    missing = missing_publication_fields(campos)

    assert "Fiscal técnico" not in missing
    assert "Gestor do contrato" in missing
    assert "Fiscal requisitante" in missing


def test_run_report_missing_fiscais_is_rascunho_nao_publicavel(tmp_path: Path) -> None:
    # Ticket 02: fiscais ausentes na capa não impedem o processamento, mas
    # marcam o relatório como rascunho — nunca publicável.
    comum = _scaffold_capas(tmp_path)
    roms_dir = tmp_path / "roms"
    _write_summary(roms_dir, "2026-06", "INMS-1.1", "INMS 1.1")
    output_path = tmp_path / "reports" / "relatorio.xlsx"

    result = run_report(
        "2026-06",
        comum,
        roms_dir,
        output_path,
        config_dir=tmp_path / "configs",
        expected_orgao="MinC",
        data_dir=tmp_path,
    )

    assert result.status == "done"
    assert result.publicable is False  # capa_MinC.csv recém-bootstrapped: fiscais vazios


def test_run_report_campos_orfaos_da_capa_sao_ignorados(tmp_path: Path) -> None:
    """Spec competencia-cli-equipe §4 — capa antiga com Competência/períodos/
    responsáveis preenchidos à mão: lida e ignorada; os valores exibidos vêm
    da CLI + equipe.csv (sem warning de divergência — não há mais o que
    divergir)."""
    comum = _scaffold_capas(tmp_path)
    (tmp_path / "capa_MinC.csv").write_text(
        "Capa e controle do contrato;\n;\nCampo;Valor\n"
        "Competência;2026-05\n"
        "Período inicial da aferição;01/01/2000\n"
        "Fiscal técnico;Hand-fill esquecido\n",
        encoding="utf-8-sig",
    )
    roms_dir = tmp_path / "roms"
    _write_summary(roms_dir, "2026-06", "INMS-1.1", "INMS 1.1")
    output_path = tmp_path / "reports" / "relatorio.xlsx"

    result = run_report(
        "2026-06",
        comum,
        roms_dir,
        output_path,
        config_dir=tmp_path / "configs",
        expected_orgao="MinC",
        data_dir=tmp_path,
    )

    assert result.status == "done"
    assert not any("2026-05" in w for w in result.warnings)
    # sem equipe.csv, responsáveis ficam vazios → rascunho
    assert result.publicable is False


def test_run_report_valor_mensal_zero_is_glosa_calculada(tmp_path: Path) -> None:
    # Ticket 01: 0,00 é um valor legítimo — nunca confundido com "não
    # calculada" (isso é reservado à ausência do arquivo objetos.csv inteiro).
    comum = _scaffold_capas(tmp_path)
    zerado = OBJETOS_CSV
    for item_valor in (
        "148.205,54",
        "77.654,90",
        "43.888,89",
        "59.694,54",
        "21.035,21",
        "16.145,94",
        "31.382,28",
        "34.143,44",
        "28.912,84",
    ):
        zerado = zerado.replace(f'"R$ {item_valor}"', '"R$ 0,00"')
    (tmp_path / "objetos.csv").write_text(zerado, encoding="utf-8-sig")
    roms_dir = tmp_path / "roms"
    _write_summary(roms_dir, "2026-06", "INMS-1.1", "INMS 1.1")
    output_path = tmp_path / "reports" / "relatorio.xlsx"

    result = run_report(
        "2026-06",
        comum,
        roms_dir,
        output_path,
        config_dir=tmp_path / "configs",
        expected_orgao="MinC",
        data_dir=tmp_path,
    )

    assert result.status == "done"
    assert result.glosa_calculada is True  # valor_base=0.0, não None
    workbook = load_workbook(output_path)
    glosas_sheet = workbook[GLOSAS_SHEET]
    assert glosas_sheet.cell(row=2, column=5).value == 0.0  # Valor-base — numérico, não vazio

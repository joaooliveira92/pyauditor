from pathlib import Path
from typing import TypedDict

from openpyxl import load_workbook

from pyauditor.cli.run import run_run


class _Paths(TypedDict):
    config_dir: Path
    data_dir: Path
    output_dir: Path
    report_dir: Path
    capa_path: Path
    runs_dir: Path


_CONFIG_YAML = """\
indicator:
  id: INMS-TEST
  contractual_id: "INMS TEST"
  name: Indicador sintético
scope:
  contract: "40/2022 - Ministério da Cultura"
  orgao: MinC
source:
  csv: data.csv
  delimiter: ";"
  encoding: utf-8
  period_column: "DataHoraFim"
quality_gates:
  checks:
    - type: not_null
      column: "DataHoraFim"
calculation:
  shape: ratio
  aggregation: count_distinct
  numerator_filter:
    column: "No prazo"
    equals: "S"
target:
  operator: ">="
  value: 98.0
penalty:
  base_points: 100
  step_points: 10
  step_size_pct: 1.0
"""

_GOOD_CSV = "Nº Solicitacao;DataHoraFim;No prazo\n1;2026-06-01;S\n2;2026-06-02;N\n"
_HARD_FAILURE_CSV = "Nº Solicitacao;DataHoraFim;No prazo\n1;;S\n2;;N\n"

_CATEGORIAS_YAML = """\
categorias:
  DUMMY:
    label: "dummy"
    inms:
      "1.99": {mode: whole_indicator}
"""


def _scaffold_single(tmp_path: Path, *, csv_body: str) -> _Paths:
    (tmp_path / "configs" / "MinC").mkdir(parents=True)
    (tmp_path / "input" / "MinC" / "2026" / "06").mkdir(parents=True)
    (tmp_path / "configs" / "MinC" / "inms-test.yaml").write_text(_CONFIG_YAML, encoding="utf-8")
    (tmp_path / "configs" / "MinC" / "categorias.yaml").write_text(
        _CATEGORIAS_YAML, encoding="utf-8"
    )
    (tmp_path / "input" / "MinC" / "2026" / "06" / "data.csv").write_text(
        csv_body, encoding="utf-8"
    )
    return {
        "config_dir": tmp_path / "configs",
        "data_dir": tmp_path / "input",
        "output_dir": tmp_path / "roms",
        "report_dir": tmp_path / "reports",
        "capa_path": tmp_path / "input" / "capa.csv",
        "runs_dir": tmp_path / ".pyauditor" / "runs",
    }


def _scaffold_both(tmp_path: Path, *, minc_csv: str, mtur_csv: str) -> _Paths:
    for orgao, csv_body in (("MinC", minc_csv), ("MTur", mtur_csv)):
        (tmp_path / "configs" / orgao).mkdir(parents=True)
        (tmp_path / "input" / orgao / "2026" / "06").mkdir(parents=True)
        (tmp_path / "configs" / orgao / "inms-test.yaml").write_text(
            _CONFIG_YAML.replace("orgao: MinC", f"orgao: {orgao}"), encoding="utf-8"
        )
        (tmp_path / "configs" / orgao / "categorias.yaml").write_text(
            _CATEGORIAS_YAML, encoding="utf-8"
        )
        (tmp_path / "input" / orgao / "2026" / "06" / "data.csv").write_text(
            csv_body, encoding="utf-8"
        )
    return {
        "config_dir": tmp_path / "configs",
        "data_dir": tmp_path / "input",
        "output_dir": tmp_path / "roms",
        "report_dir": tmp_path / "reports",
        "capa_path": tmp_path / "input" / "capa.csv",
        "runs_dir": tmp_path / ".pyauditor" / "runs",
    }


def test_run_run_executes_pipeline_and_returns_4_for_unfilled_capa(tmp_path: Path) -> None:
    (tmp_path / "configs" / "MinC").mkdir(parents=True)
    (tmp_path / "input" / "MinC" / "2026" / "06").mkdir(parents=True)
    (tmp_path / "configs" / "MinC" / "inms-test.yaml").write_text(_CONFIG_YAML, encoding="utf-8")
    (tmp_path / "configs" / "MinC" / "categorias.yaml").write_text(
        _CATEGORIAS_YAML, encoding="utf-8"
    )
    (tmp_path / "input" / "MinC" / "2026" / "06" / "data.csv").write_text(
        "Nº Solicitacao;DataHoraFim;No prazo\n1;2026-06-01;S\n2;2026-06-02;N\n", encoding="utf-8"
    )

    exit_code = run_run(
        competencia="2026-06",
        orgao="MinC",
        config_dir=tmp_path / "configs",
        data_dir=tmp_path / "input",
        output_dir=tmp_path / "roms",
        report_dir=tmp_path / "reports",
        capa_path=tmp_path / "input" / "capa.csv",
        runs_dir=tmp_path / ".pyauditor" / "runs",
    )
    # Capa criada pelo bootstrap sem objetos.csv: relatório rascunho (3) + glosa
    # não calculada (4) → 4 vence pela precedência; nunca 0.
    assert exit_code == 4
    assert (tmp_path / "reports" / "relatorio_2026-06_MinC.xlsx").exists()


def test_run_run_default_resumes_and_does_not_rerun_done_steps(tmp_path: Path) -> None:
    # Ticket 08: `force=False` é o novo default — uma segunda chamada não
    # reprocessa etapas já `done`. Prova: corrompemos o CSV depois da 1ª
    # chamada; se `measure` fosse reexecutado, o resultado mudaria (a coluna
    # de decisão de erro difere entre boa e "hard failure" CSV).
    paths = _scaffold_single(tmp_path, csv_body=_GOOD_CSV)

    first = run_run(competencia="2026-06", orgao="MinC", **paths)
    assert first == 4  # capa em branco: rascunho + glosa não calculada

    (paths["data_dir"] / "MinC" / "2026" / "06" / "data.csv").write_text(
        _HARD_FAILURE_CSV, encoding="utf-8"
    )
    second = run_run(competencia="2026-06", orgao="MinC", **paths)

    assert second == 4  # inalterado — measure/report já "done", não reprocessados


def test_run_run_force_flag_reprocesses_done_steps(tmp_path: Path) -> None:
    # Ticket 08: `--force` restaura o reprocessamento total.
    paths = _scaffold_single(tmp_path, csv_body=_GOOD_CSV)

    first = run_run(competencia="2026-06", orgao="MinC", **paths)
    assert first == 4

    (paths["data_dir"] / "MinC" / "2026" / "06" / "data.csv").write_text(
        _HARD_FAILURE_CSV, encoding="utf-8"
    )
    second = run_run(competencia="2026-06", orgao="MinC", force=True, **paths)

    assert second == 1  # `measure` reexecutado, agora falha tecnicamente


def test_run_run_isolates_orgao_failure_both_orgaos(tmp_path: Path) -> None:
    # Ticket 08: falha técnica no MinC não impede o MTur de completar
    # (bootstrap→measure→report) — só o `consolidate` compartilhado é
    # pulado em cascata. O código final é `1` (falha vence a precedência),
    # mas o relatório do MTur existe.
    paths = _scaffold_both(tmp_path, minc_csv=_HARD_FAILURE_CSV, mtur_csv=_GOOD_CSV)

    exit_code = run_run(competencia="2026-06", orgao="both", **paths)

    assert exit_code == 1
    assert (paths["report_dir"] / "relatorio_2026-06_MTur.xlsx").exists()
    assert not (paths["report_dir"] / "relatorio_2026-06_consolidado.xlsx").exists()


def test_run_run_repeated_execution_is_idempotent(tmp_path: Path) -> None:
    # Cenário "execução repetida e idempotente" (review.md §Operação):
    # rodar `run` de novo sem mudar nada produz o mesmo resultado e o mesmo
    # conteúdo — resume (ticket 08) não deixa nada pela metade nem duplicado.
    # `report` é sempre redespachado no resume (force_commands, ticket 08),
    # então o `.xlsx` é regravado a cada chamada — comparar bytes crus é
    # frágil (openpyxl embute timestamp de criação/modificação nos metadados
    # do zip); comparamos o conteúdo relevante em vez disso.
    paths = _scaffold_single(tmp_path, csv_body=_GOOD_CSV)

    first = run_run(competencia="2026-06", orgao="MinC", **paths)
    report_path = paths["report_dir"] / "relatorio_2026-06_MinC.xlsx"
    first_sheets = load_workbook(report_path).sheetnames

    second = run_run(competencia="2026-06", orgao="MinC", **paths)

    assert second == first
    assert load_workbook(report_path).sheetnames == first_sheets
